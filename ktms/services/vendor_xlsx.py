"""Vendor RFQ 견적 요청 Excel 양식 생성 — openpyxl only (reportlab 의존성 없음)."""
from __future__ import annotations

import io
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _num(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _normalize_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = []
    for i, raw in enumerate(items, start=1):
        result.append({
            "item_no": raw.get("item_no") or i,
            "part_no": raw.get("part_no", ""),
            "description": raw.get("description", ""),
            "maker": raw.get("maker", ""),
            "qty": _num(raw.get("qty", 0)),
            "unit": raw.get("unit", "PCS"),
        })
    return result


def make_vendor_rfq_quote_xlsx(
    rfq_no: str,
    vessel_name: str,
    customer_name: str,
    enquiry_date: str,
    vendor_name: str,
    items: List[Dict[str, Any]],
    incoterms: str = "CNF Busan port",
    currency: str = "USD",
    reply_days: int = 5,
) -> bytes:
    """Vendor 견적 응답용 Excel 양식 생성. 공급사가 Unit Price, Lead Time 등을 입력해서 반환."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Quote Sheet"

    navy_fill      = PatternFill("solid", fgColor="0B1D3A")
    blue_fill      = PatternFill("solid", fgColor="0055A8")
    light_blue_fill= PatternFill("solid", fgColor="EAF3FF")
    light_gray_fill= PatternFill("solid", fgColor="F4F6F8")
    input_fill     = PatternFill("solid", fgColor="FFF9C4")
    alt_row_fill   = PatternFill("solid", fgColor="FAFBFC")

    bold           = Font(name="Calibri", bold=True)
    white_bold_lg  = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
    white_sm       = Font(name="Calibri", color="FFFFFF", size=9)
    white_bold_sm  = Font(name="Calibri", color="FFFFFF", bold=True, size=9)
    note_font      = Font(name="Calibri", bold=True, size=9, color="7B5E00")
    footer_font    = Font(name="Calibri", size=8, italic=True)

    thin = Side(style="thin", color="D8DEE6")
    bdr  = Border(top=thin, bottom=thin, left=thin, right=thin)
    center_mid = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    left_mid   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    NUM_COLS = 11

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    merge(1, 1, 1, NUM_COLS)
    c = ws.cell(1, 1, "VENDOR QUOTATION REQUEST SHEET")
    c.fill = navy_fill; c.font = white_bold_lg; c.alignment = center_mid
    ws.row_dimensions[1].height = 28

    # ── Row 2: Sub-header ─────────────────────────────────────────────────────
    merge(2, 1, 2, NUM_COLS)
    c = ws.cell(2, 1, "K-MARIS Energy & Solutions Co., Ltd.  |  sales@k-maris.com  |  www.k-maris.com")
    c.fill = blue_fill; c.font = white_sm; c.alignment = center_mid
    ws.row_dimensions[2].height = 16

    # ── Rows 4-7: Meta info ───────────────────────────────────────────────────
    meta = [
        ("RFQ Reference", rfq_no,       "Enquiry Date", enquiry_date),
        ("Vessel",        vessel_name,   "End Customer", customer_name),
        ("To (Vendor)",   vendor_name,   "Incoterms",    incoterms),
        ("Currency",      currency,      "Reply Within", f"{reply_days} business days"),
    ]
    for offset, (k1, v1, k2, v2) in enumerate(meta, start=4):
        for col, val, is_label in [(1, k1, True), (2, v1, False), (6, k2, True), (7, v2, False)]:
            c = ws.cell(offset, col, val)
            c.border = bdr; c.alignment = left_mid
            if is_label:
                c.fill = light_gray_fill; c.font = bold
        merge(offset, 2, offset, 5)
        merge(offset, 7, offset, NUM_COLS)
        for col in range(2, 6):
            ws.cell(offset, col).border = bdr
        for col in range(7, NUM_COLS + 1):
            ws.cell(offset, col).border = bdr
        ws.row_dimensions[offset].height = 16

    # ── Row 9: Instruction note ───────────────────────────────────────────────
    merge(9, 1, 9, NUM_COLS)
    c = ws.cell(9, 1, (
        "▶  Please fill in the yellow-highlighted cells "
        "(Unit Price, Lead Time, Country of Origin, Manufacturer, Remarks) "
        "and return this sheet to sales@k-maris.com."
    ))
    c.fill = input_fill; c.font = note_font
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = bdr
    ws.row_dimensions[9].height = 20

    # ── Row 11: Item table header ─────────────────────────────────────────────
    HEADERS = [
        "No.", "Part No.", "Description", "Maker\n(Enquiry)", "Qty", "Unit",
        "Unit Price\n(USD)", "Lead Time", "Country\nof Origin", "Manufacturer",
        "Technical Remarks\n/ Alternatives",
    ]
    COL_WIDTHS = [5, 20, 36, 18, 7, 7, 14, 14, 14, 20, 36]
    INPUT_COLS = {7, 8, 9, 10, 11}

    for c_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(11, c_idx, h)
        cell.fill = navy_fill; cell.font = white_bold_sm
        cell.alignment = center_mid; cell.border = bdr
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[11].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    normalized = _normalize_items(items)
    for r_offset, item in enumerate(normalized, start=1):
        r = 11 + r_offset
        row_values = [
            item["item_no"], item["part_no"], item["description"],
            item["maker"], item["qty"], item["unit"],
            "", "", "", "", "",
        ]
        is_alt = r_offset % 2 == 0
        for c_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(r, c_idx, val)
            cell.border = bdr
            if c_idx in INPUT_COLS:
                cell.fill = input_fill; cell.alignment = left_top
            else:
                if is_alt:
                    cell.fill = alt_row_fill
                align_h = "center" if c_idx in {1, 5, 6} else "left"
                cell.alignment = Alignment(horizontal=align_h, vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 20

    # ── Footer ────────────────────────────────────────────────────────────────
    foot = 11 + len(normalized) + 2
    merge(foot, 1, foot, NUM_COLS)
    c = ws.cell(foot, 1, (
        f"Kindly return this completed sheet within {reply_days} business days to: sales@k-maris.com  "
        "|  K-MARIS Energy & Solutions Co., Ltd.  |  Engineering Reliability. Supplying Performance."
    ))
    c.fill = light_blue_fill; c.font = footer_font
    c.alignment = center_mid; c.border = bdr
    ws.row_dimensions[foot].height = 18

    ws.freeze_panes = "A12"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
