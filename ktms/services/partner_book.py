"""거래선 명부(Customer · Vendor · Maker) 인쇄물 — 화면에 보이는 목록 그대로 Excel·PDF.

이 모듈은 '어떤 명부인가'를 모른다. 화면이 보고 있는 표(제목·칸 이름·줄 값)를 그대로
받아 종이로 옮길 뿐이다. 그래야 검색어나 열 필터로 좁혀 놓은 목록이 미리보기와
파일에서 갈라지지 않는다 — 서버가 DB 를 다시 읽어 그리면, 화면에서 걸러 낸 줄이
파일에는 그대로 살아 나와 "내가 본 것과 다른 종이"가 나온다.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Spacer, Table, TableStyle

from services.doc_xlsx import _apply_noto_font  # type: ignore
from services.kmaris_docs import (  # type: ignore
    LIGHT_GRAY, MID_GRAY, NAVY, _footer, _letterhead, _p, _styles,
)

MAX_ROWS = 5000
MAX_COLS = 20
MAX_CELL = 2000

# 인쇄물 폭(가로 A4) — _footer 가 긋는 밑줄(12mm~285mm)과 같은 자리에 맞춘다.
PAGE_LEFT = 12 * mm
PAGE_RIGHT = 285 * mm
BODY_W = PAGE_RIGHT - PAGE_LEFT


def _cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v)[:MAX_CELL]


def normalize(book: Dict[str, Any]) -> Dict[str, Any]:
    """화면이 보낸 표를 인쇄 가능한 모양으로 다듬는다(칸 수 맞추기·상한 자르기)."""
    cols = list(book.get("columns") or [])[:MAX_COLS]
    if not cols:
        cols = [{"label": ""}]
    columns = []
    for c in cols:
        d = c if isinstance(c, dict) else {"label": c}
        width = d.get("width")
        align = d.get("align") or "left"
        columns.append({
            "label": _cell(d.get("label")),
            "width": float(width) if width else 0.0,
            "align": align if align in ("left", "center", "right") else "left",
        })
    rows: List[List[str]] = []
    for r in list(book.get("rows") or [])[:MAX_ROWS]:
        vals = [_cell(v) for v in (r or [])][: len(columns)]
        vals += [""] * (len(columns) - len(vals))
        rows.append(vals)
    return {
        "title": _cell(book.get("title")) or "List",
        "subtitle": _cell(book.get("subtitle")),
        "columns": columns,
        "rows": rows,
    }


def filename(title: str, ext: str) -> str:
    stem = "".join(ch if ch.isalnum() or ch in " -_" else "_" for ch in (title or "List"))
    stamp = date.today().strftime("%Y%m%d")
    return "{}_List_{}.{}".format(stem.strip().replace(" ", "_") or "List", stamp, ext)


# ── Excel ────────────────────────────────────────────────────────────────────
def make_book_xlsx(book: Dict[str, Any]) -> bytes:
    b = normalize(book)
    cols, rows = b["columns"], b["rows"]
    last = get_column_letter(len(cols))
    today = date.today().strftime("%Y-%m-%d")

    wb = Workbook()
    ws = wb.active
    ws.title = (b["title"] or "List")[:31]

    thin = Side(style="thin", color="D8DEE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:%s1" % last)
    ws["A1"] = b["title"]
    ws["A1"].font = Font(name="Noto Sans KR", bold=True, size=16, color="0B1D3A")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:%s2" % last)
    ws["A2"] = " · ".join(x for x in (b["subtitle"], "%d rows" % len(rows), today) if x)
    ws["A2"].font = Font(name="Noto Sans KR", size=9, color="606A76")
    ws.row_dimensions[2].height = 15

    head_row = 4
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=head_row, column=i, value=c["label"])
        cell.font = Font(name="Noto Sans KR", bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B1D3A")
        cell.alignment = Alignment(horizontal=c["align"], vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[head_row].height = 20

    for ri, row in enumerate(rows):
        for ci, v in enumerate(row, start=1):
            cell = ws.cell(row=head_row + 1 + ri, column=ci, value=v)
            cell.font = Font(name="Noto Sans KR", size=9)
            cell.alignment = Alignment(horizontal=cols[ci - 1]["align"], vertical="top",
                                       wrap_text=True)
            cell.border = border
            if ri % 2:
                cell.fill = PatternFill("solid", fgColor="F4F6F8")

    # 칸 폭 — 실제 글자 길이를 보되 지나치게 벌어지지 않게 9~52자 사이로 가둔다.
    for i, c in enumerate(cols, start=1):
        longest = max([len(c["label"])] + [len(r[i - 1]) for r in rows])
        ws.column_dimensions[get_column_letter(i)].width = min(52, max(9, longest + 2))

    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    if rows:
        ws.auto_filter.ref = "A%d:%s%d" % (head_row, last, head_row + len(rows))
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "%d:%d" % (head_row, head_row)

    _apply_noto_font(wb)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────
def _company() -> Dict[str, Any]:
    try:
        from services.company_profile import read_company_profile  # type: ignore
        return read_company_profile() or {}
    except Exception:
        return {}


def _col_widths(cols: List[Dict[str, Any]], rows: List[List[str]]) -> List[float]:
    """칸 폭 — 화면이 준 가중치(width)를 쓰되, 안 주면 그 칸의 실제 글자 길이로 나눈다.

    긴 주소 한 줄 때문에 다른 칸이 실오라기처럼 눌리지 않도록 길이는 제곱근으로 눌러
    반영하고, 어떤 칸도 본문 폭의 4% 아래로는 내려가지 않게 바닥을 깔아 준다."""
    given_any = any(c["width"] > 0 for c in cols)
    weights: List[float] = []
    for i, c in enumerate(cols):
        if given_any:
            weights.append(c["width"] if c["width"] > 0 else 1.0)
            continue
        lens = [len(c["label"])] + [len(r[i]) for r in rows]
        avg = sum(lens) / len(lens)
        weights.append(max(1.0, max(lens) * 0.35 + avg * 0.65) ** 0.5)
    total = sum(weights) or 1.0
    floor = BODY_W * 0.04
    widths = [max(floor, BODY_W * w / total) for w in weights]
    # 바닥을 깔면 합이 본문 폭을 넘을 수 있다 — 넘친 만큼 넉넉한 칸에서 되돌려 받는다.
    over = sum(widths) - BODY_W
    if over > 0:
        spare = [max(0.0, w - floor) for w in widths]
        pool = sum(spare)
        if pool > 0:
            widths = [w - over * sp / pool for w, sp in zip(widths, spare)]
        else:
            widths = [BODY_W / len(widths)] * len(widths)
    return widths


def make_book_pdf(book: Dict[str, Any]) -> bytes:
    b = normalize(book)
    cols, rows = b["columns"], b["rows"]
    s = _styles()
    body = {"left": s["base"], "center": s["center"], "right": s["right"]}

    buf = io.BytesIO()
    page = landscape(A4)
    doc = BaseDocTemplate(
        buf, pagesize=page,
        leftMargin=PAGE_LEFT, rightMargin=page[0] - PAGE_RIGHT,
        topMargin=12 * mm, bottomMargin=16 * mm,
        title="%s List" % b["title"],
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, BODY_W,
                  page[1] - doc.topMargin - doc.bottomMargin, id="body",
                  leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    doc.addPageTemplates([PageTemplate(id="book", frames=[frame], onPage=_footer)])

    story: List[Any] = list(_letterhead(_company(), "%s LIST" % b["title"].upper(), s, width=BODY_W))
    caption = " · ".join(x for x in (b["subtitle"], "%d rows" % len(rows),
                                     "as of " + date.today().strftime("%Y-%m-%d")) if x)
    story.append(_p(caption, s["small"]))
    story.append(Spacer(1, 3 * mm))

    data = [[_p(c["label"], s["th"]) for c in cols]]
    for r in rows:
        data.append([_p(v, body[cols[i]["align"]]) for i, v in enumerate(r)])

    table = Table(data, colWidths=_col_widths(cols, rows), repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, MID_GRAY),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
    ]))
    story.append(table)
    if not rows:
        story.append(Spacer(1, 4 * mm))
        story.append(_p("No rows.", s["small"]))

    doc.build(story)
    return buf.getvalue()
