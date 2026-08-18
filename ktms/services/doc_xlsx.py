"""견적서·발주서 Excel 생성 — make_pdf 와 동일한 payload 를 소비한다. openpyxl only.

내장 PDF(kmaris_docs.make_pdf)와 같은 데이터(payload)를 받아 Excel 로 렌더링해,
사용자가 문서를 PDF/Excel 중 선택해 내려받거나 이메일에 첨부할 수 있게 한다.
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from services.kmaris_docs import (
    normalize_items, calc_totals, _num, DOC_TITLES, doc_parties,
    packing_totals, consignee_mark_lines, _dim_parts, _invoice_shipping_rows,
    PI_COLUMN_UNITS, PI_MIN_ITEM_ROWS, PL_COLUMN_UNITS, PL_MIN_ITEM_ROWS,
    pi_decimals, pi_charges, pi_doc_date,
)

_CONFIG_DIR = Path(__file__).resolve().parent.parent / "config"
_TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"
_REPO_DIR = Path(__file__).resolve().parents[2]


def _find_asset(*names: str) -> Optional[str]:
    """자산 이미지(로고·서명·직인)를 config/ · templates/ · 저장소 루트에서 찾는다.
    templates/ 는 git 에 커밋되므로 커밋된 아이콘 로고(logo_icon.jpg 등)를 여기 두면 배포에 반영된다.
    이름 우선순위가 폴더보다 우선 — 앞선 이름(아이콘)이 있으면 배포본 텍스트 로고보다 먼저 선택된다."""
    for n in names:
        for root in (_TEMPLATES_DIR, _CONFIG_DIR, _REPO_DIR):
            p = root / n
            if p.exists():
                return str(p)
    return None


def _apply_noto_font(wb: "Workbook") -> None:
    """워크북 전체 글꼴을 Noto Sans KR 로 통일한다(모든 문서 Excel 공통).

    핵심 주의점(견적 Excel 에서 크게 데인 부분):
    1) 이름만 바꾸면 안 된다 — scheme("minor"/"major")·family 도 지워야 한다.
       scheme 이 남으면 Excel 은 글꼴 이름을 무시하고 '테마 글꼴'을 쓰는데,
       한글판 Excel 테마 기본 글꼴이 맑은 고딕이라 화면에 계속 맑은고딕으로 뜬다.
    2) 병합 셀 비앵커 칸/빈 칸은 저장 시 개별 글꼴이 버려지고 기본 글꼴(fontId 0)
       로 되돌아간다 → 기본 글꼴 레코드 자체를 Noto 로 교체.
    3) 테마 major/minor 글꼴도 Noto 로 교체(혹시 남은 scheme 참조의 폴백까지 차단).
    """
    from copy import copy
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                f0 = cell.font
                if f0 is not None and (f0.name != "Noto Sans KR" or f0.scheme is not None):
                    f = copy(f0)
                    f.name = "Noto Sans KR"
                    f.scheme = None
                    f.family = None
                    cell.font = f
    try:
        wb._fonts[0] = Font(name="Noto Sans KR", size=11, scheme=None, family=None)
    except Exception:
        pass
    try:
        th = wb.loaded_theme
        if th:
            th = th.decode("utf-8") if isinstance(th, bytes) else th
            import re as _re
            def _fix(block):
                block = _re.sub(r'<a:latin typeface="[^"]*"', '<a:latin typeface="Noto Sans KR"', block)
                block = _re.sub(r'<a:ea typeface="[^"]*"', '<a:ea typeface="Noto Sans KR"', block)
                return block
            for tag in ("majorFont", "minorFont"):
                th = _re.sub(r'(<a:%s>)(.*?)(</a:%s>)' % (tag, tag),
                             lambda m: m.group(1) + _fix(m.group(2)) + m.group(3), th, flags=_re.S)
            wb.loaded_theme = th.encode("utf-8")
    except Exception:
        pass


def _compose_marks(sh: Dict[str, Any]) -> str:
    """구조화 Shipping Mark(sm_*)를 여러 줄 문자열로 합성 — 프론트 composeShippingMarks 와
    동일 규약(무게·치수 포함). 저장된 shipping_marks 문자열이 없거나 비어도 항상 재구성한다."""
    lines = []
    def push(v):
        if v and str(v).strip():
            lines.append(str(v).strip())
    push(sh.get("sm_type"))
    for line in consignee_mark_lines(sh):
        push(line)
    if sh.get("sm_vessel"): push(f"M/V {str(sh['sm_vessel']).upper()}")
    if sh.get("sm_po_no"): push(f"P.O. NO.: {sh['sm_po_no']}")
    if sh.get("sm_ref_no"): push(f"REF. NO.: {sh['sm_ref_no']}")
    push(sh.get("sm_desc"))
    if sh.get("sm_case_no"): push(f"CASE NO.: {sh['sm_case_no']}")
    if sh.get("sm_total_cases"): push(f"TOTAL: {sh['sm_total_cases']} CASE(S)")
    if sh.get("sm_net_weight"): push(f"N.W.: {sh['sm_net_weight']} KG")
    if sh.get("sm_gross_weight"): push(f"G.W.: {sh['sm_gross_weight']} KG")
    dim = [sh.get("sm_dim_l"), sh.get("sm_dim_w"), sh.get("sm_dim_h")]
    if any(d and str(d).strip() for d in dim):
        push("DIM.: " + " × ".join((str(d).strip() if d and str(d).strip() else "-") for d in dim) + " CM")
    if sh.get("sm_final_dest"): push(f"FINAL DESTINATION: {sh['sm_final_dest']}")
    push(sh.get("sm_origin"))
    push(sh.get("sm_handling"))
    return "\n".join(lines)


class _FormSheet:
    """송장 계열 Excel(PI · CI · PL) 공통 서식 — 레터헤드·머리띠·라벨표·서명 블록.

    같은 문서의 PDF(kmaris_docs._DocForm)와 같은 열 너비를 쓰기 때문에 두 파일의 칸이
    정확히 겹친다. blocks 는 라벨·값 네 칸의 (시작열, 끝열) 이라, 8열 문서(PI/CI)와
    13열 문서(PL)를 같은 코드로 그린다.
    """

    NAVY = "0B1D3A"
    BAND = "1F3B66"
    GRAY = "F4F6F8"
    LIGHTBLUE = "EAF3FF"
    ALT = "FAFBFC"
    BORDER = "C8D2E0"

    def __init__(self, ws, widths, blocks):
        self.ws = ws
        self.widths = widths
        self.ncol = len(widths)
        self.blocks = blocks
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.sheet_view.showGridLines = False

        self.navy = PatternFill("solid", fgColor=self.NAVY)
        self.band_fill = PatternFill("solid", fgColor=self.BAND)
        self.gray = PatternFill("solid", fgColor=self.GRAY)
        self.lightblue = PatternFill("solid", fgColor=self.LIGHTBLUE)
        self.alt = PatternFill("solid", fgColor=self.ALT)

        self.title_font = Font(name="Noto Sans KR", color=self.NAVY, bold=True, size=19)
        self.company_font = Font(name="Noto Sans KR", size=14)
        self.small = Font(name="Noto Sans KR", size=9)
        self.white_sec = Font(name="Noto Sans KR", color="FFFFFF", bold=True, size=10)
        self.white_hdr = Font(name="Noto Sans KR", color="FFFFFF", bold=True, size=9)
        self.boldsm = Font(name="Noto Sans KR", bold=True, size=9)
        self.bold = Font(name="Noto Sans KR", bold=True, size=11)
        self.normal = Font(name="Noto Sans KR", size=9)
        self.item_font = Font(name="Noto Sans KR", size=11)

        self.thin = Side(style="thin", color=self.BORDER)
        self.bdr = Border(top=self.thin, bottom=self.thin, left=self.thin, right=self.thin)
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.right = Alignment(horizontal="right", vertical="center")
        self.left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── 저수준 도우미 ─────────────────────────────────────────────────────
    def merge(self, r1, c1, r2, c2):
        if (r1, c1) != (r2, c2):
            self.ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def bd(self, r1, c1, r2, c2, fill=None):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                self.ws.cell(rr, cc).border = self.bdr
                if fill:
                    self.ws.cell(rr, cc).fill = fill

    def put(self, r, c, v="", *, fill=None, font=None, align=None, fmt=None):
        x = self.ws.cell(r, c, v)
        if fill:
            x.fill = fill
        if font:
            x.font = font
        if align:
            x.alignment = align
        if fmt:
            x.number_format = fmt
        return x

    def cell(self, r, block, value, *, font=None, fill=None, align=None, fmt=None):
        """네 칸(라벨1·값1·라벨2·값2) 중 하나를 채운다."""
        c1, c2 = self.blocks[block]
        self.merge(r, c1, r, c2)
        return self.put(r, c1, value, font=font, fill=fill, align=align, fmt=fmt)

    def add_image(self, path, anchor, w, h):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(path)
            img.width, img.height = w, h
            self.ws.add_image(img, anchor)
            return True
        except Exception:
            return False

    # ── 블록 ──────────────────────────────────────────────────────────────
    def letterhead(self, company, title):
        """1~6행 — 로고 + 가운데 회사명·주소·연락처 + 문서 제목(PDF 레터헤드와 같은 내용)."""
        logo = _find_asset("logo_icon.jpg", "logo_icon.png", "logo_K-maris.png", "logo.png", "logo.jpg")
        if logo:
            self.add_image(logo, "A1", 96, 45)
        self.cell(1, "full", company.get("company_name_en", "K-MARIS Energy & Solutions Co., Ltd."),
                  font=self.company_font, align=self.center)
        self.cell(2, "full", company.get("address_en") or company.get("address", ""),
                  font=self.small, align=self.center)
        contact = " | ".join(x for x in [
            f"Tel: {company.get('phone', '')}" if company.get("phone") else "",
            company.get("sales_email", ""), company.get("website", ""),
        ] if x)
        self.cell(3, "full", contact, font=self.small, align=self.center)
        self.cell(5, "full", title, font=self.title_font, align=self.center)
        for row, height in ((1, 37.2), (2, 14.1), (3, 14.1), (4, 15.6), (5, 30), (6, 13.8)):
            self.ws.row_dimensions[row].height = height
        return 7

    def doc_info(self, r, rows):
        """문서 정보(송장번호·일자·PO No.) — 참조 양식처럼 오른쪽 반쪽에만 그린다."""
        from datetime import date as _date
        for k, v in rows:
            self.cell(r, "label2", k, fill=self.gray, font=self.boldsm, align=self.left)
            c = self.cell(r, "value2", v, font=self.normal, align=self.left)
            if isinstance(v, _date):
                c.number_format = "dd-mmm-yyyy"
            self.bd(r, self.blocks["label2"][0], r, self.ncol)
            r += 1
        return r

    def band(self, r, *titles):
        """섹션 머리띠 — 제목 하나면 전폭, 둘이면 좌우로 나눈다(CONSIGNEE | BUYER)."""
        if len(titles) == 1:
            self.cell(r, "full", f" {titles[0]}", fill=self.band_fill, font=self.white_sec, align=self.left)
            self.bd(r, 1, r, self.ncol, self.band_fill)
        else:
            split = self.blocks["value1"][1]
            self.merge(r, 1, r, split)
            self.put(r, 1, f" {titles[0]}", fill=self.band_fill, font=self.white_sec, align=self.left)
            self.merge(r, split + 1, r, self.ncol)
            self.put(r, split + 1, f" {titles[1]}", fill=self.band_fill, font=self.white_sec, align=self.left)
            self.bd(r, 1, r, self.ncol, self.band_fill)
        self.ws.row_dimensions[r].height = 17.4
        return r + 1

    def pair_row(self, r, k1, v1, k2, v2, *, value_align=None, height=17.4):
        """'라벨 | 값 | 라벨 | 값' 한 행."""
        from datetime import date as _date
        va = value_align or self.left
        for block, value, is_label in (("label1", k1, True), ("value1", v1, False),
                                       ("label2", k2, True), ("value2", v2, False)):
            c = self.cell(r, block, value,
                          fill=self.gray if is_label else None,
                          font=self.boldsm if is_label else self.normal,
                          align=self.left if is_label else va)
            if isinstance(value, _date):
                c.number_format = "dd-mmm-yyyy"
        self.bd(r, 1, r, self.ncol)
        self.ws.row_dimensions[r].height = height
        return r + 1

    def pairs(self, r, rows, *, value_align=None, heights=None):
        for i, (k1, v1, k2, v2) in enumerate(rows):
            r = self.pair_row(r, k1, v1, k2, v2, value_align=value_align,
                              height=(heights or {}).get(i, 17.4))
        return r

    def item_header(self, r, headers):
        """품목표 머리행 — headers 는 (제목, 시작열, 끝열) 목록."""
        for title, c1, c2 in headers:
            self.merge(r, c1, r, c2)
            self.put(r, c1, title, fill=self.navy, font=self.white_hdr, align=self.center)
        self.bd(r, 1, r, self.ncol, self.navy)
        self.ws.row_dimensions[r].height = 24
        return r + 1

    def item_row_height(self, description, chars_per_line):
        """병합 셀은 Excel 이 자동 높이를 못 잡는다 — 품명 길이로 줄 수를 어림한다."""
        text = str(description or "")
        lines = max(len(text.split("\n")), -(-len(text) // chars_per_line), 1)
        return 14.5 * lines + 8

    def declaration(self, r, text):
        self.cell(r, "full", text, font=self.normal, align=self.left)
        for cc in range(1, self.ncol + 1):
            self.ws.cell(r, cc).border = Border(top=self.thin, bottom=self.thin)
        return r + 1

    def signature(self, r, cells):
        """서명 칸 — (시작열, 끝열, 라벨, 이미지파일, 폭, 높이) 목록. 3행을 차지한다."""
        for row, height in ((r, 28.05), (r + 1, 20.1), (r + 2, 21)):
            self.ws.row_dimensions[row].height = height
        for c1, c2, text, image, w, h in cells:
            self.merge(r, c1, r + 2, c2)
            self.put(r, c1, text, font=self.boldsm, align=self.left_top)
            self.bd(r, c1, r + 2, c2)
            if image:
                self.add_image(image, f"{get_column_letter(c1 + 1)}{r + 1}", w, h)
        return r + 3

    def page_setup(self, last_row, *, scale=None):
        ws = self.ws
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        if scale:
            ws.page_setup.scale = scale
        else:
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_area = f"A1:{get_column_letter(self.ncol)}{last_row}"
        ws.sheet_view.zoomScale = 85
        ws.page_margins.left = 0.24
        ws.page_margins.right = 0.24
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.32


# 라벨·값 네 칸의 열 위치 — 8열 문서(PI/CI)와 13열 문서(PL).
_BLOCKS_8 = {"label1": (1, 2), "value1": (3, 4), "label2": (5, 6), "value2": (7, 8), "full": (1, 8)}
_BLOCKS_13 = {"label1": (1, 2), "value1": (3, 5), "label2": (6, 7), "value2": (8, 13), "full": (1, 13)}

_SIGN_ASSET = ("Authorized signature_Sungyeon Cho.jpg", "signature.png", "signature.jpg", "sign.png")
_STAMP_ASSET = ("Company stamp_K-Maris Energy & Solutions.jpg", "stamp.png", "stamp.jpg", "seal.png")


def _invoice_money_format(currency: str) -> str:
    """금액 서식 — 0 은 '-' 로 찍히는 회계 서식(참조 양식과 같다)."""
    if pi_decimals(currency):
        return '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"_-;_-@_-'
    return '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'


def _invoice_doc_date(value: Any):
    """송장 일자 — 엑셀에서 날짜로 다룰 수 있게 date 로, 아니면 원문 문자열."""
    from datetime import date as _date
    try:
        return _date.fromisoformat(str(value or "")[:10])
    except ValueError:
        return pi_doc_date(value)


def _invoice_sheet(data: Dict[str, Any], company: Dict[str, Any], title: str, sheet_name: str):
    """PI·CI 공통 뼈대(레터헤드 → 문서정보) — 두 문서의 위쪽 절반은 완전히 같다."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    form = _FormSheet(ws, PI_COLUMN_UNITS, _BLOCKS_8)
    shipping = data.get("shipping", {}) or {}
    r = form.letterhead(company, title)
    r = form.doc_info(r, [("Invoice No.", data.get("doc_no", "")),
                          ("Invoice Date", _invoice_doc_date(data.get("date", ""))),
                          ("PO No.", shipping.get("po_no", ""))])
    return wb, form, r + 1


def _invoice_items_and_totals(form: "_FormSheet", data: Dict[str, Any], r: int, currency: str):
    """PI·CI 공통 품목표 + 합계 — 금액은 수식이라 받는 쪽에서 고쳐도 합계가 따라온다."""
    items = normalize_items(data.get("items", []))
    shipping = data.get("shipping", {}) or {}
    money = pi_charges(data)
    num_fmt = _invoice_money_format(currency)

    r = form.item_header(r, [("No.", 1, 1), ("Description", 2, 3), ("Part No.", 4, 4), ("HS Code", 5, 5),
                             ("Qty", 6, 6), ("Unit Price", 7, 7), (f"Amount ({currency})", 8, 8)])
    first = r
    for i in range(max(len(items), PI_MIN_ITEM_ROWS)):
        it = items[i] if i < len(items) else None
        form.put(r, 1, it["item_no"] if it else "", font=form.item_font, align=form.center)
        form.merge(r, 2, r, 3)
        form.put(r, 2, it["description"] if it else "", font=form.item_font, align=form.left)
        form.put(r, 4, it["part_no"] if it else "", font=form.item_font, align=form.left)
        form.put(r, 5, (it.get("hs_code") or shipping.get("hs_code", "")) if it else "",
                 font=form.item_font, align=form.center)
        form.put(r, 6, _num(it["qty"]) if it else "", font=form.item_font, align=form.center)
        form.put(r, 7, _num(it["unit_price"]) if it else "", font=form.item_font, align=form.right, fmt=num_fmt)
        form.put(r, 8, f'=IF(F{r}="","",F{r}*G{r})', font=form.item_font, align=form.right, fmt=num_fmt)
        form.bd(r, 1, r, form.ncol, None if it else form.alt)
        form.ws.row_dimensions[r].height = form.item_row_height(it["description"], 30) if it else 18
        r += 1
    last = r - 1

    def total_line(label, value):
        nonlocal r
        form.merge(r, 6, r, 7)
        form.put(r, 6, label, fill=form.gray, font=form.boldsm, align=form.right)
        form.put(r, 8, value, font=form.item_font, align=form.right, fmt=num_fmt)
        form.bd(r, 6, r, form.ncol)
        form.ws.row_dimensions[r].height = 18
        ref = f"H{r}"
        r += 1
        return ref

    sref = total_line("Subtotal", f"=SUM(H{first}:H{last})")
    fref = total_line("Freight", money["freight"])
    pref = total_line("Packing", money["packing"])
    iref = total_line("Insurance", money["insurance"])
    # VAT 는 부대비용까지 더한 금액에 매긴다(입력 화면·PDF 와 같은 계산).
    vref = total_line("VAT", f"=({sref}+{fref}+{pref}+{iref})*{money['vat_rate']}")
    form.merge(r, 1, r, 7)
    form.put(r, 1, "TOTAL INVOICE VALUE", fill=form.lightblue, font=form.bold, align=form.center)
    form.put(r, 8, f"={sref}+{fref}+{pref}+{iref}+{vref}", fill=form.lightblue, font=form.bold,
             align=form.right, fmt=num_fmt)
    form.bd(r, 1, r, form.ncol)
    form.ws.row_dimensions[r].height = 18
    return r + 1


def _invoice_closing(form: "_FormSheet", company: Dict[str, Any], r: int, doc_label: str,
                     extra_cell: bool = False) -> int:
    """DECLARATION + 서명·직인 — 세 문서가 같이 쓴다."""
    r += 1
    r = form.band(r, "DECLARATION")
    r = form.declaration(r, f"We hereby certify that this {doc_label} is true and correct.")
    r += 1
    sign = _find_asset(*_SIGN_ASSET)
    stamp = _find_asset(*_STAMP_ASSET)
    company_name = f"{company.get('company_name_en', '')}\n(Company Stamp)"
    if extra_cell:   # Packing List — 수령 확인 칸이 하나 더 붙는다.
        cells = [(1, 3, "Authorized Signature", sign, 130, 44),
                 (4, 7, company_name, stamp, 66, 66),
                 (8, 13, "Received by\n(Company Stamp & Date)", None, 0, 0)]
    else:
        cells = [(1, 4, "Authorized Signature", sign, 130, 44),
                 (5, 8, company_name, stamp, 66, 66)]
    return form.signature(r, cells) - 1


def make_commercial_invoice_xlsx(
    data: Dict[str, Any], company: Optional[Dict[str, Any]] = None
) -> bytes:
    """Commercial Invoice 전용 Excel — templates/'commercial invoice_sample.xlsx' 서식 그대로.

    구성: 레터헤드 → 문서정보 → CONSIGNEE | BUYER(if different) → SHIPPING INFORMATION →
    품목표(금액=수식) → 합계 → DECLARATION → 서명. 같은 payload 로 만드는 PDF 와 칸이 겹친다.
    """
    company = company or {}
    currency = (data.get("currency") or "USD").upper()
    shipping = data.get("shipping", {}) or {}
    terms = data.get("terms", {}) or {}
    consignee, buyer = doc_parties(data)

    wb, form, r = _invoice_sheet(data, company, "COMMERCIAL INVOICE", "Commercial Invoice")
    r = form.band(r, "CONSIGNEE", "BUYER (if different)")
    r = form.pairs(r, [("Company Name", consignee.get("name", ""), "Company Name", buyer.get("name", "")),
                       ("Address", consignee.get("address", ""), "Address", buyer.get("address", "")),
                       ("Contact", consignee.get("contact", ""), "Contact", buyer.get("contact", "")),
                       ("e-mail", consignee.get("email", ""), "e-mail", buyer.get("email", ""))],
                   heights={1: 34.2})
    r = form.band(r, "SHIPPING INFORMATION")
    r = form.pairs(r, _invoice_shipping_rows(data, "Incoterms® 2020") + [
        ("Packing", terms.get("packing_type", ""), "Country of Origin", shipping.get("sm_origin", "")),
        ("Currency", currency, "", ""),
    ])
    r += 1
    r = _invoice_items_and_totals(form, data, r, currency)
    last = _invoice_closing(form, company, r, "Commercial Invoice")

    form.page_setup(last)
    _apply_noto_font(wb)   # 전체 글꼴 Noto Sans KR 통일(모든 문서 Excel 공통)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def make_proforma_invoice_xlsx(
    data: Dict[str, Any], company: Optional[Dict[str, Any]] = None
) -> bytes:
    """Proforma Invoice 전용 Excel — templates/'proforma invoice_sample.xlsx' 서식 그대로.

    Commercial Invoice 와 같은 뼈대에, 당사자는 매수인 한 칸이고 합계 아래 BANK
    INFORMATION 이 붙는다는 점만 다르다. 금액 칸은 수식(수량×단가·SUM)이라 받는 쪽에서
    숫자를 고치면 합계가 따라 움직인다.
    """
    company = company or {}
    currency = (data.get("currency") or "USD").upper()
    shipping = data.get("shipping", {}) or {}
    _, buyer = doc_parties(data)

    wb, form, r = _invoice_sheet(data, company, "PROFORMA INVOICE", "Proforma Invoice")
    r = form.band(r, "BUYER")
    r = form.pairs(r, [("Company Name", buyer.get("name", ""), "Address", buyer.get("address", "")),
                       ("Contact", buyer.get("contact", ""), "e-mail", buyer.get("email", ""))],
                   heights={0: 30.6, 1: 18.6})
    r += 1
    r = form.band(r, "SHIPPING INFORMATION")
    r = form.pairs(r, _invoice_shipping_rows(data) + [
        ("Currency", currency, "Country of Origin", shipping.get("sm_origin", "")),
    ])
    r += 1
    r = _invoice_items_and_totals(form, data, r, currency)

    # ── BANK INFORMATION — 통화에 맞는 계좌(외화/원화)를 Settings 에서 가져온다 ──
    r += 1
    r = form.band(r, "BANK INFORMATION")
    foreign = currency != "KRW"
    holder = (company.get("fx_bank_holder") if foreign else company.get("bank_holder")) or company.get("company_name_en", "")
    bank_name = (company.get("fx_bank_name") if foreign else company.get("bank_name")) or ""
    account = (company.get("fx_bank_account") if foreign else company.get("bank_account")) or ""
    r = form.pairs(r, [("Remittee's name", holder, "Bank Name & Address", bank_name),
                       ("Swift Code", company.get("swift", ""), "Remittee's Account No.", account)],
                   value_align=form.center)

    last = _invoice_closing(form, company, r - 1, "Proforma Invoice")
    form.page_setup(last)
    _apply_noto_font(wb)   # 전체 글꼴 Noto Sans KR 통일(모든 문서 Excel 공통)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _numcell(value: Any):
    """숫자 칸 — 숫자로 읽히면 숫자로 넣어 Excel 이 계산·서식을 걸 수 있게 한다.
    빈 값과 0 은 빈 칸으로 둔다(서식에 0 이 줄줄이 찍히지 않게)."""
    text = str(value if value is not None else "").strip()
    if not text:
        return ""
    try:
        n = float(text)
    except ValueError:
        return text
    if n == 0:
        return ""
    return int(n) if n == int(n) else n


def _pkg_text_xlsx(it: Dict[str, Any]) -> str:
    """'No. & Kind of Packages' 셀 — 수량+종류 결합, 없으면 레거시 package."""
    q = str(it.get("pkg_qty") or "").strip()
    k = str(it.get("pkg_kind") or "").strip()
    combined = f"{q} {k}".strip()
    return combined or str(it.get("package") or "").strip()


def make_packing_list_xlsx(
    data: Dict[str, Any], company: Optional[Dict[str, Any]] = None
) -> bytes:
    """Packing List 전용 Excel — templates/'packing list_sample.xlsx' 서식 그대로.

    Commercial Invoice 와 같은 위쪽 구성(레터헤드 → 문서정보 → 당사자 → 선적정보)에,
    가격 대신 포장(No.&Kind of Packages)·중량(N.W./G.W.)·치수(L·W·H)·용적(Meas.) 열과
    TOTAL 행을 갖는다. 열 격자(PL_COLUMN_UNITS)는 같은 문서의 PDF 와 같다.
    """
    company = company or {}
    shipping = data.get("shipping", {}) or {}
    terms = data.get("terms", {}) or {}
    items = normalize_items(data.get("items", []))
    consignee, buyer = doc_parties(data)
    num_fmt = "#,##0.###"

    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"
    form = _FormSheet(ws, PL_COLUMN_UNITS, _BLOCKS_13)

    r = form.letterhead(company, "PACKING LIST")
    # Packing List 는 자체 번호·발행일이 없다 — 딸려 나가는 송장의 번호·발행일을 싣는다.
    r = form.doc_info(r, [("Invoice No.", shipping.get("ci_no", "") or data.get("doc_no", "")),
                          ("Invoice Date", _invoice_doc_date(shipping.get("ci_date", "") or data.get("date", ""))),
                          ("PO No.", shipping.get("po_no", ""))])
    r += 1
    r = form.band(r, "CONSIGNEE", "BUYER (if different)")
    r = form.pairs(r, [("Company Name", consignee.get("name", ""), "Company Name", buyer.get("name", "")),
                       ("Address", consignee.get("address", ""), "Address", buyer.get("address", "")),
                       ("Contact", consignee.get("contact", ""), "Contact", buyer.get("contact", "")),
                       ("e-mail", consignee.get("email", ""), "e-mail", buyer.get("email", ""))],
                   heights={1: 29.4})
    r = form.band(r, "SHIPPING INFORMATION")
    r = form.pairs(r, _invoice_shipping_rows(data) + [
        ("Packing", terms.get("packing_type", ""), "Country of Origin", shipping.get("sm_origin", "")),
    ])
    r += 1

    # ── 품목 표 — 머리행 2줄, Dim.(cm) 아래 L/W/H 세 칸 ──────────────────
    hrow = r
    for title, c1, c2 in [("No.", 1, 1), ("Description", 2, 3), ("Part No.", 4, 4), ("Q'ty", 5, 5),
                          ("Unit", 6, 6), ("No. & Kind of Packages", 7, 7), ("N.W. (kg)", 8, 8),
                          ("G.W. (kg)", 9, 9), ("Meas. (m³)", 13, 13)]:
        form.merge(hrow, c1, hrow + 1, c2)
        form.put(hrow, c1, title, fill=form.navy, font=form.white_hdr, align=form.center)
    form.merge(hrow, 10, hrow, 12)
    form.put(hrow, 10, "Dim. (cm)", fill=form.navy, font=form.white_hdr, align=form.center)
    for i, axis in enumerate(("L", "W", "H")):
        form.put(hrow + 1, 10 + i, axis, fill=form.navy, font=form.white_hdr, align=form.center)
    form.bd(hrow, 1, hrow + 1, form.ncol, form.navy)
    form.ws.row_dimensions[hrow].height = 20.4
    form.ws.row_dimensions[hrow + 1].height = 17.4
    r = hrow + 2

    first = r
    for i in range(max(len(items), PL_MIN_ITEM_ROWS)):
        it = items[i] if i < len(items) else None
        dims = _dim_parts(it.get("dimension")) if it else ["", "", ""]
        form.put(r, 1, it["item_no"] if it else "", font=form.item_font, align=form.center)
        form.merge(r, 2, r, 3)
        form.put(r, 2, it["description"] if it else "", font=form.item_font, align=form.left)
        form.put(r, 4, it["part_no"] if it else "", font=form.item_font, align=form.left)
        form.put(r, 5, _num(it["qty"]) if it else "", font=form.item_font, align=form.center)
        form.put(r, 6, it["unit"] if it else "", font=form.item_font, align=form.center)
        form.put(r, 7, _pkg_text_xlsx(it) if it else "", font=form.item_font, align=form.center)
        form.put(r, 8, _numcell(it.get("net_weight")) if it else "", font=form.item_font,
                 align=form.right, fmt=num_fmt)
        form.put(r, 9, _numcell(it.get("gross_weight")) if it else "", font=form.item_font,
                 align=form.right, fmt=num_fmt)
        for j, d in enumerate(dims):
            form.put(r, 10 + j, _numcell(d), font=form.item_font, align=form.center, fmt=num_fmt)
        form.put(r, 13, _numcell(it.get("measurement")) if it else "", font=form.item_font,
                 align=form.right, fmt="#,##0.####")
        form.bd(r, 1, r, form.ncol, None if it else form.alt)
        form.ws.row_dimensions[r].height = form.item_row_height(it["description"], 26) if it else 18
        r += 1
    last = r - 1

    # ── TOTAL 행 — 전체 포장 규격을 직접 적었으면 그 값, 아니면 품목별 합산 ──
    tot = packing_totals(data)
    form.merge(r, 1, r, 4)
    form.put(r, 1, "TOTAL", fill=form.lightblue, font=form.bold, align=form.center)
    form.put(r, 5, f"=SUM(E{first}:E{last})", fill=form.lightblue, font=form.bold, align=form.center)
    form.put(r, 6, items[0]["unit"] if items else "", fill=form.lightblue, font=form.bold, align=form.center)
    form.put(r, 7, tot["packages"], fill=form.lightblue, font=form.bold, align=form.right)
    form.put(r, 8, _numcell(tot["net_weight"]), fill=form.lightblue, font=form.bold, align=form.right, fmt=num_fmt)
    form.put(r, 9, _numcell(tot["gross_weight"]), fill=form.lightblue, font=form.bold, align=form.right, fmt=num_fmt)
    for j, key in enumerate(("sm_dim_l", "sm_dim_w", "sm_dim_h")):
        form.put(r, 10 + j, _numcell(shipping.get(key)), fill=form.lightblue, font=form.bold,
                 align=form.center, fmt=num_fmt)
    form.put(r, 13, _numcell(tot["measurement"]), fill=form.lightblue, font=form.bold,
             align=form.right, fmt="#,##0.####")
    form.bd(r, 1, r, form.ncol)
    form.ws.row_dimensions[r].height = 20
    r += 1

    # ── Packing Information(자유 메모) — 적었을 때만 붙인다 ────────────────
    packing_info = (data.get("packing_info") or "").strip()
    if packing_info:
        r += 1
        r = form.band(r, "PACKING INFORMATION")
        form.cell(r, "full", packing_info, font=form.normal, align=form.left)
        form.bd(r, 1, r, form.ncol)
        r += 1

    last_row = _invoice_closing(form, company, r, "Packing List", extra_cell=True)
    form.page_setup(last_row)
    _apply_noto_font(wb)   # 전체 글꼴 Noto Sans KR 통일(모든 문서 Excel 공통)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def make_shipping_mark_xlsx(
    data: Dict[str, Any], company: Optional[Dict[str, Any]] = None
) -> bytes:
    """Shipping Mark(케이스 마킹) Excel — PDF(_make_shipping_mark_pdf)와 같은 구성:
    타이틀·배너 → 참조 스트립(REF/P.O./DATE) → 주 마크 박스 → 실측표 → 취급주의.

    범용 make_document_xlsx 를 쓰지 않는 이유: 그쪽은 품목표 문서(CI/PL/PO)용이고
    Shipping Mark 은 품목이 없는 마킹 라벨이라 형태가 전혀 다르다.
    마크 박스는 스텐실로 옮겨 적는 원본이므로 PDF 와 같은 줄·같은 순서로 유지한다.
    """
    company = company or {}
    vessel = data.get("vessel", {}) or {}
    shipping = data.get("shipping", {}) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Shipping Mark"
    ws.sheet_view.showGridLines = False

    navy = PatternFill("solid", fgColor="0B1D3A")
    blue = PatternFill("solid", fgColor="0055A8")

    title_font = Font(name="Noto Sans KR", color="0B1D3A", bold=True, size=19)
    white_sm = Font(name="Noto Sans KR", color="FFFFFF", size=9)
    white_lbl = Font(name="Noto Sans KR", color="FFFFFF", bold=True, size=9)
    normal = Font(name="Noto Sans KR", size=9)
    mark_font = Font(name="Noto Sans KR", bold=True, size=13)

    thin = Side(style="thin", color="C8D2E0")
    bdr = Border(top=thin, bottom=thin, left=thin, right=thin)
    thick = Side(style="medium", color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    NCOL = 6
    for i, w in enumerate([13, 20, 12, 18, 10, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def bd(r1, c1, r2, c2, fill=None, border=bdr):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(rr, cc).border = border
                if fill:
                    ws.cell(rr, cc).fill = fill

    def put(r, c, v="", *, fill=None, font=None, align=None):
        x = ws.cell(r, c, v)
        if fill:
            x.fill = fill
        if font:
            x.font = font
        if align:
            x.alignment = align
        return x

    r = 1
    logo = _find_asset("logo_K-maris.png", "logo.png", "logo.jpg", "logo.jpeg")
    if logo:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo)
            img.width, img.height = 105, 35
            ws.add_image(img, f"A{r}")
        except Exception:
            pass   # 로고는 장식 — 없거나 실패해도 문서는 나와야 한다.
    merge(r, 1, r, NCOL); put(r, 1, "SHIPPING MARK", font=title_font, align=center)
    bd(r, 1, r, NCOL); ws.row_dimensions[r].height = 53.6; r += 1

    banner = "   |   ".join(x for x in [
        company.get("company_name_en", "K-MARIS Energy & Solutions Co., Ltd."),
        company.get("sales_email", ""), company.get("website", ""),
    ] if x)
    merge(r, 1, r, NCOL); put(r, 1, banner, fill=blue, font=white_sm, align=center)
    bd(r, 1, r, NCOL, blue); ws.row_dimensions[r].height = 16; r += 2

    # ── 참조 스트립 ──────────────────────────────────────────────────────
    for c, (label, value) in enumerate([
        ("REF. NO.", shipping.get("sm_ref_no", "")),
        ("P.O. NO.", shipping.get("sm_po_no", "")),
        ("DATE", data.get("date", "")),
    ]):
        lc = c * 2 + 1
        put(r, lc, label, fill=navy, font=white_lbl, align=left)
        put(r, lc + 1, value, font=normal, align=left)
    bd(r, 1, r, NCOL); ws.row_dimensions[r].height = 20; r += 2

    # ── 주 마크 박스 — PDF 와 같은 줄 구성. 한 셀에 줄바꿈으로 넣어야 스텐실용으로
    #    그대로 복사된다(줄마다 행을 쓰면 붙여넣을 때 셀이 쪼개진다).
    lines = []
    if shipping.get("sm_type"):
        lines.append(str(shipping["sm_type"]).upper())
    lines += consignee_mark_lines(shipping)
    if shipping.get("sm_vessel"):
        lines.append(f"M/V {str(shipping['sm_vessel']).upper()}")
    elif vessel.get("name"):
        lines.append(f"M/V {str(vessel['name']).upper()}")
    if shipping.get("sm_po_no"):
        lines.append(f"P.O. NO. : {shipping['sm_po_no']}")
    if shipping.get("sm_ref_no"):
        lines.append(f"REF. NO. : {shipping['sm_ref_no']}")
    if shipping.get("sm_desc"):
        lines.append(str(shipping["sm_desc"]).upper())
    if shipping.get("sm_final_dest"):
        lines.append(f"FINAL DESTINATION : {str(shipping['sm_final_dest']).upper()}")
    if shipping.get("sm_case_no"):
        lines.append(f"CASE NO. : {shipping['sm_case_no']}")
    if shipping.get("sm_origin"):
        lines.append(str(shipping["sm_origin"]).upper())
    if not lines:
        lines.append("(NO SHIPPING MARK DATA)")

    box_top = r
    # 16행이 기본 높이고, 줄이 그보다 많으면(수하인 주소 등) 박스를 늘린다 — 병합 셀은
    # 자동으로 커지지 않아 고정해 두면 긴 마크가 테두리 안에서 잘린다.
    box_bottom = r + max(15, len(lines) + 1)
    merge(box_top, 1, box_bottom, NCOL)
    put(box_top, 1, "\n".join(lines), font=mark_font, align=center)
    bd(box_top, 1, box_bottom, NCOL,
       border=Border(top=thick, bottom=thick, left=thick, right=thick))
    for rr in range(box_top, box_bottom + 1):
        ws.row_dimensions[rr].height = 18
    r = box_bottom + 2

    # ── 실측(중량·치수·케이스) ────────────────────────────────────────────
    dim = [shipping.get("sm_dim_l"), shipping.get("sm_dim_w"), shipping.get("sm_dim_h")]
    dim_txt = (" × ".join((str(d).strip() if d and str(d).strip() else "-") for d in dim) + " CM"
               if any(d and str(d).strip() for d in dim) else "")
    metrics = [
        [("N.W.", f"{shipping['sm_net_weight']} KG" if shipping.get("sm_net_weight") else ""),
         ("G.W.", f"{shipping['sm_gross_weight']} KG" if shipping.get("sm_gross_weight") else "")],
        [("DIMENSION", dim_txt),
         ("TOTAL CASES", f"{shipping['sm_total_cases']} CASE(S)" if shipping.get("sm_total_cases") else "")],
    ]
    for row in metrics:
        (k1, v1), (k2, v2) = row
        put(r, 1, k1, fill=navy, font=white_lbl, align=left)
        merge(r, 2, r, 3); put(r, 2, v1, font=normal, align=left)
        put(r, 4, k2, fill=navy, font=white_lbl, align=left)
        merge(r, 5, r, 6); put(r, 5, v2, font=normal, align=left)
        bd(r, 1, r, NCOL); ws.row_dimensions[r].height = 20; r += 1

    # ── 취급 주의 ────────────────────────────────────────────────────────
    handling = str(shipping.get("sm_handling") or "").strip()
    if handling:
        put(r, 1, "HANDLING", fill=navy, font=white_lbl, align=left)
        merge(r, 2, r, NCOL)
        put(r, 2, " · ".join(h.strip() for h in handling.split(",") if h.strip()),
            font=normal, align=left)
        bd(r, 1, r, NCOL); ws.row_dimensions[r].height = 20

    _apply_noto_font(wb)   # 전체 글꼴 Noto Sans KR 통일(모든 문서 Excel 공통)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def make_document_xlsx(
    doc_type: str, data: Dict[str, Any], company: Optional[Dict[str, Any]] = None
) -> bytes:
    # 라우트가 company 를 안 넘기면(대부분) 설정(company.json)에서 로드 — PDF 경로와 동일.
    if company is None:
        try:
            from services.pdf_svc import _load_company
            company = _load_company()
        except Exception:
            company = {}
    if doc_type == "quotation":
        return make_quotation_costing_xlsx(data, company)
    if doc_type == "shipping_mark":
        return make_shipping_mark_xlsx(data, company)
    if doc_type == "purchase_order":
        return make_purchase_order_xlsx(data, company)
    title = DOC_TITLES.get(doc_type, "DOCUMENT")
    currency = (data.get("currency") or "USD").upper()
    customer = data.get("customer", {})   # PO 면 공급사(Vendor)
    vessel = data.get("vessel", {})
    terms = data.get("terms", {})
    items = normalize_items(data.get("items", []))
    totals = calc_totals(
        data.get("items", []), _num(data.get("vat_rate", 0)), _num(data.get("discount_pct", 0))
    )
    num_fmt = "#,##0.00" if currency == "USD" else "#,##0"

    wb = Workbook()
    ws = wb.active
    ws.title = (title.title() or "Document")[:31]

    navy = PatternFill("solid", fgColor="0B1D3A")
    blue = PatternFill("solid", fgColor="0055A8")
    gray = PatternFill("solid", fgColor="F4F6F8")
    lightblue = PatternFill("solid", fgColor="EAF3FF")
    alt = PatternFill("solid", fgColor="FAFBFC")

    white_lg = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
    white_sm = Font(name="Calibri", color="FFFFFF", size=9)
    white_hdr = Font(name="Calibri", color="FFFFFF", bold=True, size=9)
    bold = Font(name="Calibri", bold=True)
    boldsm = Font(name="Calibri", bold=True, size=9)

    thin = Side(style="thin", color="D8DEE6")
    bdr = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    HEADERS = ["No.", "Part No.", "Description", "Maker", "Origin", "Qty", "Unit",
               "Unit Price", "Amount", "Lead Time / Remark"]
    WIDTHS = [5, 18, 40, 18, 12, 7, 7, 15, 16, 28]
    NUM_COLS = len(HEADERS)

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    # ── Title / sub-header ──────────────────────────────────────────────
    merge(1, 1, 1, NUM_COLS)
    c = ws.cell(1, 1, title); c.fill = navy; c.font = white_lg; c.alignment = center
    ws.row_dimensions[1].height = 28
    merge(2, 1, 2, NUM_COLS)
    c = ws.cell(2, 1, "K-MARIS Energy & Solutions Co., Ltd.  |  sales@k-maris.com  |  www.k-maris.com")
    c.fill = blue; c.font = white_sm; c.alignment = center
    ws.row_dimensions[2].height = 16

    # ── Meta (rows 4-7): 좌측 상대방/선박, 우측 문서정보 ─────────────────
    party = "Customer / Buyer"
    meta = [
        (party, customer.get("name", ""), "Document No.", data.get("doc_no", "")),
        ("Address", customer.get("address", ""), "Date", data.get("date", "")),
        ("Vessel", vessel.get("name", ""), "Currency", currency),
        ("Contact", customer.get("contact", ""), "Incoterms", terms.get("incoterms", "")),
    ]
    for off, (k1, v1, k2, v2) in enumerate(meta, start=4):
        for col, val, is_label in [(1, k1, True), (2, v1, False), (7, k2, True), (8, v2, False)]:
            cell = ws.cell(off, col, val); cell.border = bdr; cell.alignment = left
            if is_label:
                cell.fill = gray; cell.font = boldsm
        merge(off, 2, off, 6); merge(off, 8, off, NUM_COLS)
        for col in range(2, 7):
            ws.cell(off, col).border = bdr
        for col in range(8, NUM_COLS + 1):
            ws.cell(off, col).border = bdr
        ws.row_dimensions[off].height = 15

    # ── Item table header (row 9) ───────────────────────────────────────
    HROW = 9
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(HROW, ci, h); cell.fill = navy; cell.font = white_hdr
        cell.alignment = center; cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HROW].height = 26

    # ── Data rows ───────────────────────────────────────────────────────
    for ri, it in enumerate(items, start=1):
        r = HROW + ri
        lead_remark = f"{it.get('lead_time', '')} {it.get('remark', '')}".strip()
        vals = [it["item_no"], it["part_no"], it["description"], it["maker"], it["origin"],
                _num(it["qty"]), it["unit"], _num(it["unit_price"]), _num(it["amount"]), lead_remark]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(r, ci, val); cell.border = bdr
            if ri % 2 == 0:
                cell.fill = alt
            if ci in (6, 8, 9):  # Qty · Unit Price · Amount → 우측, 숫자서식
                cell.alignment = right
                if ci in (8, 9):
                    cell.number_format = num_fmt
            elif ci in (1, 7):
                cell.alignment = center
            else:
                cell.alignment = left
        ws.row_dimensions[r].height = 18

    # ── Totals ──────────────────────────────────────────────────────────
    trow = HROW + len(items) + 1
    lines = [("Subtotal", totals.get("subtotal", 0))]
    if totals.get("discount_pct"):
        lines.append((f"Discount ({_num(totals['discount_pct']):g}%)", -totals.get("discount", 0)))
    lines.append(("VAT", totals.get("vat", 0)))
    lines.append(("Total", totals.get("total", 0)))
    for i, (lab, val) in enumerate(lines):
        r = trow + i
        lc = ws.cell(r, 8, lab); lc.fill = gray; lc.font = boldsm; lc.alignment = right; lc.border = bdr
        vc = ws.cell(r, 9, _num(val)); vc.border = bdr; vc.alignment = right; vc.number_format = num_fmt
        ws.cell(r, 10).border = bdr
        if lab == "Total":
            lc.fill = lightblue; vc.fill = lightblue; lc.font = bold; vc.font = bold

    # ── Terms & Conditions ──────────────────────────────────────────────
    tstart = trow + len(lines) + 2
    ws.cell(tstart, 1, "Terms & Conditions").font = bold
    term_rows = [
        ("Incoterms", terms.get("incoterms", "")),
        ("Place", terms.get("delivery_place", "")),
        ("Payment Terms", terms.get("payment_terms", "")),
        ("Packing", terms.get("packing", "")),
        ("Warranty", terms.get("warranty", "")),
        ("Remarks", terms.get("remarks", "")),
    ]
    for i, (k, v) in enumerate(term_rows, start=1):
        r = tstart + i
        kc = ws.cell(r, 1, k); kc.fill = gray; kc.font = boldsm; kc.alignment = left; kc.border = bdr
        merge(r, 2, r, NUM_COLS)
        vc = ws.cell(r, 2, v); vc.alignment = left; vc.border = bdr
        for col in range(2, NUM_COLS + 1):
            ws.cell(r, col).border = bdr

    ws.freeze_panes = f"A{HROW + 1}"

    _apply_noto_font(wb)   # 전체 글꼴 Noto Sans KR 통일(모든 문서 Excel 공통)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def make_quotation_costing_xlsx(
    data: Dict[str, Any], company: Optional[Dict[str, Any]] = None
) -> bytes:
    """고객 견적서(COSTING SHEET) Excel — sales + PURCHASE(원가) + MARGIN 포함(내부용)."""
    from services.kmaris_docs import quotation_remark_lines, quotation_standard_terms

    company = company or {}
    customer = data.get("customer", {}) or {}
    vessel = data.get("vessel", {}) or {}
    terms = data.get("terms", {}) or {}
    currency = (data.get("currency") or "USD").upper()
    raw_items = data.get("items", []) or []
    num_fmt = "#,##0.00" if currency == "USD" else "#,##0"

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    navy = PatternFill("solid", fgColor="0B1D3A")
    blue = PatternFill("solid", fgColor="0055A8")
    gray = PatternFill("solid", fgColor="F4F6F8")
    lightblue = PatternFill("solid", fgColor="EAF3FF")
    cost_fill = PatternFill("solid", fgColor="FFFFCC")   # PURCHASE 열 톤(샘플: 연노랑)
    alt = PatternFill("solid", fgColor="FAFBFC")

    white_lg = Font(name="Calibri", color="FFFFFF", bold=True, size=15)
    white_sm = Font(name="Calibri", color="FFFFFF", size=9)
    white_hdr = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    bold = Font(name="Calibri", bold=True)
    boldsm = Font(name="Calibri", bold=True, size=9)

    thin = Side(style="thin", color="D8DEE6")
    bdr = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # No | Part No | Description | Qty | Unit | Cost U/P | Cost Amount | Margin% | U/Price | Amount | Lead Time | Remark
    HEADERS = ["No.", "Part No.", "Description", "Qty", "Unit",
               "Cost U/P", "Cost Amount", "Margin %", "U/Price", "Amount", "Lead Time", "Remark"]
    WIDTHS = [5, 18, 34, 7, 7, 13, 14, 9, 13, 15, 14, 20]
    NCOL = len(HEADERS)

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def add_image(path, anchor, w, h):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(path); img.width = w; img.height = h
            ws.add_image(img, anchor)
        except Exception:
            pass

    # ── 레터헤드: 아이콘 로고(좌, 텍스트 없는 심볼 우선) + 회사정보(중) + 태그라인(우) ──
    # 텍스트가 빠진 아이콘 로고를 쓰려면 config/ 또는 저장소 루트에 logo_icon.png(또는
    # logo_mark/logo_symbol) 파일을 두면 그것을 우선 사용한다. 없으면 기존 로고로 대체.
    logo = _find_asset("logo_icon.jpg", "logo_icon.png", "logo_mark.png", "logo_symbol.png",
                       "logo_K-maris.png", "logo.png", "logo.jpg")
    if logo:
        # 아이콘 로고 — B~C 열 사이에 걸치도록(colOff) 배치, 행 1-3 세로 걸침.
        try:
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            im = XLImage(logo)
            im.anchor = OneCellAnchor(
                _from=AnchorMarker(col=1, colOff=pixels_to_EMU(60), row=0, rowOff=pixels_to_EMU(4)),
                ext=XDRPositiveSize2D(pixels_to_EMU(112), pixels_to_EMU(99)),
            )
            ws.add_image(im)
        except Exception:
            add_image(logo, "B1", 112, 99)
    GRAYTX = "404040"   # 진한 회색(회사명·주소·연락처)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hd_name = Font(name="Calibri", bold=False, size=18, color=GRAYTX)
    hd_addr = Font(name="Calibri", size=10, color=GRAYTX)
    hd_tag = Font(name="Calibri", italic=True, size=10, color="0055A8")
    addr = company.get("address_en") or company.get("address") or ""
    bits = []
    if company.get("phone"): bits.append(f"Tel: {company['phone']}")
    if company.get("sales_email"): bits.append(company["sales_email"])
    if company.get("website"): bits.append(company["website"])
    contact = "   |   ".join(bits)
    # 회사명/주소/연락처 — C~K 가운데 정렬, 진한 회색, 음영 없음.
    merge(1, 3, 1, 11); cc = ws.cell(1, 3, company.get("company_name_en", "K-MARIS Energy & Solutions Co., Ltd.")); cc.font = hd_name; cc.alignment = center_wrap
    merge(2, 3, 2, 11); cc = ws.cell(2, 3, addr); cc.font = hd_addr; cc.alignment = center_wrap
    merge(3, 3, 3, 11); cc = ws.cell(3, 3, contact); cc.font = hd_addr; cc.alignment = center_wrap
    # 슬로건 — 우측 열 L1:L3, 좌측 정렬, 2줄(줄바꿈), 이탤릭 블루.
    tagline = (company.get("tagline", "") or "").replace(". ", ".\n", 1)
    merge(1, 12, 3, NCOL); cc = ws.cell(1, 12, tagline); cc.font = hd_tag; cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 17
    ws.row_dimensions[3].height = 17
    for col in range(1, NCOL + 1):
        ws.cell(3, col).border = Border(bottom=Side(style="medium", color="0055A8"))
    ws.row_dimensions[4].height = 6
    # ── 타이틀 (샘플처럼 크게) ─────────────────────────────────────────────
    merge(5, 1, 5, NCOL)
    c = ws.cell(5, 1, "QUOTATION / COSTING SHEET"); c.font = Font(name="Calibri", bold=True, size=22, color="0B1D3A"); c.alignment = center
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[6].height = 4

    # 원가(cost) 통화 → 판매 통화 환산계수. Margin 수식에서 통화가 섞일 때 사용.
    cost_cur = (data.get("cost_currency") or currency).upper()
    fx = _num(data.get("fx_rate")) or 0.0
    if cost_cur == currency or fx <= 0:
        factor = 1.0
    elif cost_cur == "KRW" and currency == "USD":
        factor = 1.0 / fx
    elif cost_cur == "USD" and currency == "KRW":
        factor = fx
    else:
        factor = 1.0
    fx_str = f"{factor:.10g}"
    cost_fmt = "#,##0.00" if cost_cur == "USD" else "#,##0"
    # 판매가 올림 자릿수 — 편집창 "Round unit price up to"(1000→-3, 100→-2 …). None 이면 -2.
    _rd = data.get("round_digits")
    rd = -2 if _rd is None else int(_rd)

    # ── Meta (rows 4-8) — 미리보기(PDF) 와 동일한 순서·구성 ────────────
    vat_label = "VAT excluded" if _num(data.get("vat_rate", 0)) == 0 else f"VAT {int(_num(data.get('vat_rate', 0)) * 100)}%"
    meta = [
        ("User", customer.get("name", ""), "Quotation No.", data.get("doc_no", "")),
        ("Messrs", data.get("messrs", ""), "Ref. No.", data.get("ref_no", "")),
        ("Attn.", data.get("attn", "") or customer.get("contact", ""), "Date", data.get("date", "")),
        ("Ship Name", vessel.get("name", ""), "Currency", currency),
        ("Project", data.get("project_title", ""), "VAT", vat_label),
    ]
    # 원가열(F·G·H)은 기본 숨김이므로 메타는 그 열을 피해 좌(1-5)·우(9-12)에 배치.
    for off, (k1, v1, k2, v2) in enumerate(meta, start=7):
        merge(off, 1, off, 2); merge(off, 3, off, 5)
        merge(off, 9, off, 10); merge(off, 11, off, NCOL)
        for col, val, is_label in [(1, k1, True), (3, v1, False), (9, k2, True), (11, v2, False)]:
            cell = ws.cell(off, col, val); cell.alignment = left
            if is_label:
                cell.fill = gray; cell.font = boldsm
        for col in (1, 2, 3, 4, 5, 9, 10, 11, 12):
            ws.cell(off, col).border = bdr
        ws.row_dimensions[off].height = 15

    # ── PURCHASE 그룹 라벨(원가 열 위) ─────────────────────────────────
    GROUP_ROW = 12
    merge(GROUP_ROW, 6, GROUP_ROW, 8)
    gc = ws.cell(GROUP_ROW, 6, f"PURCHASE (internal, {cost_cur})"); gc.fill = cost_fill; gc.font = boldsm; gc.alignment = center
    for col in range(1, NCOL + 1):
        ws.cell(GROUP_ROW, col).border = bdr
        if col < 6 or col > 8:
            ws.cell(GROUP_ROW, col).fill = gray
    ws.row_dimensions[GROUP_ROW].height = 14

    # ── Item header (row 13) ───────────────────────────────────────────
    HROW = 13
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(HROW, ci, h); cell.fill = navy; cell.font = white_hdr
        cell.alignment = center; cell.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HROW].height = 24

    # 컬럼: A No · B Part · C Desc · D Qty · E Unit · F Cost U/P · G Cost Amount
    #       · H Margin% · I U/Price · J Amount · K Lead · L Remark
    first = HROW + 1
    for ri, it in enumerate(raw_items, start=1):
        r = HROW + ri
        qty = _num(it.get("qty", 0))
        unit_price = _num(it.get("unit_price", 0))
        cost = _num(it.get("cost_price", 0))
        # 마진(H)은 입력값(분수), U/Price(I)는 원가·마진으로 계산하는 수식 — 샘플과 동일.
        # margin_pct(예: 35 또는 0.35)를 분수로 정규화. 없으면 원가·판매가에서 유도.
        mp = _num(it.get("margin_pct", 0))
        if mp:
            margin_frac = mp / 100.0 if mp > 1 else mp
        else:
            csell = cost * factor
            margin_frac = (1 - csell / unit_price) if unit_price else 0.0
        cells = {
            1: ri, 2: it.get("part_no", ""), 3: it.get("description", ""),
            4: qty, 5: it.get("unit", ""),
            6: cost,                                   # Cost U/P (입력, 원가통화)
            7: f"=D{r}*F{r}",                          # Cost Amount = Qty × Cost
            8: margin_frac,                            # Margin % (입력, 분수)
            # U/Price = 원가(판매통화 환산) ÷ (1−마진), 100단위 올림 — 샘플 수식.
            9: f"=IF(OR(F{r}=0,H{r}>=1),0,ROUNDUP(F{r}*{fx_str}/(1-H{r}),{rd}))",
            10: f"=D{r}*I{r}",                         # Amount = Qty × U/Price
            11: str(it.get("lead_time", "") or ""), 12: it.get("remark", ""),
        }
        for ci, val in cells.items():
            cell = ws.cell(r, ci, val); cell.border = bdr
            if ci in (6, 7, 8):   # 원가열만 연노랑 음영(샘플엔 판매열 zebra 없음)
                cell.fill = cost_fill
            if ci in (4, 9, 10):
                cell.alignment = right
                if ci in (9, 10):
                    cell.number_format = num_fmt
            elif ci in (6, 7):
                cell.alignment = right; cell.number_format = cost_fmt
            elif ci == 8:
                cell.alignment = right; cell.number_format = '0.0%'
            elif ci in (1, 5):
                cell.alignment = center
            else:
                cell.alignment = left
        # 줄바꿈 텍스트가 잘리지 않도록 Description(C, ~40자/줄)·Remark(L, ~26자/줄)
        # 내용에 맞춰 행 높이를 늘린다(고정 18은 2줄 이상에서 겹침).
        _desc = str(it.get("description", "") or ""); _rmk = str(it.get("remark", "") or "")
        _dl = sum(max(1, (len(x) + 39) // 40) for x in _desc.split("\n")) if _desc else 1
        _rl = sum(max(1, (len(x) + 25) // 26) for x in _rmk.split("\n")) if _rmk else 1
        ws.row_dimensions[r].height = 14 * max(_dl, _rl, 1) + 5
    # 샘플처럼 최소 5줄의 폼 형태 — 품목이 적어도 빈 줄로 표 높이를 유지(Total 위치 고정).
    MIN_ITEM_ROWS = 5
    for ri in range(len(raw_items) + 1, MIN_ITEM_ROWS + 1):
        r = HROW + ri
        for ci in range(1, NCOL + 1):
            cell = ws.cell(r, ci); cell.border = bdr
            if ci in (6, 7, 8):
                cell.fill = cost_fill
        ws.row_dimensions[r].height = 18
    last = HROW + max(len(raw_items), MIN_ITEM_ROWS)

    # ── Totals (수식) ──────────────────────────────────────────────────
    trow = last + 1
    has_rows = len(raw_items) > 0
    # "Total" 라벨 — A~E 병합, 가운데. 행 전체 글자 크기 12.
    bold16 = Font(name="Noto Sans KR", bold=True, size=12)
    merge(trow, 1, trow, 5)
    tc = ws.cell(trow, 1, "Total"); tc.font = bold16; tc.alignment = center
    for col in (1, 2, 3, 4, 5, 11, 12):
        ws.cell(trow, col).border = bdr
    ws.row_dimensions[trow].height = 16
    cost_sum = f"=SUM(G{first}:G{last})" if has_rows else 0
    amt_sum = f"=SUM(J{first}:J{last})" if has_rows else 0
    margin_tot = f"=IF(J{trow}=0,0,(J{trow}-G{trow}*{fx_str})/J{trow})" if has_rows else 0
    for col, val, fill in [(6, "", cost_fill), (7, cost_sum, cost_fill), (8, margin_tot, cost_fill),
                           (9, "", lightblue), (10, amt_sum, lightblue)]:
        cell = ws.cell(trow, col, val); cell.border = bdr; cell.fill = fill; cell.font = bold16; cell.alignment = right
        if col == 7:
            cell.number_format = cost_fmt
        if col == 10:
            cell.number_format = num_fmt
        if col == 8:
            cell.number_format = '0.0%'

    # 섹션 헤더(네이비 바) 헬퍼.
    def section_bar(r, title):
        merge(r, 1, r, NCOL)
        c = ws.cell(r, 1, title); c.fill = navy; c.font = white_hdr; c.alignment = left
        for col in range(1, NCOL + 1):
            ws.cell(r, col).fill = navy
        ws.row_dimensions[r].height = 16

    # ── Remark ─────────────────────────────────────────────────────────
    # 4단계 Remarks 입력 — PDF 와 같은 자리(품목표와 T&C 사이)에 같은 형식으로 둔다.
    r = trow + 2
    remark_lines = quotation_remark_lines(terms)
    if remark_lines:
        section_bar(r, "Remark"); r += 1
        for line in remark_lines:
            merge(r, 1, r, NCOL)
            ws.cell(r, 1, f"• {line}").alignment = left
            ws.row_dimensions[r].height = 13
            r += 1
        r += 1

    # ── Terms & Conditions ─────────────────────────────────────────────
    section_bar(r, "Terms & Conditions"); r += 1
    for line in quotation_standard_terms(terms):
        merge(r, 1, r, NCOL)
        ws.cell(r, 1, f"• {line}").alignment = left
        ws.row_dimensions[r].height = 13
        r += 1

    # ── Payment ────────────────────────────────────────────────────────
    r += 1
    section_bar(r, "Payment"); r += 1
    for line in (terms.get("payment_terms") or "T/T in advance",):
        merge(r, 1, r, NCOL)
        ws.cell(r, 1, f"• {line}").alignment = left
        ws.row_dimensions[r].height = 13
        r += 1
    r += 1
    merge(r, 1, r, NCOL)
    ws.cell(r, 1, "We hope this quotation meets your requirement and to receive your order "
            "confirmation at your earliest convenience.").alignment = left
    r += 2

    # ── 서명 ───────────────────────────────────────────────────────────
    merge(r, 1, r, NCOL); ws.cell(r, 1, "Your sincerely").alignment = left
    sig = _find_asset("Authorized signature_Sungyeon Cho.jpg", "signature.png", "signature.jpg")
    if sig:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(sig); img.width = 150; img.height = 52
            # 서명이 아래 서명란(밑줄) 바로 위에 앉도록 한 행 내려서 배치.
            ws.add_image(img, f"A{r + 2}")
        except Exception:
            pass
    r += 4
    merge(r, 1, r, 2); ws.cell(r, 1, "________________________").alignment = left; r += 1
    ws.cell(r, 1, "Sam Cho, Managing Director").font = bold; r += 1
    merge(r, 1, r, NCOL)
    fc = ws.cell(r, 1, "K-MARIS Energy & Solutions | Seoul, Korea | www.k-maris.com")
    fc.font = Font(name="Noto Sans KR", size=9); fc.alignment = center
    last_row = r

    # ── 원가/마진 열(F·G·H)은 기본 숨김(내부 코스팅용) — 필요시 사용자가 펼침 ──
    for col in ("F", "G", "H"):
        ws.column_dimensions[col].hidden = True
    # A4 세로 1페이지 폭에 맞춰 인쇄(PDF 미리보기와 동일한 세로 규격).
    # 숨긴 원가열은 인쇄 폭 계산에서 제외되어 판매 열만 세로로 깔끔히 맞는다.
    ws.print_area = f"A1:{get_column_letter(NCOL)}{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    _apply_noto_font(wb)   # 전체 글꼴 Noto Sans KR 통일(모든 문서 Excel 공통)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def make_purchase_order_xlsx(
    data: Dict[str, Any], company: Optional[Dict[str, Any]] = None
) -> bytes:
    """공급사 발주서(PURCHASE ORDER) Excel — kmaris_docs._make_purchase_order_pdf 와 같은 서식.

    미리보기(PDF)와 칸·순서·문구가 그대로 겹치도록 구성한다:
    레터헤드 → 타이틀 → 정보박스 2단(좌 공급사 / 우 문서정보) → 품목표(+Total 행)
    → 합계 문장 → Terms & Conditions → Payment → 서명 → 푸터.
    """
    from services.kmaris_docs import _money

    company = company or {}
    vendor = data.get("customer", {}) or {}      # build_po_payload: Supplier/Seller = Vendor
    vessel = data.get("vessel", {}) or {}
    terms = data.get("terms", {}) or {}
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "USD").upper()
    total = sum(_num(it.get("amount", 0)) for it in items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    ws.sheet_view.showGridLines = False

    navy = PatternFill("solid", fgColor="0B1D3A")
    gray = PatternFill("solid", fgColor="F4F6F8")
    lightblue = PatternFill("solid", fgColor="EAF3FF")
    alt = PatternFill("solid", fgColor="FAFBFC")

    white_hdr = Font(name="Calibri", color="FFFFFF", bold=True, size=9)
    bold = Font(name="Calibri", bold=True)
    boldsm = Font(name="Calibri", bold=True, size=9)
    normal = Font(name="Calibri", size=9)
    small = Font(name="Calibri", size=8)

    thin = Side(style="thin", color="C8D2E0")
    bdr = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # PDF 품목표와 같은 열 구성(Origin 열 없음 — 미리보기와 동일).
    HEADERS = ["No.", "Part No.", "Description", "Maker", "Qty", "Unit",
               "Unit Price", "Amount", "Lead Time / Remark"]
    WIDTHS = [5, 15, 29, 15, 7, 8, 14, 14, 16]
    NCOL = len(HEADERS)
    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    def merge(r1, c1, r2, c2):
        if (r1, c1) != (r2, c2):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def put(r, c, v="", *, fill=None, font=None, align=None, fmt=None):
        cell = ws.cell(r, c, v)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if fmt:
            cell.number_format = fmt
        return cell

    def box(r1, c1, r2, c2, fill=None):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(rr, cc).border = bdr
                if fill:
                    ws.cell(rr, cc).fill = fill

    def numfmt(value):
        """PDF _qnum 과 같은 표기 — 정수면 소수 생략, 아니면 두 자리."""
        try:
            return "#,##0" if float(value) == int(float(value)) else "#,##0.00"
        except (TypeError, ValueError):
            return "#,##0"

    # ── 레터헤드(로고 · 회사정보 · 슬로건 + 파란 구분선) ────────────────
    logo = _find_asset("logo_icon.jpg", "logo_icon.png", "logo_K-maris.png", "logo.png", "logo.jpg")
    if logo:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo)
            img.width, img.height = 96, 58
            ws.add_image(img, "A1")
        except Exception:
            pass
    GRAYTX = "404040"
    addr = company.get("address_en") or company.get("address") or ""
    bits = []
    if company.get("phone"):
        bits.append(f"Tel: {company['phone']}")
    if company.get("sales_email"):
        bits.append(company["sales_email"])
    if company.get("website"):
        bits.append(company["website"])
    merge(1, 2, 1, 7)
    put(1, 2, company.get("company_name_en", "K-MARIS Energy & Solutions Co., Ltd."),
        font=Font(name="Calibri", size=16, color=GRAYTX), align=center)
    merge(2, 2, 2, 7)
    put(2, 2, addr, font=Font(name="Calibri", size=8, color=GRAYTX), align=center)
    merge(3, 2, 3, 7)
    put(3, 2, "   |   ".join(bits), font=Font(name="Calibri", size=8, color=GRAYTX), align=center)
    merge(1, 8, 3, NCOL)
    put(1, 8, (company.get("tagline", "") or "").replace(". ", ".\n", 1),
        font=Font(name="Calibri", italic=True, size=9, color="0055A8"),
        align=Alignment(horizontal="right", vertical="center", wrap_text=True))
    for row, height in ((1, 30), (2, 14), (3, 14), (4, 6)):
        ws.row_dimensions[row].height = height
    for col in range(1, NCOL + 1):
        ws.cell(3, col).border = Border(bottom=Side(style="medium", color="0055A8"))

    # ── 타이틀 ─────────────────────────────────────────────────────────
    merge(5, 1, 5, NCOL)
    put(5, 1, "PURCHASE ORDER",
        font=Font(name="Calibri", bold=True, size=18, color="0B1D3A"), align=center)
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 5

    # ── 정보 박스(2단) — PDF 와 같은 항목·순서 ──────────────────────────
    incoterms = terms.get("incoterms", "")
    place = terms.get("delivery_place", "")
    incoterms_line = " · ".join([x for x in (incoterms, place) if x])
    left_rows = [
        ("Supplier / Seller", vendor.get("name", "")),
        ("Address", vendor.get("address", "")),
        ("Contact", vendor.get("contact", "")),
        ("Email", vendor.get("email", "")),
        ("Ship Name", vessel.get("name", "")),
        ("Engine Type", vessel.get("engine_type", "")),
    ]
    right_rows = [
        ("P/O No.", data.get("doc_no", "")),
        ("Date", data.get("date", "")),
        ("Currency", currency),
        ("IMO No.", vessel.get("imo", "")),
        ("Incoterms", incoterms_line),
        ("Payment", terms.get("payment_terms", "")),
    ]
    META_ROW = 7
    for i in range(6):
        r = META_ROW + i
        for (c1, c2), value, is_label in (
            ((1, 2), left_rows[i][0], True), ((3, 4), left_rows[i][1], False),
            ((5, 6), right_rows[i][0], True), ((7, NCOL), right_rows[i][1], False),
        ):
            merge(r, c1, r, c2)
            put(r, c1, value, fill=gray if is_label else None,
                font=boldsm if is_label else normal, align=left if is_label else left_top)
        box(r, 1, r, NCOL)
        # 주소처럼 긴 값은 여러 줄로 접히므로 행 높이를 내용 길이에 맞춘다(최대 3줄).
        lines = max(1, -(-len(str(left_rows[i][1] or "")) // 44), -(-len(str(right_rows[i][1] or "")) // 44))
        ws.row_dimensions[r].height = 13.5 * min(lines, 3) + 2

    # ── 품목표 ─────────────────────────────────────────────────────────
    HROW = META_ROW + 6 + 1          # 정보박스 아래 한 줄 띄우고 머리행
    ws.row_dimensions[HROW - 1].height = 8
    for ci, h in enumerate(HEADERS, start=1):
        put(HROW, ci, h, fill=navy, font=white_hdr, align=center).border = bdr
    ws.row_dimensions[HROW].height = 24

    first = HROW + 1
    for ri, it in enumerate(items, start=1):
        r = HROW + ri
        lead_remark = "\n".join(x for x in (str(it.get("lead_time", "") or "").strip(),
                                            str(it.get("remark", "") or "").strip()) if x)
        vals = [it["item_no"], it["part_no"], it["description"], it.get("maker", ""),
                _num(it["qty"]), it.get("unit", ""), _num(it["unit_price"]),
                _num(it["amount"]), lead_remark]
        for ci, val in enumerate(vals, start=1):
            cell = put(r, ci, val)
            cell.border = bdr
            cell.font = normal
            if ri % 2 == 0:
                cell.fill = alt
            if ci in (5, 7, 8):                     # Qty · Unit Price · Amount
                cell.alignment = right
                cell.number_format = numfmt(val)
            elif ci in (1, 6):                      # No. · Unit
                cell.alignment = center
            else:
                cell.alignment = left_top
        desc = str(it.get("description", "") or "")
        dl = sum(max(1, -(-len(x) // 29)) for x in desc.split("\n")) if desc else 1
        rl = sum(max(1, -(-len(x) // 16)) for x in lead_remark.split("\n")) if lead_remark else 1
        ws.row_dimensions[r].height = 13 * max(dl, rl, 1) + 5
    last = HROW + len(items)

    # Total 행 — PDF 와 같이 Description 칸에 'Total', Amount 칸에 합계.
    trow = last + 1
    for ci in range(1, NCOL + 1):
        put(trow, ci, "", fill=lightblue).border = bdr
    put(trow, 3, "Total", fill=lightblue, font=boldsm, align=left)
    put(trow, 8, (f"=SUM(H{first}:H{last})" if items else 0), fill=lightblue, font=boldsm,
        align=right, fmt=numfmt(total))
    ws.row_dimensions[trow].height = 16

    # ── 합계 문장(우측 정렬) ────────────────────────────────────────────
    r = trow + 1
    merge(r, 1, r, NCOL)
    put(r, 1, f"Total: {_money(total, currency)}", font=bold, align=right)
    ws.row_dimensions[r].height = 18

    def section_bar(row, title):
        merge(row, 1, row, NCOL)
        put(row, 1, f" {title}", fill=navy, font=white_hdr, align=left)
        for col in range(1, NCOL + 1):
            ws.cell(row, col).fill = navy
        ws.row_dimensions[row].height = 16

    def bullet(row, text):
        merge(row, 1, row, NCOL)
        put(row, 1, f"• {text}", font=normal, align=left)
        ws.row_dimensions[row].height = 13

    # ── Terms & Conditions ─────────────────────────────────────────────
    r += 2
    section_bar(r, "Terms & Conditions")
    r += 1
    term_lines = []
    if incoterms_line:
        term_lines.append(f"Incoterms: {incoterms_line}")
    if terms.get("packing"):
        term_lines.append(f"Packing: {terms.get('packing')}")
    if terms.get("warranty"):
        term_lines.append(f"Warranty: {terms.get('warranty')}")
    if terms.get("remarks"):
        term_lines.append(f"Remarks: {terms.get('remarks')}")
    if not term_lines:
        term_lines.append("As per the terms agreed between K-MARIS and the supplier.")
    for line in term_lines:
        bullet(r, line)
        r += 1

    # ── Payment ────────────────────────────────────────────────────────
    r += 1
    section_bar(r, "Payment")
    r += 1
    bullet(r, terms.get("payment_terms") or "T/T after delivery")
    r += 1
    bullet(r, "Please confirm this purchase order and proceed with delivery per the agreed schedule.")
    r += 2

    # ── 서명 ───────────────────────────────────────────────────────────
    merge(r, 1, r, NCOL)
    put(r, 1, "For and on behalf of K-MARIS Energy & Solutions Co., Ltd.", font=normal, align=left)
    sign = _find_asset(*_SIGN_ASSET)
    if sign:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(sign)
            img.width, img.height = 140, 48
            ws.add_image(img, f"C{r + 1}")
        except Exception:
            pass
    for i in range(1, 4):
        ws.row_dimensions[r + i].height = 16
    r += 4
    merge(r, 1, r, 2)
    put(r, 1, "________________________", font=normal, align=left)
    r += 1
    put(r, 1, "Sam Cho, Managing Director", font=bold, align=left)
    r += 1
    merge(r, 1, r, NCOL)
    put(r, 1, "K-MARIS Energy & Solutions | Seoul, Korea | www.k-maris.com",
        font=small, align=center)
    last_row = r

    ws.print_area = f"A1:{get_column_letter(NCOL)}{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    _apply_noto_font(wb)   # 전체 글꼴 Noto Sans KR 통일(모든 문서 Excel 공통)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
