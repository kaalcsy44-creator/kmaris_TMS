from __future__ import annotations

import io
import json
import re
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)

def _register_default_font() -> tuple[str, str]:
    bundled_fonts = Path(__file__).resolve().parent.parent / "config" / "fonts"
    candidates = [
        (str(bundled_fonts / "NotoSansKR-Regular.ttf"), str(bundled_fonts / "NotoSansKR-Bold.ttf")),
        ("/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf", "/usr/share/fonts/truetype/nanum/NanumBarunGothicBold.ttf"),
        ("/usr/share/fonts/truetype/nanum/NanumGothic.ttf", "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
        ("C:/Windows/Fonts/malgun.ttf", "C:/Windows/Fonts/malgunbd.ttf"),
        ("/Library/Fonts/AppleGothic.ttf", "/Library/Fonts/AppleGothic.ttf"),
    ]
    for regular, bold in candidates:
        try:
            if Path(regular).exists():
                pdfmetrics.registerFont(TTFont("KMBaseFont", regular))
                if Path(bold).exists():
                    pdfmetrics.registerFont(TTFont("KMBoldFont", bold))
                    return "KMBaseFont", "KMBoldFont"
                return "KMBaseFont", "KMBaseFont"
        except Exception:
            continue
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
        return "HYGothic-Medium", "HYGothic-Medium"
    except Exception:
        return "Helvetica", "Helvetica-Bold"


DEFAULT_FONT, DEFAULT_BOLD_FONT = _register_default_font()

# ReportLab의 인라인 <b> 마크업은 폰트 "패밀리"에 bold 매핑이 등록돼 있어야만
# 실제 굵게 렌더된다. KMBaseFont/KMBoldFont는 개별 폰트로만 등록돼 있어서 매핑이
# 없으면 <b>가 무시된다(정보 박스 라벨 등이 안 굵어짐). 여기서 패밀리를 묶는다.
try:
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    registerFontFamily(
        DEFAULT_FONT,
        normal=DEFAULT_FONT,
        bold=DEFAULT_BOLD_FONT,
        italic=DEFAULT_FONT,
        boldItalic=DEFAULT_BOLD_FONT,
    )
except Exception:
    pass

NAVY = colors.HexColor("#0B1D3A")
BLUE = colors.HexColor("#0055A8")
LIGHT_BLUE = colors.HexColor("#EAF3FF")
LIGHT_GRAY = colors.HexColor("#F4F6F8")
MID_GRAY = colors.HexColor("#D8DEE6")
DARK_GRAY = colors.HexColor("#3A3F44")

DOC_TITLES = {
    "quotation": "QUOTATION",
    "vendor_rfq": "REQUEST FOR QUOTATION",
    "purchase_order": "PURCHASE ORDER",
    "proforma_invoice": "PROFORMA INVOICE",
    "commercial_invoice": "COMMERCIAL INVOICE",
    "tax_invoice": "TAX INVOICE",
    "shipping_mark": "SHIPPING MARK",
    "packing_list": "PACKING LIST",
    "shipping_advice": "SHIPPING ADVICE",
    "credit_note": "CREDIT NOTE",
}

DOC_PREFIX = {
    "quotation": "QTN",
    "purchase_order": "PO",
    "proforma_invoice": "PI",
    "commercial_invoice": "CI",
    "shipping_mark": "SM",
    "packing_list": "PL",
    "shipping_advice": "SA",
    "tax_invoice_data": "TAX",
    "credit_note": "CN",
}


def load_json(path: str | Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def make_doc_no(doc_type: str, sequence: int = 1, company_prefix: str = "KMS") -> str:
    year = date.today().year
    prefix = DOC_PREFIX.get(doc_type, "DOC")
    return f"{company_prefix}-{prefix}-{year}-{sequence:04d}"


def _money(value: Any, currency: str = "USD") -> str:
    try:
        q = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{currency} {q:,.2f}"
    except Exception:
        return f"{currency} 0.00"


def _num(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def normalize_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """발행용 품목 목록 — 화면에서 "문서에서 제외"(excluded)한 행은 여기서 빠진다.
    PDF·Excel·합계(calc_totals)가 모두 이 함수를 거치므로 제외 규칙은 이 한 곳에만 둔다.
    빠진 행이 있으면 남은 행의 No. 를 1..n 으로 다시 매겨 문서에 번호 구멍이 없게 한다."""
    kept = [raw for raw in (items or []) if not (isinstance(raw, dict) and raw.get("excluded"))]
    renumber = len(kept) != len(items or [])
    normalized: List[Dict[str, Any]] = []
    for i, raw in enumerate(kept, start=1):
        qty = _num(raw.get("qty", 0))
        unit_price = _num(raw.get("unit_price", 0))
        amount = raw.get("amount")
        if amount in (None, "", 0):
            amount = qty * unit_price
        normalized.append(
            {
                "item_no": i if renumber else (raw.get("item_no") or i),
                "part_no": raw.get("part_no", ""),
                "description": raw.get("description", ""),
                "maker": raw.get("maker", ""),
                "origin": raw.get("origin", ""),
                "qty": qty,
                "unit": raw.get("unit", "PCS"),
                "unit_price": unit_price,
                "amount": _num(amount),
                "lead_time": raw.get("lead_time", ""),
                "remark": raw.get("remark", ""),
                "gross_weight": raw.get("gross_weight", ""),
                "net_weight": raw.get("net_weight", ""),
                "package": raw.get("package", ""),
                "pkg_qty": raw.get("pkg_qty", ""),
                "pkg_kind": raw.get("pkg_kind", ""),
                "measurement": raw.get("measurement", ""),
                "dimension": raw.get("dimension", ""),
                "hs_code": raw.get("hs_code", ""),
            }
        )
    return normalized


def calc_totals(
    items: List[Dict[str, Any]], vat_rate: float = 0.0, discount_pct: float = 0.0
) -> Dict[str, float]:
    subtotal = sum(_num(item.get("amount", 0)) for item in normalize_items(items))
    discount = subtotal * (_num(discount_pct) / 100.0)
    discounted = subtotal - discount
    vat = discounted * _num(vat_rate)
    total = discounted + vat
    return {
        "subtotal": subtotal,
        "discount_pct": _num(discount_pct),
        "discount": discount,
        "discounted": discounted,
        "vat": vat,
        "total": total,
    }


def _styles() -> Dict[str, ParagraphStyle]:
    styles = getSampleStyleSheet()
    base = ParagraphStyle(
        "KMBase",
        parent=styles["Normal"],
        fontName=DEFAULT_FONT,
        fontSize=8.2,
        leading=10.2,
        textColor=colors.black,
    )
    return {
        "base": base,
        "small": ParagraphStyle("KMSmall", parent=base, fontSize=7.3, leading=9.0),
        "tiny": ParagraphStyle("KMTiny", parent=base, fontSize=6.6, leading=8.2),
        # 품목 표 헤더 — NAVY 배경 위 글자. tiny(검정)를 쓰면 검정 on 남색으로 안 보이므로
        # 반드시 흰색 볼드 스타일을 별도로 둔다.
        "th": ParagraphStyle(
            "KMTh",
            parent=base,
            fontName=DEFAULT_BOLD_FONT,
            fontSize=6.6,
            leading=8.2,
            textColor=colors.white,
        ),
        "title": ParagraphStyle(
            "KMTitle",
            parent=base,
            fontName=DEFAULT_BOLD_FONT,
            fontSize=20,
            leading=24,
            alignment=TA_RIGHT,
            textColor=NAVY,
        ),
        "subtitle": ParagraphStyle(
            "KMSubtitle",
            parent=base,
            fontSize=8.8,
            leading=11,
            alignment=TA_RIGHT,
            textColor=DARK_GRAY,
        ),
        "section": ParagraphStyle(
            "KMSection",
            parent=base,
            fontName=DEFAULT_BOLD_FONT,
            fontSize=9,
            leading=11,
            textColor=colors.white,
        ),
        "right": ParagraphStyle("KMRight", parent=base, alignment=TA_RIGHT),
        "center": ParagraphStyle("KMCenter", parent=base, alignment=TA_CENTER),
    }


def _p(text: Any, style: ParagraphStyle) -> Paragraph:
    safe = "" if text is None else str(text)
    safe = (
        safe.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    # Allow only minimal internal markup used by this template.
    safe = safe.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
    return Paragraph(safe, style)


_ASSET_ROOTS = [Path(__file__).resolve().parent.parent / "templates",
                Path(__file__).resolve().parent.parent / "config",
                Path(__file__).resolve().parents[2]]


def _pdf_asset(*names: str):
    """문서 자산(로고·서명)을 templates/·config/·저장소 루트에서 찾는다.
    이름 우선순위가 폴더보다 우선(앞선 아이콘 이름이 있으면 먼저 선택)."""
    for name in names:
        for root in _ASSET_ROOTS:
            cand = root / name
            if cand.exists():
                return cand
    return None


def _pdf_image(path, max_w, max_h):
    if not path:
        return ""
    try:
        from PIL import Image as PILImage
        with PILImage.open(path) as src:
            w, h = src.size
        scale = min(max_w / w, max_h / h)
        return Image(str(path), width=w * scale, height=h * scale)
    except Exception:
        return ""


def _letterhead(company: Dict[str, Any], doc_title: str, s: Dict[str, ParagraphStyle],
                width: float = 190 * mm) -> List[Any]:
    """모든 문서 PDF 공통 레터헤드 + 중앙 타이틀(견적서와 동일한 비주얼).

    좌: 아이콘 로고 / 중: 회사명(가운데·진회색)·주소·연락처 / 우: 슬로건(블루·우측),
    그 아래 파란 구분선. 이어서 중앙 정렬 문서 타이틀.
    width 는 본문 폭(세로 A4=190mm, 가로 A4≈273mm) — 로고·슬로건 폭은 고정하고
    가운데(회사정보) 칸만 늘려 어떤 방향에서도 꽉 차게 한다.
    """
    def _esc(t: str) -> str:
        return (t or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    title_style = ParagraphStyle("KMHTitle", parent=s["section"], fontName=DEFAULT_BOLD_FONT,
                                 fontSize=17, leading=20, alignment=TA_CENTER, textColor=NAVY)
    logo = _pdf_image(_pdf_asset("logo_icon.jpg", "logo_icon.png", "logo_K-maris.png", "logo.png", "logo.jpg"), 24 * mm, 15 * mm)
    name_style = ParagraphStyle("KMHName", parent=s["base"], fontSize=16, leading=19,
                                alignment=TA_CENTER, textColor=colors.HexColor("#404040"))
    addr_style = ParagraphStyle("KMHAddr", parent=s["base"], fontSize=7.5, leading=10,
                                alignment=TA_CENTER, textColor=colors.HexColor("#404040"))
    tag_style = ParagraphStyle("KMHTag", parent=s["base"], fontSize=8.5, leading=11,
                               alignment=TA_RIGHT, textColor=BLUE)
    _addr = company.get("address_en") or company.get("address") or ""
    _bits = []
    if company.get("phone"): _bits.append(f"Tel: {company['phone']}")
    if company.get("sales_email"): _bits.append(company["sales_email"])
    if company.get("website"): _bits.append(company["website"])
    org_block = [
        Paragraph(_esc(company.get("company_name_en", "K-MARIS Energy & Solutions Co., Ltd.")), name_style),
        Paragraph(_esc(_addr), addr_style),
        Paragraph(_esc("   |   ".join(_bits)), addr_style),
    ]
    tagline = _esc(company.get("tagline", "")).replace(". ", ".<br/>")
    head = Table([[logo, org_block, Paragraph(tagline, tag_style)]],
                 colWidths=[24 * mm, width - 62 * mm, 38 * mm], rowHeights=[17 * mm])
    head.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 1.2, BLUE),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
    ]))
    return [head, Spacer(1, 4 * mm), Paragraph(doc_title, title_style), Spacer(1, 4 * mm)]


def _footer_center(s: Dict[str, ParagraphStyle]) -> Paragraph:
    """모든 문서 PDF 공통 하단 회사 푸터 — 가운데 정렬(엑셀 푸터와 동일)."""
    fc = ParagraphStyle("KMHFoot", parent=s["small"], alignment=TA_CENTER)
    return _p("K-MARIS Energy & Solutions | Seoul, Korea | www.k-maris.com", fc)


def _header(company: Dict[str, Any], doc_title: str, logo_path: Optional[str] = None):
    s = _styles()
    left_lines = [
        f"<b>{company.get('company_name_en', 'K-MARIS Energy & Solutions Co., Ltd.')}</b>",
        company.get("company_name_kr", ""),
        company.get("address", ""),
        f"Tel: {company.get('phone', '')} | Email: {company.get('general_email', '')}",
        f"Website: {company.get('website', '')}",
        f"{company.get('tagline', '')}",
    ]
    left = _p("\n".join([x for x in left_lines if x]), s["base"])
    if logo_path and Path(logo_path).exists():
        try:
            logo = Image(logo_path, width=32 * mm, height=18 * mm)
            left = Table([[logo, left]], colWidths=[36 * mm, 100 * mm])
            left.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        except Exception:
            pass
    right = [
        _p(f"<b>{doc_title}</b>", s["title"]),
        _p("Marine Equipment | Engine Parts | Bunkering | Technical Solutions", s["subtitle"]),
    ]
    table = Table([[left, right]], colWidths=[160 * mm, 110 * mm])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LINEBELOW", (0, 0), (-1, -1), 1, BLUE),
            ]
        )
    )
    return table


def _info_tables(data: Dict[str, Any], doc_type: str):
    s = _styles()
    customer = data.get("customer", {})
    vessel = data.get("vessel", {})
    terms = data.get("terms", {})
    shipping = data.get("shipping", {})

    is_po = doc_type == "purchase_order"
    is_rfq = doc_type == "vendor_rfq"
    party_label = (
        "To (Vendor / Supplier)" if is_rfq
        else "Supplier / Seller" if is_po
        else "Customer / Buyer"
    )
    # 송장류는 문서에 적어 넣은 매수인이 고객 마스터보다 우선하고, 수하인이 매수인과
    # 다르면 한 줄 덧붙인다(수하인·매수인이 다른 거래를 문서에서 구분해 보이게).
    consignee_party = None
    if doc_type in {"proforma_invoice", "commercial_invoice", "packing_list", "shipping_advice"}:
        consignee_party, buyer_party = doc_parties(data)
        customer = {**customer, **buyer_party}
    left_rows = [
        [_p(f"<b>{party_label}</b>", s["base"]), _p(customer.get("name", ""), s["base"])],
        [_p("Address", s["base"]), _p(customer.get("address", ""), s["base"])],
        [_p("Contact", s["base"]), _p(customer.get("contact", ""), s["base"])],
        [_p("Email", s["base"]), _p(customer.get("email", ""), s["base"])],
    ]
    if consignee_party and consignee_party.get("name") and consignee_party["name"] != customer.get("name"):
        left_rows.append([_p("Consignee", s["base"]), _p(consignee_party["name"], s["base"])])
    mid_rows = [
        [_p("<b>Vessel</b>", s["base"]), _p(vessel.get("name", ""), s["base"])],
        [_p("IMO No.", s["base"]), _p(vessel.get("imo", ""), s["base"])],
        [_p("Engine Type", s["base"]), _p(vessel.get("engine_type", ""), s["base"])],
        [_p("Hull No.", s["base"]), _p(vessel.get("hull_no", ""), s["base"])],
    ]

    right_rows = [
        [_p("<b>Document No.</b>", s["base"]), _p(data.get("doc_no", ""), s["base"])],
        [_p("Date", s["base"]), _p(data.get("date", ""), s["base"])],
        [_p("Currency", s["base"]), _p(data.get("currency", "USD"), s["base"])],
        [_p("Incoterms", s["base"]), _p(terms.get("incoterms", ""), s["base"])],
    ]
    if doc_type == "quotation":
        right_rows.append([_p("Validity", s["base"]), _p(data.get("valid_until", ""), s["base"])])
    if doc_type in {"commercial_invoice", "packing_list", "shipping_advice"}:
        right_rows.extend(
            [
                [_p("PO No.", s["base"]), _p(shipping.get("po_no", ""), s["base"])],
                [_p("Export Ref.", s["base"]), _p(shipping.get("export_ref", ""), s["base"])],
            ]
        )

    def box(title, rows):
        table = Table(rows, colWidths=[28 * mm, 58 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.3, MID_GRAY),
                    ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        return table

    outer = Table(
        [[box("Supplier" if is_po else "Customer", left_rows),
          box("Vessel", mid_rows), box("Doc", right_rows)]],
        colWidths=[90 * mm, 90 * mm, 90 * mm],
    )
    outer.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    return outer


def _items_table(data: Dict[str, Any], doc_type: str):
    s = _styles()
    currency = data.get("currency", "USD")
    items = normalize_items(data.get("items", []))

    if doc_type == "packing_list":
        headers = ["No.", "Part No.", "Description", "Qty", "Unit", "Package", "N.W.", "G.W.", "Dimension", "Remark"]
        widths = [10, 30, 65, 15, 16, 28, 20, 20, 34, 32]
        rows = [[_p(h, s["th"]) for h in headers]]
        for item in items:
            rows.append(
                [
                    _p(item["item_no"], s["tiny"]),
                    _p(item["part_no"], s["tiny"]),
                    _p(item["description"], s["tiny"]),
                    _p(str(item["qty"]), s["tiny"]),
                    _p(item["unit"], s["tiny"]),
                    _p(item.get("package", ""), s["tiny"]),
                    _p(item.get("net_weight", ""), s["tiny"]),
                    _p(item.get("gross_weight", ""), s["tiny"]),
                    _p(item.get("dimension", ""), s["tiny"]),
                    _p(item.get("remark", ""), s["tiny"]),
                ]
            )
    elif doc_type == "vendor_rfq":
        # 견적요청서 — 단가/납기/원산지는 공급사가 채우도록 빈칸으로 둔다.
        headers = ["No.", "Part No.", "Description", "Maker", "Qty", "Unit",
                   "Unit Price\n(to quote)", "Lead Time", "Country\nof Origin", "Remark"]
        widths = [10, 30, 64, 32, 15, 16, 26, 24, 28, 43]
        rows = [[_p(h, s["th"]) for h in headers]]
        for item in items:
            rows.append(
                [
                    _p(item["item_no"], s["tiny"]),
                    _p(item["part_no"], s["tiny"]),
                    _p(item["description"], s["tiny"]),
                    _p(item["maker"], s["tiny"]),
                    _p(str(item["qty"]), s["tiny"]),
                    _p(item["unit"], s["tiny"]),
                    _p("", s["tiny"]),  # Unit Price — 공급사 입력
                    _p("", s["tiny"]),  # Lead Time — 공급사 입력
                    _p("", s["tiny"]),  # Origin — 공급사 입력
                    _p(item.get("remark", ""), s["tiny"]),
                ]
            )
    else:
        headers = ["No.", "Part No.", "Description", "Maker", "Origin", "Qty", "Unit", "Unit Price", "Amount", "Lead Time / Remark"]
        widths = [10, 30, 58, 35, 24, 15, 16, 25, 28, 47]
        rows = [[_p(h, s["th"]) for h in headers]]
        for item in items:
            lead_remark = f"{item.get('lead_time', '')}\n{item.get('remark', '')}".strip()
            rows.append(
                [
                    _p(item["item_no"], s["tiny"]),
                    _p(item["part_no"], s["tiny"]),
                    _p(item["description"], s["tiny"]),
                    _p(item["maker"], s["tiny"]),
                    _p(item["origin"], s["tiny"]),
                    _p(str(item["qty"]), s["tiny"]),
                    _p(item["unit"], s["tiny"]),
                    _p(_money(item["unit_price"], currency), s["tiny"]),
                    _p(_money(item["amount"], currency), s["tiny"]),
                    _p(lead_remark, s["tiny"]),
                ]
            )

    table = Table(rows, colWidths=[w * mm for w in widths], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (5, 1), (8, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for r in range(1, len(rows)):
        if r % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FAFBFC")))
    table.setStyle(TableStyle(style_cmds))
    return table


def _totals_table(data: Dict[str, Any], doc_type: str | None = None):
    s = _styles()
    currency = data.get("currency", "USD")
    base = calc_totals(
        data.get("items", []), 0.0, _num(data.get("discount_pct", 0))
    )
    subtotal = base["subtotal"]
    discount = base["discount"]
    # Freight/Packing/Insurance — Proforma Invoice 의 부대비용(terms 에 보관).
    # PI 는 참조 양식처럼 값이 없어도 4개 항목(Freight/Packing/Insurance/VAT)을 0으로 항상 표기,
    # 그 외 문서는 값이 있는 항목만 표기한다.
    force_charges = doc_type == "proforma_invoice"
    terms = data.get("terms", {}) or {}

    def _extra(key: str) -> float:
        try:
            return float(terms.get(key) or 0)
        except (TypeError, ValueError):
            return 0.0

    freight = _extra("freight")
    packing = _extra("packing")
    insurance = _extra("insurance")
    extras = freight + packing + insurance
    # VAT 율 정규화 — 프론트는 퍼센트(10), 이 모듈은 분수(0.1) 규약. >1 이면 퍼센트로 간주.
    rate = _num(data.get("vat_rate", 0))
    if rate > 1:
        rate /= 100.0
    taxable = subtotal - discount + extras
    vat = taxable * rate
    total = taxable + vat

    rows = [
        [_p("Subtotal", s["base"]), _p(_money(subtotal, currency), s["right"])],
    ]
    if base.get("discount_pct"):
        rows.append(
            [
                _p(f"Discount ({_num(base['discount_pct']):g}%)", s["base"]),
                _p(f"-{_money(discount, currency)}", s["right"]),
            ]
        )
    if force_charges or freight:
        rows.append([_p("Freight", s["base"]), _p(_money(freight, currency), s["right"])])
    if force_charges or packing:
        rows.append([_p("Packing", s["base"]), _p(_money(packing, currency), s["right"])])
    if force_charges or insurance:
        rows.append([_p("Insurance", s["base"]), _p(_money(insurance, currency), s["right"])])
    rows.append([_p("VAT", s["base"]), _p(_money(vat, currency), s["right"])])
    rows.append(
        [_p("Total", s["base"]), _p(f"<b>{_money(total, currency)}</b>", s["right"])]
    )
    table = Table(rows, colWidths=[35 * mm, 45 * mm])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, MID_GRAY),
                ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                ("BACKGROUND", (0, -1), (-1, -1), LIGHT_BLUE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return Table([["", table]], colWidths=[190 * mm, 80 * mm])


def _terms_block(data: Dict[str, Any], doc_type: str):
    s = _styles()
    terms = data.get("terms", {})
    company = data.get("company", {})
    shipping = data.get("shipping", {})

    rows = []
    if doc_type in {"quotation", "proforma_invoice", "commercial_invoice"}:
        rows.extend(
            [
                ["Payment Terms", terms.get("payment_terms", "")],
                ["Delivery Place", terms.get("delivery_place", "")],
                # 포장 방법(Carton Box 등) — 하단 합계의 Packing(포장비)과 다른 값.
                ["Packing", terms.get("packing_type") or terms.get("packing", "")],
                ["Warranty", terms.get("warranty", "")],
                ["Remarks", terms.get("remarks", "")],
            ]
        )
        if doc_type == "proforma_invoice":
            # 선적 정보 — 입력 화면(Proforma Invoice 탭)의 Shipping information 과 같은 이름.
            rows.extend(
                [
                    ["Vessel / IMO No.", _vessel_imo(shipping, data.get("vessel"))],
                    ["Carrier", shipping.get("carrier", "")],
                    ["Place of Departure", shipping.get("port_loading", "")],
                    ["Place of Destination", shipping.get("port_discharge", "")],
                    ["ETD", shipping.get("etd", "")],
                    ["ETA", shipping.get("eta", "")],
                    ["Country of Origin", shipping.get("sm_origin", "")],
                ]
            )
            rows.extend(
                [
                    ["Bank", company.get("bank_name", "")],
                    ["Account", company.get("bank_account", "")],
                    ["Account Holder", company.get("bank_holder", "")],
                    ["SWIFT", company.get("swift", "")],
                ]
            )
    if doc_type == "vendor_rfq":
        rows.extend(
            [
                ["Requested Incoterms", terms.get("incoterms", "") or "CNF Busan port"],
                ["Instructions", "Please provide Unit Price, Lead Time and Country of Origin, then return this sheet to sales@k-maris.com."],
                ["Remarks", terms.get("remarks", "")],
            ]
        )
    if doc_type in {"packing_list", "shipping_advice"}:
        rows.extend(
            [
                ["Place of Departure", shipping.get("port_loading", "")],
                ["Place of Destination", shipping.get("port_discharge", "")],
                ["Carrier", shipping.get("carrier", "")],
                ["B/L or AWB No.", shipping.get("bl_awb_no", "")],
                ["ETD", shipping.get("etd", "")],
                ["ETA", shipping.get("eta", "")],
                ["Shipping Mark", shipping.get("shipping_marks", "")],
            ]
        )

    if not rows:
        return Spacer(1, 1)

    table = Table([[_p("<b>Terms / Instructions</b>", s["section"]), ""]] + [[_p(a, s["small"]), _p(b, s["small"])] for a, b in rows], colWidths=[42 * mm, 228 * mm])
    table.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (-1, 0)),
                ("BACKGROUND", (0, 0), (-1, 0), BLUE),
                ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
                ("BACKGROUND", (0, 1), (0, -1), LIGHT_GRAY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def doc_parties(data: Dict[str, Any]) -> tuple[Dict[str, str], Dict[str, str]]:
    """문서에 인쇄할 (수하인 CONSIGNEE, 매수인 BUYER).

    수하인과 매수인은 다른 회사일 수 있어 문서(shipping)에 따로 담는다.
      · 매수인 — buyer_* 를 적어 넣었으면 그 값, 비어 있으면 고객 마스터 값.
      · 수하인 — 회사명(sm_consignee)이 있으면 consignee_* 값, 없으면 매수인과 같다.
    """
    customer = data.get("customer", {}) or {}
    shipping = data.get("shipping", {}) or {}

    def pick(key: str, fallback: str) -> str:
        return (shipping.get(key) or "").strip() or fallback

    buyer = {
        "name": pick("buyer_name", customer.get("name", "") or ""),
        "address": pick("buyer_address", customer.get("address", "") or ""),
        "contact": pick("buyer_contact", customer.get("contact", "") or ""),
        "email": pick("buyer_email", customer.get("email", "") or ""),
    }
    if not (shipping.get("sm_consignee") or "").strip():
        return dict(buyer), buyer
    consignee = {
        "name": (shipping.get("sm_consignee") or "").strip(),
        "address": (shipping.get("consignee_address") or "").strip(),
        "contact": (shipping.get("consignee_contact") or "").strip(),
        "email": (shipping.get("consignee_email") or "").strip(),
    }
    return consignee, buyer


def party_lines(party: Dict[str, str]) -> list[str]:
    """당사자 블록 3줄 — 회사명 / 주소 / 담당자·이메일(문서 서식과 같은 구성)."""
    contact = "    ".join(x for x in (party.get("contact", ""), party.get("email", "")) if x)
    return [party.get("name", ""), party.get("address", ""),
            f"Contact: {contact}" if contact else ""]


def _vessel_imo(shipping: Dict[str, Any], vessel: Dict[str, Any] | None) -> str:
    """문서의 'Vessel / IMO No.' 칸 — 문서에 적어 넣은 선명이 우선, 없으면 선박 마스터."""
    vessel = vessel or {}
    name = (shipping.get("sm_vessel") or vessel.get("name", "") or "").strip()
    imo = (vessel.get("imo", "") or "").strip()
    return " / ".join(x for x in (name, imo) if x)


def _commercial_shipping_block(data: Dict[str, Any]):
    """Render the CI footer in the same section order and field grouping as the XLSX."""
    s = _styles()
    shipping = data.get("shipping", {})
    terms = data.get("terms", {})

    dim = " x ".join(str(shipping.get(k, "") or "-") for k in ("sm_dim_l", "sm_dim_w", "sm_dim_h"))
    marks = (shipping.get("shipping_marks") or "").strip()
    if not marks:
        mark_lines = []
        for value in (
            shipping.get("sm_type"),
            f"C/O {shipping.get('sm_consignee')}" if shipping.get("sm_consignee") else "",
            f"M/V {shipping.get('sm_vessel')}" if shipping.get("sm_vessel") else "",
            f"P.O. NO.: {shipping.get('sm_po_no')}" if shipping.get("sm_po_no") else "",
            f"REF. NO.: {shipping.get('sm_ref_no')}" if shipping.get("sm_ref_no") else "",
            shipping.get("sm_desc"),
            f"CASE NO.: {shipping.get('sm_case_no')}" if shipping.get("sm_case_no") else "",
        ):
            if value:
                mark_lines.append(str(value))
        marks = "\n".join(mark_lines)

    rows = [
        [_p("<b>SHIPPING INFORMATION</b>", s["section"]), "", _p("<b>SHIPPING INFORMATION</b>", s["section"]), ""],
        [_p("Vessel / IMO No.", s["small"]), _p(_vessel_imo(shipping, data.get("vessel")), s["small"]), _p("Carrier", s["small"]), _p(shipping.get("carrier", ""), s["small"])],
        [_p("Place of Departure", s["small"]), _p(shipping.get("port_loading", ""), s["small"]), _p("Place of Destination", s["small"]), _p(shipping.get("port_discharge", ""), s["small"])],
        [_p("B/L or AWB No.", s["small"]), _p(shipping.get("bl_awb_no", ""), s["small"]), _p("ETD / ETA", s["small"]), _p(f"{shipping.get('etd', '')} / {shipping.get('eta', '')}", s["small"])],
        [_p("Incoterms", s["small"]), _p(terms.get("incoterms", ""), s["small"]), _p("Payment Terms", s["small"]), _p(terms.get("payment_terms", ""), s["small"])],
        [_p("Packing", s["small"]), _p(terms.get("packing_type", ""), s["small"]), _p("Country of Origin", s["small"]), _p(shipping.get("sm_origin", ""), s["small"])],
        [_p("<b>SHIPPING MARKS</b>", s["section"]), "", "", ""],
        [_p(marks, s["small"]), "", "", ""],
        [_p("<b>PACKING &amp; DECLARATION</b>", s["section"]), "", "", ""],
        [_p("Total Packages", s["small"]), _p(shipping.get("sm_total_cases", ""), s["small"]), _p("N.W. / G.W. (kg)", s["small"]), _p(f"{shipping.get('sm_net_weight', '')} / {shipping.get('sm_gross_weight', '')}", s["small"])],
        [_p("Dimension (mm)", s["small"]), _p(dim if dim != "- x - x -" else "", s["small"]), _p("Country of Origin", s["small"]), _p(shipping.get("sm_origin", ""), s["small"])],
    ]
    # 행 인덱스: 0 머리 / 1-5 선적정보 / 6 마크 머리 / 7 마크 / 8 포장·선언 머리 / 9-10 포장.
    table = Table(rows, colWidths=[38 * mm, 97 * mm, 38 * mm, 97 * mm])
    table.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)), ("SPAN", (2, 0), (3, 0)),
        ("SPAN", (0, 6), (3, 6)), ("SPAN", (0, 7), (3, 7)), ("SPAN", (0, 8), (3, 8)),
        ("BACKGROUND", (0, 0), (-1, 0), BLUE), ("BACKGROUND", (0, 6), (-1, 6), BLUE),
        ("BACKGROUND", (0, 8), (-1, 8), BLUE),
        ("BACKGROUND", (0, 1), (0, 5), LIGHT_GRAY), ("BACKGROUND", (2, 1), (2, 5), LIGHT_GRAY),
        ("BACKGROUND", (0, 9), (0, 10), LIGHT_GRAY), ("BACKGROUND", (2, 9), (2, 10), LIGHT_GRAY),
        ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return table


def _footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(MID_GRAY)
    canvas.setLineWidth(0.5)
    canvas.line(12 * mm, 11 * mm, 285 * mm, 11 * mm)
    canvas.setFont(DEFAULT_FONT, 7)
    canvas.setFillColor(DARK_GRAY)
    canvas.drawString(12 * mm, 7 * mm, "K-MARIS Energy & Solutions Co., Ltd.")
    canvas.drawRightString(285 * mm, 7 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _make_commercial_invoice_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """Commercial Invoice PDF — templates/'commercial invoice_sample' 서식 그대로.

    구성: 레터헤드 → 문서정보(우측 반쪽) → CONSIGNEE | BUYER(if different) →
    SHIPPING INFORMATION → 품목표 → 합계 → DECLARATION → 서명·직인.
    Proforma Invoice 와 같은 격자(_DocForm)를 쓰고, 당사자 칸이 둘이라는 점과
    은행 정보가 없다는 점만 다르다.
    """
    s = _styles()
    shipping = data.get("shipping", {}) or {}
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "USD").upper()
    consignee, buyer = doc_parties(data)
    money = pi_charges(data)
    dec = pi_decimals(currency)

    def amount(value: Any) -> str:
        v = _num(value)
        return "-" if abs(v) < 0.5 / (10 ** dec) else f"{v:,.{dec}f}"

    buffer, doc, page_width = _form_doc("COMMERCIAL INVOICE")
    form = _DocForm(PI_COLUMN_UNITS, page_width, s)
    quad = form.cols(2, 2, 2, 2)
    item_w = form.cols(1, 2, 1, 1, 1, 1, 1)
    half = form.cols(4, 4)

    story: List[Any] = []
    story += _letterhead(company, "COMMERCIAL INVOICE", s)
    story += [form.doc_info([("Invoice No.", data.get("doc_no", "")),
                             ("Invoice Date", pi_doc_date(data.get("date", ""))),
                             ("PO No.", shipping.get("po_no", ""))], quad), Spacer(1, 3 * mm)]

    # 수하인·매수인은 다른 회사일 수 있어 두 칸으로 나눠 인쇄한다.
    story += [form.band("CONSIGNEE", "BUYER (if different)", widths=form.cols(4, 4)),
              form.pairs([("Company Name", consignee.get("name", ""), "Company Name", buyer.get("name", "")),
                          ("Address", consignee.get("address", ""), "Address", buyer.get("address", "")),
                          ("Contact", consignee.get("contact", ""), "Contact", buyer.get("contact", "")),
                          ("e-mail", consignee.get("email", ""), "e-mail", buyer.get("email", ""))], quad)]

    story += [form.band("SHIPPING INFORMATION"),
              form.pairs(_invoice_shipping_rows(data, "Incoterms® 2020")
                         + [("Packing", (data.get("terms", {}) or {}).get("packing_type", ""),
                             "Country of Origin", shipping.get("sm_origin", "")),
                            ("Currency", currency, "", "")], quad),
              Spacer(1, 3 * mm)]

    story.append(_invoice_item_table(form, data, items, currency, item_w, amount))
    story += _invoice_totals_tables(form, item_w, money, amount)
    story += [Spacer(1, 3 * mm), form.band("DECLARATION"),
              form.declaration("We hereby certify that this Commercial Invoice is true and correct."),
              Spacer(1, 2 * mm)]
    sign_img, stamp_img = _form_signature_images()
    story += [form.signature([("Authorized Signature", sign_img, 40 * mm),
                              (f"{company.get('company_name_en', '')}\n(Company Stamp)", stamp_img, 22 * mm)], half),
              Spacer(1, 2 * mm), _footer_center(s)]

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


PI_COLUMN_UNITS = [6, 14.5, 19.296875, 13.5, 12, 8, 14, 18.296875]
"""Proforma Invoice 8열 격자 — templates/'proforma invoice_sample.xlsx' 의 열 너비.
PDF(_make_proforma_invoice_pdf)와 Excel(doc_xlsx.make_proforma_invoice_xlsx)이 같은
격자를 쓰기 때문에 두 파일의 칸 위치가 정확히 겹친다. 여기만 고치면 둘 다 바뀐다."""

PI_MIN_ITEM_ROWS = 5
"""품목이 적어도 표를 이 행수만큼 빈 줄로 채운다(참조 양식과 같은 인상)."""

PL_COLUMN_UNITS = [6, 13.1, 18.6, 13, 5.9, 6, 14.6, 8.1, 8.1, 4.4, 4.4, 4.4, 10.7]
"""Packing List 13열 격자 — templates/'packing list_sample.xlsx' 의 열 너비.
A=No. / B+C=Description / D=Part No. / E=Q'ty / F=Unit / G=No.&Kind of Packages /
H=N.W. / I=G.W. / J·K·L=Dim.(L·W·H) / M=Meas. — PDF·Excel 이 함께 쓴다."""

PL_MIN_ITEM_ROWS = 10


def _dim_parts(text: Any) -> List[str]:
    """품목 치수 문자열('30 x 23 x 21')을 L·W·H 세 칸으로 나눈다.
    세 조각으로 갈라지지 않으면(자유 형식) 세 칸 모두 비운다 — 서식이 어긋나느니 비우는 편이 낫다."""
    parts = [x.strip() for x in re.split(r"[x×*]", str(text or "")) if x.strip()]
    return parts if len(parts) == 3 else ["", "", ""]


def pi_decimals(currency: str) -> int:
    """송장 금액 소수 자리 — 소수 단위가 없는 통화(원·엔)만 정수로 찍는다."""
    return 0 if (currency or "").upper() in {"KRW", "JPY"} else 2


def pi_charges(data: Dict[str, Any]) -> Dict[str, float]:
    """Proforma Invoice 합계 — 입력 화면(PI 탭)과 같은 계산.
    Total invoice value = 품목 소계 + Freight + Packing + Insurance + VAT,
    VAT 는 부대비용까지 더한 금액에 매긴다."""
    terms = data.get("terms", {}) or {}

    def charge(key: str) -> float:
        try:
            return float(terms.get(key) or 0)
        except (TypeError, ValueError):
            return 0.0

    subtotal = sum(_num(it.get("amount", 0)) for it in normalize_items(data.get("items", [])))
    freight, packing, insurance = charge("freight"), charge("packing"), charge("insurance")
    rate = _num(data.get("vat_rate", 0))
    if rate > 1:  # 프론트는 퍼센트(10), 이 모듈은 분수(0.1) 규약.
        rate /= 100.0
    taxable = subtotal + freight + packing + insurance
    vat = taxable * rate
    return {
        "subtotal": subtotal, "freight": freight, "packing": packing,
        "insurance": insurance, "vat_rate": rate, "vat": vat, "total": taxable + vat,
    }


def pi_doc_date(value: Any) -> str:
    """송장 일자 표기 — 참조 양식과 같은 '05-Aug-2026'. 날짜가 아니면 원문 그대로."""
    text = str(value or "").strip()
    try:
        return date.fromisoformat(text[:10]).strftime("%d-%b-%Y")
    except ValueError:
        return text


class _DocForm:
    """송장 계열 문서(Proforma Invoice · Commercial Invoice · Packing List) 공통 서식 조립기.

    templates/*_sample 세 문서는 같은 격자·머리띠·라벨표·서명 블록을 쓴다. units 는 짝이 되는
    Excel 시트의 열 너비 목록이라, 같은 units 를 쓰는 Excel 과 PDF 의 칸 위치가 정확히 겹친다.
    폭 계산·색·여백이 전부 여기 한 곳에 있어, 서식을 손보면 세 문서가 함께 따라온다.
    """

    CELL = [("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]
    BAND = colors.HexColor("#1F3B66")
    ALT_ROW = colors.HexColor("#FAFBFC")

    def __init__(self, units: List[float], page_width: float, s: Dict[str, ParagraphStyle]):
        self.s = s
        self.page_width = page_width
        self.w = [page_width * u / sum(units) for u in units]
        self.label = ParagraphStyle("KMFormLabel", parent=s["small"], fontName=DEFAULT_BOLD_FONT)
        self.th = ParagraphStyle("KMFormTh", parent=s["small"], fontName=DEFAULT_BOLD_FONT,
                                 textColor=colors.white, alignment=TA_CENTER)
        self.center = ParagraphStyle("KMFormCenter", parent=s["base"], alignment=TA_CENTER)
        self.right = ParagraphStyle("KMFormRight", parent=s["base"], alignment=TA_RIGHT)
        self.grand = ParagraphStyle("KMFormGrand", parent=s["base"], fontName=DEFAULT_BOLD_FONT,
                                    alignment=TA_CENTER)
        self.grand_right = ParagraphStyle("KMFormGrandR", parent=self.right, fontName=DEFAULT_BOLD_FONT)

    def cols(self, *spans: int) -> List[float]:
        """열 묶음 크기 → 폭 목록. cols(2, 2, 2, 2) 는 [A+B, C+D, E+F, G+H] 를 준다."""
        out, i = [], 0
        for n in spans:
            out.append(sum(self.w[i:i + n]))
            i += n
        return out

    def p(self, value: Any, style: Any = "small") -> Paragraph:
        return _p(value, style if isinstance(style, ParagraphStyle) else self.s[style])

    def band(self, *titles: str, widths: Optional[List[float]] = None) -> Table:
        """섹션 머리띠 — 제목 하나면 전폭, 둘이면 좌우로 나눈다(CONSIGNEE | BUYER)."""
        t = Table([[self.p(x, "section") for x in titles]], colWidths=widths or [self.page_width])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), self.BAND),
                               ("GRID", (0, 0), (-1, -1), .35, MID_GRAY)] + self.CELL))
        return t

    def pairs(self, rows: List[tuple], widths: List[float], value_style: Any = "small") -> Table:
        """'라벨 | 값 | 라벨 | 값' 표 — 라벨 칸은 회색 볼드."""
        t = Table([[self.p(a, self.label), self.p(b, value_style),
                    self.p(c, self.label), self.p(d, value_style)] for a, b, c, d in rows],
                  colWidths=widths)
        t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY),
                               ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                               ("BACKGROUND", (2, 0), (2, -1), LIGHT_GRAY)] + self.CELL))
        return t

    def doc_info(self, rows: List[tuple], widths: List[float]) -> Table:
        """문서 정보(송장번호·일자·PO No.) — 참조 양식처럼 오른쪽 반쪽에만 그린다."""
        t = Table([["", "", self.p(k, self.label), self.p(v)] for k, v in rows], colWidths=widths)
        t.setStyle(TableStyle([("GRID", (2, 0), (-1, -1), .35, MID_GRAY),
                               ("BACKGROUND", (2, 0), (2, -1), LIGHT_GRAY)] + self.CELL))
        return t

    def declaration(self, text: str) -> Table:
        t = Table([[self.p(text)]], colWidths=[self.page_width])
        t.setStyle(TableStyle([("LINEABOVE", (0, 0), (-1, -1), .35, MID_GRAY),
                               ("LINEBELOW", (0, 0), (-1, -1), .35, MID_GRAY)] + self.CELL))
        return t

    def signature(self, cells: List[tuple], widths: List[float], height: float = 20 * mm) -> Table:
        """서명 칸들 — (라벨, 이미지, 이미지칸 폭) 목록. 이미지는 라벨 오른쪽에 놓는다."""
        inner_style = TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                                  ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                                  ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)])
        row = []
        for (text, img, img_w), width in zip(cells, widths):
            img_w = min(img_w, width * 0.5)
            inner = Table([[self.p(text, self.label), img]], colWidths=[width - img_w, img_w])
            inner.setStyle(inner_style)
            row.append(inner)
        t = Table([row], colWidths=widths, rowHeights=[height])
        t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY),
                               ("VALIGN", (0, 0), (-1, -1), "TOP"),
                               ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                               ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
        return t

    def item_style(self, row_count: int, extra: Optional[List[tuple]] = None) -> TableStyle:
        """품목표 공통 스타일 — 남색 머리행 + 격자 + 짝수행 옅은 바탕."""
        cmds = [("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("GRID", (0, 0), (-1, -1), .35, MID_GRAY)] + self.CELL
        for i in range(2, row_count, 2):
            cmds.append(("BACKGROUND", (0, i), (-1, i), self.ALT_ROW))
        return TableStyle(cmds + (extra or []))


def _form_signature_images() -> tuple:
    """서명·직인 이미지(없으면 빈 칸) — 세 문서가 같은 크기로 쓴다."""
    sign = _pdf_image(_pdf_asset("Authorized signature_Sungyeon Cho.jpg", "signature.png", "signature.jpg"),
                      35 * mm, 11 * mm)
    stamp = _pdf_image(_pdf_asset("Company stamp_K-Maris Energy & Solutions.jpg", "stamp.png", "stamp.jpg"),
                       16 * mm, 16 * mm)
    return sign, stamp


def _form_doc(title: str) -> tuple:
    """송장 계열 PDF 의 공통 페이지 설정(세로 A4·10mm 여백)."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=7 * mm, bottomMargin=12 * mm, title=title,
                            author="K-MARIS Energy & Solutions Co., Ltd.")
    return buffer, doc, 190 * mm


def _invoice_shipping_rows(data: Dict[str, Any], incoterms_label: str = "Incoterms") -> List[tuple]:
    """SHIPPING INFORMATION 블록의 공통 앞 네 줄 — 이름·순서는 입력 화면(7단계 각 탭)과 같다.
    마지막 줄(Packing·Currency·Country of Origin)은 문서마다 달라 각 문서가 덧붙인다."""
    shipping = data.get("shipping", {}) or {}
    terms = data.get("terms", {}) or {}
    return [
        ("Vessel / IMO No.", _vessel_imo(shipping, data.get("vessel")), "Carrier", shipping.get("carrier", "")),
        ("Place of Departure", shipping.get("port_loading", ""),
         "Place of Destination", shipping.get("port_discharge", "")),
        ("ETD", pi_doc_date(shipping.get("etd", "")), "ETA", pi_doc_date(shipping.get("eta", ""))),
        (incoterms_label, terms.get("incoterms", ""), "Payment Terms", terms.get("payment_terms", "")),
    ]


def _invoice_totals_tables(form: _DocForm, item_w: List[float], money: Dict[str, float],
                           amount) -> List[Any]:
    """합계 블록 — 라벨은 Unit Price 칸, 값은 Amount 칸. PI·CI 가 같이 쓴다."""
    rows = [("Subtotal", money["subtotal"]), ("Freight", money["freight"]),
            ("Packing", money["packing"]), ("Insurance", money["insurance"]), ("VAT", money["vat"])]
    totals = Table([["", form.p(k, form.label), form.p(amount(v), form.right)] for k, v in rows],
                   colWidths=[sum(item_w[:5]), item_w[5], item_w[6]])
    totals.setStyle(TableStyle([("GRID", (1, 0), (-1, -1), .35, MID_GRAY),
                                ("LINEBEFORE", (0, 0), (0, -1), .35, MID_GRAY),
                                ("LINEAFTER", (0, 0), (0, -1), .35, MID_GRAY),
                                ("BACKGROUND", (1, 0), (1, -1), LIGHT_GRAY)] + form.CELL))
    grand = Table([[form.p("TOTAL INVOICE VALUE", form.grand),
                    form.p(amount(money["total"]), form.grand_right)]],
                  colWidths=[sum(item_w[:6]), item_w[6]])
    grand.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY),
                               ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BLUE)] + form.CELL))
    return [totals, grand]


def _invoice_item_table(form: _DocForm, data: Dict[str, Any], items: List[Dict[str, Any]],
                        currency: str, item_w: List[float], amount) -> Table:
    """PI·CI 품목표 — No./Description/Part No./HS Code/Qty/Unit Price/Amount 7열.
    품목이 적으면 참조 양식처럼 빈 줄로 채워 표 높이를 지킨다."""
    shipping = data.get("shipping", {}) or {}
    headers = ["No.", "Description", "Part No.", "HS Code", "Qty", "Unit Price", f"Amount ({currency})"]
    rows: List[List[Any]] = [[form.p(h, form.th) for h in headers]]
    for it in items:
        rows.append([form.p(it["item_no"], form.center), form.p(it["description"], "base"),
                     form.p(it["part_no"], "base"),
                     form.p(it.get("hs_code") or shipping.get("hs_code", ""), form.center),
                     form.p(f"{it['qty']:g}", form.center), form.p(amount(it["unit_price"]), form.right),
                     form.p(amount(it["amount"]), form.right)])
    filler = max(0, PI_MIN_ITEM_ROWS - len(items))
    rows += [[""] * 7 for _ in range(filler)]
    table = Table(rows, colWidths=item_w, repeatRows=1,
                  rowHeights=[None] * (len(items) + 1) + [16] * filler)
    table.setStyle(form.item_style(len(rows)))
    return table


def _bank_block(form: _DocForm, company: Dict[str, Any], currency: str, quad: List[float]) -> List[Any]:
    """BANK INFORMATION — 통화에 맞는 계좌(외화/원화)를 Settings 에서 가져온다."""
    foreign = currency != "KRW"
    holder = (company.get("fx_bank_holder") if foreign else company.get("bank_holder")) or company.get("company_name_en", "")
    bank = (company.get("fx_bank_name") if foreign else company.get("bank_name")) or ""
    account = (company.get("fx_bank_account") if foreign else company.get("bank_account")) or ""
    rows = [("Remittee's name", holder, "Bank Name & Address", bank),
            ("Swift Code", company.get("swift", ""), "Remittee's Account No.", account)]
    return [form.band("BANK INFORMATION"), form.pairs(rows, quad, value_style=form.center)]


SERVICE_PI_DECIMALS = 0
"""서비스 Proforma Invoice 금액 자리수 — 출장비·기술료는 원 단위라 소수점을 쓰지 않는다.
PDF·Excel 이 같은 값을 봐야 두 파일의 숫자가 어긋나지 않는다."""


def is_service_doc(data: Dict[str, Any]) -> bool:
    """서비스 딜의 문서인지 — 발행 서식이 물품용과 갈린다(선적정보 대신 서비스정보)."""
    return (data.get("doc_variant") or "") == "service"


def service_info_rows(data: Dict[str, Any]) -> List[tuple]:
    """SERVICE INFORMATION 블록 — templates/'proforma invoice_sample_service.xlsx' 의 다섯 줄.
    라벨·순서는 입력 화면(서비스 7단계 Proforma Invoice 탭)과 같다."""
    sh = data.get("shipping", {}) or {}
    terms = data.get("terms", {}) or {}
    return [
        ("Vessel / IMO No.", _vessel_imo(sh, data.get("vessel")),
         "Service Description", sh.get("service_description", "")),
        ("Service Location", sh.get("service_location", ""), "Man Power", sh.get("man_power", "")),
        ("Service Date", sh.get("service_date", ""), "Estimated Duration", sh.get("duration", "")),
        ("Vessel Schedule", sh.get("vessel_schedule", ""), "Local Agent", sh.get("local_agent", "")),
        ("Currency", (data.get("currency") or "USD").upper(),
         "Payment Terms", terms.get("payment_terms", "")),
    ]


def _make_service_proforma_invoice_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """서비스 딜의 Proforma Invoice PDF — templates/'proforma invoice_sample_service' 서식.

    물품용과 다른 점만 셋이다: SHIPPING 대신 SERVICE INFORMATION, 품목표에 품번·HS 코드가
    없고(용역엔 없는 칸이다) 대신 Unit 이 서고, 합계는 TOTAL INVOICE VALUE 한 줄뿐이다.
    레터헤드·BUYER·은행정보·서명은 물품용과 같은 블록을 그대로 쓴다."""
    s = _styles()
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "USD").upper()
    shipping = data.get("shipping", {}) or {}
    _, buyer = doc_parties(data)
    money = pi_charges(data)
    dec = SERVICE_PI_DECIMALS

    def amount(value: Any) -> str:
        v = _num(value)
        return "-" if abs(v) < 0.5 / (10 ** dec) else f"{v:,.{dec}f}"

    buffer, doc, page_width = _form_doc("PROFORMA INVOICE")
    form = _DocForm(PI_COLUMN_UNITS, page_width, s)
    quad = form.cols(2, 2, 2, 2)                       # 라벨·값 4칸
    item_w = form.cols(1, 3, 1, 1, 1, 1)               # No./Description/Qty/Unit/Unit Price/Amount
    half = form.cols(4, 4)                             # 서명 2칸

    story: List[Any] = []
    story += _letterhead(company, "PROFORMA INVOICE", s)
    story += [form.doc_info([("Invoice No.", data.get("doc_no", "")),
                             ("Invoice Date", pi_doc_date(data.get("date", ""))),
                             ("PO No.", shipping.get("po_no", ""))], quad), Spacer(1, 3 * mm)]
    story += [form.band("BUYER"),
              form.pairs([("Company Name", buyer.get("name", ""), "Address", buyer.get("address", "")),
                          ("Contact", buyer.get("contact", ""), "e-mail", buyer.get("email", ""))], quad),
              Spacer(1, 3 * mm)]
    story += [form.band("SERVICE INFORMATION"),
              form.pairs(service_info_rows(data), quad), Spacer(1, 3 * mm)]

    headers = ["No.", "Description", "Qty", "Unit", "Unit Price", f"Amount ({currency})"]
    rows: List[List[Any]] = [[form.p(h, form.th) for h in headers]]
    for it in items:
        rows.append([form.p(it["item_no"], form.center), form.p(it["description"], "base"),
                     form.p(f"{it['qty']:g}", form.center), form.p(it.get("unit", ""), form.center),
                     form.p(amount(it["unit_price"]), form.right),
                     form.p(amount(it["amount"]), form.right)])
    filler = max(0, PI_MIN_ITEM_ROWS - len(items))
    rows += [[""] * 6 for _ in range(filler)]
    table = Table(rows, colWidths=item_w, repeatRows=1,
                  rowHeights=[None] * (len(items) + 1) + [16] * filler)
    table.setStyle(form.item_style(len(rows)))
    story.append(table)

    grand = Table([[form.p("TOTAL INVOICE VALUE", form.grand),
                    form.p(amount(money["total"]), form.grand_right)]],
                  colWidths=[sum(item_w[:5]), item_w[5]])
    grand.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY),
                               ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BLUE)] + form.CELL))
    story.append(grand)

    story += [Spacer(1, 3 * mm)] + _bank_block(form, company, currency, quad)
    story += [form.band("DECLARATION"),
              form.declaration("We hereby certify that this Proforma Invoice is true and correct."),
              Spacer(1, 2 * mm)]
    sign_img, stamp_img = _form_signature_images()
    story += [form.signature([("Authorized Signature", sign_img, 40 * mm),
                              (f"{company.get('company_name_en', '')}\n(Company Stamp)", stamp_img, 22 * mm)], half),
              Spacer(1, 2 * mm), _footer_center(s)]

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _make_proforma_invoice_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """Proforma Invoice PDF — templates/'proforma invoice_sample' 서식 그대로.

    구성: 레터헤드 → 문서정보(우측 반쪽) → BUYER → SHIPPING INFORMATION → 품목표 →
    합계(Subtotal/Freight/Packing/Insurance/VAT/TOTAL INVOICE VALUE) → BANK INFORMATION →
    DECLARATION → 서명·직인.
    """
    s = _styles()
    shipping = data.get("shipping", {}) or {}
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "USD").upper()
    _, buyer = doc_parties(data)
    money = pi_charges(data)
    dec = pi_decimals(currency)

    def amount(value: Any) -> str:
        """0 은 참조 양식처럼 '-' 로 — 부대비용이 없는 줄을 0.00 으로 채우지 않는다."""
        v = _num(value)
        return "-" if abs(v) < 0.5 / (10 ** dec) else f"{v:,.{dec}f}"

    buffer, doc, page_width = _form_doc("PROFORMA INVOICE")
    form = _DocForm(PI_COLUMN_UNITS, page_width, s)
    quad = form.cols(2, 2, 2, 2)                       # 라벨·값 4칸
    item_w = form.cols(1, 2, 1, 1, 1, 1, 1)            # 품목표 7칸
    half = form.cols(4, 4)                             # 서명 2칸

    story: List[Any] = []
    story += _letterhead(company, "PROFORMA INVOICE", s)
    story += [form.doc_info([("Invoice No.", data.get("doc_no", "")),
                             ("Invoice Date", pi_doc_date(data.get("date", ""))),
                             ("PO No.", shipping.get("po_no", ""))], quad), Spacer(1, 3 * mm)]
    story += [form.band("BUYER"),
              form.pairs([("Company Name", buyer.get("name", ""), "Address", buyer.get("address", "")),
                          ("Contact", buyer.get("contact", ""), "e-mail", buyer.get("email", ""))], quad),
              Spacer(1, 3 * mm)]
    story += [form.band("SHIPPING INFORMATION"),
              form.pairs(_invoice_shipping_rows(data)
                         + [("Currency", currency, "Country of Origin", shipping.get("sm_origin", ""))], quad),
              Spacer(1, 3 * mm)]
    story.append(_invoice_item_table(form, data, items, currency, item_w, amount))
    story += _invoice_totals_tables(form, item_w, money, amount)
    story += [Spacer(1, 3 * mm)] + _bank_block(form, company, currency, quad)
    story += [form.band("DECLARATION"),
              form.declaration("We hereby certify that this Proforma Invoice is true and correct."),
              Spacer(1, 2 * mm)]
    sign_img, stamp_img = _form_signature_images()
    story += [form.signature([("Authorized Signature", sign_img, 40 * mm),
                              (f"{company.get('company_name_en', '')}\n(Company Stamp)", stamp_img, 22 * mm)], half),
              Spacer(1, 2 * mm), _footer_center(s)]

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _kor_won_amount(value: Any) -> str:
    """숫자 금액 → 한글 정식 표기 '일금...원정' (예: 471438 → 일금사십칠만일천사백삼십팔원정)."""
    n = int(round(_num(value)))
    if n == 0:
        return "일금영원정"
    neg = n < 0
    n = abs(n)
    digits = "영일이삼사오육칠팔구"
    small_units = ["", "십", "백", "천"]
    big_units = ["", "만", "억", "조", "경"]
    groups: List[int] = []
    while n > 0:
        groups.append(n % 10000)
        n //= 10000
    parts: List[str] = []
    for gi in range(len(groups) - 1, -1, -1):
        g = groups[gi]
        if g == 0:
            continue
        gstr = ""
        for pos in range(3, -1, -1):
            d = (g // (10 ** pos)) % 10
            if d:
                gstr += digits[d] + small_units[pos]
        parts.append(gstr + big_units[gi])
    body = "".join(parts)
    return f"일금{'마이너스' if neg else ''}{body}원정"


def _make_tax_invoice_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """세금계산서(대금청구서) 성격의 TAX INVOICE PDF — 국내(KRW) 청구서 양식.
    상단 정보 · BILL TO/SUPPLIER · 한글 총액 · 품목표 · 소계/VAT/합계 · 은행정보 구성."""
    s = _styles()
    customer = data.get("customer", {}) or {}
    vessel = data.get("vessel", {}) or {}
    shipping = data.get("shipping", {}) or {}
    tax = data.get("tax_invoice", {}) or {}
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "KRW").upper()
    totals = calc_totals(items, _num(data.get("vat_rate", 0)))
    # 부대비용(Freight/Packing/Insurance) — PI/CI 와 같이 terms 에 담겨 온다.
    # 소계에 더한 뒤 VAT 를 매기고, 값이 있는 항목만 합계표에 표기한다.
    charge_terms = data.get("terms", {}) or {}

    def _charge(key: str) -> float:
        try:
            return float(charge_terms.get(key) or 0)
        except (TypeError, ValueError):
            return 0.0

    freight, packing, insurance = _charge("freight"), _charge("packing"), _charge("insurance")
    extras = freight + packing + insurance
    if extras:
        rate = _num(data.get("vat_rate", 0))
        if rate > 1:
            rate /= 100.0
        taxable = totals["subtotal"] + extras
        totals = {**totals, "vat": taxable * rate, "total": taxable * (1 + rate)}

    buffer = io.BytesIO()
    page_width = 190 * mm
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=10 * mm,
                            rightMargin=10 * mm, topMargin=7 * mm, bottomMargin=12 * mm,
                            title="TAX INVOICE", author="K-MARIS Energy & Solutions Co., Ltd.")

    def p(value, style="small"):
        return _p(value, s[style])

    def section(title, width=page_width):
        t = Table([[p(title, "section")]], colWidths=[width])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1F3B66")),
                               ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                               ("FONTNAME", (0, 0), (-1, -1), DEFAULT_BOLD_FONT),
                               ("LEFTPADDING", (0, 0), (-1, -1), 4),
                               ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        return t

    story = []
    story += _letterhead(company, "TAX INVOICE", s)

    # ── 상단 정보(송장번호·일자·통화·PO) ──
    half = page_width / 2
    lab = page_width * 0.16
    val = half - lab
    info_rows = [
        [p("Invoice No.", "small"), p(data.get("doc_no", ""), "small"),
         p("PO / Ref. No.", "small"), p(shipping.get("po_no", ""), "small")],
        [p("Invoice Date", "small"), p(data.get("date", ""), "small"),
         p("Currency", "small"), p(currency, "small")],
        [p("Due Date", "small"), p(tax.get("due_date", ""), "small"), p("", "small"), p("", "small")],
    ]
    info = Table(info_rows, colWidths=[lab, val, lab, val])
    info.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                              ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY), ("BACKGROUND", (2, 0), (2, -1), LIGHT_GRAY),
                              ("FONTNAME", (0, 0), (0, -1), DEFAULT_BOLD_FONT), ("FONTNAME", (2, 0), (2, -1), DEFAULT_BOLD_FONT),
                              ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                              ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
    story += [info, Spacer(1, 3 * mm)]

    # ── BILL TO / CUSTOMER + SUPPLIER ──
    address = company.get("address_en") or company.get("address", "")
    biz = "Trading / Marine Spare Parts, Equipment & Services"
    rep = company.get("representative", "") or "Sungyeon Cho"
    l = [("Company Name", customer.get("name", "")),
         ("Business Reg. No.", customer.get("tax_id", "")),
         ("Contact", customer.get("contact", "")),
         ("Email", customer.get("email", "")), ("Tel", customer.get("phone", "") or customer.get("tel", "")),
         ("Vessel", vessel.get("name", "")), ("Project", data.get("project_title", ""))]
    r = [("Company Name", company.get("company_name_en", "")), ("Address", address),
         ("Business Reg. No.", company.get("business_no", "")), ("Business", biz),
         ("Representative", rep), ("Tel", company.get("phone", "")), ("", "")]
    party_rows = [[p("BILL TO / CUSTOMER", "section"), "", p("SUPPLIER", "section"), ""]]
    for i in range(len(l)):
        party_rows.append([p(l[i][0], "small"), p(l[i][1], "small"), p(r[i][0], "small"), p(r[i][1], "small")])
    party = Table(party_rows, colWidths=[lab, val, lab, val])
    party.setStyle(TableStyle([("SPAN", (0, 0), (1, 0)), ("SPAN", (2, 0), (3, 0)),
                               ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3B66")),
                               ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                               ("FONTNAME", (0, 0), (-1, 0), DEFAULT_BOLD_FONT),
                               ("BACKGROUND", (0, 1), (0, -1), LIGHT_GRAY), ("BACKGROUND", (2, 1), (2, -1), LIGHT_GRAY),
                               ("FONTNAME", (0, 1), (0, -1), DEFAULT_BOLD_FONT), ("FONTNAME", (2, 1), (2, -1), DEFAULT_BOLD_FONT),
                               ("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                               ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                               ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
    story += [party, Spacer(1, 3 * mm)]

    # ── 총액(VAT 포함) 한글 표기 배너 ──
    total_style = ParagraphStyle("KMTaxTotal", parent=s["base"], fontName=DEFAULT_BOLD_FONT, fontSize=11, leading=14)
    banner = Table([[p("TOTAL VALUE (VAT included)", "section"),
                     _p(_kor_won_amount(totals["total"]), total_style),
                     _p(f"₩{totals['total']:,.0f}", total_style)]],
                   colWidths=[page_width * 0.30, page_width * 0.45, page_width * 0.25])
    banner.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#1F3B66")),
                                ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
                                ("BACKGROUND", (1, 0), (-1, 0), LIGHT_BLUE),
                                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                                ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story += [banner]

    # ── 품목표 ──
    headers = ["No.", "Description", "Part No.", "Qty", "Unit Price", f"Amount ({currency})"]
    col_w = [page_width * w for w in (0.06, 0.44, 0.14, 0.08, 0.14, 0.14)]
    item_rows = [[p(h, "th") for h in headers]]
    for it in items:
        item_rows.append([p(it["item_no"], "tiny"), p(it["description"], "tiny"), p(it["part_no"], "tiny"),
                          p(f"{it['qty']:g}", "tiny"), p(f"{it['unit_price']:,.0f}", "tiny"),
                          p(f"{it['amount']:,.0f}", "tiny")])
    item_table = Table(item_rows, colWidths=col_w, repeatRows=1)
    cmds = [("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"), ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3), ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]
    for i in range(2, len(item_rows), 2):
        cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FAFBFC")))
    item_table.setStyle(TableStyle(cmds))
    story += [item_table]

    # ── 소계 / VAT / 합계 (우측 정렬) ──
    tot_lines = [("Subtotal", totals["subtotal"])]
    for lab_, v in (("Freight", freight), ("Packing", packing), ("Insurance", insurance)):
        if v:
            tot_lines.append((lab_, v))
    tot_lines += [("VAT", totals["vat"]), ("TOTAL INVOICE VALUE", totals["total"])]
    tot_rows = [[p(lab_, "small"), _p(f"{v:,.0f}", ParagraphStyle('r', parent=s['small'], alignment=TA_RIGHT))]
                for lab_, v in tot_lines]
    tot_inner = Table(tot_rows, colWidths=[col_w[3] + col_w[4], col_w[5]])
    tot_inner.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY),
                                   ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                                   ("BACKGROUND", (0, -1), (-1, -1), LIGHT_BLUE),
                                   ("FONTNAME", (0, -1), (-1, -1), DEFAULT_BOLD_FONT),
                                   ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                                   ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                                   ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
    tot_wrap = Table([["", tot_inner]], colWidths=[sum(col_w[:3]), col_w[3] + col_w[4] + col_w[5]])
    tot_wrap.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                                  ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                                  ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story += [tot_wrap, Spacer(1, 4 * mm)]

    # ── 은행 정보 ──
    story += [section("BANK INFORMATION")]
    bank_rows = [
        [p("Remittee's name", "small"), p(company.get("company_name_kr", ""), "small"),
         p("Bank Name & Address", "small"), p(company.get("bank_name", ""), "small")],
        [p("Currency", "small"), p(currency, "small"),
         p("Remittee's Account No.", "small"), p(company.get("bank_account", ""), "small")],
        [p("Remarks", "small"), p(data.get("remarks", "") or tax.get("remarks", ""), "small"), "", ""],
    ]
    bank = Table(bank_rows, colWidths=[lab, val, lab, val])
    bank.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                              ("SPAN", (1, 2), (3, 2)),
                              ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY), ("BACKGROUND", (2, 0), (2, 1), LIGHT_GRAY),
                              ("FONTNAME", (0, 0), (0, -1), DEFAULT_BOLD_FONT), ("FONTNAME", (2, 0), (2, 1), DEFAULT_BOLD_FONT),
                              ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                              ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
    story += [bank, Spacer(1, 3 * mm), _footer_center(s)]

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _make_credit_note_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """CREDIT NOTE — 클레임으로 고객 청구서를 깎아 준 사실을 한 장으로 증명하는 문서.

    청구서(TAX INVOICE)와 같은 서식을 쓰되 품목표가 없다. 깎아 준 것은 물건이 아니라
    금액이고, 이 문서가 답해야 하는 질문은 셋뿐이다 — 어느 청구서를, 얼마나, 왜.

    통화가 두 개일 수 있다: 현장 비용은 USD 로 났는데 깎아 줄 청구서는 KRW 인 식이다.
    그래서 발행 금액·적용 환율·청구서 통화 상계액을 나란히 적는다. 하나만 적으면 나중에
    "이 금액이 어디서 나왔나"를 되짚을 수 없다.
    """
    s = _styles()
    customer = data.get("customer", {}) or {}
    vessel = data.get("vessel", {}) or {}
    claim = data.get("claim", {}) or {}
    cur = (data.get("currency") or "USD").upper()          # 발행 통화
    inv_cur = (data.get("invoice_currency") or cur).upper()  # 상계 대상 청구서 통화
    amount = _num(data.get("amount"))
    applied = _num(data.get("applied_amount")) or amount
    fx = _num(data.get("fx_rate")) or 1.0
    vat = _num(data.get("vat_amount"))
    dec = 0 if inv_cur == "KRW" else 2
    buffer, doc, page_width = _form_doc("CREDIT NOTE")
    form = _DocForm([1] * 8, page_width, s)
    quad = form.cols(2, 2, 2, 2)
    half = form.cols(4, 4)

    def p(v, style="small"):
        return _p(v, s[style])

    story = []
    story += _letterhead(company, "CREDIT NOTE", s)
    story += [form.pairs([
        ("Credit Note No.", data.get("doc_no", ""), "Issue Date", pi_doc_date(data.get("date", ""))),
        ("Against Invoice No.", data.get("invoice_no", ""), "Invoice Date", pi_doc_date(data.get("invoice_date", ""))),
        ("Customer P/O", data.get("po_no", ""), "Project No.", data.get("export_ref", "")),
    ], quad), Spacer(1, 3 * mm)]

    # ── 받는 쪽 / 보내는 쪽 ──
    story += [form.band("TO / CUSTOMER", "FROM / SUPPLIER", widths=half)]
    left = [("Company Name", customer.get("name", "")),
            ("Business Reg. No.", customer.get("tax_id", "")),
            ("Contact", customer.get("contact", "")),
            ("Vessel", vessel.get("name", ""))]
    right = [("Company Name", company.get("company_name_en", "")),
             ("Business Reg. No.", company.get("business_no", "")),
             ("Representative", company.get("representative", "") or "Sungyeon Cho"),
             ("Tel", company.get("phone", ""))]
    story += [form.pairs([(left[i][0], left[i][1], right[i][0], right[i][1]) for i in range(len(left))], quad),
              Spacer(1, 3 * mm)]

    # ── 금액 배너 — 청구서에서 실제로 빠지는 금액이 이 장의 결론이다. ──
    banner_style = ParagraphStyle("KMCNTotal", parent=s["base"], fontName=DEFAULT_BOLD_FONT,
                                  fontSize=11, leading=14)
    words = _kor_won_amount(applied) if inv_cur == "KRW" else ""
    banner = Table([[p("CREDIT AMOUNT", "section"), _p(words, banner_style),
                     _p(f"{inv_cur} {applied:,.{dec}f}",
                        ParagraphStyle("KMCNTotalR", parent=banner_style, alignment=TA_RIGHT))]],
                   colWidths=[page_width * 0.30, page_width * 0.45, page_width * 0.25])
    banner.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY),
                                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#1F3B66")),
                                ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
                                ("BACKGROUND", (1, 0), (-1, 0), LIGHT_BLUE),
                                ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story += [banner, Spacer(1, 3 * mm)]

    # ── 산출 근거 — 발행 통화 금액 → 환율 → 상계액. 환율은 통화가 다를 때만 적는다. ──
    rows = [("Reason", data.get("reason", "") or "Claim settlement", "Claim", claim.get("title", ""))]
    if cur != inv_cur:
        rows.append(("Issued Amount", f"{cur} {amount:,.2f}",
                     "Exchange Rate", f"1 {cur} = {fx:,.2f} {inv_cur}"))
    rows.append(("Applied to Invoice", f"{inv_cur} {applied:,.{dec}f}",
                 "of which VAT", f"{inv_cur} {vat:,.{dec}f}" if vat else "—"))
    rows.append(("Site · Vessel", claim.get("site", ""), "Occurred", pi_doc_date(claim.get("occurred_date", ""))))
    story += [form.band("DETAILS"), form.pairs(rows, quad), Spacer(1, 3 * mm)]

    story += [form.declaration(
        "This credit note is applied against the invoice stated above; the outstanding balance is "
        "reduced by the credit amount. No separate remittance is required for the credited portion."
    ), Spacer(1, 2 * mm)]
    sign_img, stamp_img = _form_signature_images()
    story += [form.signature([("Issued by · K-MARIS Energy & Solutions Co., Ltd.", sign_img, 40 * mm),
                              ("Company Seal", stamp_img, 20 * mm)], half),
              Spacer(1, 3 * mm), _footer_center(s)]

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def consignee_mark_lines(sh: Dict[str, Any]) -> List[str]:
    """마크의 수하인 블록 — 'C/O 회사명' 다음에 그 회사 주소를 줄마다 이어 붙인다.
    주소는 Commercial Invoice 의 CONSIGNEE 주소(consignee_address)와 같은 칸이라
    한쪽에서 고치면 다른 쪽도 따라간다. 화물이 어디로 가는지가 마크의 핵심이라
    회사명만 찍고 주소를 빼면 현장에서 인도처를 알 수 없다."""
    name = str(sh.get("sm_consignee") or "").strip()
    if not name:
        return []
    lines = [f"C/O {name}"]
    lines += [x.strip() for x in str(sh.get("consignee_address") or "").splitlines() if x.strip()]
    return lines


def compose_shipping_marks(sh: Dict[str, Any]) -> str:
    """구조화 Shipping Mark(sm_*)를 여러 줄 문자열로 재구성 — 프론트 composeShippingMarks·
    doc_xlsx._compose_marks 와 동일 규약. PL 은 CI 상속값+PL 수정값이 병합된 sh 를 받아
    저장된 shipping_marks 문자열 대신 항상 최신 sm_* 로 재구성한다."""
    lines: List[str] = []

    def push(v):
        if v and str(v).strip():
            lines.append(str(v).strip())

    push(sh.get("sm_type"))
    for line in consignee_mark_lines(sh):
        push(line)
    if sh.get("sm_vessel"): push(f"M/V {str(sh['sm_vessel']).upper()}")
    if sh.get("sm_po_no"): push(f"P.O. NO.: {sh['sm_po_no']}")
    if sh.get("sm_ref_no"): push(f"REF. NO.: {sh['sm_ref_no']}")
    push(sh.get("sm_desc"))
    if sh.get("sm_case_no"): push(f"CASE NO.: {sh['sm_case_no']}")
    if sh.get("sm_total_cases"): push(f"TOTAL: {sh['sm_total_cases']} CASE(S)")
    if sh.get("sm_net_weight"): push(f"N.W.: {sh['sm_net_weight']} KG")
    if sh.get("sm_gross_weight"): push(f"G.W.: {sh['sm_gross_weight']} KG")
    dim = [sh.get("sm_dim_l"), sh.get("sm_dim_w"), sh.get("sm_dim_h")]
    if any(d and str(d).strip() for d in dim):
        push("DIM.: " + " × ".join((str(d).strip() if d and str(d).strip() else "-") for d in dim) + " CM")
    if sh.get("sm_final_dest"): push(f"FINAL DESTINATION: {sh['sm_final_dest']}")
    push(sh.get("sm_origin"))
    push(sh.get("sm_handling"))
    return "\n".join(lines)


def _make_shipping_mark_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """Shipping Mark(케이스 마킹) PDF — 수출 화물에 스텐실로 찍는 통상적인 마킹 라벨 양식.
    가운데 큰 마크 박스(주 마크) + 하단 실측(중량·치수·케이스) + 취급주의 라인."""
    s = _styles()
    vessel = data.get("vessel", {}) or {}
    shipping = data.get("shipping", {}) or {}
    buffer = io.BytesIO()
    page_width = 190 * mm
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=10 * mm,
                            rightMargin=10 * mm, topMargin=7 * mm, bottomMargin=12 * mm,
                            title="SHIPPING MARK")

    asset_roots = [Path(__file__).resolve().parents[2], Path(__file__).resolve().parent.parent / "config"]

    def asset(*names):
        for root in asset_roots:
            for name in names:
                candidate = root / name
                if candidate.exists():
                    return candidate
        return None

    def image(path, max_width, max_height):
        if not path:
            return ""
        from PIL import Image as PILImage
        with PILImage.open(path) as source:
            width, height = source.size
        scale = min(max_width / width, max_height / height)
        return Image(str(path), width=width * scale, height=height * scale)

    def p(value, style="small"):
        text = str(value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(text.replace("\n", "<br/>"), s[style])

    story = []
    # ── 헤더: 공통 레터헤드 + 중앙 타이틀 ────────────────────────────────
    story += _letterhead(company, "SHIPPING MARK", s)

    # ── 참조 정보 스트립 ─────────────────────────────────────────────────
    ref_rows = [[p("REF. NO.", "section"), p(shipping.get("sm_ref_no", "")),
                 p("P.O. NO.", "section"), p(shipping.get("sm_po_no", "")),
                 p("DATE", "section"), p(data.get("date", ""))]]
    ref = Table(ref_rows, colWidths=[page_width * w for w in (.13, .27, .12, .22, .10, .16)])
    ref.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                             ("BACKGROUND", (0, 0), (0, -1), NAVY), ("BACKGROUND", (2, 0), (2, -1), NAVY),
                             ("BACKGROUND", (4, 0), (4, -1), NAVY),
                             ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                             ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story += [ref, Spacer(1, 5 * mm)]

    # ── 주 마크 박스 — 스텐실 스타일(가운데 정렬, 볼드) ────────────────────
    mark_style = ParagraphStyle("KMMark", parent=s["base"], fontName=DEFAULT_BOLD_FONT,
                                fontSize=13, leading=20, alignment=TA_CENTER, textColor=colors.black)
    head_style = ParagraphStyle("KMMarkHead", parent=mark_style, fontSize=15, leading=22)
    desc_style = ParagraphStyle("KMMarkDesc", parent=mark_style, fontSize=13, leading=20)

    def mline(text, style=mark_style):
        safe = str(text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(safe, style)

    flow = []
    if shipping.get("sm_type"):
        flow += [mline(str(shipping["sm_type"]).upper(), head_style), Spacer(1, 3 * mm)]
    for line in consignee_mark_lines(shipping):
        flow.append(mline(line))
    if shipping.get("sm_vessel"):
        flow.append(mline(f"M/V {str(shipping['sm_vessel']).upper()}"))
    elif vessel.get("name"):
        flow.append(mline(f"M/V {str(vessel['name']).upper()}"))
    flow.append(Spacer(1, 2 * mm))
    if shipping.get("sm_po_no"):
        flow.append(mline(f"P.O. NO. : {shipping['sm_po_no']}"))
    if shipping.get("sm_ref_no"):
        flow.append(mline(f"REF. NO. : {shipping['sm_ref_no']}"))
    if shipping.get("sm_desc"):
        flow += [Spacer(1, 2 * mm), mline(str(shipping["sm_desc"]).upper(), desc_style)]
    if shipping.get("sm_final_dest"):
        flow.append(Spacer(1, 2 * mm))
    if shipping.get("sm_final_dest"):
        flow.append(mline(f"FINAL DESTINATION : {str(shipping['sm_final_dest']).upper()}"))
    flow.append(Spacer(1, 2 * mm))
    if shipping.get("sm_case_no"):
        flow.append(mline(f"CASE NO. : {shipping['sm_case_no']}"))
    if shipping.get("sm_origin"):
        flow.append(mline(str(shipping["sm_origin"]).upper()))
    if not flow:
        flow.append(mline("(NO SHIPPING MARK DATA)"))

    # 마크 박스는 92mm 를 기본 높이로 두되, 줄이 많으면(수하인 주소 등) 내용에 맞춰 늘린다.
    # 높이를 92mm 로 고정하면 긴 마크의 첫 줄·끝 줄이 테두리 밖으로 잘려 나간다.
    box_style = TableStyle([("BOX", (0, 0), (-1, -1), 1.4, colors.black),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                            ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)])
    mark_box = Table([[flow]], colWidths=[page_width])
    mark_box.setStyle(box_style)
    if mark_box.wrap(page_width, 10 ** 6)[1] < 92 * mm:
        mark_box = Table([[flow]], colWidths=[page_width], rowHeights=[92 * mm])
        mark_box.setStyle(box_style)
    story += [mark_box, Spacer(1, 5 * mm)]

    # ── 실측(중량·치수·케이스) 표 ────────────────────────────────────────
    dim = [shipping.get("sm_dim_l"), shipping.get("sm_dim_w"), shipping.get("sm_dim_h")]
    dim_txt = " × ".join((str(d).strip() if d and str(d).strip() else "-") for d in dim) + " CM" if any(d and str(d).strip() for d in dim) else ""
    m_rows = [[p("N.W.", "section"), p(f"{shipping.get('sm_net_weight', '')} KG" if shipping.get("sm_net_weight") else ""),
               p("G.W.", "section"), p(f"{shipping.get('sm_gross_weight', '')} KG" if shipping.get("sm_gross_weight") else "")],
              [p("DIMENSION", "section"), p(dim_txt),
               p("TOTAL CASES", "section"), p(f"{shipping.get('sm_total_cases', '')} CASE(S)" if shipping.get("sm_total_cases") else "")]]
    metrics = Table(m_rows, colWidths=[page_width * w for w in (.16, .34, .16, .34)])
    metrics.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                 ("BACKGROUND", (0, 0), (0, -1), NAVY), ("BACKGROUND", (2, 0), (2, -1), NAVY),
                                 ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                                 ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.append(metrics)

    # ── 취급 주의 라인 ───────────────────────────────────────────────────
    handling = str(shipping.get("sm_handling") or "").strip()
    if handling:
        parts = " · ".join(h.strip() for h in handling.split(",") if h.strip())
        h_rows = [[p("HANDLING", "section"), p(parts)]]
        h_tbl = Table(h_rows, colWidths=[page_width * .16, page_width * .84])
        h_tbl.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                   ("BACKGROUND", (0, 0), (0, -1), NAVY),
                                   ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                                   ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
        story += [Spacer(1, 2 * mm), h_tbl]

    story.append(Spacer(1, 2 * mm))
    story.append(_footer_center(s))
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _pkg_text(it: Dict[str, Any]) -> str:
    """'No. & Kind of Packages' 셀 — 수량+종류 결합, 없으면 레거시 package 문자열."""
    q = str(it.get("pkg_qty") or "").strip()
    k = str(it.get("pkg_kind") or "").strip()
    combined = f"{q} {k}".strip()
    return combined or str(it.get("package") or "").strip()


def _g(v: float) -> str:
    """합계 숫자 표기 — 0/빈값은 빈 칸으로 둔다(서식에 0 이 줄줄이 찍히지 않게)."""
    return f"{v:g}" if v else ""


def packing_totals(data: Dict[str, Any]) -> Dict[str, str]:
    """Packing List 합계행 값 — 화면에서 직접 적은 "전체 포장 규격"이 있으면 그 값을 쓰고,
    없으면 지금까지처럼 품목별 입력을 합산한다.

    여러 품목이 상자 하나에 함께 들어가는 선적은 무게·치수를 품목별로 나눌 수 없어 전체를
    한 번만 적는다. 그 값은 Shipping Marks 와 같은 칸(shipping.sm_*)에 저장되므로 케이스 수·
    중량·치수가 마크와 늘 같은 값이 된다. 포장 종류는 따로 적지 않으면 거래조건의 Packing 을 쓴다.
    반환값은 서식에 그대로 넣을 문자열이며, 빈 문자열이면 그 칸은 비워 둔다."""
    shipping = data.get("shipping") or {}
    terms = data.get("terms") or {}
    items = normalize_items(data.get("items", []))

    def sm(key: str) -> str:
        return str(shipping.get(key, "") or "").strip()

    # 포장 개수를 적었을 때만 "1 Carton Box" 처럼 종류를 붙인다 — 개수 없이 종류만 있으면
    # 품목별 합계가 곧 개수이고, 품목마다 종류가 다를 수 있어 숫자만 적는다(기존 동작).
    cases = sm("sm_total_cases")
    kind = sm("sm_pkg_kind") or str(terms.get("packing_type", "") or "").strip()
    packages = f"{cases} {kind}".strip() if cases else _g(sum(_num(it.get("pkg_qty")) for it in items))
    net = sm("sm_net_weight") or _g(sum(_num(it.get("net_weight")) for it in items))
    gross = sm("sm_gross_weight") or _g(sum(_num(it.get("gross_weight")) for it in items))
    dims = [sm("sm_dim_l"), sm("sm_dim_w"), sm("sm_dim_h")]
    measurement = sm("sm_measurement")
    if not measurement and all(dims):
        # 치수는 cm 로 적는다 — 용적(m³)은 세 변의 곱을 100³ 으로 나눈 값.
        measurement = _g(round(_num(dims[0]) * _num(dims[1]) * _num(dims[2]) / 1_000_000, 4))
    if not measurement:
        measurement = _g(sum(_num(it.get("measurement")) for it in items))
    return {
        "packages": packages,
        "net_weight": net,
        "gross_weight": gross,
        "measurement": measurement,
        "dimension": " × ".join(dims) if all(dims) else "",
    }


def _make_packing_list_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """Packing List PDF — templates/'packing list_sample' 서식 그대로.

    구성: 레터헤드 → 문서정보(우측 반쪽) → CONSIGNEE | BUYER(if different) →
    SHIPPING INFORMATION → 품목표(가격 없음·포장/중량/치수/용적) → TOTAL 행 →
    DECLARATION → 서명·직인·수령 확인. 송장(PI/CI)과 같은 조립기(_DocForm)를 쓰되
    열 격자만 Packing List 전용(PL_COLUMN_UNITS, 13열)이다.
    """
    s = _styles()
    shipping = data.get("shipping", {}) or {}
    terms = data.get("terms", {}) or {}
    items = normalize_items(data.get("items", []))

    buffer, doc, page_width = _form_doc("PACKING LIST")
    form = _DocForm(PL_COLUMN_UNITS, page_width, s)
    quad = form.cols(2, 3, 2, 6)                                  # 라벨·값 4칸
    item_w = form.cols(1, 2, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1)        # 품목표 12칸
    signs = form.cols(3, 4, 6)                                    # 서명 3칸
    consignee, buyer = doc_parties(data)

    def num(v: Any) -> str:
        """0·빈값은 빈 칸 — 서식에 0 이 줄줄이 찍히지 않게."""
        return str(v).strip() if v not in (None, "", 0, 0.0) else ""

    story: List[Any] = []
    story += _letterhead(company, "PACKING LIST", s)
    # Packing List 는 자체 번호·발행일이 없다 — 딸려 나가는 송장의 번호·발행일을 싣는다.
    story += [form.doc_info([("Invoice No.", shipping.get("ci_no", "") or data.get("doc_no", "")),
                             ("Invoice Date", pi_doc_date(shipping.get("ci_date", "") or data.get("date", ""))),
                             ("PO No.", shipping.get("po_no", ""))], quad), Spacer(1, 3 * mm)]
    story += [form.band("CONSIGNEE", "BUYER (if different)", widths=form.cols(5, 8)),
              form.pairs([("Company Name", consignee.get("name", ""), "Company Name", buyer.get("name", "")),
                          ("Address", consignee.get("address", ""), "Address", buyer.get("address", "")),
                          ("Contact", consignee.get("contact", ""), "Contact", buyer.get("contact", "")),
                          ("e-mail", consignee.get("email", ""), "e-mail", buyer.get("email", ""))], quad)]
    story += [form.band("SHIPPING INFORMATION"),
              form.pairs(_invoice_shipping_rows(data)
                         + [("Packing", terms.get("packing_type", ""),
                             "Country of Origin", shipping.get("sm_origin", ""))], quad),
              Spacer(1, 3 * mm)]

    # ── 품목 표(가격 없음) — 머리행 2줄, Dim.(cm) 아래 L/W/H 세 칸 ───────────
    head1 = [form.p(h, form.th) for h in
             ["No.", "Description", "Part No.", "Q'ty", "Unit", "No. & Kind of Packages",
              "N.W. (kg)", "G.W. (kg)", "Dim. (cm)", "", "", "Meas. (m³)"]]
    head2 = ["", "", "", "", "", "", "", ""] + [form.p(x, form.th) for x in ("L", "W", "H")] + [""]
    rows: List[List[Any]] = [head1, head2]
    for it in items:
        dims = _dim_parts(it.get("dimension"))
        rows.append([form.p(it["item_no"], form.center), form.p(it["description"], "base"),
                     form.p(it["part_no"], "base"), form.p(f"{it['qty']:g}", form.center),
                     form.p(it["unit"], form.center), form.p(_pkg_text(it), form.center),
                     form.p(num(it.get("net_weight")), form.right), form.p(num(it.get("gross_weight")), form.right)]
                    + [form.p(d, form.center) for d in dims]
                    + [form.p(num(it.get("measurement")), form.right)])
    filler = max(0, PL_MIN_ITEM_ROWS - len(items))
    rows += [[""] * 12 for _ in range(filler)]

    # 합계행 — 전체 포장 규격을 직접 적었으면 그 값, 아니면 품목별 합산(packing_totals).
    tot = packing_totals(data)
    dims = [num(shipping.get("sm_dim_l")), num(shipping.get("sm_dim_w")), num(shipping.get("sm_dim_h"))]
    unit = items[0]["unit"] if items else ""
    total_row = [form.p("TOTAL", form.grand), "", "",
                 form.p(_g(sum(_num(it["qty"]) for it in items)), form.grand),
                 form.p(unit, form.grand), form.p(tot["packages"], form.grand_right),
                 form.p(tot["net_weight"], form.grand_right), form.p(tot["gross_weight"], form.grand_right)] \
        + [form.p(d, form.grand) for d in dims] + [form.p(tot["measurement"], form.grand_right)]
    rows.append(total_row)

    last = len(rows) - 1
    table = Table(rows, colWidths=item_w, repeatRows=2,
                  rowHeights=[None, None] + [None] * len(items) + [15] * filler + [None])
    table.setStyle(form.item_style(len(rows) - 1, extra=[
        ("BACKGROUND", (0, 0), (-1, 1), NAVY),
        ("SPAN", (0, 0), (0, 1)), ("SPAN", (1, 0), (1, 1)), ("SPAN", (2, 0), (2, 1)),
        ("SPAN", (3, 0), (3, 1)), ("SPAN", (4, 0), (4, 1)), ("SPAN", (5, 0), (5, 1)),
        ("SPAN", (6, 0), (6, 1)), ("SPAN", (7, 0), (7, 1)), ("SPAN", (8, 0), (10, 0)),
        ("SPAN", (11, 0), (11, 1)),
        ("BACKGROUND", (0, last), (-1, last), LIGHT_BLUE), ("SPAN", (0, last), (2, last)),
    ]))
    story.append(table)

    # ── Packing Information(자유 메모) — 적었을 때만 붙인다 ────────────────
    packing_info = (data.get("packing_info") or "").strip()
    if packing_info:
        note = Table([[form.p(packing_info)]], colWidths=[page_width])
        note.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .35, MID_GRAY),
                                  ("VALIGN", (0, 0), (-1, -1), "TOP")] + form.CELL))
        story += [Spacer(1, 3 * mm), form.band("PACKING INFORMATION"), note]

    story += [Spacer(1, 3 * mm), form.band("DECLARATION"),
              form.declaration("We hereby certify that this Packing List is true and correct."),
              Spacer(1, 2 * mm)]
    sign_img, stamp_img = _form_signature_images()
    story += [form.signature([("Authorized Signature", sign_img, 40 * mm),
                              (f"{company.get('company_name_en', '')}\n(Company Stamp)", stamp_img, 22 * mm),
                              ("Received by\n(Company Stamp & Date)", "", 0)], signs),
              Spacer(1, 2 * mm), _footer_center(s)]

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _qnum(value: Any) -> str:
    """견적서 금액 표기 — 천단위 구분, 정수면 소수 생략(첨부 양식과 동일)."""
    try:
        q = Decimal(str(value or 0))
    except Exception:
        return "0"
    if q == q.to_integral_value():
        return f"{int(q):,}"
    return f"{q:,.2f}"


def quotation_clause_catalog(terms: Dict[str, Any], validity_days: int = 30) -> List[tuple]:
    """QUOTATION / COSTING SHEET 표준 Terms & Conditions 문장 목록 [(id, 문장)].

    id 는 견적에 저장되는 선택값(terms["clauses"])의 계약이다 — 화면(4단계)에서 문장을
    골라 넣을 수 있게 웹의 QUOTATION_CLAUSES(web/lib/terms.ts)가 같은 id·같은 문장을
    들고 있으니, 문구를 고치면 양쪽을 함께 고쳐야 화면과 문서가 어긋나지 않는다.
    거래조건 필드(Incoterms·결제·보증)는 문장에 그때그때 끼워 넣어 항상 최신값으로 찍힌다."""
    incoterms = terms.get("incoterms") or "EXW (Ex Works)"
    place = terms.get("delivery_place") or "Busan, Republic of Korea"
    payment = terms.get("payment_terms") or "T/T in advance"
    warranty = terms.get("warranty") or "6 months from delivery"
    return [
        ("validity", f"Quotation validity: {validity_days} days from quotation date."),
        ("confirmation", "Price, availability, and delivery time are subject to final confirmation upon order placement."),
        ("delivery_term", f"Delivery term: {incoterms} {place}, Incoterms 2020."),
        ("charges_excluded", "Freight, customs duty, local tax, and other logistics charges are excluded unless otherwise stated."),
        ("payment_term", f"Payment term: {payment}"),
        ("buyer_check", "Buyer to confirm part number, description, quantity, engine type, and technical suitability before order."),
        ("certificates", "Certificates are excluded unless specifically stated."),
        ("cancellation", "Cancellation or return may not be accepted after order confirmation, especially for specially ordered or non-stock items."),
        ("warranty", f"Warranty follows {warranty}."),
        ("complete_order", "The unit price suggested is based on the complete order with complete quantities. In case of reduction for qty, "
                           "it may constitute a variation to the contract, subject to mutual agreement."),
    ]


def quotation_standard_terms(terms: Dict[str, Any], validity_days: int = 30) -> List[str]:
    """문서에 찍을 Terms & Conditions 문장들.

    terms["clauses"] 가 있으면 그 id 들만 골라 찍는다(4단계에서 선택한 문장). 값이 없는
    옛 견적은 선택 이력이 없다는 뜻이라 전부 찍는다 — 기존 견적서 내용이 바뀌지 않게.
    terms["extra_clauses"] 는 사용자가 직접 추가한 문장으로 표준 문장 뒤에 붙인다.
    Remarks 는 여기 섞지 않는다 — 품목표와 T&C 사이의 별도 Remark 섹션에서 찍는다."""
    catalog = quotation_clause_catalog(terms, validity_days)
    picked = terms.get("clauses")
    if isinstance(picked, list):
        chosen = set(str(c) for c in picked)
        lines = [text for cid, text in catalog if cid in chosen]
    else:
        lines = [text for _cid, text in catalog]
    for extra in (terms.get("extra_clauses") or []):
        if str(extra).strip():
            lines.append(str(extra).strip())
    return lines


def quotation_remark_lines(terms: Dict[str, Any]) -> List[str]:
    """Remark 섹션 본문 — 4단계 Remarks 입력을 줄 단위로 나눈 것(빈 줄 제외)."""
    raw = str(terms.get("remarks") or "").strip()
    return [line.strip() for line in raw.splitlines() if line.strip()]


def _make_quotation_costing_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """고객 견적서(QUOTATION / COSTING SHEET) — sales only. 첨부 양식을 따른다."""
    s = _styles()
    customer = data.get("customer", {}) or {}
    vessel = data.get("vessel", {}) or {}
    terms = data.get("terms", {}) or {}
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "USD").upper()
    total = sum(_num(it.get("amount", 0)) for it in items)

    page_width = 190 * mm
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=8 * mm, bottomMargin=14 * mm, title="QUOTATION / COSTING SHEET",
        author="K-MARIS Energy & Solutions Co., Ltd.",
    )
    asset_roots = [Path(__file__).resolve().parent.parent / "templates",
                   Path(__file__).resolve().parent.parent / "config",
                   Path(__file__).resolve().parents[2]]

    def asset(*names):
        # 이름 우선순위가 폴더보다 우선 — 앞선 이름(아이콘)이 있으면 텍스트 로고보다 먼저 선택.
        for name in names:
            for root in asset_roots:
                cand = root / name
                if cand.exists():
                    return cand
        return None

    def image(path, max_w, max_h):
        if not path:
            return ""
        try:
            from PIL import Image as PILImage
            with PILImage.open(path) as src:
                w, h = src.size
            scale = min(max_w / w, max_h / h)
            return Image(str(path), width=w * scale, height=h * scale)
        except Exception:
            return ""

    def section(title):
        t = Table([[_p(f"<b>{title}</b>", s["th"])]], colWidths=[page_width])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY), ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return t

    story: List[Any] = []

    # ── 헤더: 공통 레터헤드(로고 + 회사정보 + 슬로건) + 중앙 타이틀 ──────────
    story += _letterhead(company, "QUOTATION / COSTING SHEET", s)

    # ── 정보 박스(2단) ─────────────────────────────────────────────────
    left_rows = [
        ("User", customer.get("name", "")),
        ("Messrs", data.get("messrs", "")),
        ("Attn.", data.get("attn", "") or customer.get("contact", "")),
        ("Ship Name", vessel.get("name", "")),
        ("Project", data.get("project_title", "")),
    ]
    vat_label = "VAT excluded" if _num(data.get("vat_rate", 0)) == 0 else f"VAT {int(_num(data.get('vat_rate', 0)) * 100)}%"
    right_rows = [
        ("Quotation No.", data.get("doc_no", "")),
        ("Ref. No.", data.get("ref_no", "")),
        ("Date", data.get("date", "")),
        ("Currency", currency),
        ("VAT", vat_label),
    ]

    def meta_box(rows):
        body = [[_p(f"<b>{k}</b>", s["small"]), _p(v, s["small"])] for k, v in rows]
        t = Table(body, colWidths=[28 * mm, 65 * mm])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]))
        return t

    info = Table([[meta_box(left_rows), meta_box(right_rows)]], colWidths=[95 * mm, 95 * mm])
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story += [info, Spacer(1, 4 * mm)]

    # ── 품목표 (sales only) — Unit 열 포함, 샘플처럼 최소 6줄의 폼 형태 ──────
    headers = ["No.", "Part No.", "Description", "Qty", "Unit", "U/Price", "Amount", "Lead Time", "Remark"]
    widths = [8, 24, 50, 11, 12, 22, 22, 18, 23]
    rows = [[_p(h, s["th"]) for h in headers]]
    for it in items:
        rows.append([
            _p(it["item_no"], s["tiny"]),
            _p(it["part_no"], s["tiny"]),
            _p(it["description"], s["tiny"]),
            _p(_qnum(it["qty"]), s["tiny"]),
            _p(it.get("unit", ""), s["tiny"]),
            _p(_qnum(it["unit_price"]), s["tiny"]),
            _p(_qnum(it["amount"]), s["tiny"]),
            _p(it.get("lead_time", ""), s["tiny"]),
            _p(it.get("remark", ""), s["tiny"]),
        ])
    for _pad in range(max(0, 6 - len(items))):
        rows.append([_p("", s["tiny"]) for _ in range(9)])
    total_row = len(rows)
    rows.append([
        _p("<b>Total</b>", s["tiny"]), _p("", s["tiny"]), _p("", s["tiny"]),
        _p("", s["tiny"]), _p("", s["tiny"]), _p("", s["tiny"]),
        _p(f"<b>{_qnum(total)}</b>", s["tiny"]), _p("", s["tiny"]), _p("", s["tiny"]),
    ])
    items_table = Table(rows, colWidths=[w * mm for w in widths], repeatRows=1)
    tcmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"), ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ("ALIGN", (4, 1), (4, -1), "CENTER"), ("ALIGN", (5, 1), (6, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BACKGROUND", (0, total_row), (-1, total_row), LIGHT_BLUE),
        ("SPAN", (0, total_row), (5, total_row)),
        ("ALIGN", (0, total_row), (0, total_row), "CENTER"),
    ]
    for r in range(1, total_row):
        if r % 2 == 0:
            tcmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FAFBFC")))
    items_table.setStyle(TableStyle(tcmds))
    story += [items_table, Spacer(1, 5 * mm)]

    # ── Remark ────────────────────────────────────────────────────────
    # 4단계에서 입력한 견적 전체 비고. 품목표 바로 아래(=T&C 위)에 T&C 와 같은 형식으로
    # 둔다 — 이 견적에만 해당하는 이야기를 표준 약관보다 먼저 읽히게 하기 위함.
    remark_lines = quotation_remark_lines(terms)
    if remark_lines:
        story.append(section("Remark"))
        story.append(Spacer(1, 2 * mm))
        for line in remark_lines:
            story.append(_p(f"• {line}", s["small"]))
            story.append(Spacer(1, 1 * mm))
        story.append(Spacer(1, 3 * mm))

    # ── Terms & Conditions ────────────────────────────────────────────
    story.append(section("Terms & Conditions"))
    story.append(Spacer(1, 2 * mm))
    for line in quotation_standard_terms(terms):
        story.append(_p(f"• {line}", s["small"]))
        story.append(Spacer(1, 1 * mm))
    story.append(Spacer(1, 3 * mm))

    # ── Payment ───────────────────────────────────────────────────────
    story.append(section("Payment"))
    story.append(Spacer(1, 2 * mm))
    story.append(_p(f"• {terms.get('payment_terms') or 'T/T in advance'}", s["small"]))
    story.append(Spacer(1, 3 * mm))
    story.append(_p("We hope this quotation meets your requirement and to receive your order confirmation "
                    "at your earliest convenience.", s["base"]))
    story.append(Spacer(1, 6 * mm))

    # ── 서명 ──────────────────────────────────────────────────────────
    # 서명 이미지가 서명란(밑줄) 바로 위에 얹히도록: 이미지를 한 칸 표의 하단에
    # 정렬하고 그 칸의 아래 테두리를 서명선으로 쓴다(이름은 선 바로 아래).
    sign_img = image(asset("Authorized signature_Sungyeon Cho.jpg", "signature.png", "signature.jpg"), 40 * mm, 16 * mm)
    story.append(_p("Your sincerely", s["base"]))
    story.append(Spacer(1, 1 * mm))
    sig_line = Table([[sign_img or _p("", s["base"])]], colWidths=[62 * mm], rowHeights=[16 * mm])
    sig_line.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.7, colors.black),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    sig_line.hAlign = "LEFT"
    story.append(sig_line)
    story.append(_p("<b>Sam Cho, Managing Director</b>", s["base"]))
    # 하단 회사 푸터 — 공통(엑셀 푸터처럼 가운데 정렬).
    story.append(Spacer(1, 2 * mm))
    story.append(_footer_center(s))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _make_vendor_rfq_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """공급사 견적요청서(REQUEST FOR QUOTATION) — 4단계 견적서(QUOTATION / COSTING SHEET)·
    발주서(PURCHASE ORDER)와 같은 비주얼 시스템(세로 A4·로고 헤더·중앙 타이틀·남색 섹션바·
    지브라 품목표·서명)을 따른다. 단가·납기·원산지는 공급사가 채우도록 빈칸으로 둔다.
    payload['customer'] 에는 build_po_payload 가 넣은 Vendor(공급사) 정보가 들어 있다."""
    s = _styles()
    vendor = data.get("customer", {}) or {}   # build_po_payload: To(Vendor) = Vendor
    vessel = data.get("vessel", {}) or {}
    terms = data.get("terms", {}) or {}
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "USD").upper()

    page_width = 190 * mm
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=8 * mm, bottomMargin=14 * mm, title="REQUEST FOR QUOTATION",
        author="K-MARIS Energy & Solutions Co., Ltd.",
    )
    asset_roots = [Path(__file__).resolve().parent.parent / "templates",
                   Path(__file__).resolve().parent.parent / "config",
                   Path(__file__).resolve().parents[2]]

    def asset(*names):
        # 이름 우선순위가 폴더보다 우선 — 앞선 이름(아이콘)이 있으면 텍스트 로고보다 먼저 선택.
        for name in names:
            for root in asset_roots:
                cand = root / name
                if cand.exists():
                    return cand
        return None

    def image(path, max_w, max_h):
        if not path:
            return ""
        try:
            from PIL import Image as PILImage
            with PILImage.open(path) as src:
                w, h = src.size
            scale = min(max_w / w, max_h / h)
            return Image(str(path), width=w * scale, height=h * scale)
        except Exception:
            return ""

    def section(title):
        t = Table([[_p(f"<b>{title}</b>", s["th"])]], colWidths=[page_width])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY), ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return t

    story: List[Any] = []

    # ── 헤더: 공통 레터헤드 + 중앙 타이틀 ────────────────────────────────
    story += _letterhead(company, "REQUEST FOR QUOTATION", s)

    # ── 정보 박스(2단) — 발주서와 동일 스타일 ─────────────────────────────
    incoterms = terms.get("incoterms", "") or "CNF Busan port"
    left_rows = [
        ("To (Vendor)", vendor.get("name", "")),
        ("Address", vendor.get("address", "")),
        ("Contact", vendor.get("contact", "")),
        ("Email", vendor.get("email", "")),
        ("Ship Name", vessel.get("name", "")),
        ("Engine Type", vessel.get("engine_type", "")),
    ]
    right_rows = [
        ("RFQ No.", data.get("doc_no", "")),
        ("Date", data.get("date", "")),
        ("Currency", currency),
        ("IMO No.", vessel.get("imo", "")),
        ("Incoterms", incoterms),
        ("Reply to", "sales@k-maris.com"),
    ]

    def meta_box(rows):
        body = [[_p(f"<b>{k}</b>", s["small"]), _p(v, s["small"])] for k, v in rows]
        t = Table(body, colWidths=[28 * mm, 65 * mm])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]))
        return t

    info = Table([[meta_box(left_rows), meta_box(right_rows)]], colWidths=[95 * mm, 95 * mm])
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story += [info, Spacer(1, 4 * mm)]

    # ── 품목표 — 단가/납기/원산지는 공급사 기입용 빈칸 ─────────────────────
    headers = ["No.", "Part No.", "Description", "Maker", "Qty", "Unit",
               "Unit Price", "Lead Time", "Country of Origin", "Remark"]
    widths = [8, 22, 42, 22, 10, 12, 22, 16, 16, 20]
    rows = [[_p(h, s["th"]) for h in headers]]
    for it in items:
        rows.append([
            _p(it["item_no"], s["tiny"]),
            _p(it["part_no"], s["tiny"]),
            _p(it["description"], s["tiny"]),
            _p(it.get("maker", ""), s["tiny"]),
            _p(_qnum(it["qty"]), s["tiny"]),
            _p(it.get("unit", ""), s["tiny"]),
            _p("", s["tiny"]),   # Unit Price — 공급사 입력
            _p("", s["tiny"]),   # Lead Time — 공급사 입력
            _p("", s["tiny"]),   # Country of Origin — 공급사 입력
            _p(it.get("remark", ""), s["tiny"]),
        ])
    items_table = Table(rows, colWidths=[w * mm for w in widths], repeatRows=1)
    tcmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"), ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("ALIGN", (6, 1), (6, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for r in range(1, len(rows)):
        if r % 2 == 0:
            tcmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FAFBFC")))
    items_table.setStyle(TableStyle(tcmds))
    story += [items_table, Spacer(1, 5 * mm)]

    # ── Instructions(견적서의 Terms 자리) ─────────────────────────────────
    story.append(section("Instructions"))
    story.append(Spacer(1, 2 * mm))
    instructions = [
        "Please quote your best Unit Price, Lead Time (delivery) and Country of Origin for each item above.",
        f"Quotation currency: {currency}. Requested Incoterms: {incoterms}.",
        "Kindly advise the validity of your quotation and the minimum order quantity, if any.",
        "Please return the completed sheet to sales@k-maris.com at your earliest convenience.",
    ]
    if terms.get("remarks"):
        instructions.append(str(terms.get("remarks")))
    for line in instructions:
        story.append(_p(f"• {line}", s["small"]))
        story.append(Spacer(1, 1 * mm))
    story.append(Spacer(1, 4 * mm))

    story.append(_p("We look forward to receiving your best quotation. Thank you for your kind cooperation.", s["base"]))
    story.append(Spacer(1, 6 * mm))

    # ── 서명 ──────────────────────────────────────────────────────────
    sign_img = image(asset("Authorized signature_Sungyeon Cho.jpg", "signature.png", "signature.jpg"), 40 * mm, 16 * mm)
    story.append(_p("Best regards,", s["base"]))
    if sign_img:
        story.append(sign_img)
    story.append(_p("________________________", s["base"]))
    story.append(_p("<b>Sam Cho, Managing Director</b>", s["base"]))
    story.append(Spacer(1, 2 * mm))
    story.append(_footer_center(s))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _make_purchase_order_pdf(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    """공급사 발주서(PURCHASE ORDER) — 4단계 견적서(QUOTATION / COSTING SHEET)와 같은
    비주얼 시스템(세로 A4·로고 헤더·중앙 타이틀·남색 섹션바·지브라 품목표·서명)을 따른다.
    payload['customer'] 에는 build_po_payload 가 넣은 Vendor(공급사) 정보가 들어 있다."""
    s = _styles()
    vendor = data.get("customer", {}) or {}   # build_po_payload: Supplier/Seller = Vendor
    vessel = data.get("vessel", {}) or {}
    terms = data.get("terms", {}) or {}
    items = normalize_items(data.get("items", []))
    currency = (data.get("currency") or "USD").upper()
    total = sum(_num(it.get("amount", 0)) for it in items)

    page_width = 190 * mm
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=8 * mm, bottomMargin=14 * mm, title="PURCHASE ORDER",
        author="K-MARIS Energy & Solutions Co., Ltd.",
    )
    asset_roots = [Path(__file__).resolve().parent.parent / "templates",
                   Path(__file__).resolve().parent.parent / "config",
                   Path(__file__).resolve().parents[2]]

    def asset(*names):
        # 이름 우선순위가 폴더보다 우선 — 앞선 이름(아이콘)이 있으면 텍스트 로고보다 먼저 선택.
        for name in names:
            for root in asset_roots:
                cand = root / name
                if cand.exists():
                    return cand
        return None

    def image(path, max_w, max_h):
        if not path:
            return ""
        try:
            from PIL import Image as PILImage
            with PILImage.open(path) as src:
                w, h = src.size
            scale = min(max_w / w, max_h / h)
            return Image(str(path), width=w * scale, height=h * scale)
        except Exception:
            return ""

    def section(title):
        t = Table([[_p(f"<b>{title}</b>", s["th"])]], colWidths=[page_width])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY), ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return t

    story: List[Any] = []

    # ── 헤더: 공통 레터헤드 + 중앙 타이틀 ────────────────────────────────
    story += _letterhead(company, "PURCHASE ORDER", s)

    # ── 정보 박스(2단) ─────────────────────────────────────────────────
    incoterms = terms.get("incoterms", "")
    place = terms.get("delivery_place", "")
    incoterms_line = " · ".join([x for x in (incoterms, place) if x])
    # 발주서 양식(1. Order information / 2. Supplier information)과 같은 항목·순서.
    # 6단계 편집 화면의 입력칸이 이 두 상자를 그대로 채운다.
    left_rows = [
        ("Supplier / Seller", vendor.get("name", "")),
        ("Address", vendor.get("address", "")),
        ("Contact", vendor.get("contact", "")),
        ("Tel.", vendor.get("phone", "")),
        ("Email", vendor.get("email", "")),
        ("Ship Name", vessel.get("name", "")),
        ("IMO No.", vessel.get("imo", "")),
        ("Project", data.get("project_title", "") or ""),
    ]
    right_rows = [
        ("P/O No.", data.get("doc_no", "")),
        ("Date", data.get("date", "")),
        ("Quotation No.", terms.get("vendor_quote_no", "")),
        ("Requested Date", terms.get("requested_date", "")),
        ("Currency", currency),
        ("Incoterms", incoterms_line),
        ("Payment", terms.get("payment_terms", "")),
        ("Engine Type", vessel.get("engine_type", "")),
    ]

    def meta_box(rows):
        body = [[_p(f"<b>{k}</b>", s["small"]), _p(v, s["small"])] for k, v in rows]
        t = Table(body, colWidths=[28 * mm, 65 * mm])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]))
        return t

    info = Table([[meta_box(left_rows), meta_box(right_rows)]], colWidths=[95 * mm, 95 * mm])
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story += [info, Spacer(1, 4 * mm)]

    # ── 품목표 ─────────────────────────────────────────────────────────
    headers = ["No.", "Part No.", "Description", "Maker", "Qty", "Unit", "Unit Price", "Amount", "Lead Time / Remark"]
    widths = [8, 24, 44, 26, 12, 14, 22, 22, 18]
    rows = [[_p(h, s["th"]) for h in headers]]
    for it in items:
        lead_remark = f"{it.get('lead_time', '')}\n{it.get('remark', '')}".strip()
        rows.append([
            _p(it["item_no"], s["tiny"]),
            _p(it["part_no"], s["tiny"]),
            _p(it["description"], s["tiny"]),
            _p(it.get("maker", ""), s["tiny"]),
            _p(_qnum(it["qty"]), s["tiny"]),
            _p(it.get("unit", ""), s["tiny"]),
            _p(_qnum(it["unit_price"]), s["tiny"]),
            _p(_qnum(it["amount"]), s["tiny"]),
            _p(lead_remark, s["tiny"]),
        ])
    rows.append([
        _p("", s["tiny"]), _p("", s["tiny"]), _p("<b>Total</b>", s["tiny"]),
        _p("", s["tiny"]), _p("", s["tiny"]), _p("", s["tiny"]), _p("", s["tiny"]),
        _p(f"<b>{_qnum(total)}</b>", s["tiny"]), _p("", s["tiny"]),
    ])
    items_table = Table(rows, colWidths=[w * mm for w in widths], repeatRows=1)
    tcmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (4, 1), (7, -1), "RIGHT"), ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BACKGROUND", (0, len(rows) - 1), (-1, len(rows) - 1), LIGHT_BLUE),
    ]
    for r in range(1, len(rows) - 1):
        if r % 2 == 0:
            tcmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FAFBFC")))
    items_table.setStyle(TableStyle(tcmds))
    story += [items_table, Spacer(1, 2 * mm)]
    story.append(_p(f"<b>Total: {_money(total, currency)}</b>", s["right"]))
    story.append(Spacer(1, 5 * mm))

    # ── Terms & Conditions ────────────────────────────────────────────
    story.append(section("Terms & Conditions"))
    story.append(Spacer(1, 2 * mm))
    term_lines = []
    if incoterms_line:
        term_lines.append(f"Incoterms: {incoterms_line}")
    if terms.get("packing"):
        term_lines.append(f"Packing: {terms.get('packing')}")
    if terms.get("warranty"):
        term_lines.append(f"Warranty: {terms.get('warranty')}")
    if terms.get("remarks"):
        term_lines.append(f"Remarks: {terms.get('remarks')}")
    if not term_lines:
        term_lines.append("As per the terms agreed between K-MARIS and the supplier.")
    for line in term_lines:
        story.append(_p(f"• {line}", s["small"]))
        story.append(Spacer(1, 1 * mm))
    story.append(Spacer(1, 3 * mm))

    # ── Payment ───────────────────────────────────────────────────────
    story.append(section("Payment"))
    story.append(Spacer(1, 2 * mm))
    story.append(_p(f"• {terms.get('payment_terms') or 'T/T after delivery'}", s["small"]))
    story.append(Spacer(1, 1 * mm))
    story.append(_p("• Please confirm this purchase order and proceed with delivery per the agreed schedule.", s["small"]))
    story.append(Spacer(1, 6 * mm))

    # ── 서명 ──────────────────────────────────────────────────────────
    sign_img = image(asset("Authorized signature_Sungyeon Cho.jpg", "signature.png", "signature.jpg"), 40 * mm, 16 * mm)
    story.append(_p("For and on behalf of K-MARIS Energy & Solutions Co., Ltd.", s["base"]))
    if sign_img:
        story.append(sign_img)
    story.append(_p("________________________", s["base"]))
    story.append(_p("<b>Sam Cho, Managing Director</b>", s["base"]))
    story.append(Spacer(1, 2 * mm))
    story.append(_footer_center(s))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def make_pdf(doc_type: str, data: Dict[str, Any], company: Optional[Dict[str, Any]] = None, logo_path: Optional[str] = None) -> bytes:
    if doc_type not in DOC_TITLES:
        raise ValueError(f"Unsupported document type: {doc_type}")
    payload = dict(data)
    payload["company"] = company or data.get("company", {})
    if doc_type == "quotation":
        return _make_quotation_costing_pdf(payload, payload["company"])
    if doc_type == "purchase_order":
        return _make_purchase_order_pdf(payload, payload["company"])
    if doc_type == "commercial_invoice":
        return _make_commercial_invoice_pdf(payload, payload["company"])
    if doc_type == "proforma_invoice":
        if is_service_doc(payload):
            return _make_service_proforma_invoice_pdf(payload, payload["company"])
        return _make_proforma_invoice_pdf(payload, payload["company"])
    if doc_type == "tax_invoice":
        return _make_tax_invoice_pdf(payload, payload["company"])
    if doc_type == "credit_note":
        return _make_credit_note_pdf(payload, payload["company"])
    if doc_type == "shipping_mark":
        return _make_shipping_mark_pdf(payload, payload["company"])
    if doc_type == "packing_list":
        return _make_packing_list_pdf(payload, payload["company"])
    if doc_type == "vendor_rfq":
        return _make_vendor_rfq_pdf(payload, payload["company"])
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=15 * mm,
        title=DOC_TITLES[doc_type],
        author="K-MARIS Energy & Solutions Co., Ltd.",
    )
    s = _styles()
    story = []
    story += _letterhead(payload["company"], DOC_TITLES[doc_type], s, width=273 * mm)
    story.append(_info_tables(payload, doc_type))
    story.append(Spacer(1, 5 * mm))
    story.append(_items_table(payload, doc_type))
    if doc_type not in {"packing_list", "shipping_advice", "vendor_rfq"}:
        story.append(Spacer(1, 4 * mm))
        story.append(_totals_table(payload, doc_type))
    story.append(Spacer(1, 5 * mm))
    story.append(_commercial_shipping_block(payload) if doc_type == "commercial_invoice" else _terms_block(payload, doc_type))
    story.append(Spacer(1, 2 * mm))
    sign = Table(
        [[
            _p("Prepared by\n\n________________", s["base"]),
            _p("Approved by\n\n________________", s["base"]),
            _p("For and on behalf of K-MARIS Energy & Solutions Co., Ltd.\n\nAuthorized Signature", s["base"]),
        ]],
        colWidths=[70 * mm, 70 * mm, 130 * mm],
        rowHeights=[12 * mm],
    )
    sign.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
                ("BACKGROUND", (0, 0), (-1, 0), LIGHT_GRAY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(sign)
    story.append(Spacer(1, 2 * mm))
    story.append(_footer_center(s))
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def make_tax_invoice_xlsx(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Invoice Data"
    currency = data.get("currency", "KRW")
    items = normalize_items(data.get("items", []))
    vat_rate = _num(data.get("vat_rate", 0.1))
    totals = calc_totals(items, vat_rate)
    customer = data.get("customer", {})
    tax = data.get("tax_invoice", {})

    title_fill = PatternFill("solid", fgColor="0B1D3A")
    section_fill = PatternFill("solid", fgColor="EAF3FF")
    header_fill = PatternFill("solid", fgColor="D8DEE6")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    thin = Side(style="thin", color="D8DEE6")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = "K-MARIS TAX INVOICE DATA SHEET / 세금계산서 발행용 데이터"
    ws["A1"].fill = title_fill
    ws["A1"].font = white_font
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Issue Date / 작성일자", tax.get("issue_date", data.get("date", "")), "Document No.", data.get("doc_no", "")),
        ("Supply Type / 공급유형", tax.get("supply_type", ""), "Currency", currency),
        ("Supplier / 공급자", company.get("company_name_kr", ""), "Supplier Business No.", tax.get("supplier_business_no", company.get("business_no", ""))),
        ("Supplier Email", company.get("tax_email", company.get("general_email", "")), "Supplier Address", company.get("address", "")),
        ("Buyer / 공급받는 자", customer.get("name", ""), "Buyer Business No.", tax.get("buyer_business_no", customer.get("tax_id", ""))),
        ("Buyer Email", customer.get("email", ""), "Buyer Address", customer.get("address", "")),
    ]
    start = 3
    for r_idx, row in enumerate(rows, start=start):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx in [1, 3]:
                cell.fill = section_fill
                cell.font = bold
    item_start = start + len(rows) + 2
    headers = ["No.", "Part No.", "Description", "Maker", "Qty", "Unit", "Unit Price", "Amount", "HS Code", "Remark"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(item_start, c_idx, h)
        cell.fill = header_fill
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for r_offset, item in enumerate(items, start=1):
        values = [
            item["item_no"],
            item["part_no"],
            item["description"],
            item["maker"],
            item["qty"],
            item["unit"],
            item["unit_price"],
            item["amount"],
            item.get("hs_code", ""),
            item.get("remark", ""),
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(item_start + r_offset, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx in [7, 8]:
                cell.number_format = "#,##0.00"

    total_row = item_start + len(items) + 2
    total_values = [
        ("Supply Amount / 공급가액", totals["subtotal"]),
        ("VAT / 부가세", totals["vat"]),
        ("Total / 합계", totals["total"]),
    ]
    for i, (label, value) in enumerate(total_values):
        r = total_row + i
        ws.cell(r, 7, label).fill = section_fill
        ws.cell(r, 7).font = bold
        ws.cell(r, 7).border = border
        ws.cell(r, 8, value).border = border
        ws.cell(r, 8).number_format = "#,##0.00"

    note_row = total_row + 5
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=10)
    ws.cell(note_row, 1).value = (
        "Note: This sheet is a data-preparation document only. Actual electronic tax invoice issuance must be processed "
        "through Hometax or an authorized e-tax invoice provider/ERP after tax review. / 본 시트는 발행용 데이터이며 실제 전자세금계산서 발행은 홈택스 또는 공인 발급 시스템에서 세무 검토 후 진행하십시오."
    )
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(note_row, 1).border = border

    widths = [10, 20, 35, 24, 10, 10, 16, 18, 16, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
