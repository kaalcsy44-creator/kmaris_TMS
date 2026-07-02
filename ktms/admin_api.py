"""
KTMS Admin API — internal RFQ/Quotation overview for the Next.js admin UI.

This is the backend for the Vercel(Next.js) migration pilot. It reuses the
existing SQLAlchemy models/engine and re-implements the 12-step pipeline logic
WITHOUT Streamlit's cache decorator so it can run under FastAPI/uvicorn.

Run (dev):    uvicorn admin_api:app --reload --port 8001
Auth:         send  Authorization: Bearer <ADMIN_API_TOKEN>
              (set env ADMIN_API_TOKEN; defaults to "dev-token" for local dev)
"""
from __future__ import annotations

import io
import os
import re
import secrets
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import bcrypt
import jwt
from fastapi import Depends, FastAPI, File, Header, HTTPException, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from sqlalchemy import text

from db.engine import get_session, get_engine
from services.tracking_status import (
    rfq_tracking_step, order_tracking_step, RFQ_STEPS, ORDER_STEPS,
)
from services.email_svc import (
    quotation_email_body, quotation_email_subject, send_email,
    shipping_advice_email_body,
)
from services.pdf_svc import (
    build_payload, build_po_payload, generate_pdf, generate_po_pdf,
    generate_tax_xlsx,
)
from services.pdf_parser import (
    extract_text_from_pdf, parse_order_fields, parse_rfq_fields,
    parse_rfq_image, parse_order_image,
    parse_vendor_quote_text, parse_vendor_quote_image,
    parse_vendor_quote_pdf_document,
)
from services.vendor_xlsx import make_vendor_rfq_quote_xlsx
from services.quote_response_parser import parse_vendor_quote_bytes, excel_to_text
from db.models import (
    RFQ, Customer, Vessel, Vendor, User, UserRole, RolePermission, ItemMaster, DocSequence,
    VendorRFQ, VendorQuote, Quotation, QuotationStatus, FollowUpLevel,
    Order, PurchaseOrder, ShippingAdvice, CommercialInvoice,
    PackingList, TaxInvoiceData, ARRecord, DeliveryProof,
    RFQStatus, OrderStatus, ARStatus, WorkType, MarketingActivity, ScheduleEvent,
)

# ── App / CORS ────────────────────────────────────────────────────────────────
app = FastAPI(title="KTMS Admin API", docs_url=None, redoc_url=None)

_ALLOWED_ORIGINS = {"http://localhost:3000", "http://127.0.0.1:3000"}
_ALLOWED_ORIGIN_RE = re.compile(r"https://.*\.vercel\.app$")
USD_KRW_RATE = 1543.41
API_BUILD = "vendor-quote-currency-sql-update"


def _allow_origin(origin: str | None) -> str | None:
    """요청 Origin 이 허용 대상이면 그대로 돌려준다(에러 응답에 CORS 헤더용)."""
    if not origin:
        return None
    if origin in _ALLOWED_ORIGINS or _ALLOWED_ORIGIN_RE.match(origin):
        return origin
    return None


def _dual_money(value, currency: str = "USD") -> str:
    try:
        amount = float(value or 0)
    except Exception:
        amount = 0.0
    cur = (currency or "USD").upper()
    if cur == "KRW":
        return f"KRW {amount:,.0f} USD {amount / USD_KRW_RATE:,.0f}"
    if cur == "USD":
        return f"USD {amount:,.0f} KRW {round(amount * USD_KRW_RATE):,}"
    return f"{cur} {amount:,.0f}"


app.add_middleware(
    CORSMiddleware,
    allow_origins=list(_ALLOWED_ORIGINS),
    allow_origin_regex=r"https://.*\.vercel\.app",
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _sync_schema() -> None:
    """배포된 DB 스키마를 모델과 동기화한다.

    모델에 추가된 신규 컬럼(예: vendor_rfqs.sent_at / sent_to_email)이 운영 DB에
    누락되면 INSERT 시 500이 나고, CORSMiddleware가 500 응답에 CORS 헤더를 붙이지
    않아 프런트엔드에는 "Failed to fetch"로만 보인다. 시작 시 누락 컬럼을 자동
    추가해 스키마 드리프트를 방지한다."""
    try:
        from db.engine import Base
        from init_db import migrate_columns

        Base.metadata.create_all(bind=get_engine())
        migrate_columns()
    except Exception as exc:  # 스키마 동기화 실패가 앱 기동을 막지 않도록 로그만 남긴다.
        print(f"[WARN] startup schema sync skipped: {exc}", file=sys.stderr)
    try:
        _seed_perms()   # 역할 권한 기본값 시드 + 캐시 로드
    except Exception as exc:
        print(f"[WARN] permission seed skipped: {exc}", file=sys.stderr)


@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: Request, exc: Exception):
    """처리되지 않은 예외를 JSON 500으로 변환한다.

    catch-all 예외 핸들러의 응답은 CORSMiddleware 바깥(ServerErrorMiddleware)에서
    생성되어 Access-Control-Allow-Origin 헤더가 자동으로 붙지 않는다. 그러면
    프런트엔드는 진짜 500 메시지 대신 "Failed to fetch"만 보게 되므로, 여기서
    Origin 을 검증해 CORS 헤더를 직접 부착한다."""
    print(f"[ERROR] {request.method} {request.url.path}: {exc!r}", file=sys.stderr)
    headers = {}
    origin = _allow_origin(request.headers.get("origin"))
    if origin:
        headers["Access-Control-Allow-Origin"] = origin
        headers["Vary"] = "Origin"
    return JSONResponse(
        status_code=500, content={"detail": f"서버 오류: {exc}"}, headers=headers
    )


ADMIN_API_TOKEN = os.environ.get("ADMIN_API_TOKEN", "dev-token")
SECRET_KEY = os.environ.get("SECRET_KEY", "dev-secret-change-me")
JWT_ALGO = "HS256"
TOKEN_TTL_HOURS = 12

INTERNAL_STEPS = [
    "Customer RFQ Received",
    "Vendor RFQ Sent",
    "Vendor Quote Received",
    "Customer Quote Sent",
    "Customer P/O Received",
    "Vendor P/O Sent",
    "Delivery Readiness",
    "Delivery Arrangement",
    "Delivery Complete · POD",
    "Tax Invoice · Billing",
    "Tax Invoice Issued",
    "Payment Completed",
]

# 업무타입 "서비스"는 7·8·9단계(운송)를 서비스 관점 명칭으로 별도 관리한다.
SERVICE_STEP_OVERRIDES = {
    7: "Service Readiness",
    8: "Service Arrangement",
    9: "Service Complete · Report",
}


def steps_for(work_type) -> list[str]:
    """업무타입에 맞는 12단계 명칭. 서비스면 7·8·9단계를 서비스 명칭으로 치환."""
    wt = _enum_val(work_type) if work_type else WorkType.PARTS.value
    if wt == WorkType.SERVICE.value:
        return [SERVICE_STEP_OVERRIDES.get(i, name)
                for i, name in enumerate(INTERNAL_STEPS, start=1)]
    return list(INTERNAL_STEPS)


# ── Auth ──────────────────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


def _bearer(authorization: str | None) -> str:
    if authorization and authorization.lower().startswith("bearer "):
        return authorization[7:].strip()
    return ""


def _make_jwt(user: dict) -> str:
    payload = {
        "sub": str(user["id"]),
        "username": user["username"],
        "role": user["role"],
        "exp": datetime.now(timezone.utc) + timedelta(hours=TOKEN_TTL_HOURS),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=JWT_ALGO)


def _decode_jwt(token: str) -> dict | None:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[JWT_ALGO])
    except jwt.PyJWTError:
        return None


def _user_from_auth(authorization: str | None) -> dict | None:
    """Authorization 헤더에서 사용자(dict) 또는 None 을 돌려준다.
    - 유효한 JWT → 토큰 claims 기반 사용자
    - 파일럿 정적 ADMIN_API_TOKEN → admin 권한의 'dev' 사용자(id=0)
    예외를 던지지 않으므로 미들웨어에서 안전하게 쓸 수 있다."""
    token = _bearer(authorization)
    if not token:
        return None
    claims = _decode_jwt(token)
    if claims:
        return {
            "id": int(claims.get("sub", 0)),
            "username": claims.get("username", ""),
            "role": claims.get("role", ""),
        }
    if token == ADMIN_API_TOKEN:
        return {"id": 0, "username": "dev", "role": UserRole.ADMIN.value}
    return None


def require_token(authorization: str | None = Header(default=None)) -> None:
    """Guard: accept a valid JWT, or the pilot static ADMIN_API_TOKEN."""
    if _user_from_auth(authorization) is not None:
        return
    raise HTTPException(status_code=401, detail="Unauthorized")


def get_current_user(authorization: str | None = Header(default=None)) -> dict:
    user = _user_from_auth(authorization)
    if user is not None:
        return user
    raise HTTPException(status_code=401, detail="Unauthorized")


def _authz_error(request: Request, status: int, detail: str) -> JSONResponse:
    """권한 오류 응답에 CORS 헤더를 직접 부착해 돌려준다.
    (role 가드 미들웨어는 CORSMiddleware 바깥에서 동작하므로 수동 부착이 필요하다.)"""
    headers: dict[str, str] = {}
    origin = _allow_origin(request.headers.get("origin"))
    if origin:
        headers["Access-Control-Allow-Origin"] = origin
        headers["Vary"] = "Origin"
    return JSONResponse(status_code=status, content={"detail": detail}, headers=headers)


# ── 권한 매트릭스 (역할 × 페이지 × 동작) ───────────────────────────────────────
# 페이지(모듈)와 동작의 정본. 프런트 매트릭스 UI와 동일한 순서.
PERM_MODULES = ["dashboard", "progress", "rfq", "po", "documents", "ar", "marketing", "settings"]
PERM_ACTIONS = ["view", "create", "edit", "delete"]
# dashboard 는 열람만 의미가 있다(입력/수정/삭제 없음) — UI에서 view만 노출.
PERM_VIEW_ONLY = {"dashboard"}


def _perm_grid(value_map) -> dict:
    """{module: {action: bool}} 전체 그리드 생성. value_map(module,action)->bool."""
    return {m: {a: bool(value_map(m, a)) for a in PERM_ACTIONS} for m in PERM_MODULES}


def _full_perms(value: bool = True) -> dict:
    return _perm_grid(lambda m, a: value)


# 기본 권한(시드) — 기존 동작과 동일하게 맞춘다.
def _default_perms(role: str) -> dict:
    biz = ["progress", "rfq", "po", "documents", "ar", "marketing"]
    if role == UserRole.SALES.value:
        return _perm_grid(lambda m, a: (
            (m == "dashboard" and a == "view") or
            (m in biz)  # 거래 모듈 전체(열람·입력·수정·삭제)
        ))
    if role == UserRole.VIEWER.value:
        return _perm_grid(lambda m, a: (
            a == "view" and (m == "dashboard" or m in biz)  # 읽기 전용
        ))
    # 알 수 없는 역할: 대시보드 열람만.
    return _perm_grid(lambda m, a: (m == "dashboard" and a == "view"))


def _default_scope(role: str) -> str:
    return "own" if role == UserRole.SALES.value else "all"


def _normalize_perms(perms: dict | None) -> dict:
    """저장값을 전체 그리드로 정규화(누락 키는 False, dashboard 는 view만)."""
    perms = perms or {}
    def val(m, a):
        if m in PERM_VIEW_ONLY and a != "view":
            return False
        return bool((perms.get(m) or {}).get(a, False))
    return _perm_grid(val)


# 역할별 권한 캐시(요청마다 DB 조회를 피한다). PUT 시 _reload_perms() 로 갱신.
_PERM_CACHE: dict[str, dict] = {}


def _reload_perms() -> None:
    """DB(role_permissions)에서 sales/viewer 등 편집 가능한 역할 권한을 캐시에 로드."""
    cache: dict[str, dict] = {}
    try:
        s = get_session()
        try:
            for rp in s.query(RolePermission).all():
                cache[rp.role] = {
                    "perms": _normalize_perms(rp.perms),
                    "scope": rp.scope or _default_scope(rp.role),
                }
        finally:
            s.close()
    except Exception as exc:  # DB 미준비 시 기본값으로 동작.
        print(f"[WARN] permission cache load failed: {exc}", file=sys.stderr)
    _PERM_CACHE.clear()
    _PERM_CACHE.update(cache)


def _seed_perms() -> None:
    """sales/viewer 기본 권한 행이 없으면 시드(관리자가 편집할 베이스라인)."""
    s = get_session()
    try:
        for role in (UserRole.SALES.value, UserRole.VIEWER.value):
            if not s.query(RolePermission).filter_by(role=role).first():
                s.add(RolePermission(role=role, perms=_default_perms(role),
                                     scope=_default_scope(role)))
        s.commit()
    finally:
        s.close()
    _reload_perms()


def _perms_for(role: str) -> dict:
    """역할의 효과적 권한 그리드. admin 은 항상 전체 권한."""
    if role == UserRole.ADMIN.value:
        return _full_perms(True)
    entry = _PERM_CACHE.get(role)
    if entry:
        return entry["perms"]
    return _default_perms(role)


def _scope_for(role: str) -> str:
    if role == UserRole.ADMIN.value:
        return "all"
    entry = _PERM_CACHE.get(role)
    return entry["scope"] if entry else _default_scope(role)


def _can(role: str, module: str, action: str) -> bool:
    if role == UserRole.ADMIN.value:
        return True
    return bool(_perms_for(role).get(module, {}).get(action, False))


# ── 엔드포인트 → (모듈, 동작) 매핑 ─────────────────────────────────────────────
# 권한 검사에서 제외(공개/공통 참조). 로그인·본인정보·드롭다운용 마스터 조회 등.
_PERM_EXEMPT = {
    "/api/admin/login", "/api/admin/me", "/api/admin/me/permissions",
    "/api/admin/me/password", "/api/admin/customers", "/api/admin/vendors",
    "/api/admin/po-work-options", "/api/admin/health",
}
# 항상 admin 전용(권한 부여 대상 아님 — 권한 상승 방지).
_ADMIN_ONLY_PREFIXES = ("/api/admin/settings/users", "/api/admin/settings/permissions")
# 신규 레코드 생성(POST) — 그 외 POST 는 edit 으로 본다.
_CREATE_POST_EXACT = {
    "/api/admin/rfq", "/api/admin/orders", "/api/admin/vendor-pos", "/api/admin/ar",
    "/api/admin/marketing",
}
_CREATE_POST_SUFFIX = ("/vendor-rfq", "/customer-quote", "/vendor-quote",
                       "/ci", "/pl", "/sa", "/tax")


def _route_module(path: str) -> str | None:
    if path == "/api/admin/dashboard":
        return "dashboard"
    if path == "/api/admin/pipeline":
        return "progress"
    if path.startswith(("/api/admin/rfq", "/api/admin/quotation", "/api/admin/vrfq",
                        "/api/admin/vendor-rfq", "/api/admin/vendor-quote")):
        return "rfq"
    if path.startswith(("/api/admin/po-", "/api/admin/orders", "/api/admin/order/",
                        "/api/admin/vendor-po")):
        return "po"
    if path.startswith("/api/admin/documents"):
        return "documents"
    if path.startswith("/api/admin/ar"):
        return "ar"
    if path.startswith("/api/admin/marketing"):
        return "marketing"
    if path.startswith("/api/admin/settings"):
        return "settings"
    return None


def _route_action(method: str, path: str) -> str:
    if method == "GET":
        return "view"
    if method in ("PUT", "PATCH"):
        return "edit"
    if method == "DELETE":
        return "delete"
    if method == "POST":
        if path in _CREATE_POST_EXACT or path.endswith(_CREATE_POST_SUFFIX):
            return "create"
        if path.startswith("/api/admin/settings/"):
            rest = path[len("/api/admin/settings/"):]
            return "create" if "/" not in rest else "edit"
        return "edit"
    return "view"


def _route_perm(method: str, path: str):
    """요청에 필요한 (module, action) 반환. None=검사 제외, ('__admin__','')=admin 전용."""
    if method == "OPTIONS":
        return None
    if path in _PERM_EXEMPT or not path.startswith("/api/admin/"):
        return None
    if path.startswith(_ADMIN_ONLY_PREFIXES):
        return ("__admin__", "")
    module = _route_module(path)
    if module is None:
        return None
    # 설정 페이지의 마스터 데이터 조회(GET)는 화면 공통 참조이므로 열람 검사 제외.
    if module == "settings" and method == "GET":
        return None
    return (module, _route_action(method, path))


_PERM_DENY_MSG = {
    "view": "이 페이지를 열람할 권한이 없습니다.",
    "create": "등록(입력) 권한이 없습니다.",
    "edit": "수정 권한이 없습니다.",
    "delete": "삭제 권한이 없습니다.",
}

# 경로 첫 리소스/ID → 그 딜의 담당자(PIC=RFQ.created_by) 조회용. (소유권 게이트)
_DEAL_PATH_RE = re.compile(r"^/api/admin/([a-z-]+)/(\d+)")


def _deal_owner_from_path(path: str) -> int | None:
    """요청 경로가 가리키는 딜의 담당자(PIC=RFQ.created_by)를 반환. 판별 불가 시 None.

    모든 하위 문서(Vendor RFQ/Quote·Quotation·Order·Vendor PO·Documents·AR)는
    rfq_id/order_id 로 RFQ 에 연결되므로 소유권은 항상 RFQ.created_by 로 귀결된다.
    None(신규 등록처럼 대상 딜이 없거나, PIC 미지정)이면 소유권 검사를 건너뛴다.
    """
    m = _DEAL_PATH_RE.match(path)
    if not m:
        return None
    res, rid = m.group(1), int(m.group(2))
    s = get_session()
    try:
        rfq_id: int | None = None
        if res == "rfq":
            r = s.query(RFQ).filter_by(id=rid).first()
            return r.created_by if r else None
        if res == "marketing":
            # 마케팅 활동은 RFQ 딜이 아니라 owner_id(작성 담당자)로 소유권을 판별한다.
            m = s.query(MarketingActivity).filter_by(id=rid).first()
            return m.owner_id if m else None
        if res == "vendor-rfq":
            v = s.query(VendorRFQ).filter_by(id=rid).first()
            rfq_id = v.rfq_id if v else None
        elif res == "vendor-quote":
            vq = s.query(VendorQuote).filter_by(id=rid).first()
            if vq:
                vr = s.query(VendorRFQ).filter_by(id=vq.vendor_rfq_id).first()
                rfq_id = vr.rfq_id if vr else None
        elif res == "quotation":
            q = s.query(Quotation).filter_by(id=rid).first()
            rfq_id = q.rfq_id if q else None
        elif res in ("orders", "order", "documents"):
            o = s.query(Order).filter_by(id=rid).first()
            rfq_id = o.rfq_id if o else None
        elif res == "vendor-pos":
            vp = s.query(PurchaseOrder).filter_by(id=rid).first()
            if vp:
                o = s.query(Order).filter_by(id=vp.order_id).first()
                rfq_id = o.rfq_id if o else None
        elif res == "ar":
            ar = s.query(ARRecord).filter_by(id=rid).first()
            if ar:
                o = s.query(Order).filter_by(id=ar.order_id).first()
                rfq_id = o.rfq_id if o else None
        else:
            return None
        if rfq_id is None:
            return None
        r = s.query(RFQ).filter_by(id=rfq_id).first()
        return r.created_by if r else None
    finally:
        s.close()


@app.middleware("http")
async def _perm_guard(request: Request, call_next):
    """역할×페이지×동작 권한 매트릭스를 모든 /api/admin 요청에 적용한다.

    admin 은 항상 전체 허용(잠금 방지). 정적 ADMIN_API_TOKEN 도 admin 으로 취급.
    settings/users·settings/permissions 는 매트릭스와 무관하게 admin 전용.
    인증(토큰 유효성)은 별도로 각 엔드포인트의 require_token 도 검사한다."""
    need = _route_perm(request.method, request.url.path)
    if need is None:
        return await call_next(request)
    user = _user_from_auth(request.headers.get("authorization"))
    if user is None:
        return _authz_error(request, 401, "Unauthorized")
    role = user.get("role", "")
    if role == UserRole.ADMIN.value:
        return await call_next(request)
    module, action = need
    if module == "__admin__":
        return _authz_error(request, 403, "관리자 권한이 필요합니다.")
    if not _can(role, module, action):
        return _authz_error(request, 403, _PERM_DENY_MSG.get(action, "권한이 없습니다."))
    # 담당(PIC) 소유권 게이트: 비관리자는 본인이 담당인 딜만 편집/삭제(및 하위 등록) 가능.
    # 조회(view)는 기존 데이터 범위 그대로. 대상 딜이 없거나 PIC 미지정이면 통과.
    if action != "view" and module in ("rfq", "po", "documents", "ar", "marketing"):
        owner = _deal_owner_from_path(request.url.path)
        if owner is not None and owner != (user.get("id") or 0):
            return _authz_error(request, 403, "담당자(PIC)만 이 건을 수정·삭제할 수 있습니다.")
    return await call_next(request)


def _apply_owner_filter(q, model, user: dict, mine: int, assignee: int | None):
    """담당자(소유자=created_by) 필터.
    - 데이터 범위가 'own' 인 역할: 항상 본인 담당 건으로 강제 제한(파라미터 무시).
    - 'all' 역할(admin 포함): assignee 지정 시 해당 담당자, mine=1 이면 본인, 아니면 전체.
    """
    role = user.get("role", "")
    uid = user.get("id") or 0
    if role != UserRole.ADMIN.value and _scope_for(role) == "own":
        return q.filter(model.created_by == uid)
    if assignee:
        return q.filter(model.created_by == assignee)
    if mine:
        return q.filter(model.created_by == uid)
    return q


@app.post("/api/admin/login")
def admin_login(body: LoginRequest):
    s = get_session()
    try:
        user = s.query(User).filter_by(username=body.username, is_active=True).first()
        ok = bool(user) and bcrypt.checkpw(
            body.password.encode(), user.password_hash.encode()
        )
        if not ok:
            raise HTTPException(status_code=401, detail="사용자명 또는 비밀번호가 올바르지 않습니다.")
        role = user.role.value
        u = {"id": user.id, "username": user.username,
             "role": role, "email": user.email or ""}
        return {"token": _make_jwt(u), "user": u,
                "permissions": _perms_for(role), "scope": _scope_for(role)}
    finally:
        s.close()


@app.get("/api/admin/me")
def me(user: dict = Depends(get_current_user)):
    return user


@app.get("/api/admin/me/permissions")
def my_permissions(user: dict = Depends(get_current_user)):
    """현재 사용자의 효과적 권한 그리드 + 데이터 범위 + 메타데이터."""
    role = user.get("role", "")
    return {
        "role": role,
        "permissions": _perms_for(role),
        "scope": _scope_for(role),
        "modules": PERM_MODULES,
        "actions": PERM_ACTIONS,
        "view_only": sorted(PERM_VIEW_ONLY),
    }


# ── Helpers (decoupled from Streamlit) ────────────────────────────────────────
def _kst(dt) -> str:
    if not dt:
        return ""
    return (dt + timedelta(hours=9)).strftime("%y-%m-%d %H:%M")


def _items_cost_total(items) -> float:
    tot = 0.0
    for it in (items or []):
        try:
            tot += float(it.get("cost_price", 0) or 0) * float(it.get("qty", 1) or 1)
        except (TypeError, ValueError):
            pass
    return tot


def _total_amount(items) -> float:
    return sum(float(i.get("amount", 0) or 0) for i in (items or []))


def _enum_val(v) -> str:
    return v.value if hasattr(v, "value") else str(v)


def _coerce_work_type(v) -> WorkType | None:
    """필터 파라미터(한글 값 '부품공급'/'서비스' 또는 이름 'PARTS'/'SERVICE')를 WorkType 으로.
    빈값/전체/미인식은 None(필터 없음)."""
    if not v or v == "전체":
        return None
    try:
        return WorkType(v)            # 값('부품공급')으로 조회
    except ValueError:
        try:
            return WorkType[v]        # 이름('PARTS')으로 조회
        except KeyError:
            return None


def _pipeline_stage(s, rfq_id: int) -> int:
    """RFQ 1건의 내부 진행 단계(1~12) — helpers.internal_pipeline_stage 와 동일 로직.
    (FastAPI에서 돌도록 st.cache_data 제거, 세션은 호출측에서 공유)"""
    stage = 1
    rfq_obj = s.query(RFQ).filter_by(id=rfq_id).first()
    is_service = bool(rfq_obj and _enum_val(rfq_obj.work_type) == "서비스")

    vrfqs = s.query(VendorRFQ).filter_by(rfq_id=rfq_id).all()
    if vrfqs:
        stage = max(stage, 2)
        vrfq_ids = [v.id for v in vrfqs]
        if s.query(VendorQuote).filter(VendorQuote.vendor_rfq_id.in_(vrfq_ids)).first():
            stage = max(stage, 3)

    if (s.query(Quotation)
            .filter(Quotation.rfq_id == rfq_id, Quotation.status != QuotationStatus.DRAFT)
            .first()):
        stage = max(stage, 4)

    order = (s.query(Order).filter(Order.rfq_id == rfq_id)
             .order_by(Order.created_at.desc()).first())
    if not order:
        order = (s.query(Order).join(Quotation, Order.quotation_id == Quotation.id)
                 .filter(Quotation.rfq_id == rfq_id)
                 .order_by(Order.created_at.desc()).first())

    if order:
        stage = max(stage, 5)
        if s.query(PurchaseOrder).filter_by(order_id=order.id).first():
            stage = max(stage, 6)

        ost = _enum_val(order.status)
        stage = max(stage, {
            "오더 수주": 5,
            "발주 완료": 6,
            "제조/준비중": 7,
            "출고완료": 8,
            "운송중": 8,
            "목적지 하차 완료": 9,
        }.get(ost, 5))

        is_domestic = (getattr(order, "trade_type", "수출") == "내수")
        if is_domestic and not is_service:
            # 내수 부품공급: 7·8·9단계(CI/PL/SA/POD)는 해당 없음 → 건너뛴다.
            # 발주(6) 이후에는 곧바로 대금청구(10) 준비 단계로 본다.
            # (서비스는 7·8·9가 실제 작업 단계이므로 수동 완료로만 진행)
            if stage >= 6:
                stage = max(stage, 9)
        else:
            if getattr(order, "consignee_confirmed_date", None):
                stage = max(stage, 8)
            if s.query(ShippingAdvice).filter_by(order_id=order.id).first():
                stage = max(stage, 8)
            if getattr(order, "vendor_docs_sent_date", None):
                stage = max(stage, 8)
            # 9) 운송 완료 · POD 수취 — POD 파일 업로드 시 완료
            if s.query(DeliveryProof).filter_by(order_id=order.id).first():
                stage = max(stage, 9)
            ci = s.query(CommercialInvoice).filter_by(order_id=order.id).first()
            if ci:
                stage = max(stage, 8)
                # 10) Tax Invoice 작성 · 대금 청구 — Tax Invoice Data 생성 시
                if s.query(TaxInvoiceData).filter_by(ci_id=ci.id).first():
                    stage = max(stage, 10)

        ars = s.query(ARRecord).filter_by(order_id=order.id).all()
        if ars:
            stage = max(stage, 10)
            if any(_enum_val(a.status) == "완납" for a in ars):
                stage = max(stage, 12)

    # 수동 완료(완료 버튼/POD)로 stage_dates 에 표시된 단계를 반영.
    # (자동 근거가 약하거나 없는 단계만 — 의도치 않은 점프 방지)
    sd = (getattr(rfq_obj, "stage_dates", None) or {}) if rfq_obj else {}
    # 서비스 업무는 7·8단계(Service Readiness/arrangement)도 수동 완료로 진행한다.
    manual_keys = ("7", "8", "9", "11", "12") if is_service else ("9", "11", "12")
    for k in manual_keys:
        if sd.get(k):
            stage = max(stage, int(k))
    return stage


def _kst_iso(dt) -> str:
    """UTC datetime → KST 'YYYY-MM-DDTHH:MM' (datetime-local 입력과 호환)."""
    if not dt:
        return ""
    return (dt + timedelta(hours=9)).strftime("%Y-%m-%dT%H:%M")


def _fmt_received(iso: str) -> str:
    """'YYYY-MM-DDTHH:MM' → 'yy-mm-dd HH:MM' (목록 표시용). 빈값이면 ''."""
    if not iso or len(iso) < 16:
        return ""
    return f"{iso[2:10]} {iso[11:16]}"


def _first_rfq_iso(rfq) -> str:
    """RFQ 최초 수신 일시(iso 'YYYY-MM-DDTHH:MM') — 수동 received_at 우선, 없으면 생성시각.
    모든 단계 목록의 공통 식별 컬럼('First RFQ at')에서 정렬·필터·표시에 쓴다."""
    if not rfq:
        return ""
    return (getattr(rfq, "received_at", None) or "") or _kst_iso(rfq.created_at)


def _project_no_map(s) -> dict[int, str]:
    """프로젝트(=RFQ)별 내부 관리번호 {rfq_id: 'yymmdd-nn'}.
    최초 RFQ 수신 일시 기준으로, 같은 날짜 안에서 수신 순서대로 01,02,… 부여한다.
    (수신 일시 동률은 RFQ id 순. 저장값이 아니라 매 조회 시 결정적으로 산출.)
    같은 세션 안에서는 한 번만 계산하고 캐시한다(요청마다 세션은 새로 생성)."""
    cached = getattr(s, "_proj_no_cache", None)
    if cached is not None:
        return cached
    triples = [(_first_rfq_iso(r), r.id) for r in s.query(RFQ).all()]
    triples.sort(key=lambda t: (t[0] or "9999-99-99T99:99", t[1]))
    counters: dict[str, int] = {}
    out: dict[int, str] = {}
    for iso, rid in triples:
        yymmdd = (iso[2:4] + iso[5:7] + iso[8:10]) if len(iso) >= 10 else "000000"
        counters[yymmdd] = counters.get(yymmdd, 0) + 1
        out[rid] = f"{yymmdd}-{counters[yymmdd]:02d}"
    try:
        s._proj_no_cache = out
    except Exception:
        pass
    return out


def _vrfq_sent_iso(v) -> str:
    """Vendor RFQ 발신 일시(iso) — 수동 입력(sent_at) 우선, 없으면 생성 시각."""
    return (getattr(v, "sent_at", None) or "") or _kst_iso(v.created_at)


def _date_iso(d: str | None) -> str:
    """'YYYY-MM-DD' 문자열 → 'YYYY-MM-DDT00:00' (시각 정보가 없는 단계용)."""
    if not d:
        return ""
    d = d.strip()
    return f"{d}T00:00" if len(d) == 10 else ""


def _stage_auto_times(s, rfq, order) -> dict[str, str]:
    """내부 12단계 중, 근거 레코드가 존재하는 단계의 완료 일시를 자동 추출.
    수동 입력(stage_dates)이 없을 때 표시·기본값으로 사용된다. (7·12단계는 근거 없음)"""
    auto: dict[str, str] = {}

    def _set(stage: int, val: str):
        if val:
            auto[str(stage)] = val

    # 1) Customer RFQ 수신 — 수신 일시(received_at) 우선, 없으면 생성 시각
    _set(1, (getattr(rfq, "received_at", None) or "") or _kst_iso(rfq.created_at))

    # 2) Vendor RFQ 발신 · 3) Vendor Quot. 수신
    vrfqs = s.query(VendorRFQ).filter_by(rfq_id=rfq.id).all()
    if vrfqs:
        _set(2, min((_vrfq_sent_iso(v) for v in vrfqs), default=""))
        vrfq_ids = [v.id for v in vrfqs]
        vqs = (s.query(VendorQuote)
               .filter(VendorQuote.vendor_rfq_id.in_(vrfq_ids)).all())
        if vqs:
            # 3단계 일시 = 실제 견적 수신일시(received_at 수동입력) 우선,
            # 없으면 수신일(received_date), 그래도 없으면 레코드 생성시각.
            def _vq_recv(q) -> str:
                return ((getattr(q, "received_at", None) or "").strip()
                        or _date_iso(q.received_date)
                        or _kst_iso(q.created_at))
            _set(3, min((r for r in (_vq_recv(q) for q in vqs) if r), default=""))

    # 4) Customer Quot. 발신
    quo = (s.query(Quotation)
           .filter(Quotation.rfq_id == rfq.id, Quotation.status != QuotationStatus.DRAFT)
           .order_by(Quotation.created_at.asc()).first())
    if quo:
        _set(4, _date_iso(quo.sent_date) or _kst_iso(quo.created_at))

    if order:
        # 5) Customer P/O 수신
        _set(5, _kst_iso(order.created_at))
        # 6) Vendor P/O 발신
        po = (s.query(PurchaseOrder).filter_by(order_id=order.id)
              .order_by(PurchaseOrder.created_at.asc()).first())
        if po:
            _set(6, _date_iso(po.sent_date) or _kst_iso(po.created_at))
        # 8) Delivery arrangement
        sa = (s.query(ShippingAdvice).filter_by(order_id=order.id)
              .order_by(ShippingAdvice.created_at.asc()).first())
        _set(8, (_kst_iso(sa.created_at) if sa else "")
             or _date_iso(getattr(order, "consignee_confirmed_date", None))
             or _date_iso(getattr(order, "vendor_docs_sent_date", None))
             or _date_iso(getattr(order, "shipped_date", None)))
        # 9) 운송 완료 · POD 수취 — POD 업로드 일시 우선, 없으면 인도일
        pod = (s.query(DeliveryProof).filter_by(order_id=order.id)
               .order_by(DeliveryProof.created_at.asc()).first())
        _set(9, (getattr(pod, "uploaded_at", "") if pod else "")
             or _date_iso(getattr(order, "delivered_date", None)))
        # 10) Tax Invoice 작성 · 대금 청구 — Tax Invoice Data 생성 시점 우선
        ci = (s.query(CommercialInvoice).filter_by(order_id=order.id)
              .order_by(CommercialInvoice.created_at.asc()).first())
        ars = s.query(ARRecord).filter_by(order_id=order.id).all()
        tax = (s.query(TaxInvoiceData).filter_by(ci_id=ci.id)
               .order_by(TaxInvoiceData.created_at.asc()).first()) if ci else None
        _set(10, (_date_iso(tax.date) or _kst_iso(tax.created_at) if tax else "")
             or _kst_iso(min((a.created_at for a in ars if a.created_at), default=None))
             or (_kst_iso(ci.created_at) if ci else ""))
        # 11) 세금계산서 발행 · 12) 대금 결제 완료 — 수동 완료(stage_dates)로만 표시

    return auto


def _status_label(stage: int, work_type=None) -> str:
    steps = steps_for(work_type)
    return f"{stage}/{len(steps)} {steps[stage - 1]}"


# ── Endpoints ─────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    # DB 백엔드 종류만 노출(자격증명 X). sqlite 면 임시 디스크일 수 있어 재배포 시
    # 데이터가 사라질 수 있으므로 persistent=false 로 경고.
    backend = get_engine().url.get_backend_name()
    return {
        "status": "ok",
        "db": backend,
        "persistent": backend != "sqlite",
        "build": API_BUILD,
    }


@app.get("/api/admin/pipeline", dependencies=[Depends(require_token)])
def pipeline_overview(customer_id: int | None = None, work_type: str | None = None,
                      mine: int = 0, assignee: int | None = None,
                      user: dict = Depends(get_current_user)):
    """거래(RFQ) 1건 = 1행으로, RFQ→Quote(1~4)와 Order→Vendor PO(5~6) 체인을 한 번에
    합친 통합 파이프라인. 진행현황(내부확인용)이 RFQ표·PO표를 대체하는 단일 목록으로 쓴다."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}

        q = s.query(RFQ).order_by(RFQ.id.desc())
        if customer_id:
            q = q.filter(RFQ.customer_id == customer_id)
        wt = _coerce_work_type(work_type)
        if wt is not None:
            q = q.filter(RFQ.work_type == wt)
        q = _apply_owner_filter(q, RFQ, user, mine, assignee)
        rfqs = q.all()

        rows = []
        for r in rfqs:
            stage = _pipeline_stage(s, r.id)

            # 2) Vendor RFQ 발신
            vrfqs = (s.query(VendorRFQ).filter_by(rfq_id=r.id)
                     .order_by(VendorRFQ.id.desc()).all())
            if vrfqs:
                _vnames = []
                for x in vrfqs:
                    nm = vendor_names.get(x.vendor_id, "—")
                    if nm not in _vnames:
                        _vnames.append(nm)
                # 복수 벤더는 모두 줄바꿈으로 기재(프런트 white-space:pre-line).
                vrfq_vendors = "\n".join(_vnames)
                vrfq_at = _fmt_received(_vrfq_sent_iso(vrfqs[0]))
            else:
                vrfq_vendors, vrfq_at = "", ""

            # 3) Vendor Quot. 수신
            vrfq_ids = [x.id for x in vrfqs]
            vqs = (s.query(VendorQuote)
                   .filter(VendorQuote.vendor_rfq_id.in_(vrfq_ids))
                   .order_by(VendorQuote.id.desc()).all() if vrfq_ids else [])
            if vqs:
                vq0 = vqs[0]
                _vq_no = getattr(vq0, "vendor_quote_no", None) or "—"
                vquote_no = str(_vq_no) + (f"  (외 {len(vqs) - 1}건)" if len(vqs) > 1 else "")
                vquote_at = _kst(vq0.created_at)
                _cur = getattr(vq0, "currency", None) or "USD"
                vendor_amount = _dual_money(_items_cost_total(vq0.items), _cur)
            else:
                vquote_no, vquote_at, vendor_amount = "", "", ""

            # 4) Customer Quot. 발신
            qtn = (s.query(Quotation).filter_by(rfq_id=r.id)
                   .order_by(Quotation.id.desc()).first())
            if qtn:
                cquote_no, cquote_at = qtn.qtn_no, _kst(qtn.created_at)
                customer_amount = _dual_money(_total_amount(qtn.items or []), qtn.currency)
            else:
                cquote_no, cquote_at, customer_amount = "", "", ""

            # 5) Customer P/O 수신 · 6) Vendor P/O 발신
            o = _order_for_rfq(s, r.id)
            vpos = (s.query(PurchaseOrder).filter_by(order_id=o.id)
                    .order_by(PurchaseOrder.id.desc()).all()) if o else []
            if vpos:
                vp0 = vpos[0]
                vendor_po_no = (vp0.po_no or "—") + (f"  (외 {len(vpos) - 1}건)" if len(vpos) > 1 else "")
                # 발주서가 복수 벤더로 나간 경우 모두 줄바꿈으로 기재.
                _po_vnames: list[str] = []
                for vp in vpos:
                    nm = vendor_names.get(vp.vendor_id, "—")
                    if nm not in _po_vnames:
                        _po_vnames.append(nm)
                vendor_po_vendor = "\n".join(_po_vnames)
                vendor_po_email = vp0.sent_to_email or "—"
                vendor_po_at = _kst(vp0.created_at)
            else:
                vendor_po_no = vendor_po_vendor = vendor_po_email = vendor_po_at = ""

            rows.append({
                "rfq_id": r.id,
                "order_id": o.id if o else 0,
                # 식별
                "customer_rfq_no": r.customer_rfq_no or "",
                "kmaris_rfq_no": _rfq_no_disp(r.rfq_no),
                "work_type": _enum_val(r.work_type) if r.work_type else "부품공급",
                "trade_type": (o.trade_type if o else "수출") or "수출",
                "customer": cust_names.get(r.customer_id, "—"),
                "customer_id": r.customer_id or 0,
                "vessel": (vessel_names.get(r.vessel_id) if r.vessel_id else "") or "",
                "vessel_id": r.vessel_id or 0,
                "project_title": getattr(r, "project_title", None) or "",
                "received_at": getattr(r, "received_at", None) or "",
                "first_rfq_at": _first_rfq_iso(r),
                "project_no": _project_no_map(s).get(r.id, ""),
                # 담당자(PIC) = created_by 직원. 직접 지정 가능(설정 시 created_by 갱신).
                "assignee": user_names.get(getattr(r, "created_by", None), "") or "",
                "assignee_id": getattr(r, "created_by", None) or 0,
                "item_count": len((o.items if o else None) or r.items or []),
                "crfq_at": _fmt_received(getattr(r, "received_at", None) or "") or _kst(r.created_at),
                # 1~4 RFQ 체인
                "vrfq_vendors": vrfq_vendors,
                "vrfq_at": vrfq_at,
                "vquote_no": vquote_no,
                "vquote_at": vquote_at,
                "vendor_amount": vendor_amount,
                "cquote_no": cquote_no,
                "cquote_at": cquote_at,
                "customer_amount": customer_amount,
                # 5~6 PO 체인
                "customer_po_no": (o.po_no if o else "") or "",
                "customer_po_at": _kst(o.created_at) if o else "",
                "vendor_po_no": vendor_po_no,
                "vendor_po_at": vendor_po_at,
                "vendor": vendor_po_vendor,
                "vendor_email": vendor_po_email,
                # 상태 · 단계 일시
                "stage": stage,
                "status": _status_label(stage, r.work_type),
                "stage_dates": getattr(r, "stage_dates", None) or {},
                "stage_auto": _stage_auto_times(s, r, o),
                "stage_notes": getattr(r, "stage_notes", None) or {},
            })

        return {"steps": INTERNAL_STEPS, "rows": rows}
    finally:
        s.close()


def _search_href(stage: int, rfq_id: int, order_id: int, is_service: bool) -> str:
    """검색 결과 클릭 시 이동할 화면. ProgressScreen.stageHref 와 동일한 단계→페이지 규칙."""
    view = "service" if is_service else "parts"
    if stage <= 4:
        tab = {1: "new", 2: "vrfq", 3: "vquote", 4: "cquote"}.get(stage, "new")
        return f"/rfq?rfq={rfq_id}&tab={tab}"
    if order_id <= 0:
        return f"/rfq?rfq={rfq_id}"
    if stage == 5:
        return f"/po?order={order_id}&tab=customer"
    if stage == 6:
        return f"/po?order={order_id}&tab=vendor"
    if 7 <= stage <= 9:
        return f"/documents?order={order_id}&view={view}&stage={stage}"
    if stage == 10 and is_service:
        return f"/ar?order={order_id}"
    return f"/documents?order={order_id}&view={view}"


@app.get("/api/admin/search", dependencies=[Depends(require_token)])
def global_search(q: str = "", limit: int = 40, user: dict = Depends(get_current_user)):
    """전역 통합 검색 — RFQ(프로젝트) 1건을 단위로 식별자·품목·연락처·관련 문서번호까지
    훑어 매칭 결과를 반환한다. 본인 담당 스코프(own)는 파이프라인과 동일하게 강제된다."""
    term = (q or "").strip().lower()
    if len(term) < 2:
        return {"results": [], "query": q}
    tokens = [t for t in term.split() if t]

    s = get_session()
    try:
        cust = {c.id: c for c in s.query(Customer).all()}
        vessels = {v.id: v for v in s.query(Vessel).all()}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        proj_map = _project_no_map(s)

        base = s.query(RFQ).order_by(RFQ.id.desc())
        base = _apply_owner_filter(base, RFQ, user, 0, None)
        rfqs = base.all()

        results = []
        for r in rfqs:
            # 검색 대상 텍스트를 (분류 라벨, 텍스트) 쌍으로 모은다.
            fields: list[tuple[str, str]] = []

            def add(label: str, *vals) -> None:
                for v in vals:
                    if v:
                        fields.append((label, str(v)))

            c = cust.get(r.customer_id)
            ves = vessels.get(r.vessel_id) if r.vessel_id else None
            proj_no = proj_map.get(r.id, "")
            add("Customer RFQ No.", r.customer_rfq_no)
            add("K-Maris RFQ No.", _rfq_no_disp(r.rfq_no))
            add("Project No.", proj_no)
            add("Project title", getattr(r, "project_title", None))
            add("Customer", c.name if c else None)
            add("Contact", getattr(c, "contact", None), getattr(c, "email", None))
            add("Vessel", ves.name if ves else None, getattr(ves, "imo", None) if ves else None)
            add("PIC", user_names.get(getattr(r, "created_by", None)))
            add("Notes", getattr(r, "notes", None))
            for it in (r.items or []):
                add("Item", it.get("part_no"), it.get("description"))

            # 관련 문서(벤더 RFQ·벤더 견적·고객 견적·오더·벤더 PO)
            vrfqs = s.query(VendorRFQ).filter_by(rfq_id=r.id).all()
            for x in vrfqs:
                add("Vendor", vendor_names.get(x.vendor_id), x.sent_to_email)
            vrfq_ids = [x.id for x in vrfqs]
            vqs = (s.query(VendorQuote).filter(VendorQuote.vendor_rfq_id.in_(vrfq_ids)).all()
                   if vrfq_ids else [])
            for vq in vqs:
                add("Vendor quote No.", getattr(vq, "vendor_quote_no", None))
                for it in (vq.items or []):
                    add("Item", it.get("part_no"), it.get("description"))
            qtn = s.query(Quotation).filter_by(rfq_id=r.id).order_by(Quotation.id.desc()).first()
            if qtn:
                add("Quotation No.", qtn.qtn_no)
            o = _order_for_rfq(s, r.id)
            if o:
                add("Customer PO No.", o.po_no)
                for vp in s.query(PurchaseOrder).filter_by(order_id=o.id).all():
                    add("Vendor PO No.", vp.po_no)
                    add("Vendor", vendor_names.get(vp.vendor_id))

            blob = "\n".join(t.lower() for _, t in fields)
            if not all(tok in blob for tok in tokens):
                continue

            # 매칭 필드/스니펫 = 첫 토큰을 포함하는 첫 항목.
            primary = tokens[0]
            matched_label, matched_text = "", ""
            for label, text in fields:
                if primary in text.lower():
                    matched_label, matched_text = label, text
                    break

            stage = _pipeline_stage(s, r.id)
            is_service = (_enum_val(r.work_type) if r.work_type else "부품공급") == "서비스"
            results.append({
                "rfq_id": r.id,
                "order_id": o.id if o else 0,
                "project_no": proj_no,
                "customer": c.name if c else "—",
                "vessel": (ves.name if ves else "") or "",
                "project_title": getattr(r, "project_title", None) or "",
                "status": _status_label(stage, r.work_type),
                "stage": stage,
                "matched_label": matched_label,
                "matched_text": matched_text,
                "href": _search_href(stage, r.id, o.id if o else 0, is_service),
            })
            if len(results) >= limit:
                break

        return {"results": results, "query": q}
    finally:
        s.close()


@app.get("/api/admin/rfq-overview", dependencies=[Depends(require_token)])
def rfq_overview(customer_id: int | None = None, work_type: str | None = None,
                 mine: int = 0, assignee: int | None = None,
                 user: dict = Depends(get_current_user)):
    """RFQ 거래별 통합 현황 — Streamlit render_overview 와 동일한 행 데이터를 JSON으로."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}

        q = s.query(RFQ).order_by(RFQ.id.desc())
        if customer_id:
            q = q.filter(RFQ.customer_id == customer_id)
        wt = _coerce_work_type(work_type)
        if wt is not None:
            q = q.filter(RFQ.work_type == wt)
        q = _apply_owner_filter(q, RFQ, user, mine, assignee)
        rfqs = q.all()

        rows = []
        for r in rfqs:
            stage = _pipeline_stage(s, r.id)

            vrfqs = (s.query(VendorRFQ).filter_by(rfq_id=r.id)
                     .order_by(VendorRFQ.id.desc()).all())
            if vrfqs:
                vr0 = vrfqs[0]
                vrfq_at = _fmt_received(_vrfq_sent_iso(vr0))
                # "2. Vendor RFQ 발신" 칼럼은 발송한 Vendor사 이름을 표시한다.
                _vnames = []
                for x in vrfqs:
                    nm = vendor_names.get(x.vendor_id, "—")
                    if nm not in _vnames:
                        _vnames.append(nm)
                vrfq_vendors = _vnames[0] + (f"  (외 {len(_vnames) - 1}곳)" if len(_vnames) > 1 else "")
            else:
                vrfq_at, vrfq_vendors = "", ""

            vrfq_ids = [x.id for x in vrfqs]
            vqs = (s.query(VendorQuote)
                   .filter(VendorQuote.vendor_rfq_id.in_(vrfq_ids))
                   .order_by(VendorQuote.id.desc()).all()
                   if vrfq_ids else [])
            if vqs:
                vq0 = vqs[0]
                _vq_no = getattr(vq0, "vendor_quote_no", None) or "—"
                vq_main = str(_vq_no) + (f"  (외 {len(vqs) - 1}건)" if len(vqs) > 1 else "")
                vq_at = _kst(vq0.created_at)
                _cur = getattr(vq0, "currency", None) or "USD"
                vendor_amount = _dual_money(_items_cost_total(vq0.items), _cur)
            else:
                vq_main, vq_at, vendor_amount = "", "", ""

            qtn = (s.query(Quotation).filter_by(rfq_id=r.id)
                   .order_by(Quotation.id.desc()).first())
            if qtn:
                qtn_main, qtn_at = qtn.qtn_no, _kst(qtn.created_at)
                customer_amount = _dual_money(_total_amount(qtn.items or []), qtn.currency)
            else:
                qtn_main, qtn_at, customer_amount = "", "", ""

            rows.append({
                "id": r.id,
                "customer_rfq_no": r.customer_rfq_no or "",
                "project_title": getattr(r, "project_title", None) or "",
                "contact_person": getattr(r, "contact_person", None) or "",
                "assignee": user_names.get(getattr(r, "created_by", None), "") or "",
                "assignee_id": getattr(r, "created_by", None) or 0,
                "level": _enum_val(r.follow_up_level) if r.follow_up_level else "B",
                "work_type": _enum_val(r.work_type) if r.work_type else "부품공급",
                "customer": cust_names.get(r.customer_id, "—"),
                "vessel": vessel_names.get(r.vessel_id, "") if r.vessel_id else "",
                "item_count": len(r.items or []),
                "crfq_no": _rfq_no_disp(r.rfq_no),
                "crfq_at": _fmt_received(getattr(r, "received_at", None) or "") or _kst(r.created_at),
                "first_rfq_at": _first_rfq_iso(r),
                "project_no": _project_no_map(s).get(r.id, ""),
                # K-Maris RFQ No.는 Vendor RFQ 발신 시점에 부여된다. 발신한 거래에서만 표시.
                "vrfq_kmaris_no": (_rfq_no_disp(r.rfq_no) if vrfqs else ""),
                "vrfq_vendors": vrfq_vendors,
                "vrfq_at": vrfq_at,
                "vquote_no": vq_main,
                "vquote_at": vq_at,
                "vendor_amount": vendor_amount,
                "cquote_no": qtn_main,
                "cquote_at": qtn_at,
                "customer_amount": customer_amount,
                "stage": stage,
                "status": _status_label(stage, r.work_type),
            })

        return {"steps": INTERNAL_STEPS, "rows": rows}
    finally:
        s.close()


def _item_view(it: dict) -> dict:
    qty = it.get("qty", 1) or 1
    amount = it.get("amount")
    unit = it.get("unit_price", it.get("price"))
    if unit is None and amount is not None:
        try:
            unit = float(amount) / float(qty or 1)
        except (TypeError, ValueError, ZeroDivisionError):
            unit = None
    return {
        "part_no": it.get("part_no") or "",
        "description": it.get("description") or "",
        "qty": qty,
        "unit": it.get("unit") or "",
        "unit_price": unit,
        "amount": amount,
    }


def _po_item_lines(items, korean: bool) -> str:
    qty_label = "수량" if korean else "Qty"
    desc_label = "품명" if korean else "Desc"
    return "\n".join(
        f"  {i+1:>2}. Part No.: {str(it.get('part_no','—')):<20s}"
        f"  {qty_label}: {it.get('qty','—')} {str(it.get('unit','')):<5s}"
        f"  Maker: {it.get('maker','—')}\n"
        f"       {desc_label}: {it.get('description','—')}"
        for i, it in enumerate(items or [])
    )


def _vendor_po_email_body(po, vendor, order, vessel, notes: str, lang: str, project_no: str = "") -> str:
    vendor_name = vendor.name if vendor else "Vendor"
    vessel_str = vessel.name if vessel else "—"
    if lang == "ko":
        body = f"""{vendor_name} 귀중

안녕하세요,
항상 협조해 주셔서 감사드립니다.

아래 선박용 부품에 대한 발주서를 첨부와 같이 송부드립니다.

발주번호 : {po.po_no}
프로젝트 : {project_no or '—'}
선박명   : {vessel_str}
발주일   : {po.date or date.today().isoformat()}

──────────────────────── 품목 리스트 ────────────────────────
{_po_item_lines(po.items, korean=True)}
──────────────────────────────────────────────────────────────

수령 후 아래 사항을 확인·회신해 주시기 바랍니다:
  • 본 발주 수락 여부
  • 확정 납기 (출고 예정일)
  • 품번·수량·단가 상이 여부

"""
        if notes:
            body += f"추가 사항:\n{notes}\n\n"
        body += """영업일 기준 3일 이내 수령 확인 및 회신 부탁드립니다.

감사합니다.
K-MARIS Energy & Solutions Co., Ltd.
Email: sales@k-maris.com  |  www.k-maris.com
Engineering Reliability. Supplying Performance.
"""
        return body

    body = f"""Dear {vendor_name},

Please find attached our official Purchase Order for the following marine spare parts.

PO No.        : {po.po_no}
Project No.   : {project_no or '—'}
Vessel        : {vessel_str}
Order Date    : {po.date or date.today().isoformat()}

──────────────────────── ITEM LIST ────────────────────────
{_po_item_lines(po.items, korean=False)}
────────────────────────────────────────────────────────────

Please confirm the following upon receipt:
  • Acceptance of this Purchase Order
  • Confirmed delivery schedule (ex-works / shipment date)
  • Any discrepancy in part number, quantity, or price

"""
    if notes:
        body += f"Additional Notes:\n{notes}\n\n"
    body += """Kindly acknowledge receipt and confirm within 3 business days.

Best regards,
K-MARIS Energy & Solutions Co., Ltd.
Email: sales@k-maris.com  |  www.k-maris.com
Engineering Reliability. Supplying Performance.
"""
    return body


def _sanitize_vendor_rfq_items(raw) -> list[dict]:
    """발신 화면에서 넘어온 품목(선택·편집본)을 저장/문서용 dict 리스트로 정규화.
    빈 행(부품번호·품명·수량이 모두 비어 있음)은 제거한다."""
    out: list[dict] = []
    for it in (raw or []):
        part_no = str(it.get("part_no", "") or "").strip()
        desc = str(it.get("description", "") or "").strip()
        unit = str(it.get("unit", "") or "").strip()
        remark = str(it.get("remark", "") or "").strip()
        try:
            qty = float(it.get("qty") or 0)
        except (TypeError, ValueError):
            qty = 0
        if not part_no and not desc and not qty:
            continue
        row = {"part_no": part_no, "description": desc, "qty": qty, "unit": unit}
        if remark:
            row["remark"] = remark
        out.append(row)
    return out


def _vendor_rfq_email_body(rfq, cust, vessel, vendor, notes: str, lang: str,
                           items=None) -> str:
    # items 를 명시적으로 주면(발신 화면에서 선택·편집한 품목) 그것을 쓰고,
    # 없으면 RFQ 원본 품목을 사용한다.
    items = rfq.items if items is None else items
    items = items or []
    if lang == "ko":
        item_lines = "\n".join(
            f"  {i+1:>2}. Part No.: {str(item.get('part_no','—')):<20s}"
            f"  수량: {item.get('qty','—')} {item.get('unit',''):<5s}"
            f"  Maker: {item.get('maker','—')}\n"
            f"       품명: {item.get('description','—')}"
            for i, item in enumerate(items)
        )
        body = f"""{vendor.name if vendor else 'Vendor'} 귀중

안녕하세요,
항상 협조해 주셔서 감사드립니다.

아래 선박용 부품에 대한 견적을 요청드립니다.

RFQ 번호 : {rfq.rfq_no}
선박명    : {vessel.name if vessel else '—'}
발주처    : {cust.name if cust else '—'}
문의일    : {rfq.date or date.today().isoformat()}

──────────────────────── 품목 리스트 ────────────────────────
{item_lines}
──────────────────────────────────────────────────────────────

각 품목에 대해 아래 사항을 포함하여 견적을 회신해 주시기 바랍니다:
  • 단가 (USD, CNF 부산항 기준)
  • 납기
  • 원산지 / 제조사
  • 기술적 비고 또는 대체품 (해당 시)

"""
        if notes:
            body += f"추가 사항:\n{notes}\n\n"
        body += """영업일 기준 5일 이내 회신 부탁드립니다.

감사합니다.
K-MARIS Energy & Solutions Co., Ltd.
Email: sales@k-maris.com  |  www.k-maris.com
Engineering Reliability. Supplying Performance.
"""
        return body

    item_lines = "\n".join(
        f"  {i+1:>2}. Part No.: {str(item.get('part_no','—')):<20s}"
        f"  Qty: {item.get('qty','—')} {item.get('unit',''):<5s}"
        f"  Maker: {item.get('maker','—')}\n"
        f"       Desc: {item.get('description','—')}"
        for i, item in enumerate(items)
    )
    body = f"""Dear {vendor.name if vendor else 'Vendor'},

We would like to request your best quotation for the following marine spare parts.

RFQ Reference : {rfq.rfq_no}
Vessel        : {vessel.name if vessel else '—'}
End Customer  : {cust.name if cust else '—'}
Enquiry Date  : {rfq.date or date.today().isoformat()}

──────────────────────── ITEM LIST ────────────────────────
{item_lines}
────────────────────────────────────────────────────────────

Please quote for each item:
  • Unit price (USD, CNF Busan port)
  • Lead time
  • Country of origin / Manufacturer
  • Technical remarks or alternatives (if any)

"""
    if notes:
        body += f"Additional Notes:\n{notes}\n\n"
    body += """Kindly reply within 5 business days.

Best regards,
K-MARIS Energy & Solutions Co., Ltd.
Email: sales@k-maris.com  |  www.k-maris.com
Engineering Reliability. Supplying Performance.
"""
    return body


@app.get("/api/admin/rfq/{rfq_id}", dependencies=[Depends(require_token)])
def rfq_detail(rfq_id: int):
    """RFQ 1건 상세 — 품목, 12단계 진행, 연결 문서(Vendor RFQ/Quote/Quotation)."""
    s = get_session()
    try:
        r = s.query(RFQ).filter_by(id=rfq_id).first()
        if not r:
            raise HTTPException(status_code=404, detail="RFQ not found")

        cust = s.query(Customer).filter_by(id=r.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=r.vessel_id).first() if r.vessel_id else None
        stage = _pipeline_stage(s, r.id)

        vrfqs = (s.query(VendorRFQ).filter_by(rfq_id=r.id)
                 .order_by(VendorRFQ.id.desc()).all())
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        vrfq_view = [{
            "id": v.id,
            "vendor": vendor_names.get(v.vendor_id, "—"),
            "at": _fmt_received(_vrfq_sent_iso(v)),
        } for v in vrfqs]

        vrfq_ids = [v.id for v in vrfqs]
        vqs = (s.query(VendorQuote)
               .filter(VendorQuote.vendor_rfq_id.in_(vrfq_ids))
               .order_by(VendorQuote.id.desc()).all() if vrfq_ids else [])
        vquote_view = [{
            "vendor_quote_no": getattr(q, "vendor_quote_no", None) or "—",
            "amount": _dual_money(_items_cost_total(q.items), getattr(q, "currency", None) or "USD"),
            "at": _kst(q.created_at),
        } for q in vqs]

        qtn = (s.query(Quotation).filter_by(rfq_id=r.id)
               .order_by(Quotation.id.desc()).first())
        qtn_view = None
        if qtn:
            qtn_view = {
                "qtn_no": qtn.qtn_no,
                "amount": _dual_money(_total_amount(qtn.items or []), qtn.currency),
                "status": _enum_val(qtn.status),
                "at": _kst(qtn.created_at),
            }

        steps = [{
            "no": i,
            "name": name,
            "state": ("done" if i < stage else "current" if i == stage else "todo"),
        } for i, name in enumerate(steps_for(r.work_type), start=1)]

        return {
            "id": r.id,
            "rfq_no": _rfq_no_disp(r.rfq_no),
            "assignee_id": r.created_by or 0,   # 담당자(PIC)
            "customer_rfq_no": r.customer_rfq_no or "",
            "contact_person": getattr(r, "contact_person", None) or "",
            "customer": cust.name if cust else "—",
            "customer_id": r.customer_id or 0,
            "customer_contact": (cust.contact if cust else "") or "",
            "customer_email": (cust.email if cust else "") or "",
            "vessel": vessel.name if vessel else "",
            "vessel_id": r.vessel_id or 0,
            "project_title": getattr(r, "project_title", None) or "",
            "work_type": _enum_val(r.work_type) if r.work_type else "부품공급",
            "received_at": getattr(r, "received_at", None) or "",
            "first_rfq_at": _first_rfq_iso(r),
            "project_no": _project_no_map(s).get(r.id, ""),
            "date": r.date or "",
            "notes": r.notes or "",
            "request_channel": getattr(r, "request_channel", None) or "",
            "follow_up_level": _enum_val(r.follow_up_level) if r.follow_up_level else "B",
            "stage": stage,
            "status": _status_label(stage, r.work_type),
            "steps": steps,
            "items": [_item_view(it) for it in (r.items or [])],
            "vendor_rfqs": vrfq_view,
            "vendor_quotes": vquote_view,
            "quotation": qtn_view,
        }
    finally:
        s.close()


@app.get("/api/admin/dashboard", dependencies=[Depends(require_token)])
def dashboard():
    """운영 현황 요약 — 핵심 KPI + 12단계 분포 + 최근 RFQ."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        rfqs = s.query(RFQ).all()
        orders = s.query(Order).all()
        quotes = s.query(Quotation).all()
        ars = s.query(ARRecord).all()

        today_iso = date.today().isoformat()
        soon_iso = (date.today() + timedelta(days=7)).isoformat()
        urgent_cutoff = (date.today() + timedelta(days=3)).isoformat()

        # ── 운영 KPI ──────────────────────────────────────────────────────────
        open_rfq = sum(1 for r in rfqs if r.status in
                       {RFQStatus.RECEIVED, RFQStatus.SOURCING, RFQStatus.QUOTING})
        active_orders = sum(1 for o in orders if o.status in
                            {OrderStatus.RECEIVED, OrderStatus.PO_SENT, OrderStatus.PREPARING,
                             OrderStatus.SHIPPED, OrderStatus.IN_TRANSIT})

        now = datetime.now(timezone.utc)
        monthly_quotes = sum(
            1 for q in quotes
            if q.created_at and q.created_at.year == now.year
            and q.created_at.month == now.month
        )
        ar_outstanding = sum(
            (a.invoice_amount or 0) - (a.paid_amount or 0)
            for a in ars
            if a.status in {ARStatus.OUTSTANDING, ARStatus.PARTIAL, ARStatus.OVERDUE}
            and (a.currency or "USD") == "USD"
        )

        urgent = [q for q in quotes
                  if q.status == QuotationStatus.SENT
                  and q.follow_up_level == FollowUpLevel.A
                  and q.valid_until and q.valid_until <= urgent_cutoff]
        overdue = [a for a in ars if a.status == ARStatus.OVERDUE]
        pending_po = sum(1 for o in orders if o.status == OrderStatus.RECEIVED)
        expiring = sum(
            1 for q in quotes
            if q.status in (QuotationStatus.SENT, QuotationStatus.NEGOTIATING)
            and q.valid_until and today_iso <= q.valid_until <= soon_iso
        )

        # ── 영업 성과 KPI ─────────────────────────────────────────────────────
        total_rfq = len(rfqs)
        sent_quote_rfq_ids = {q.rfq_id for q in quotes
                              if q.rfq_id and q.status != QuotationStatus.DRAFT}
        handling_rate = (len(sent_quote_rfq_ids) / total_rfq * 100) if total_rfq else 0.0

        rfq_created = {r.id: r.created_at for r in rfqs}
        _tat = []
        for q in quotes:
            base = rfq_created.get(q.rfq_id) if q.rfq_id else None
            if base and q.created_at and q.status != QuotationStatus.DRAFT:
                h = (q.created_at - base).total_seconds() / 3600
                if h >= 0:
                    _tat.append(h)
        quotation_tat_h = (sum(_tat) / len(_tat)) if _tat else None

        _sent_like = {QuotationStatus.SENT, QuotationStatus.NEGOTIATING,
                      QuotationStatus.WON, QuotationStatus.LOST, QuotationStatus.EXPIRED}
        sent_quotes = [q for q in quotes if q.status in _sent_like]
        won_quotes = [q for q in quotes if q.status == QuotationStatus.WON]
        hit_rate = (len(won_quotes) / len(sent_quotes) * 100) if sent_quotes else 0.0

        margin_basis = won_quotes or sent_quotes
        _rev = _cost = 0.0
        for q in margin_basis:
            for it in (q.items or []):
                qty = float(it.get("qty", 1) or 1)
                _rev += float(it.get("unit_price", 0) or 0) * qty
                _cost += float(it.get("cost_price", 0) or 0) * qty
        gross_margin_pct = ((_rev - _cost) / _rev * 100) if _rev else 0.0

        negotiating_value_usd = 0.0
        for q in quotes:
            if q.status == QuotationStatus.NEGOTIATING and (q.currency or "USD") == "USD":
                for it in (q.items or []):
                    amt = it.get("amount")
                    if amt in (None, ""):
                        amt = float(it.get("unit_price", 0) or 0) * float(it.get("qty", 1) or 1)
                    negotiating_value_usd += float(amt or 0)

        dist = [0] * len(INTERNAL_STEPS)
        for r in rfqs:
            dist[_pipeline_stage(s, r.id) - 1] += 1

        recent = []
        for r in sorted(rfqs, key=lambda x: x.id, reverse=True)[:8]:
            stage = _pipeline_stage(s, r.id)
            recent.append({
                "rfq_no": _rfq_no_disp(r.rfq_no),
                "customer": cust_names.get(r.customer_id, "—"),
                "stage": stage,
                "status": _status_label(stage, r.work_type),
                "at": _kst(r.created_at),
            })

        # ── Snapshot: 고객 추적(RFQ/Order) + 내부 12단계 (per-RFQ) ───────────────
        def _cv(cid, vid) -> str:
            nm = cust_names.get(cid, "—")
            vn = vessel_names.get(vid) if vid else None
            return f"{nm} · {vn}" if vn else nm

        snapshot = []
        for r in sorted(rfqs, key=lambda x: x.id, reverse=True)[:20]:
            o = _order_for_rfq(s, r.id)
            order_row = None
            if o:
                order_row = {
                    "customer_vessel": _cv(o.customer_id, o.vessel_id),
                    "status": _enum_val(o.status),
                    "item_count": len(o.items or []),
                    "date": o.date or "—",
                    "step": order_tracking_step(_enum_val(o.status))[0],
                }
            _lvl = getattr(r, "follow_up_level", None)
            snapshot.append({
                "id": r.id,
                "rfq_no": _rfq_no_disp(r.rfq_no),
                "customer_rfq_no": r.customer_rfq_no or "",
                "project_title": getattr(r, "project_title", None) or "",
                "customer": cust_names.get(r.customer_id, "—"),
                "vessel": (vessel_names.get(r.vessel_id) if r.vessel_id else "") or "",
                "customer_vessel": _cv(r.customer_id, r.vessel_id),
                "stage_dates": getattr(r, "stage_dates", None) or {},
                "stage_auto": _stage_auto_times(s, r, o),
                "status": _enum_val(r.status),
                "item_count": len(r.items or []),
                "follow_up_level": _enum_val(_lvl) if _lvl else "—",
                "date": r.date or "—",
                "step": rfq_tracking_step(_enum_val(r.status))[0],
                "stage": _pipeline_stage(s, r.id),
                "order": order_row,
            })

        return {
            "kpi": {
                "open_rfq": open_rfq,
                "total_rfq": len(rfqs),
                "active_orders": active_orders,
                "monthly_quotes": monthly_quotes,
                "ar_outstanding_usd": round(ar_outstanding, 2),
            },
            "ops": {
                "urgent": len(urgent),
                "pending_po": pending_po,
                "overdue": len(overdue),
                "expiring": expiring,
            },
            "perf": {
                "handling_rate": round(handling_rate, 0),
                "quotation_tat_h": round(quotation_tat_h, 0) if quotation_tat_h is not None else None,
                "hit_rate": round(hit_rate, 0),
                "gross_margin_pct": round(gross_margin_pct, 1),
                "negotiating_value_usd": round(negotiating_value_usd, 0),
            },
            "alerts": {
                "urgent_quotes": [
                    {"qtn_no": q.qtn_no, "valid_until": q.valid_until or "",
                     "status": _enum_val(q.status)} for q in urgent
                ],
                "overdue_ar": [
                    {"ci_no": a.ci_no or "", "currency": a.currency or "USD",
                     "outstanding": round((a.invoice_amount or 0) - (a.paid_amount or 0), 2),
                     "due_date": a.due_date or ""} for a in overdue
                ],
            },
            "steps": INTERNAL_STEPS,
            "stage_distribution": dist,
            "recent": recent,
            "snapshot": snapshot,
            "rfq_steps": RFQ_STEPS,
            "order_steps": ORDER_STEPS,
        }
    finally:
        s.close()


def _cur2(c: str | None) -> str:
    """통계 집계용 통화 정규화 — USD/KRW 만 구분, 그 외는 USD 로 취급."""
    return c if c in ("USD", "KRW") else "USD"


def _month_key(v: str | None) -> str:
    """'YYYY-MM-DD…' 또는 'YYYY-MM…' → 'YYYY-MM'. 비정상값이면 ''."""
    if not v or len(v) < 7:
        return ""
    return v[:7]


@app.get("/api/admin/statistics", dependencies=[Depends(require_token)])
def statistics(months: int = 12):
    """통계 대시보드 — 월별 시계열(매출·견적·수주), 랭킹(고객·품목), KPI, 업무알림.

    금액은 USD/KRW 통화별로 분리 집계(환산 없음, 프런트 토글로 전환).
    매출 인식 시점은 세금계산서 발행일(RFQ.stage_dates["11"]) 기준.
    """
    months = max(1, min(int(months or 12), 36))
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        orders_all = s.query(Order).all()
        order_map = {o.id: o for o in orders_all}
        rfq_map = {r.id: r for r in s.query(RFQ).all()}

        # ── 월 버킷(최근 N개월, KST) ───────────────────────────────────────────
        now_kst = datetime.now(timezone.utc) + timedelta(hours=9)
        month_labels: list[str] = []
        y, m = now_kst.year, now_kst.month
        for _ in range(months):
            month_labels.append(f"{y:04d}-{m:02d}")
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        month_labels.reverse()
        month_set = set(month_labels)
        cur_month = month_labels[-1]
        prev_month = month_labels[-2] if len(month_labels) >= 2 else ""

        CURS = ("USD", "KRW")
        # series[currency][metric] = {month: amount}
        def _blank():
            return {cur: {mo: 0.0 for mo in month_labels} for cur in CURS}
        rev_series = _blank()
        quote_series = _blank()
        order_series = _blank()
        cust_rev: dict[str, dict[str, float]] = {c: {} for c in CURS}
        item_rev: dict[str, dict[str, dict]] = {c: {} for c in CURS}

        # RFQ.id → 세금계산서 발행월(stage_dates["11"]) 매핑
        def _issue_month(rfq) -> str:
            sd = (getattr(rfq, "stage_dates", None) or {}) if rfq else {}
            return _month_key(sd.get("11") or "")

        # ── 매출(세금계산서 발행 기준) + 고객사 랭킹 ─────────────────────────────
        for a in s.query(ARRecord).all():
            o = order_map.get(a.order_id)
            rfq = _rfq_for_order(s, o) if o else None
            mo = _issue_month(rfq)
            if not mo or mo not in month_set:
                continue
            cur = _cur2(a.currency)
            amt = float(a.invoice_amount or 0)
            rev_series[cur][mo] += amt
            cname = cust_names.get(o.customer_id, "—") if o else "—"
            cust_rev[cur][cname] = cust_rev[cur].get(cname, 0.0) + amt

        # ── 견적금액(발송월 기준) ───────────────────────────────────────────────
        for q in s.query(Quotation).all():
            mo = _month_key(getattr(q, "sent_at", None) or q.sent_date or q.date or "")
            if not mo or mo not in month_set:
                continue
            cur = _cur2(q.currency)
            quote_series[cur][mo] += _quotation_total(q.items or [], getattr(q, "discount_pct", 0) or 0)

        # ── 수주금액(오더 수주월 기준) ──────────────────────────────────────────
        for o in orders_all:
            mo = _month_key(o.date or (o.created_at.isoformat() if o.created_at else ""))
            if not mo or mo not in month_set:
                continue
            qtn = s.query(Quotation).filter_by(id=o.quotation_id).first() if o.quotation_id else None
            if qtn:
                cur = _cur2(qtn.currency)
                amt = _quotation_total(qtn.items or [], getattr(qtn, "discount_pct", 0) or 0)
            else:
                cur = "USD"
                amt = _total_amount(o.items or [])
            order_series[cur][mo] += amt

        # ── 품목별 매출(청구 품목=CI items 기준, 기간 내) ────────────────────────
        for ci in s.query(CommercialInvoice).all():
            mo = _month_key(ci.date or "")
            if not mo or mo not in month_set:
                continue
            cur = _cur2(ci.currency)
            for it in (ci.items or []):
                pn = (it.get("part_no") or "").strip() or "—"
                amt = it.get("amount")
                if amt in (None, ""):
                    amt = float(it.get("unit_price", 0) or 0) * float(it.get("qty", 1) or 1)
                rec = item_rev[cur].setdefault(pn, {"part_no": pn, "description": it.get("description") or "", "amount": 0.0})
                rec["amount"] += float(amt or 0)

        def _series_list(series):
            return {cur: [round(series[cur][mo], 2) for mo in month_labels] for cur in CURS}

        def _top(dmap, key_field, n=10):
            out = {}
            for cur in CURS:
                if key_field == "customer":
                    items = [{"name": k, "amount": round(v, 2)} for k, v in dmap[cur].items()]
                else:
                    items = [{"part_no": r["part_no"], "description": r["description"], "amount": round(r["amount"], 2)} for r in dmap[cur].values()]
                items.sort(key=lambda x: x["amount"], reverse=True)
                out[cur] = items[:n]
            return out

        # ── KPI(이번 달 + 전월대비) ─────────────────────────────────────────────
        def _kpi():
            out = {}
            for cur in CURS:
                out[cur] = {
                    "revenue": round(rev_series[cur][cur_month], 2),
                    "revenue_prev": round(rev_series[cur].get(prev_month, 0.0), 2) if prev_month else 0.0,
                    "order": round(order_series[cur][cur_month], 2),
                    "order_prev": round(order_series[cur].get(prev_month, 0.0), 2) if prev_month else 0.0,
                    "quote": round(quote_series[cur][cur_month], 2),
                    "quote_prev": round(quote_series[cur].get(prev_month, 0.0), 2) if prev_month else 0.0,
                }
            return out

        # ── 업무 알림 ───────────────────────────────────────────────────────────
        today_iso = now_kst.strftime("%Y-%m-%d")
        week_iso = (now_kst + timedelta(days=7)).strftime("%Y-%m-%d")
        cutoff60 = (now_kst - timedelta(days=60)).strftime("%Y-%m-%d")

        def _cust_of(o):
            return cust_names.get(o.customer_id, "—") if o else "—"

        today_delivery, week_delivery, uninvoiced, unreceived_po = [], [], [], []
        for o in orders_all:
            pd = o.promised_delivery or ""
            delivered = bool(o.delivered_date)
            rfq = _rfq_for_order(s, o)
            pno = _project_no_map(s).get(rfq.id, "") if rfq else ""
            if pd and not delivered:
                if pd[:10] == today_iso:
                    today_delivery.append({"order_id": o.id, "project_no": pno, "customer": _cust_of(o), "date": pd})
                elif today_iso < pd[:10] <= week_iso:
                    week_delivery.append({"order_id": o.id, "project_no": pno, "customer": _cust_of(o), "date": pd})
            # 미청구: 인도완료(delivered_date) 이나 세금계산서(11단계) 미발행
            sd = (getattr(rfq, "stage_dates", None) or {}) if rfq else {}
            if delivered and not sd.get("11"):
                uninvoiced.append({"order_id": o.id, "project_no": pno, "customer": _cust_of(o), "date": o.delivered_date or ""})

        # 미입고 발주: 발송된 Vendor P/O 인데 해당 오더가 아직 미인도
        for po in s.query(PurchaseOrder).all():
            o = order_map.get(po.order_id)
            sent = (po.status == "이메일 발송완료") or bool(po.sent_date)
            if sent and o and not o.delivered_date:
                rfq = _rfq_for_order(s, o)
                pno = _project_no_map(s).get(rfq.id, "") if rfq else ""
                unreceived_po.append({"order_id": o.id, "po_no": po.po_no or "", "project_no": pno, "customer": _cust_of(o), "date": po.sent_date or ""})

        # 미회신 견적: 발송/협상 상태 & 발송 후 7일 경과(수주·실주 아님)
        unanswered = []
        wk_ago = (now_kst - timedelta(days=7)).strftime("%Y-%m-%d")
        for q in s.query(Quotation).all():
            if q.status in (QuotationStatus.SENT, QuotationStatus.NEGOTIATING):
                sd = (getattr(q, "sent_at", None) or q.sent_date or q.date or "")[:10]
                if sd and sd <= wk_ago:
                    unanswered.append({"rfq_id": q.rfq_id, "qtn_no": q.qtn_no or "", "customer": cust_names.get(q.customer_id, "—"), "date": sd, "status": _enum_val(q.status)})

        # 장기 미수금: 연체 & 만기 60일 초과
        long_overdue = []
        for a in s.query(ARRecord).all():
            outstanding = (a.invoice_amount or 0) - (a.paid_amount or 0)
            if a.status != ARStatus.PAID and a.due_date and a.due_date < cutoff60 and outstanding > 0:
                o = order_map.get(a.order_id)
                long_overdue.append({"order_id": a.order_id, "ci_no": a.ci_no or "", "customer": _cust_of(o), "due_date": a.due_date, "currency": _cur2(a.currency), "outstanding": round(outstanding, 2)})

        # 납기 지연 건수: 약속납기 지난 미인도 오더
        delivery_delays = sum(
            1 for o in orders_all
            if o.promised_delivery and not o.delivered_date and o.promised_delivery[:10] < today_iso
        )

        return {
            "months": month_labels,
            "currencies": list(CURS),
            "series": {
                "revenue": _series_list(rev_series),
                "quote": _series_list(quote_series),
                "order": _series_list(order_series),
            },
            "customer_top": _top(cust_rev, "customer"),
            "item_top": _top(item_rev, "item"),
            "kpi": _kpi(),
            "delivery_delays": delivery_delays,
            "alerts": {
                "today_delivery": today_delivery,
                "week_delivery": week_delivery,
                "unanswered_quotes": unanswered,
                "unreceived_po": unreceived_po,
                "uninvoiced": uninvoiced,
                "long_overdue_ar": long_overdue,
            },
        }
    finally:
        s.close()


def _order_for_rfq(s, rfq_id: int):
    """RFQ에 연결된 Order — 직접 연결 우선, 없으면 Quotation 경유."""
    order = (s.query(Order).filter(Order.rfq_id == rfq_id)
             .order_by(Order.created_at.desc()).first())
    if not order:
        order = (s.query(Order).join(Quotation, Order.quotation_id == Quotation.id)
                 .filter(Quotation.rfq_id == rfq_id)
                 .order_by(Order.created_at.desc()).first())
    return order


def _rfq_for_order(s, order: Order):
    """Order에 연결된 RFQ — 직접 연결 우선, 없으면 Quotation 경유."""
    if order.rfq_id:
        return s.query(RFQ).filter_by(id=order.rfq_id).first()
    if order.quotation_id:
        qtn = s.query(Quotation).filter_by(id=order.quotation_id).first()
        if qtn and qtn.rfq_id:
            return s.query(RFQ).filter_by(id=qtn.rfq_id).first()
    return None


def _project_no_for_order(s, order) -> str:
    """오더의 Project No.(yymmdd-nn). 연결 RFQ 기준, 없으면 ''. (ord_no 대체)"""
    rfq = _rfq_for_order(s, order) if order else None
    return _project_no_map(s).get(rfq.id, "") if rfq else ""


def _base_meta(s, rfq, order=None) -> dict:
    """모든 상세 팝업 공통 기본정보.
    Project No.·최초 RFQ 수신일시·고객·선박·업무타입·거래구분(오더 있을 때만)."""
    customer = vessel = None
    if rfq and rfq.customer_id:
        customer = s.query(Customer).filter_by(id=rfq.customer_id).first()
    if order and not customer and getattr(order, "customer_id", None):
        customer = s.query(Customer).filter_by(id=order.customer_id).first()
    vid = (rfq.vessel_id if rfq else None) or (getattr(order, "vessel_id", None) if order else None)
    if vid:
        vessel = s.query(Vessel).filter_by(id=vid).first()
    return {
        "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
        "first_rfq_at": _first_rfq_iso(rfq) if rfq else "",
        "customer": customer.name if customer else "—",
        "vessel": vessel.name if vessel else "",
        "work_type": _enum_val(rfq.work_type) if (rfq and rfq.work_type) else "부품공급",
        "trade_type": (getattr(order, "trade_type", "") or "") if order else "",
        "project_title": (rfq.project_title or "") if rfq else "",
    }


@app.get("/api/admin/po-overview", dependencies=[Depends(require_token)])
def po_overview():
    """고객 P/O · Vendor P/O 현황 — RFQ → Order → PurchaseOrder 체인."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}

        rows = []
        for r in s.query(RFQ).order_by(RFQ.id.desc()).all():
            o = _order_for_rfq(s, r.id)
            vpos = (s.query(PurchaseOrder).filter_by(order_id=o.id)
                    .order_by(PurchaseOrder.id.desc()).all()) if o else []
            if vpos:
                vp0 = vpos[0]
                vendor_po_no = (vp0.po_no or "—") + (
                    f"  (외 {len(vpos) - 1}건)" if len(vpos) > 1 else "")
                vendor_nm = vendor_names.get(vp0.vendor_id, "—")
                vendor_email = vp0.sent_to_email or "—"
                # Vendor P/O 발신 일시 (시·분) — created_at 기준, 앱 전반 규칙과 동일
                vendor_po_at = _kst(vp0.created_at)
            else:
                vendor_po_no = vendor_nm = vendor_email = vendor_po_at = ""

            # Vendor RFQ 발신 일시 (시·분) — Vendor RFQ를 보낸 거래에서만
            vrfqs = (s.query(VendorRFQ).filter_by(rfq_id=r.id)
                     .order_by(VendorRFQ.id.desc()).all())
            vrfq_at = _fmt_received(_vrfq_sent_iso(vrfqs[0])) if vrfqs else ""

            stage = _pipeline_stage(s, r.id)
            rows.append({
                "id": o.id if o else 0,
                "customer_rfq_no": r.customer_rfq_no or "",
                "crfq_at": _kst(r.created_at),
                "kmaris_rfq_no": _rfq_no_disp(r.rfq_no),
                "vrfq_at": vrfq_at,
                "customer": cust_names.get(r.customer_id, "—"),
                "vessel": vessel_names.get(r.vessel_id, "") if r.vessel_id else "",
                "customer_po_no": (o.po_no if o else "") or "",
                # 고객 P/O 수신 일시 (시·분) — 시스템 수신(created_at) 기준
                "customer_po_at": _kst(o.created_at) if o else "",
                "item_count": len((o.items if o else None) or r.items or []),
                "vendor_po_no": vendor_po_no,
                "vendor_po_at": vendor_po_at,
                "vendor": vendor_nm,
                "vendor_email": vendor_email,
                "stage": stage,
                "status": _status_label(stage, r.work_type),
            })
        return {"rows": rows}
    finally:
        s.close()


@app.get("/api/admin/order/{order_id}", dependencies=[Depends(require_token)])
def order_detail(order_id: int):
    """Order 1건 상세 — 고객 P/O, Vendor P/O, 품목, 연결 문서."""
    s = get_session()
    try:
        o = s.query(Order).filter_by(id=order_id).first()
        if not o:
            raise HTTPException(status_code=404, detail="Order not found")

        cust = s.query(Customer).filter_by(id=o.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=o.vessel_id).first() if o.vessel_id else None
        rfq = _rfq_for_order(s, o)
        qtn = s.query(Quotation).filter_by(id=o.quotation_id).first() if o.quotation_id else None
        stage = _pipeline_stage(s, rfq.id) if rfq else 5

        steps = [{
            "no": i,
            "name": name,
            "state": ("done" if i < stage else "current" if i == stage else "todo"),
        } for i, name in enumerate(steps_for(rfq.work_type if rfq else None), start=1)]

        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        vendor_emails = {v.id: v.email for v in s.query(Vendor).all()}
        vpos = (s.query(PurchaseOrder).filter_by(order_id=o.id)
                .order_by(PurchaseOrder.id.desc()).all())
        vendor_po_view = [{
            "id": po.id,
            "po_no": po.po_no or "",
            "vendor": vendor_names.get(po.vendor_id, "—"),
            "vendor_email": po.sent_to_email or vendor_emails.get(po.vendor_id, "") or "",
            "date": po.date or "",
            "sent_date": po.sent_date or "",
            "status": po.status or "",
            "item_count": len(po.items or []),
        } for po in vpos]

        ci = s.query(CommercialInvoice).filter_by(order_id=o.id).order_by(CommercialInvoice.id.desc()).first()
        sa = s.query(ShippingAdvice).filter_by(order_id=o.id).order_by(ShippingAdvice.id.desc()).first()
        pl = (s.query(PackingList).filter_by(ci_id=ci.id).order_by(PackingList.id.desc()).first()
              if ci else None)
        tax = (s.query(TaxInvoiceData).filter_by(ci_id=ci.id).order_by(TaxInvoiceData.id.desc()).first()
               if ci else None)
        ars = s.query(ARRecord).filter_by(order_id=o.id).order_by(ARRecord.id.desc()).all()

        return {
            "id": o.id,
            "assignee_id": (rfq.created_by or 0) if rfq else 0,
            "customer_po_no": o.po_no or "",
            "customer_po_at": o.date or "",
            "rfq_no": _rfq_no_disp(rfq.rfq_no) if rfq else "",
            "customer_rfq_no": (rfq.customer_rfq_no or _rfq_no_disp(rfq.rfq_no)) if rfq else "",
            "quotation_no": qtn.qtn_no if qtn else "",
            "currency": (qtn.currency if qtn else "USD") or "USD",
            "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
            "first_rfq_at": _first_rfq_iso(rfq),
            "customer": cust.name if cust else "—",
            "customer_contact": (cust.contact if cust else "") or "",
            "customer_email": (cust.email if cust else "") or "",
            "vessel": vessel.name if vessel else "",
            "work_type": _enum_val(rfq.work_type) if (rfq and rfq.work_type) else "부품공급",
            "trade_type": o.trade_type or "수출",
            "project_title": (rfq.project_title or "") if rfq else "",
            "status": _status_label(stage, rfq.work_type) if rfq else _enum_val(o.status),
            "order_status": _enum_val(o.status),
            "stage": stage,
            "promised_delivery": o.promised_delivery or "",
            "shipped_date": o.shipped_date or "",
            "delivered_date": o.delivered_date or "",
            "tracking_token": o.tracking_token or "",
            "steps": steps,
            "items": [_item_view(it) for it in (o.items or [])],
            "vendor_pos": vendor_po_view,
            "documents": {
                "ci_no": ci.ci_no if ci else "",
                "pl_no": pl.pl_no if pl else "",
                "sa_no": sa.sa_no if sa else "",
                "tax_no": tax.tax_no if tax else "",
                "ar": [{
                    "ci_no": ar.ci_no or "",
                    "currency": ar.currency or "USD",
                    "invoice_amount": round(ar.invoice_amount or 0, 2),
                    "paid_amount": round(ar.paid_amount or 0, 2),
                    "due_date": ar.due_date or "",
                    "status": _enum_val(ar.status),
                } for ar in ars],
            },
        }
    finally:
        s.close()


@app.get("/api/admin/po-work-options", dependencies=[Depends(require_token)])
def po_work_options():
    """P/O 작업 탭용 옵션 — Streamlit Customer P/O / Vendor P/O 탭 데이터."""
    s = get_session()
    try:
        customers = [{"id": c.id, "name": c.name} for c in s.query(Customer).order_by(Customer.name).all()]
        vessels = [{
            "id": v.id,
            "name": v.name,
            "customer_id": v.customer_id,
        } for v in s.query(Vessel).order_by(Vessel.name).all()]
        vendors = [{
            "id": v.id,
            "name": v.name,
            "email": v.email or "",
        } for v in s.query(Vendor).order_by(Vendor.name).all()]

        cust_names = {c["id"]: c["name"] for c in customers}
        vessel_names = {v["id"]: v["name"] for v in vessels}
        user_names = {u.id: u.username for u in s.query(User).all()}

        rfqs = []
        for r in s.query(RFQ).order_by(RFQ.id.desc()).all():
            stage = _pipeline_stage(s, r.id)
            rfqs.append({
                "id": r.id,
                "rfq_no": _rfq_no_disp(r.rfq_no),
                "customer_rfq_no": r.customer_rfq_no or "",
                "customer_id": r.customer_id,
                "customer": cust_names.get(r.customer_id, "—"),
                "vessel_id": r.vessel_id,
                "vessel": vessel_names.get(r.vessel_id, "") if r.vessel_id else "",
                "status": _status_label(stage, r.work_type),
                "items": [_item_view(it) for it in (r.items or [])],
            })

        quotations = []
        for q in s.query(Quotation).order_by(Quotation.id.desc()).all():
            quotations.append({
                "id": q.id,
                "qtn_no": q.qtn_no,
                "rfq_id": q.rfq_id,
                "customer_id": q.customer_id,
                "customer": cust_names.get(q.customer_id, "—"),
                "vessel_id": q.vessel_id,
                "vessel": vessel_names.get(q.vessel_id, "") if q.vessel_id else "",
                "status": _enum_val(q.status),
                "currency": q.currency or "USD",
                "amount": round(_total_amount(q.items or []), 2),
                "items": [_item_view(it) for it in (q.items or [])],
            })

        orders = []
        for o in s.query(Order).order_by(Order.id.desc()).all():
            rfq = _rfq_for_order(s, o)
            qtn = s.query(Quotation).filter_by(id=o.quotation_id).first() if o.quotation_id else None
            stage = _pipeline_stage(s, rfq.id) if rfq else 5
            orders.append({
                "id": o.id,
                "customer_id": o.customer_id,
                "customer": cust_names.get(o.customer_id, "—"),
                "vessel_id": o.vessel_id,
                "vessel": vessel_names.get(o.vessel_id, "") if o.vessel_id else "",
                "po_no": o.po_no or "",
                "date": o.date or "",
                "trade_type": o.trade_type or "수출",
                "currency": (qtn.currency if qtn else "USD") or "USD",
                "status": _status_label(stage, rfq.work_type) if rfq else _enum_val(o.status),
                "items": [_item_view(it) for it in (o.items or [])],
                # 공통 식별 컬럼
                "project_title": (getattr(rfq, "project_title", None) or "") if rfq else "",
                "contact_person": (getattr(rfq, "contact_person", None) or "") if rfq else "",
                "assignee": (user_names.get(rfq.created_by, "") or "") if rfq else "",
                "assignee_id": (rfq.created_by or 0) if rfq else 0,
                "work_type": (_enum_val(rfq.work_type) if rfq and rfq.work_type else "부품공급"),
                "first_rfq_at": _first_rfq_iso(rfq),
                "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
            })

        purchase_orders = []
        for po in s.query(PurchaseOrder).order_by(PurchaseOrder.id.desc()).all():
            o = s.query(Order).filter_by(id=po.order_id).first()
            vendor = s.query(Vendor).filter_by(id=po.vendor_id).first()
            rfq = _rfq_for_order(s, o) if o else None
            qtn = s.query(Quotation).filter_by(id=o.quotation_id).first() if o and o.quotation_id else None
            purchase_orders.append({
                "id": po.id,
                "po_no": po.po_no or "",
                "order_id": po.order_id,
                "customer_po_no": (o.po_no if o else "") or "",
                "vendor_id": po.vendor_id,
                "vendor": vendor.name if vendor else "—",
                "vendor_email": po.sent_to_email or (vendor.email if vendor else "") or "",
                "date": po.date or "",
                "sent_date": po.sent_date or "",
                "status": po.status or "",
                "sent": po.status == "이메일 발송완료",
                "items": [_item_view(it) for it in (po.items or [])],
                "currency": (qtn.currency if qtn else "USD") or "USD",
                # 공통 식별 컬럼
                "customer": cust_names.get(o.customer_id, "—") if o else "—",
                "project_title": (getattr(rfq, "project_title", None) or "") if rfq else "",
                "contact_person": (getattr(rfq, "contact_person", None) or "") if rfq else "",
                "assignee": (user_names.get(rfq.created_by, "") or "") if rfq else "",
                "assignee_id": (rfq.created_by or 0) if rfq else 0,
                "vessel": (vessel_names.get(o.vessel_id, "") if o and o.vessel_id else ""),
                "trade_type": (o.trade_type or "수출") if o else "수출",
                "work_type": (_enum_val(rfq.work_type) if rfq and rfq.work_type else "부품공급"),
                "first_rfq_at": _first_rfq_iso(rfq),
                "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
            })

        return {
            "customers": customers,
            "vessels": vessels,
            "vendors": vendors,
            "rfqs": rfqs,
            "quotations": quotations,
            "orders": orders,
            "purchase_orders": purchase_orders,
            "smtp_configured": bool(os.getenv("SMTP_USER") and os.getenv("SMTP_PASSWORD")),
        }
    finally:
        s.close()


_IMAGE_MEDIA = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".webp": "image/webp", ".gif": "image/gif",
}


def _ocr_image_media_type(file: UploadFile) -> str | None:
    """업로드가 이미지면 Claude 비전용 media_type 반환, 아니면 None."""
    fname = (file.filename or "").lower()
    for ext, mt in _IMAGE_MEDIA.items():
        if fname.endswith(ext):
            return mt
    ct = (file.content_type or "").lower()
    if ct.startswith("image/") and ct in _IMAGE_MEDIA.values():
        return ct
    return None


@app.post("/api/admin/ocr/rfq", dependencies=[Depends(require_token)])
def ocr_rfq_pdf(file: UploadFile = File(...)):
    """Customer RFQ 자동 입력 — PDF(텍스트 추출) 또는 이미지/캡쳐(Claude 비전) 지원."""
    s = get_session()
    try:
        customer_names = [c.name for c in s.query(Customer).order_by(Customer.name).all()]
    finally:
        s.close()
    fname = (file.filename or "").lower()
    img_media = _ocr_image_media_type(file)
    try:
        file.file.seek(0)
        if img_media:
            return parse_rfq_image(file.file.read(), img_media, customer_names)
        if fname.endswith(".pdf"):
            raw_text = extract_text_from_pdf(file.file)
            if not raw_text:
                raise HTTPException(status_code=400, detail="PDF에서 텍스트를 추출할 수 없습니다.")
            return parse_rfq_fields(raw_text, customer_names)
        raise HTTPException(status_code=400, detail="PDF 또는 이미지(PNG·JPG·WEBP) 파일만 업로드할 수 있습니다.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"OCR 추출 실패: {exc}") from exc


@app.post("/api/admin/ocr/order", dependencies=[Depends(require_token)])
def ocr_order_pdf(file: UploadFile = File(...)):
    """Customer P/O 자동 입력 — PDF 또는 이미지/캡쳐 지원."""
    s = get_session()
    try:
        customer_names = [c.name for c in s.query(Customer).order_by(Customer.name).all()]
    finally:
        s.close()
    fname = (file.filename or "").lower()
    img_media = _ocr_image_media_type(file)
    try:
        file.file.seek(0)
        if img_media:
            return parse_order_image(file.file.read(), img_media, customer_names)
        if fname.endswith(".pdf"):
            raw_text = extract_text_from_pdf(file.file)
            if not raw_text:
                raise HTTPException(status_code=400, detail="PDF에서 텍스트를 추출할 수 없습니다.")
            return parse_order_fields(raw_text, customer_names)
        raise HTTPException(status_code=400, detail="PDF 또는 이미지(PNG·JPG·WEBP) 파일만 업로드할 수 있습니다.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"OCR 추출 실패: {exc}") from exc



class PoWorkItem(BaseModel):
    part_no: str = ""
    description: str = ""
    maker: str = ""
    qty: float = 1
    unit: str = "PCS"
    unit_price: float | None = 0
    amount: float | None = None
    remark: str | None = ""


class OrderCreate(BaseModel):
    customer_id: int
    vessel_id: int | None = None
    quotation_id: int | None = None
    rfq_id: int | None = None
    po_no: str = ""
    date: str | None = None
    trade_type: str = "수출"
    promised_delivery: str | None = None
    items: list[PoWorkItem] = []


@app.post("/api/admin/orders", dependencies=[Depends(require_token)])
def create_order(body: OrderCreate):
    """Customer P/O 수신 탭 — 신규 오더 등록."""
    s = get_session()
    try:
        cust = s.query(Customer).filter_by(id=body.customer_id).first()
        if not cust:
            raise HTTPException(status_code=400, detail="Customer를 선택하세요.")
        qtn = s.query(Quotation).filter_by(id=body.quotation_id).first() if body.quotation_id else None
        rfq_id = body.rfq_id or (qtn.rfq_id if qtn else None)
        items = []
        for it in body.items:
            if not (it.part_no or it.description):
                continue
            qty = it.qty or 1
            unit_price = it.unit_price or 0
            items.append({
                "part_no": it.part_no.strip(),
                "description": it.description.strip(),
                "maker": it.maker.strip(),
                "qty": qty,
                "unit": it.unit or "PCS",
                "unit_price": unit_price,
                "amount": it.amount if it.amount is not None else qty * unit_price,
                "remark": (it.remark or "").strip(),
            })

        order = Order(
            quotation_id=qtn.id if qtn else None,
            rfq_id=rfq_id,
            customer_id=cust.id,
            vessel_id=body.vessel_id,
            po_no=(body.po_no or "").strip(),
            date=body.date or date.today().isoformat(),
            trade_type=(body.trade_type or "수출").strip() or "수출",
            promised_delivery=body.promised_delivery or None,
            status=OrderStatus.RECEIVED,
            items=items,
        )
        s.add(order)
        if qtn:
            qtn.status = QuotationStatus.WON
        s.commit()
        return {"ok": True, "id": order.id, "project_no": _project_no_for_order(s, order)}
    finally:
        s.close()


class OrderUpdate(BaseModel):
    customer_id: int | None = None
    vessel_id: int | None = None       # 0 = 선박 미지정 해제
    po_no: str | None = None
    date: str | None = None
    trade_type: str | None = None
    promised_delivery: str | None = None
    items: list[PoWorkItem] | None = None


@app.put("/api/admin/orders/{order_id}", dependencies=[Depends(require_token)])
def update_order(order_id: int, body: OrderUpdate):
    """오더 수정 — 헤더 필드 + 품목 리스트 교체."""
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="오더를 찾을 수 없습니다.")
        if body.customer_id is not None:
            order.customer_id = body.customer_id
        if body.vessel_id is not None:
            order.vessel_id = body.vessel_id or None
        if body.po_no is not None:
            order.po_no = body.po_no.strip()
        if body.date is not None:
            order.date = body.date or order.date
        if body.trade_type is not None:
            order.trade_type = (body.trade_type or "수출").strip() or "수출"
        if body.promised_delivery is not None:
            order.promised_delivery = body.promised_delivery or None
        if body.items is not None:
            items = []
            for it in body.items:
                if not (it.part_no or it.description):
                    continue
                qty = it.qty or 1
                unit_price = it.unit_price or 0
                items.append({
                    "part_no": it.part_no.strip(),
                    "description": it.description.strip(),
                    "maker": it.maker.strip(),
                    "qty": qty,
                    "unit": it.unit or "PCS",
                    "unit_price": unit_price,
                    "amount": it.amount if it.amount is not None else qty * unit_price,
                    "remark": (it.remark or "").strip(),
                })
            order.items = items
        s.commit()
        return {"ok": True, "id": order.id, "project_no": _project_no_for_order(s, order)}
    finally:
        s.close()


@app.delete("/api/admin/orders/{order_id}", dependencies=[Depends(require_token)])
def delete_order(order_id: int):
    """오더 삭제 — 발주서·문서·AR 등 다운스트림이 있으면 데이터 보호를 위해 거부한다."""
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="오더를 찾을 수 없습니다.")
        if s.query(PurchaseOrder).filter_by(order_id=order_id).first():
            raise HTTPException(status_code=400,
                detail="발주서(Vendor P/O)가 연결된 오더입니다. 먼저 발주서를 삭제하세요.")
        if s.query(CommercialInvoice).filter_by(order_id=order_id).first() or \
           s.query(ShippingAdvice).filter_by(order_id=order_id).first():
            raise HTTPException(status_code=400,
                detail="선적/송장 문서가 연결된 오더입니다. 삭제할 수 없습니다.")
        if s.query(ARRecord).filter_by(order_id=order_id).first():
            raise HTTPException(status_code=400,
                detail="AR(미수금) 기록이 연결된 오더입니다. 삭제할 수 없습니다.")
        project_no = _project_no_for_order(s, order)
        # 연결된 견적 상태(수주확정)는 되돌리지 않는다(별도 화면에서 관리).
        s.query(DeliveryProof).filter_by(order_id=order_id).delete(synchronize_session=False)
        s.query(Order).filter_by(id=order_id).delete(synchronize_session=False)
        s.commit()
        return {"ok": True, "project_no": project_no}
    finally:
        s.close()


class PurchaseOrderCreate(BaseModel):
    order_id: int
    vendor_id: int
    po_no: str | None = None
    date: str | None = None
    items: list[PoWorkItem] = []


@app.post("/api/admin/vendor-pos", dependencies=[Depends(require_token)])
def create_purchase_order(body: PurchaseOrderCreate):
    """Vendor P/O 발신 탭 — 발주서 생성."""
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=body.order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="오더를 찾을 수 없습니다.")
        vendor = s.query(Vendor).filter_by(id=body.vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=400, detail="Vendor를 선택하세요.")

        items = []
        for it in body.items:
            if not (it.part_no or it.description):
                continue
            qty = it.qty or 1
            unit_price = it.unit_price or 0
            items.append({
                "part_no": it.part_no.strip(),
                "description": it.description.strip(),
                "maker": it.maker.strip(),
                "qty": qty,
                "unit": it.unit or "PCS",
                "unit_price": unit_price,
                "amount": it.amount if it.amount is not None else qty * unit_price,
            })

        # 수동 입력. 비우면 번호 없이 저장(나중에 편집에서 채움).
        po_no = (body.po_no or "").strip() or None
        if po_no and s.query(PurchaseOrder).filter_by(po_no=po_no).first():
            raise HTTPException(status_code=400, detail=f"이미 존재하는 PO No. 입니다: {po_no}")
        po = PurchaseOrder(
            po_no=po_no,
            order_id=order.id,
            vendor_id=vendor.id,
            date=body.date or date.today().isoformat(),
            items=items,
            status="발주완료",
        )
        s.add(po)
        order.status = OrderStatus.PO_SENT
        s.commit()
        return {"ok": True, "id": po.id, "po_no": po.po_no}
    finally:
        s.close()


class PurchaseOrderUpdate(BaseModel):
    vendor_id: int | None = None
    po_no: str | None = None
    date: str | None = None
    sent_date: str | None = None
    status: str | None = None
    items: list[PoWorkItem] | None = None


@app.get("/api/admin/vendor-pos/{po_id}", dependencies=[Depends(require_token)])
def vendor_po_detail(po_id: int):
    """Vendor P/O(발주서) 1건 상세."""
    s = get_session()
    try:
        po = s.query(PurchaseOrder).filter_by(id=po_id).first()
        if not po:
            raise HTTPException(status_code=404, detail="발주서를 찾을 수 없습니다.")
        order = s.query(Order).filter_by(id=po.order_id).first()
        vendor = s.query(Vendor).filter_by(id=po.vendor_id).first()
        rfq = _rfq_for_order(s, order) if order else None
        qtn = s.query(Quotation).filter_by(id=order.quotation_id).first() if order and order.quotation_id else None
        return {
            "id": po.id,
            "po_no": po.po_no or "",
            "order_id": po.order_id,
            "assignee_id": (rfq.created_by or 0) if rfq else 0,
            "customer_po_no": (order.po_no if order else "") or "",
            **_base_meta(s, rfq, order),   # 공통 기본정보
            "vendor_id": po.vendor_id or 0,
            "vendor": vendor.name if vendor else "—",
            "vendor_email": po.sent_to_email or (vendor.email if vendor else "") or "",
            "date": po.date or "",
            "sent_date": po.sent_date or "",
            "status": po.status or "",
            "sent": po.status == "이메일 발송완료",
            "currency": (qtn.currency if qtn else "USD") or "USD",
            "items": [_item_view(it) for it in (po.items or [])],
        }
    finally:
        s.close()


@app.put("/api/admin/vendor-pos/{po_id}", dependencies=[Depends(require_token)])
def update_purchase_order(po_id: int, body: PurchaseOrderUpdate):
    """발주서 수정 — Vendor·발주일·상태·품목 교체."""
    s = get_session()
    try:
        po = s.query(PurchaseOrder).filter_by(id=po_id).first()
        if not po:
            raise HTTPException(status_code=404, detail="발주서를 찾을 수 없습니다.")
        if body.vendor_id is not None:
            po.vendor_id = body.vendor_id
        if body.po_no is not None:
            new_no = body.po_no.strip()
            if new_no and new_no != po.po_no:
                if s.query(PurchaseOrder).filter(PurchaseOrder.po_no == new_no, PurchaseOrder.id != po.id).first():
                    raise HTTPException(status_code=400, detail=f"이미 존재하는 PO No. 입니다: {new_no}")
                po.po_no = new_no
        if body.date is not None:
            po.date = body.date or po.date
        if body.sent_date is not None:
            # 메일 발송 없이도 수동으로 발송일 입력 가능 (빈 값이면 해제)
            po.sent_date = body.sent_date.strip() or None
        if body.status is not None:
            po.status = body.status.strip() or po.status
        if body.items is not None:
            items = []
            for it in body.items:
                if not (it.part_no or it.description):
                    continue
                qty = it.qty or 1
                unit_price = it.unit_price or 0
                items.append({
                    "part_no": it.part_no.strip(),
                    "description": it.description.strip(),
                    "maker": it.maker.strip(),
                    "qty": qty,
                    "unit": it.unit or "PCS",
                    "unit_price": unit_price,
                    "amount": it.amount if it.amount is not None else qty * unit_price,
                    "remark": (it.remark or "").strip(),
                })
            po.items = items
        s.commit()
        return {"ok": True, "id": po.id, "po_no": po.po_no}
    finally:
        s.close()


@app.delete("/api/admin/vendor-pos/{po_id}", dependencies=[Depends(require_token)])
def delete_purchase_order(po_id: int):
    """발주서 삭제."""
    s = get_session()
    try:
        po = s.query(PurchaseOrder).filter_by(id=po_id).first()
        if not po:
            raise HTTPException(status_code=404, detail="발주서를 찾을 수 없습니다.")
        po_no = po.po_no
        s.query(PurchaseOrder).filter_by(id=po_id).delete(synchronize_session=False)
        s.commit()
        return {"ok": True, "po_no": po_no}
    finally:
        s.close()


class VendorPoPreview(BaseModel):
    lang: str = "en"
    notes: str = ""


@app.post("/api/admin/vendor-pos/{po_id}/preview", dependencies=[Depends(require_token)])
def vendor_po_preview(po_id: int, body: VendorPoPreview):
    s = get_session()
    try:
        po = s.query(PurchaseOrder).filter_by(id=po_id).first()
        if not po:
            raise HTTPException(status_code=404, detail="발주서를 찾을 수 없습니다.")
        vendor = s.query(Vendor).filter_by(id=po.vendor_id).first()
        order = s.query(Order).filter_by(id=po.order_id).first()
        vessel = s.query(Vessel).filter_by(id=order.vessel_id).first() if order and order.vessel_id else None
        lang = "ko" if body.lang == "ko" else "en"
        subject = (
            f"[K-MARIS] 발주서 송부 — {po.po_no} / {vessel.name if vessel else po.po_no}"
            if lang == "ko"
            else f"[K-MARIS] Purchase Order — {po.po_no} / {vessel.name if vessel else po.po_no}"
        )
        return {
            "to": (vendor.email if vendor else "") or "",
            "subject": subject,
            "body": _vendor_po_email_body(po, vendor, order, vessel, body.notes, lang, _project_no_for_order(s, order)),
            "pdf_filename": f"{po.po_no}_PurchaseOrder.pdf",
            "smtp_configured": bool(os.getenv("SMTP_USER") and os.getenv("SMTP_PASSWORD")),
        }
    finally:
        s.close()


@app.get("/api/admin/vendor-pos/{po_id}/pdf", dependencies=[Depends(require_token)])
def vendor_po_pdf(po_id: int):
    s = get_session()
    try:
        po = s.query(PurchaseOrder).filter_by(id=po_id).first()
        if not po:
            raise HTTPException(status_code=404, detail="발주서를 찾을 수 없습니다.")
        vendor = s.query(Vendor).filter_by(id=po.vendor_id).first()
        order = s.query(Order).filter_by(id=po.order_id).first()
        vessel = s.query(Vessel).filter_by(id=order.vessel_id).first() if order and order.vessel_id else None
        payload = build_po_payload(
            po_no=po.po_no,
            date=po.date or date.today().isoformat(),
            vendor=vendor,
            vessel=vessel,
            items=po.items or [],
        )
        pdf = generate_po_pdf(payload)
        return Response(
            content=pdf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{po.po_no}_PurchaseOrder.pdf"'},
        )
    finally:
        s.close()


class VendorPoSend(BaseModel):
    to: str
    subject: str
    body: str


@app.post("/api/admin/vendor-pos/{po_id}/send", dependencies=[Depends(require_token)])
def vendor_po_send(po_id: int, body: VendorPoSend):
    s = get_session()
    try:
        po = s.query(PurchaseOrder).filter_by(id=po_id).first()
        if not po:
            raise HTTPException(status_code=404, detail="발주서를 찾을 수 없습니다.")
        if not body.to.strip():
            raise HTTPException(status_code=400, detail="수신자 이메일을 입력하세요.")
        vendor = s.query(Vendor).filter_by(id=po.vendor_id).first()
        order = s.query(Order).filter_by(id=po.order_id).first()
        vessel = s.query(Vessel).filter_by(id=order.vessel_id).first() if order and order.vessel_id else None
        payload = build_po_payload(
            po_no=po.po_no,
            date=po.date or date.today().isoformat(),
            vendor=vendor,
            vessel=vessel,
            items=po.items or [],
        )
        pdf = generate_po_pdf(payload)
        sent = send_email(
            to=body.to.strip(),
            subject=body.subject,
            body=body.body,
            attachments=[(f"{po.po_no}_PurchaseOrder.pdf", pdf)],
        )
        if not sent:
            raise HTTPException(status_code=400, detail="이메일 발송 실패 — SMTP 설정 또는 서버 상태를 확인하세요.")
        po.status = "이메일 발송완료"
        po.sent_to_email = body.to.strip()
        po.sent_date = date.today().isoformat()
        s.commit()
        return {"ok": True, "sent_date": po.sent_date}
    finally:
        s.close()


@app.get("/api/admin/ar-overview", dependencies=[Depends(require_token)])
def ar_overview():
    """미수금(AR) 현황 — 청구/수금/연체."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        ord_map = {o.id: o for o in s.query(Order).all()}
        # order_id → 발주 Vendor 이름들(중복 제거, 발주 순)
        po_vendors_by_order: dict[int, list[str]] = {}
        for po in s.query(PurchaseOrder).order_by(PurchaseOrder.id).all():
            nm = vendor_names.get(po.vendor_id)
            if not nm:
                continue
            lst = po_vendors_by_order.setdefault(po.order_id, [])
            if nm not in lst:
                lst.append(nm)
        today_str = date.today().isoformat()

        rows = []
        out_usd = 0.0
        overdue_usd = 0.0
        for r in s.query(ARRecord).order_by(ARRecord.id.desc()).all():
            o = ord_map.get(r.order_id)
            cust = cust_names.get(o.customer_id, "—") if o else "—"
            outstanding = (r.invoice_amount or 0) - (r.paid_amount or 0)
            overdue = (r.status != ARStatus.PAID and r.due_date
                       and r.due_date < today_str)
            status = "연체" if overdue else _enum_val(r.status)
            if (r.currency or "USD") == "USD" and r.status != ARStatus.PAID:
                out_usd += outstanding
                if overdue:
                    overdue_usd += outstanding
            # 11) 세금계산서 발행 · 12) 대금 결제 완료 — RFQ.stage_dates 의 수동 완료 표시
            rfq = _rfq_for_order(s, o) if o else None
            sd = (getattr(rfq, "stage_dates", None) or {}) if rfq else {}
            rows.append({
                "id": r.id,
                "order_id": r.order_id,
                "assignee_id": (rfq.created_by or 0) if rfq else 0,
                "assignee": (user_names.get(rfq.created_by, "") or "") if rfq else "",
                "project_title": (getattr(rfq, "project_title", None) or "") if rfq else "",
                "contact_person": (getattr(rfq, "contact_person", None) or "") if rfq else "",
                "ci_no": r.ci_no or "",
                "customer": cust,
                "currency": r.currency or "USD",
                "invoice_amount": round(r.invoice_amount or 0, 2),
                "paid_amount": round(r.paid_amount or 0, 2),
                "outstanding": round(outstanding, 2),
                "due_date": r.due_date or "",
                "status": status,
                "overdue": bool(overdue),
                "notes": r.notes or "",
                "tax_issued": bool(sd.get("11")),
                "tax_issued_date": sd.get("11", "") or "",
                "paid_done": bool(sd.get("12")),
                "paid_date": sd.get("12", "") or "",
                # 공통 식별 컬럼
                "vessel": (vessel_names.get(o.vessel_id, "") if o and o.vessel_id else ""),
                "trade_type": (o.trade_type or "수출") if o else "수출",
                "work_type": (_enum_val(rfq.work_type) if rfq and rfq.work_type else "부품공급"),
                "vendor": (lambda v: (v[0] + (f"  (외 {len(v) - 1}곳)" if len(v) > 1 else "")) if v else "")(po_vendors_by_order.get(r.order_id, [])),
                "first_rfq_at": _first_rfq_iso(rfq),
                "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
            })
        return {
            "kpi": {
                "outstanding_usd": round(out_usd, 2),
                "overdue_usd": round(overdue_usd, 2),
                "count": len(rows),
            },
            "rows": rows,
        }
    finally:
        s.close()


@app.get("/api/admin/ar/soa.xlsx", dependencies=[Depends(require_token)])
def ar_soa_xlsx(status: str | None = None, currency: str | None = None):
    """Statement of Account (SOA) XLSX 내보내기 — AR 현황을 엑셀로 추출한다.
    AR 페이지의 status/currency 필터를 그대로 적용하고 통화별 합계를 덧붙인다."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        ord_map = {o.id: o for o in s.query(Order).all()}
        today_str = date.today().isoformat()

        wb = Workbook()
        ws = wb.active
        ws.title = "SOA"
        ws.append(["Statement of Account (Accounts Receivable)"])
        ws.append([f"Generated: {today_str}"])
        active_filters = []
        if status:
            active_filters.append(f"Status={status}")
        if currency:
            active_filters.append(f"Currency={currency}")
        ws.append(["Filter: " + (", ".join(active_filters) if active_filters else "All")])
        ws.append([])

        headers = ["CI No.", "Customer", "Project No.", "Currency", "Invoice",
                   "Paid", "Outstanding", "Due Date", "Status"]
        ws.append(headers)
        head_row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=head_row, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3A5F")

        totals: dict[str, list[float]] = {}
        for r in s.query(ARRecord).order_by(ARRecord.id.desc()).all():
            o = ord_map.get(r.order_id)
            cust = cust_names.get(o.customer_id, "—") if o else "—"
            cur = r.currency or "USD"
            overdue = (r.status != ARStatus.PAID and r.due_date and r.due_date < today_str)
            st_label = "연체" if overdue else _enum_val(r.status)
            if status and status not in (st_label, _enum_val(r.status)):
                continue
            if currency and cur != currency:
                continue
            invoice = round(r.invoice_amount or 0, 2)
            paid = round(r.paid_amount or 0, 2)
            outstanding = round(invoice - paid, 2)
            ws.append([r.ci_no or "—", cust, _project_no_for_order(s, o) if o else "—", cur,
                       invoice, paid, outstanding, r.due_date or "—", st_label])
            t = totals.setdefault(cur, [0.0, 0.0, 0.0])
            t[0] += invoice
            t[1] += paid
            t[2] += outstanding

        ws.append([])
        for cur, (inv, paid, out) in sorted(totals.items()):
            ws.append([f"TOTAL ({cur})", "", "", cur, round(inv, 2),
                       round(paid, 2), round(out, 2), "", ""])
            total_row = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(row=total_row, column=c).font = Font(bold=True)

        for col, width in zip("ABCDEFGHI", [16, 22, 14, 9, 14, 14, 14, 12, 10]):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="SOA_{today_str}.xlsx"'},
        )
    finally:
        s.close()


class ARPayment(BaseModel):
    amount: float
    due_date: str | None = None


class ARSave(BaseModel):
    order_id: int
    ci_no: str | None = ""
    invoice_amount: float = 0.0
    paid_amount: float = 0.0
    currency: str = "USD"
    due_date: str | None = None
    status: str = ""
    notes: str | None = ""


def _ar_status_from_text(value: str | None, paid: float, invoice: float) -> ARStatus:
    if paid >= invoice and invoice > 0:
        return ARStatus.PAID
    if paid > 0:
        return ARStatus.PARTIAL
    if value:
        for status in ARStatus:
            if value in {status.value, status.name}:
                return status
    return ARStatus.OUTSTANDING


@app.post("/api/admin/ar", dependencies=[Depends(require_token)])
def create_ar(body: ARSave):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=body.order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        ar = ARRecord(
            order_id=body.order_id,
            ci_no=body.ci_no or "",
            invoice_amount=body.invoice_amount or 0.0,
            paid_amount=body.paid_amount or 0.0,
            currency=body.currency or "USD",
            due_date=body.due_date,
            status=_ar_status_from_text(body.status, body.paid_amount or 0.0, body.invoice_amount or 0.0),
            notes=body.notes or "",
        )
        s.add(ar)
        s.commit()
        return {"ok": True, "id": ar.id}
    finally:
        s.close()


@app.put("/api/admin/ar/{ar_id}", dependencies=[Depends(require_token)])
def update_ar(ar_id: int, body: ARSave):
    s = get_session()
    try:
        ar = s.query(ARRecord).filter_by(id=ar_id).first()
        if not ar:
            raise HTTPException(status_code=404, detail="AR 레코드를 찾을 수 없습니다.")
        if not s.query(Order).filter_by(id=body.order_id).first():
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        ar.order_id = body.order_id
        ar.ci_no = body.ci_no or ""
        ar.invoice_amount = body.invoice_amount or 0.0
        ar.paid_amount = body.paid_amount or 0.0
        ar.currency = body.currency or "USD"
        ar.due_date = body.due_date
        ar.status = _ar_status_from_text(body.status, ar.paid_amount or 0.0, ar.invoice_amount or 0.0)
        ar.notes = body.notes or ""
        s.commit()
        return {"ok": True, "id": ar.id, "status": _enum_val(ar.status)}
    finally:
        s.close()


@app.delete("/api/admin/ar/{ar_id}", dependencies=[Depends(require_token)])
def delete_ar(ar_id: int):
    s = get_session()
    try:
        ar = s.query(ARRecord).filter_by(id=ar_id).first()
        if not ar:
            raise HTTPException(status_code=404, detail="AR 레코드를 찾을 수 없습니다.")
        s.delete(ar)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.post("/api/admin/ar/{ar_id}/payment", dependencies=[Depends(require_token)])
def ar_payment(ar_id: int, body: ARPayment):
    """수금 등록 — paid_amount 누적 후 상태 자동 갱신."""
    s = get_session()
    try:
        ar = s.query(ARRecord).filter_by(id=ar_id).first()
        if not ar:
            raise HTTPException(status_code=404, detail="AR 레코드를 찾을 수 없습니다.")
        if body.amount <= 0:
            raise HTTPException(status_code=400, detail="수금액은 0보다 커야 합니다.")
        ar.paid_amount = (ar.paid_amount or 0) + body.amount
        if body.due_date:
            ar.due_date = body.due_date
        if ar.paid_amount >= (ar.invoice_amount or 0):
            ar.status = ARStatus.PAID
        elif ar.paid_amount > 0:
            ar.status = ARStatus.PARTIAL
        else:
            ar.status = ARStatus.OUTSTANDING
        s.commit()
        return {"ok": True, "paid_amount": ar.paid_amount, "status": _enum_val(ar.status)}
    finally:
        s.close()


def _quotation_total(items, discount_pct: float = 0.0) -> float:
    """견적 최종 총액 — amount 합계(없으면 unit_price*qty 보정)에 할인율 적용."""
    amt = _total_amount(items)
    if not amt:
        tot = 0.0
        for it in (items or []):
            try:
                tot += float(it.get("unit_price", 0) or 0) * float(it.get("qty", 1) or 1)
            except (TypeError, ValueError):
                pass
        amt = tot
    try:
        disc = float(discount_pct or 0)
    except (TypeError, ValueError):
        disc = 0.0
    return amt * (1 - disc / 100.0)


@app.get("/api/admin/quotation-overview", dependencies=[Depends(require_token)])
def quotation_overview(customer_id: int | None = None,
                       mine: int = 0, assignee: int | None = None,
                       user: dict = Depends(get_current_user)):
    """Customer Quotation 현황 — 견적 목록(고객/선박/금액/상태/파이프라인)."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        rfqs_all = s.query(RFQ).all()
        rfq_map = {r.id: r for r in rfqs_all}
        rfq_nos = {r.id: _rfq_no_disp(r.rfq_no) for r in rfqs_all}
        rfq_wt = {r.id: r.work_type for r in rfqs_all}

        q = s.query(Quotation)
        if customer_id:
            q = q.filter(Quotation.customer_id == customer_id)
        q = _apply_owner_filter(q, Quotation, user, mine, assignee)

        rows = []
        for qt in q.order_by(Quotation.id.desc()).all():
            stage = _pipeline_stage(s, qt.rfq_id) if qt.rfq_id else 0
            rows.append({
                "id": qt.id,
                "rfq_id": qt.rfq_id,
                "qtn_no": qt.qtn_no,
                "rfq_no": rfq_nos.get(qt.rfq_id, "") if qt.rfq_id else "",
                "customer": cust_names.get(qt.customer_id, "—"),
                "assignee": (user_names.get(getattr(rfq_map.get(qt.rfq_id), "created_by", None), "") or "") if qt.rfq_id else "",
                "assignee_id": (getattr(rfq_map.get(qt.rfq_id), "created_by", None) or 0) if qt.rfq_id else 0,
                "project_title": (getattr(rfq_map.get(qt.rfq_id), "project_title", None) or "") if qt.rfq_id else "",
                "contact_person": (getattr(rfq_map.get(qt.rfq_id), "contact_person", None) or "") if qt.rfq_id else "",
                "vessel": vessel_names.get(qt.vessel_id, "") if qt.vessel_id else "",
                "currency": qt.currency or "USD",
                "amount": round(_quotation_total(qt.items or [], getattr(qt, "discount_pct", 0) or 0), 2),
                "item_count": len(qt.items or []),
                "status": _enum_val(qt.status),
                "level": _enum_val(qt.follow_up_level) if qt.follow_up_level else "",
                "valid_until": qt.valid_until or "",
                "sent_at": getattr(qt, "sent_at", None) or "",
                "sent_date": getattr(qt, "sent_at", None) or qt.sent_date or "",
                "date": qt.date or "",
                "stage": stage,
                "pipeline": _status_label(stage, rfq_wt.get(qt.rfq_id)) if stage else "",
                # 공통 식별 컬럼
                "work_type": (_enum_val(rfq_wt.get(qt.rfq_id)) if rfq_wt.get(qt.rfq_id) else "부품공급"),
                "first_rfq_at": _first_rfq_iso(rfq_map.get(qt.rfq_id)),
                "project_no": _project_no_map(s).get(qt.rfq_id, "") if qt.rfq_id else "",
            })
        return {"rows": rows}
    finally:
        s.close()


@app.get("/api/admin/vrfq-overview", dependencies=[Depends(require_token)])
def vrfq_overview():
    """Vendor RFQ 발신 내역 — VendorRFQ 1건당 1행(고객 RFQ·Vendor·수신 견적 수)."""
    s = get_session()
    try:
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        vendor_emails = {v.id: (v.email or "") for v in s.query(Vendor).all()}
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        rfq_map = {r.id: r for r in s.query(RFQ).all()}

        quote_counts: dict[int, int] = {}
        for vq in s.query(VendorQuote).all():
            quote_counts[vq.vendor_rfq_id] = quote_counts.get(vq.vendor_rfq_id, 0) + 1

        rows = []
        for vr in s.query(VendorRFQ).order_by(VendorRFQ.id.desc()).all():
            rfq = rfq_map.get(vr.rfq_id)
            rows.append({
                "id": vr.id,
                "rfq_id": vr.rfq_id,
                # 고객 RFQ No.는 1단계의 고객 참조번호(없으면 "—"). K-Maris RFQ No.가 아님.
                "customer_rfq_no": (rfq.customer_rfq_no or "—") if rfq else "—",
                "vendor": vendor_names.get(vr.vendor_id, "—"),
                "vendor_email": vr.sent_to_email or vendor_emails.get(vr.vendor_id, "") or "",
                "sent_date": vr.sent_date or "",
                "status": vr.status or "",
                "item_count": len(vr.items or []),
                "quote_count": quote_counts.get(vr.id, 0),
                # 공통 식별 컬럼
                "customer": cust_names.get(rfq.customer_id, "—") if rfq else "—",
                "project_title": (getattr(rfq, "project_title", None) or "") if rfq else "",
                "contact_person": (getattr(rfq, "contact_person", None) or "") if rfq else "",
                "assignee": (user_names.get(rfq.created_by, "") or "") if rfq else "",
                "assignee_id": (rfq.created_by or 0) if rfq else 0,
                "level": (_enum_val(rfq.follow_up_level) if rfq and rfq.follow_up_level else "B"),
                "vessel": (vessel_names.get(rfq.vessel_id, "") if rfq and rfq.vessel_id else ""),
                "work_type": (_enum_val(rfq.work_type) if rfq and rfq.work_type else "부품공급"),
                "first_rfq_at": _first_rfq_iso(rfq),
                "project_no": _project_no_map(s).get(vr.rfq_id, "") if vr.rfq_id else "",
            })
        return {"rows": rows}
    finally:
        s.close()


@app.get("/api/admin/vendor-quote-overview", dependencies=[Depends(require_token)])
def vendor_quote_overview():
    """Vendor Quote 수신 내역 — VendorQuote 1건당 1행(전체 프로젝트)."""
    s = get_session()
    try:
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        rfq_map = {r.id: r for r in s.query(RFQ).all()}
        # vendor_rfq_id → (rfq_id, vendor_id, vrfq_no)
        vrfq_map = {vr.id: vr for vr in s.query(VendorRFQ).all()}

        rows = []
        for q in s.query(VendorQuote).order_by(VendorQuote.id.desc()).all():
            vr = vrfq_map.get(q.vendor_rfq_id)
            rfq = rfq_map.get(vr.rfq_id) if vr else None
            items = q.items or []
            amount = 0.0
            for it in items:
                amt = it.get("amount")
                if amt is None:
                    amt = float(it.get("cost_price", 0) or 0) * float(it.get("qty", 1) or 1)
                amount += float(amt or 0)
            rows.append({
                "id": q.id,
                "rfq_id": vr.rfq_id if vr else None,
                "vendor_quote_no": getattr(q, "vendor_quote_no", None) or "—",
                "customer_rfq_no": (rfq.customer_rfq_no or "—") if rfq else "—",
                "vendor": vendor_names.get(vr.vendor_id, "—") if vr else "—",
                "received_at": getattr(q, "received_at", None) or "",
                "received_date": q.received_date or "",
                "item_count": len(items),
                "amount": round(amount, 2),
                "currency": getattr(q, "currency", None) or "USD",
                # 공통 식별 컬럼
                "customer": cust_names.get(rfq.customer_id, "—") if rfq else "—",
                "project_title": (getattr(rfq, "project_title", None) or "") if rfq else "",
                "contact_person": (getattr(rfq, "contact_person", None) or "") if rfq else "",
                "assignee": (user_names.get(rfq.created_by, "") or "") if rfq else "",
                "assignee_id": (rfq.created_by or 0) if rfq else 0,
                "level": (_enum_val(rfq.follow_up_level) if rfq and rfq.follow_up_level else "B"),
                "status": (_status_label(_pipeline_stage(s, rfq.id), rfq.work_type) if rfq else ""),
                "vessel": (vessel_names.get(rfq.vessel_id, "") if rfq and rfq.vessel_id else ""),
                "work_type": (_enum_val(rfq.work_type) if rfq and rfq.work_type else "부품공급"),
                "first_rfq_at": _first_rfq_iso(rfq),
                "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
            })
        return {"rows": rows}
    finally:
        s.close()


@app.get("/api/admin/documents-overview", dependencies=[Depends(require_token)])
def documents_overview():
    """문서 현황 — 오더별 CI/PL/SA/Tax 생성 여부와 문서번호."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        # order_id → 발주 Vendor 이름들(중복 제거, 발주 순). 표시는 첫 벤더 + 외 N곳.
        po_vendors_by_order: dict[int, list[str]] = {}
        for po in s.query(PurchaseOrder).order_by(PurchaseOrder.id).all():
            nm = vendor_names.get(po.vendor_id)
            if not nm:
                continue
            lst = po_vendors_by_order.setdefault(po.order_id, [])
            if nm not in lst:
                lst.append(nm)

        # ci_id → pl/tax 존재 매핑
        ci_by_order: dict[int, CommercialInvoice] = {}
        for ci in s.query(CommercialInvoice).all():
            # 오더당 최신 CI 1건
            if ci.order_id not in ci_by_order or ci.id > ci_by_order[ci.order_id].id:
                ci_by_order[ci.order_id] = ci
        pl_ci_ids = {pl.ci_id for pl in s.query(PackingList).all()}
        pl_no_by_ci = {pl.ci_id: pl.pl_no for pl in s.query(PackingList).all()}
        tax_ci_ids = {tx.ci_id for tx in s.query(TaxInvoiceData).all()}
        tax_no_by_ci = {tx.ci_id: tx.tax_no for tx in s.query(TaxInvoiceData).all()}
        sa_by_order: dict[int, ShippingAdvice] = {}
        for sa in s.query(ShippingAdvice).all():
            if sa.order_id not in sa_by_order or sa.id > sa_by_order[sa.order_id].id:
                sa_by_order[sa.order_id] = sa
        pod_by_order: dict[int, DeliveryProof] = {}
        for pod in s.query(DeliveryProof).all():
            if pod.order_id not in pod_by_order or pod.id > pod_by_order[pod.order_id].id:
                pod_by_order[pod.order_id] = pod

        rows = []
        for o in s.query(Order).order_by(Order.id.desc()).all():
            ci = ci_by_order.get(o.id)
            sa = sa_by_order.get(o.id)
            pod = pod_by_order.get(o.id)
            ci_id = ci.id if ci else None
            # 업무타입(서비스)·서비스 단계(7·8 수동완료) 상태 — RFQ.stage_dates 기준
            rfq = _rfq_for_order(s, o)
            wt = _enum_val(rfq.work_type) if (rfq and rfq.work_type) else "부품공급"
            sd = (getattr(rfq, "stage_dates", None) or {}) if rfq else {}
            svc = getattr(o, "service_info", None) or {}
            rows.append({
                "id": o.id,
                "customer": cust_names.get(o.customer_id, "—"),
                "project_title": (getattr(rfq, "project_title", None) or "") if rfq else "",
                "contact_person": (getattr(rfq, "contact_person", None) or "") if rfq else "",
                "assignee": (user_names.get(rfq.created_by, "") or "") if rfq else "",
                "assignee_id": (rfq.created_by or 0) if rfq else 0,
                "vessel": vessel_names.get(o.vessel_id, "") if o.vessel_id else "",
                "po_no": o.po_no or "",
                "trade_type": o.trade_type or "수출",
                "work_type": wt,
                "vendor": (lambda v: (v[0] + (f"  (외 {len(v) - 1}곳)" if len(v) > 1 else "")) if v else "")(po_vendors_by_order.get(o.id, [])),
                "first_rfq_at": _first_rfq_iso(rfq),
                "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
                "ci_no": ci.ci_no if ci else "",
                "pl_no": pl_no_by_ci.get(ci_id, "") if ci_id else "",
                "sa_no": sa.sa_no if sa else "",
                "sa_sent_date": (sa.sent_date or "") if sa else "",
                "tax_no": tax_no_by_ci.get(ci_id, "") if ci_id else "",
                "pod_filename": (pod.filename or "") if pod else "",
                "has_ci": bool(ci),
                "has_pl": bool(ci_id and ci_id in pl_ci_ids),
                "has_sa": bool(sa),
                "has_pod": bool(pod),
                "has_tax": bool(ci_id and ci_id in tax_ci_ids),
                # 서비스 단계 수동 완료 상태(7·8). 9(리포트)는 has_pod 로 표시.
                "svc_ready_done": bool(sd.get("7")),
                "svc_arr_done": bool(sd.get("8")),
                "svc_billed": bool(svc.get("10")),
            })
        return {"rows": rows}
    finally:
        s.close()


def _customer_for_order(session, order: Order):
    return session.query(Customer).filter_by(id=order.customer_id).first()


def _vessel_for_order(session, order: Order):
    if not order.vessel_id:
        return None
    return session.query(Vessel).filter_by(id=order.vessel_id).first()


def _latest_ci(session, order_id: int):
    return (
        session.query(CommercialInvoice)
        .filter_by(order_id=order_id)
        .order_by(CommercialInvoice.id.desc())
        .first()
    )


def _latest_pl(session, ci_id: int | None):
    if not ci_id:
        return None
    return (
        session.query(PackingList)
        .filter_by(ci_id=ci_id)
        .order_by(PackingList.id.desc())
        .first()
    )


def _latest_sa(session, order_id: int):
    return (
        session.query(ShippingAdvice)
        .filter_by(order_id=order_id)
        .order_by(ShippingAdvice.id.desc())
        .first()
    )


def _latest_tax(session, ci_id: int | None):
    if not ci_id:
        return None
    return (
        session.query(TaxInvoiceData)
        .filter_by(ci_id=ci_id)
        .order_by(TaxInvoiceData.id.desc())
        .first()
    )


def _missing_items(order_items: list[dict], doc_items: list[dict]) -> list[dict]:
    def key(item: dict) -> str:
        return (
            str(item.get("part_no") or "").strip().upper()
            or str(item.get("description") or "").strip().upper()
        )

    def qty(item: dict) -> float:
        try:
            return float(item.get("qty", 0) or 0)
        except (TypeError, ValueError):
            return 0.0

    doc_qty: dict[str, float] = {}
    for item in doc_items or []:
        k = key(item)
        if k:
            doc_qty[k] = doc_qty.get(k, 0.0) + qty(item)

    missing: list[dict] = []
    for item in order_items or []:
        k = key(item)
        if not k:
            continue
        oq = qty(item)
        dq = doc_qty.get(k, 0.0)
        if dq < oq:
            missing.append({
                "part_no": item.get("part_no", ""),
                "description": item.get("description", ""),
                "order_qty": oq,
                "doc_qty": dq,
            })
    return missing


def _doc_file_response(data: bytes, filename: str, media_type: str) -> Response:
    return Response(
        content=data,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _tracking_url(kind: str, token: str | None) -> str:
    if not token:
        return ""
    base = os.getenv("TRACKING_BASE_URL", "https://www.k-maris.com/track")
    return f"{base}?type={kind}&token={token}"


def _rfq_for_order(session, order: Order):
    """오더 → 연결된 RFQ(단계 완료 표시는 RFQ.stage_dates 에 저장됨)."""
    if getattr(order, "rfq_id", None):
        rfq = session.query(RFQ).filter_by(id=order.rfq_id).first()
        if rfq:
            return rfq
    if getattr(order, "quotation_id", None):
        q = session.query(Quotation).filter_by(id=order.quotation_id).first()
        if q and q.rfq_id:
            return session.query(RFQ).filter_by(id=q.rfq_id).first()
    return None


def _document_detail_payload(session, order: Order) -> dict:
    cust = _customer_for_order(session, order)
    vessel = _vessel_for_order(session, order)
    ci = _latest_ci(session, order.id)
    pl = _latest_pl(session, ci.id if ci else None)
    sa = _latest_sa(session, order.id)
    tax = _latest_tax(session, ci.id if ci else None)
    pod = (session.query(DeliveryProof).filter_by(order_id=order.id)
           .order_by(DeliveryProof.created_at.desc()).first())
    rfq = _rfq_for_order(session, order)
    sd = (getattr(rfq, "stage_dates", None) or {}) if rfq else {}
    # 발주된 Vendor(들) — 이 오더의 PurchaseOrder vendor_id → Vendor.name (중복 제거)
    vendor_ids = [
        po.vendor_id
        for po in session.query(PurchaseOrder).filter_by(order_id=order.id).all()
        if po.vendor_id
    ]
    vendor_names: list[str] = []
    if vendor_ids:
        name_by_id = {
            v.id: v.name
            for v in session.query(Vendor).filter(Vendor.id.in_(set(vendor_ids))).all()
        }
        seen = set()
        for vid in vendor_ids:
            nm = name_by_id.get(vid)
            if nm and nm not in seen:
                seen.add(nm)
                vendor_names.append(nm)
    return {
        "order": {
            "id": order.id,
            "rfq_id": rfq.id if rfq else 0,
            "assignee_id": (rfq.created_by or 0) if rfq else 0,
            "po_no": order.po_no or "",
            "date": order.date or "",
            "status": _enum_val(order.status),
            "customer": cust.name if cust else "",
            "customer_email": cust.email if cust else "",
            "customer_tax_id": cust.tax_id if cust else "",
            "vessel": vessel.name if vessel else "",
            "project_title": (rfq.project_title or "") if rfq else "",
            "project_no": _project_no_map(session).get(rfq.id, "") if rfq else "",
            "first_rfq_at": _first_rfq_iso(rfq),
            "work_type": _enum_val(rfq.work_type) if (rfq and rfq.work_type) else "부품공급",
            "vendor": ", ".join(vendor_names),
            "trade_type": order.trade_type or "수출",
            "service_info": getattr(order, "service_info", None) or {},
            "tracking_token": order.tracking_token or "",
            "consignee_confirmed_date": order.consignee_confirmed_date or "",
            "vendor_docs_sent_date": order.vendor_docs_sent_date or "",
            "items": order.items or [],
        },
        "pod": None if not pod else {
            "id": pod.id,
            "filename": pod.filename or "POD",
            "uploaded_at": pod.uploaded_at or "",
        },
        # 수동 완료(완료 버튼) 단계 상태 — 7·8(서비스) · 9 · 11 · 12
        "stage_done": {k: bool(sd.get(k)) for k in ("7", "8", "9", "11", "12")},
        "ci": None if not ci else {
            "id": ci.id,
            "ci_no": ci.ci_no or "",
            "date": ci.date or "",
            "currency": ci.currency or "USD",
            "vat_rate": ci.vat_rate or 0.0,
            "items": ci.items or [],
            "shipping": ci.shipping or {},
            "missing": _missing_items(order.items or [], ci.items or []),
        },
        "pl": None if not pl else {
            "id": pl.id,
            "pl_no": pl.pl_no or "",
            "date": pl.date or "",
            "items": pl.items or [],
            "missing": _missing_items(order.items or [], pl.items or []),
        },
        "sa": None if not sa else {
            "id": sa.id,
            "sa_no": sa.sa_no or "",
            "date": sa.date or "",
            "shipping": sa.shipping or {},
            "sent_date": sa.sent_date or "",
        },
        "tax": None if not tax else {
            "id": tax.id,
            "tax_no": tax.tax_no or "",
            "date": tax.date or "",
            "items": tax.items or [],
        },
        "smtp_configured": bool(os.getenv("SMTP_USER") and os.getenv("SMTP_PASSWORD")),
    }


class DocumentMilestoneUpdate(BaseModel):
    field: str
    value: bool


class CommercialInvoiceSave(BaseModel):
    ci_no: str | None = None
    date: str | None = None
    currency: str = "USD"
    vat_rate: float = 0.0
    items: list[dict] = []
    shipping: dict = {}


class PackingListSave(BaseModel):
    pl_no: str | None = None
    date: str | None = None
    items: list[dict] = []


class ShippingAdviceSave(BaseModel):
    sa_no: str | None = None
    date: str | None = None
    shipping: dict = {}


class ShippingAdviceSend(BaseModel):
    to: str
    subject: str | None = None
    body: str | None = None


class TaxInvoiceSave(BaseModel):
    tax_no: str | None = None
    date: str | None = None
    supply_type: str = "Export / Zero-rated"
    buyer_business_no: str = ""
    vat_rate: float = 0.0
    items: list[dict] = []


@app.get("/api/admin/documents/{order_id}", dependencies=[Depends(require_token)])
def document_detail(order_id: int):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order瑜?李얠쓣 ???놁뒿?덈떎.")
        return _document_detail_payload(s, order)
    finally:
        s.close()


class ServiceStageSave(BaseModel):
    stage: int                      # 7~10
    data: dict = {}
    complete: bool = True


@app.post("/api/admin/documents/{order_id}/service",
          dependencies=[Depends(require_token)])
def save_service_stage(order_id: int, body: ServiceStageSave):
    """서비스 업무 7~10단계 입력값 저장 + 단계 완료 처리.
    7·8·9 는 RFQ.stage_dates 로 완료, 10 은 청구내역으로 AR 레코드를 생성/갱신한다."""
    if body.stage not in (7, 8, 9, 10):
        raise HTTPException(status_code=400, detail="잘못된 서비스 단계입니다.")
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        info = dict(getattr(order, "service_info", None) or {})
        info[str(body.stage)] = body.data
        order.service_info = info

        if body.stage in (7, 8, 9):
            rfq = _rfq_for_order(s, order)
            if rfq:
                dates = dict(getattr(rfq, "stage_dates", None) or {})
                if body.complete:
                    dates[str(body.stage)] = _kst_iso(datetime.utcnow())
                else:
                    dates.pop(str(body.stage), None)
                rfq.stage_dates = dates
        elif body.stage == 10 and body.complete:
            def _f(v) -> float:
                try:
                    return float(v or 0)
                except (TypeError, ValueError):
                    return 0.0
            d = body.data or {}
            service_items = d.get("items") if isinstance(d.get("items"), list) else []
            total = _total_amount(service_items) + sum(_f(d.get(k)) for k in ("labor_cost", "travel_cost", "material_cost", "other_cost"))
            ar = s.query(ARRecord).filter_by(order_id=order.id).first()
            if not ar:
                ar = ARRecord(order_id=order.id, ci_no="",
                              invoice_amount=total, paid_amount=0.0,
                              currency=d.get("currency", "USD"),
                              status=ARStatus.OUTSTANDING)
                s.add(ar)
            else:
                ar.invoice_amount = total
                ar.currency = d.get("currency", "USD")
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.delete("/api/admin/documents/{order_id}/service/{stage}",
            dependencies=[Depends(require_token)])
def delete_service_stage(order_id: int, stage: int):
    """서비스 단계 입력값 삭제 + 완료 해제. 10단계는 연결 AR 레코드도 삭제한다."""
    if stage not in (7, 8, 9, 10):
        raise HTTPException(status_code=400, detail="잘못된 서비스 단계입니다.")
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        info = dict(getattr(order, "service_info", None) or {})
        info.pop(str(stage), None)
        order.service_info = info
        if stage in (7, 8, 9):
            rfq = _rfq_for_order(s, order)
            if rfq:
                dates = dict(getattr(rfq, "stage_dates", None) or {})
                dates.pop(str(stage), None)
                rfq.stage_dates = dates
        if stage == 10:
            s.query(ARRecord).filter_by(order_id=order.id).delete(synchronize_session=False)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.post("/api/admin/documents/{order_id}/milestone",
          dependencies=[Depends(require_token)])
def document_milestone(order_id: int, body: DocumentMilestoneUpdate):
    if body.field not in {"consignee_confirmed_date", "vendor_docs_sent_date"}:
        raise HTTPException(status_code=400, detail="Invalid milestone field")
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order瑜?李얠쓣 ???놁뒿?덈떎.")
        setattr(order, body.field, date.today().isoformat() if body.value else None)
        s.commit()
        return {"ok": True, "value": getattr(order, body.field) or ""}
    finally:
        s.close()


def _manual_doc_no(session, Model, col, body_val, current_id):
    """수동 문서번호 처리. 비우면 None(번호 없음), 입력 시 중복 검사."""
    no = (body_val or "").strip() or None
    if no:
        dup = session.query(Model).filter(
            getattr(Model, col) == no, Model.id != (current_id or 0)).first()
        if dup:
            raise HTTPException(status_code=400, detail=f"이미 존재하는 번호입니다: {no}")
    return no


@app.post("/api/admin/documents/{order_id}/ci",
          dependencies=[Depends(require_token)])
def save_commercial_invoice(order_id: int, body: CommercialInvoiceSave):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order瑜?李얠쓣 ???놁뒿?덈떎.")
        ci = _latest_ci(s, order_id)
        if not ci:
            ci = CommercialInvoice(
                ci_no=_manual_doc_no(s, CommercialInvoice, "ci_no", body.ci_no, None),
                order_id=order.id,
                date=body.date or date.today().isoformat(),
            )
            s.add(ci)
        elif body.ci_no is not None:
            ci.ci_no = _manual_doc_no(s, CommercialInvoice, "ci_no", body.ci_no, ci.id)
        ci.date = body.date or ci.date or date.today().isoformat()
        ci.currency = body.currency or "USD"
        ci.vat_rate = body.vat_rate or 0.0
        ci.items = body.items or []
        ci.shipping = body.shipping or {}
        s.commit()
        return {"ok": True, "id": ci.id, "ci_no": ci.ci_no}
    finally:
        s.close()


@app.get("/api/admin/documents/{order_id}/ci/pdf",
         dependencies=[Depends(require_token)])
def commercial_invoice_pdf(order_id: int):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        ci = _latest_ci(s, order_id) if order else None
        if not order or not ci:
            raise HTTPException(status_code=404, detail="Commercial Invoice瑜?李얠쓣 ???놁뒿?덈떎.")
        payload = build_payload(
            doc_no=ci.ci_no, date=ci.date,
            customer=_customer_for_order(s, order),
            vessel=_vessel_for_order(s, order),
            items=ci.items or [], terms=ci.terms or {},
            currency=ci.currency or "USD", vat_rate=ci.vat_rate or 0.0,
            shipping=ci.shipping or {}, po_no=order.po_no or "",
            export_ref=_project_no_for_order(s, order),
        )
        pdf = generate_pdf("commercial_invoice", payload)
        return _doc_file_response(pdf, f"{ci.ci_no}_CI.pdf", "application/pdf")
    finally:
        s.close()


@app.post("/api/admin/documents/{order_id}/pl",
          dependencies=[Depends(require_token)])
def save_packing_list(order_id: int, body: PackingListSave):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        ci = _latest_ci(s, order_id) if order else None
        if not order or not ci:
            raise HTTPException(status_code=400, detail="먼저 Commercial Invoice를 생성하세요.")
        pl = _latest_pl(s, ci.id)
        if not pl:
            pl = PackingList(
                pl_no=_manual_doc_no(s, PackingList, "pl_no", body.pl_no, None),
                ci_id=ci.id,
                date=body.date or date.today().isoformat(),
            )
            s.add(pl)
        elif body.pl_no is not None:
            pl.pl_no = _manual_doc_no(s, PackingList, "pl_no", body.pl_no, pl.id)
        pl.date = body.date or pl.date or date.today().isoformat()
        pl.items = body.items or []
        s.commit()
        return {"ok": True, "id": pl.id, "pl_no": pl.pl_no}
    finally:
        s.close()


@app.get("/api/admin/documents/{order_id}/pl/pdf",
         dependencies=[Depends(require_token)])
def packing_list_pdf(order_id: int):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        ci = _latest_ci(s, order_id) if order else None
        pl = _latest_pl(s, ci.id if ci else None)
        if not order or not ci or not pl:
            raise HTTPException(status_code=404, detail="Packing List瑜?李얠쓣 ???놁뒿?덈떎.")
        payload = build_payload(
            doc_no=pl.pl_no, date=pl.date,
            customer=_customer_for_order(s, order),
            vessel=_vessel_for_order(s, order),
            items=pl.items or [], terms={},
            currency=ci.currency or "USD",
            shipping=ci.shipping or {}, po_no=order.po_no or "",
            export_ref=_project_no_for_order(s, order),
        )
        pdf = generate_pdf("packing_list", payload)
        return _doc_file_response(pdf, f"{pl.pl_no}_PL.pdf", "application/pdf")
    finally:
        s.close()


@app.post("/api/admin/documents/{order_id}/sa",
          dependencies=[Depends(require_token)])
def save_shipping_advice(order_id: int, body: ShippingAdviceSave):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order瑜?李얠쓣 ???놁뒿?덈떎.")
        sa = _latest_sa(s, order_id)
        if not sa:
            sa = ShippingAdvice(
                sa_no=_manual_doc_no(s, ShippingAdvice, "sa_no", body.sa_no, None),
                order_id=order.id,
                date=body.date or date.today().isoformat(),
            )
            s.add(sa)
        elif body.sa_no is not None:
            sa.sa_no = _manual_doc_no(s, ShippingAdvice, "sa_no", body.sa_no, sa.id)
        sa.date = body.date or sa.date or date.today().isoformat()
        sa.shipping = body.shipping or {}
        s.commit()
        return {"ok": True, "id": sa.id, "sa_no": sa.sa_no}
    finally:
        s.close()


@app.get("/api/admin/documents/{order_id}/sa/pdf",
         dependencies=[Depends(require_token)])
def shipping_advice_pdf(order_id: int):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        sa = _latest_sa(s, order_id) if order else None
        ci = _latest_ci(s, order_id) if order else None
        if not order or not sa:
            raise HTTPException(status_code=404, detail="Shipping Advice瑜?李얠쓣 ???놁뒿?덈떎.")
        payload = build_payload(
            doc_no=sa.sa_no, date=sa.date,
            customer=_customer_for_order(s, order),
            vessel=_vessel_for_order(s, order),
            items=(ci.items if ci else order.items) or [], terms={},
            currency=(ci.currency if ci else "USD") or "USD",
            shipping=sa.shipping or {}, po_no=order.po_no or "",
            export_ref=_project_no_for_order(s, order),
        )
        pdf = generate_pdf("shipping_advice", payload)
        return _doc_file_response(pdf, f"{sa.sa_no}_SA.pdf", "application/pdf")
    finally:
        s.close()


@app.post("/api/admin/documents/{order_id}/sa/send",
          dependencies=[Depends(require_token)])
def send_shipping_advice(order_id: int, body: ShippingAdviceSend):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        sa = _latest_sa(s, order_id) if order else None
        ci = _latest_ci(s, order_id) if order else None
        if not order or not sa:
            raise HTTPException(status_code=404, detail="Shipping Advice瑜?李얠쓣 ???놁뒿?덈떎.")
        cust = _customer_for_order(s, order)
        payload = build_payload(
            doc_no=sa.sa_no, date=sa.date, customer=cust,
            vessel=_vessel_for_order(s, order),
            items=(ci.items if ci else order.items) or [], terms={},
            currency=(ci.currency if ci else "USD") or "USD",
            shipping=sa.shipping or {}, po_no=order.po_no or "",
            export_ref=_project_no_for_order(s, order),
        )
        pdf = generate_pdf("shipping_advice", payload)
        subject = body.subject or f"[K-MARIS] Shipping Advice {sa.sa_no}"
        mail_body = body.body or shipping_advice_email_body(
            cust.name if cust else "Customer", sa.sa_no,
            _tracking_url("order", order.tracking_token),
        )
        ok = send_email(body.to, subject, mail_body, [(f"{sa.sa_no}_SA.pdf", pdf)])
        if ok:
            sa.sent_date = date.today().isoformat()
            s.commit()
        return {"ok": ok, "sent_date": sa.sent_date or ""}
    finally:
        s.close()


@app.post("/api/admin/documents/{order_id}/tax",
          dependencies=[Depends(require_token)])
def save_tax_invoice(order_id: int, body: TaxInvoiceSave):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        ci = _latest_ci(s, order_id) if order else None
        if not order or not ci:
            raise HTTPException(status_code=400, detail="먼저 Commercial Invoice를 생성하세요.")
        tax = _latest_tax(s, ci.id)
        if not tax:
            tax = TaxInvoiceData(
                tax_no=_manual_doc_no(s, TaxInvoiceData, "tax_no", body.tax_no, None),
                ci_id=ci.id,
                date=body.date or date.today().isoformat(),
            )
            s.add(tax)
        elif body.tax_no is not None:
            tax.tax_no = _manual_doc_no(s, TaxInvoiceData, "tax_no", body.tax_no, tax.id)
        tax.date = body.date or tax.date or date.today().isoformat()
        tax.items = body.items or ci.items or []

        ar = s.query(ARRecord).filter_by(order_id=order.id, ci_no=ci.ci_no).first()
        invoice_amount = _total_amount(tax.items or [])
        if not ar:
            ar = ARRecord(
                order_id=order.id,
                ci_no=ci.ci_no,
                invoice_amount=invoice_amount,
                paid_amount=0.0,
                currency=ci.currency or "USD",
                status=ARStatus.OUTSTANDING,
            )
            s.add(ar)
        else:
            ar.invoice_amount = invoice_amount
            ar.currency = ci.currency or "USD"
        s.commit()
        return {"ok": True, "id": tax.id, "tax_no": tax.tax_no, "ar_id": ar.id}
    finally:
        s.close()


@app.get("/api/admin/documents/{order_id}/tax/xlsx",
         dependencies=[Depends(require_token)])
def tax_invoice_xlsx(order_id: int):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        ci = _latest_ci(s, order_id) if order else None
        tax = _latest_tax(s, ci.id if ci else None)
        if not order or not ci or not tax:
            raise HTTPException(status_code=404, detail="Tax Invoice Data瑜?李얠쓣 ???놁뒿?덈떎.")
        cust = _customer_for_order(s, order)
        payload = build_payload(
            doc_no=tax.tax_no, date=tax.date,
            customer=cust,
            vessel=_vessel_for_order(s, order),
            items=tax.items or [], terms={},
            currency="KRW", vat_rate=0.0,
            tax_invoice={
                "issue_date": tax.date,
                "supply_type": "Export / Zero-rated",
                "buyer_business_no": cust.tax_id if cust else "",
            },
        )
        xlsx = generate_tax_xlsx(payload)
        return _doc_file_response(
            xlsx,
            f"{tax.tax_no}_tax_invoice.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        s.close()


# ── 9) 운송 완료 · POD 수취 — 인도 증빙(POD) 파일 ───────────────────────────────
@app.post("/api/admin/documents/{order_id}/pod", dependencies=[Depends(require_token)])
def upload_pod(order_id: int, file: UploadFile = File(...)):
    """POD(인도 증빙) 파일 업로드 — 오더당 1건(기존 파일 교체). 업로드 시 9단계 완료."""
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        file.file.seek(0)
        data = file.file.read()
        if not data:
            raise HTTPException(status_code=400, detail="빈 파일입니다.")
        s.query(DeliveryProof).filter_by(order_id=order_id).delete()
        proof = DeliveryProof(
            order_id=order_id,
            filename=file.filename or "POD",
            mime=file.content_type or "application/octet-stream",
            data=data,
            uploaded_at=_kst_iso(datetime.utcnow()),
        )
        s.add(proof)
        s.commit()
        return {"ok": True, "filename": proof.filename, "uploaded_at": proof.uploaded_at}
    finally:
        s.close()


@app.get("/api/admin/documents/{order_id}/pod/file", dependencies=[Depends(require_token)])
def download_pod(order_id: int):
    s = get_session()
    try:
        proof = (s.query(DeliveryProof).filter_by(order_id=order_id)
                 .order_by(DeliveryProof.created_at.desc()).first())
        if not proof or not proof.data:
            raise HTTPException(status_code=404, detail="POD 파일이 없습니다.")
        return Response(
            content=proof.data,
            media_type=proof.mime or "application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{proof.filename or "POD"}"'},
        )
    finally:
        s.close()


@app.delete("/api/admin/documents/{order_id}/pod", dependencies=[Depends(require_token)])
def delete_pod(order_id: int):
    s = get_session()
    try:
        n = s.query(DeliveryProof).filter_by(order_id=order_id).delete()
        s.commit()
        return {"ok": True, "deleted": n}
    finally:
        s.close()


# ── 단계 완료 콜 — 오더 기준으로 RFQ.stage_dates 에 완료 표시(9·11·12) ──────────
class StageCompleteBody(BaseModel):
    done: bool = True
    at: str | None = None  # 'YYYY-MM-DDTHH:MM' (KST 벽시계) — 생략 시 현재시각


@app.post("/api/admin/orders/{order_id}/stage/{stage}/complete",
          dependencies=[Depends(require_token)])
def complete_order_stage(order_id: int, stage: int, body: StageCompleteBody):
    """11·12 등 수동 완료 단계를 토글한다. 완료 시 RFQ.stage_dates[stage]=지정 시각(없으면 현재)."""
    if not (1 <= stage <= len(INTERNAL_STEPS)):
        raise HTTPException(status_code=400, detail="잘못된 단계 번호입니다.")
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        rfq = _rfq_for_order(s, order)
        if not rfq:
            raise HTTPException(status_code=400, detail="연결된 RFQ가 없습니다.")
        dates = dict(getattr(rfq, "stage_dates", None) or {})
        key = str(stage)
        if body.done:
            dates[key] = (body.at or "").strip()[:16] or _kst_iso(datetime.utcnow())
        else:
            dates.pop(key, None)
        rfq.stage_dates = dates
        s.commit()
        return {"ok": True, "stage": _pipeline_stage(s, rfq.id), "done": body.done}
    finally:
        s.close()


@app.get("/api/admin/vendor-po-overview", dependencies=[Depends(require_token)])
def vendor_po_overview():
    """Vendor P/O 발신 내역 — PurchaseOrder 1건당 1행(생성·발송 포함 전체)."""
    s = get_session()
    try:
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        ord_map = {o.id: o for o in s.query(Order).all()}
        cust_names = {c.id: c.name for c in s.query(Customer).all()}

        rows = []
        for po in s.query(PurchaseOrder).order_by(PurchaseOrder.id.desc()).all():
            o = ord_map.get(po.order_id)
            rows.append({
                "id": po.id,
                "po_no": po.po_no or "",
                "customer": cust_names.get(o.customer_id, "—") if o else "—",
                "vendor": vendor_names.get(po.vendor_id, "—"),
                "vendor_email": po.sent_to_email or "",
                "date": po.date or "",
                "sent_date": po.sent_date or "",
                "status": po.status or "",
                "item_count": len(po.items or []),
                "sent": (po.status == "이메일 발송완료"),
            })
        return {"rows": rows}
    finally:
        s.close()


# ── 마케팅 활동 (잠정 고객사 대상) ───────────────────────────────────────────
class MarketingActivityCreate(BaseModel):
    customer_id: int | None = None
    prospect_name: str | None = ""
    contact_person: str | None = ""
    recipient_email: str | None = ""
    activity_date: str | None = ""
    channel: str | None = ""
    activity_type: str | None = ""
    subject: str | None = ""
    notes: str | None = ""
    next_action_date: str | None = ""


def _marketing_target_name(m: MarketingActivity, cust_names: dict) -> str:
    """활동 대상 표기명 — 연결 고객사가 있으면 그 이름, 없으면 잠정사 자유입력."""
    if m.customer_id and m.customer_id in cust_names:
        return cust_names[m.customer_id]
    return m.prospect_name or "—"


def _marketing_row(m: MarketingActivity, cust_names: dict, user_names: dict) -> dict:
    return {
        "id": m.id,
        "customer_id": m.customer_id,
        "customer": _marketing_target_name(m, cust_names),
        "prospect_name": m.prospect_name or "",
        "is_prospect": not bool(m.customer_id),
        "contact_person": m.contact_person or "",
        "recipient_email": m.recipient_email or "",
        "activity_date": m.activity_date or "",
        "channel": m.channel or "",
        "activity_type": m.activity_type or "",
        "subject": m.subject or "",
        "notes": m.notes or "",
        "next_action_date": m.next_action_date or "",
        "owner_id": m.owner_id or 0,
        "owner": user_names.get(m.owner_id, "") if m.owner_id else "",
    }


def _marketing_scoped(s, user: dict):
    """조회 범위 적용된 MarketingActivity 쿼리. 'own' 역할은 본인 담당 건만."""
    q = s.query(MarketingActivity).order_by(MarketingActivity.id.desc())
    role = user.get("role", "")
    if role != UserRole.ADMIN.value and _scope_for(role) == "own":
        q = q.filter(MarketingActivity.owner_id == (user.get("id") or 0))
    return q


@app.get("/api/admin/marketing", dependencies=[Depends(require_token)])
def marketing_list(user: dict = Depends(get_current_user)):
    """잠정 고객사 마케팅 활동 목록."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        rows = [_marketing_row(m, cust_names, user_names)
                for m in _marketing_scoped(s, user).all()]
        return {"rows": rows}
    finally:
        s.close()


@app.post("/api/admin/marketing", dependencies=[Depends(require_token)])
def create_marketing(body: MarketingActivityCreate, user: dict = Depends(get_current_user)):
    if not (body.customer_id or (body.prospect_name or "").strip()):
        raise HTTPException(status_code=400, detail="대상 고객사(선택) 또는 잠정사 이름을 입력하세요.")
    s = get_session()
    try:
        m = MarketingActivity(
            customer_id=body.customer_id or None,
            prospect_name=(body.prospect_name or "").strip(),
            contact_person=body.contact_person or "",
            recipient_email=body.recipient_email or "",
            activity_date=body.activity_date or "",
            channel=body.channel or "",
            activity_type=body.activity_type or "",
            subject=body.subject or "",
            notes=body.notes or "",
            next_action_date=body.next_action_date or "",
            owner_id=user.get("id") or None,
        )
        s.add(m)
        s.commit()
        return {"ok": True, "id": m.id}
    finally:
        s.close()


@app.put("/api/admin/marketing/{row_id}", dependencies=[Depends(require_token)])
def update_marketing(row_id: int, body: MarketingActivityCreate):
    if not (body.customer_id or (body.prospect_name or "").strip()):
        raise HTTPException(status_code=400, detail="대상 고객사(선택) 또는 잠정사 이름을 입력하세요.")
    s = get_session()
    try:
        m = s.query(MarketingActivity).filter_by(id=row_id).first()
        if not m:
            raise HTTPException(status_code=404, detail="마케팅 활동을 찾을 수 없습니다.")
        m.customer_id = body.customer_id or None
        m.prospect_name = (body.prospect_name or "").strip()
        m.contact_person = body.contact_person or ""
        m.recipient_email = body.recipient_email or ""
        m.activity_date = body.activity_date or ""
        m.channel = body.channel or ""
        m.activity_type = body.activity_type or ""
        m.subject = body.subject or ""
        m.notes = body.notes or ""
        m.next_action_date = body.next_action_date or ""
        s.commit()
        return {"ok": True, "id": m.id}
    finally:
        s.close()


@app.delete("/api/admin/marketing/{row_id}", dependencies=[Depends(require_token)])
def delete_marketing(row_id: int):
    s = get_session()
    try:
        m = s.query(MarketingActivity).filter_by(id=row_id).first()
        if not m:
            raise HTTPException(status_code=404, detail="마케팅 활동을 찾을 수 없습니다.")
        s.delete(m)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.get("/api/admin/marketing-overview", dependencies=[Depends(require_token)])
def marketing_overview(user: dict = Depends(get_current_user)):
    """대시보드 마케팅 카드용 요약.
      - recent:      최근 활동 목록(최신순)
      - follow_ups:  후속 예정(next_action_date 있는 건, 예정일 오름차순)
      - month:       이번 달 활동 집계(총건수 + 채널별·유형별)
    """
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        items = _marketing_scoped(s, user).all()
        rows = [_marketing_row(m, cust_names, user_names) for m in items]

        today = (datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")
        month = today[:7]

        follow_ups = sorted(
            (r for r in rows if r["next_action_date"]),
            key=lambda r: r["next_action_date"],
        )
        this_month = [r for r in rows if (r["activity_date"] or "")[:7] == month]
        by_channel: dict[str, int] = {}
        by_type: dict[str, int] = {}
        for r in this_month:
            if r["channel"]:
                by_channel[r["channel"]] = by_channel.get(r["channel"], 0) + 1
            if r["activity_type"]:
                by_type[r["activity_type"]] = by_type.get(r["activity_type"], 0) + 1

        return {
            "recent": rows[:20],
            "follow_ups": follow_ups[:20],
            "month": {
                "period": month,
                "total": len(this_month),
                "by_channel": by_channel,
                "by_type": by_type,
            },
        }
    finally:
        s.close()


# ── 일정(Schedule) — 대시보드 카드 내에서 직접 관리 ──────────────────────────
class ScheduleEventCreate(BaseModel):
    date: str | None = ""
    title: str | None = ""
    event_type: str | None = ""
    notes: str | None = ""
    customer_id: int | None = None


def _schedule_row(e: ScheduleEvent, cust_names: dict, user_names: dict) -> dict:
    return {
        "id": e.id,
        "date": e.date or "",
        "title": e.title or "",
        "event_type": e.event_type or "",
        "notes": e.notes or "",
        "customer_id": e.customer_id,
        "customer": cust_names.get(e.customer_id, "") if e.customer_id else "",
        "owner_id": e.owner_id or 0,
        "owner": user_names.get(e.owner_id, "") if e.owner_id else "",
    }


@app.get("/api/admin/schedule", dependencies=[Depends(require_token)])
def schedule_list():
    """일정 목록 — 팀 공용(전체), 날짜 오름차순."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        events = s.query(ScheduleEvent).order_by(ScheduleEvent.date, ScheduleEvent.id).all()
        return {"rows": [_schedule_row(e, cust_names, user_names) for e in events]}
    finally:
        s.close()


@app.post("/api/admin/schedule", dependencies=[Depends(require_token)])
def create_schedule(body: ScheduleEventCreate, user: dict = Depends(get_current_user)):
    if not (body.title or "").strip():
        raise HTTPException(status_code=400, detail="일정 제목을 입력하세요.")
    if not (body.date or "").strip():
        raise HTTPException(status_code=400, detail="일정 날짜를 입력하세요.")
    s = get_session()
    try:
        e = ScheduleEvent(
            date=body.date or "",
            title=(body.title or "").strip(),
            event_type=body.event_type or "",
            notes=body.notes or "",
            customer_id=body.customer_id or None,
            owner_id=user.get("id") or None,
        )
        s.add(e)
        s.commit()
        return {"ok": True, "id": e.id}
    finally:
        s.close()


def _schedule_guard(e: ScheduleEvent, user: dict) -> None:
    """작성자(owner) 또는 admin 만 수정·삭제 가능."""
    if user.get("role") == UserRole.ADMIN.value:
        return
    if (e.owner_id or 0) != (user.get("id") or 0):
        raise HTTPException(status_code=403, detail="작성자(PIC)만 이 일정을 수정·삭제할 수 있습니다.")


@app.put("/api/admin/schedule/{row_id}", dependencies=[Depends(require_token)])
def update_schedule(row_id: int, body: ScheduleEventCreate, user: dict = Depends(get_current_user)):
    if not (body.title or "").strip():
        raise HTTPException(status_code=400, detail="일정 제목을 입력하세요.")
    s = get_session()
    try:
        e = s.query(ScheduleEvent).filter_by(id=row_id).first()
        if not e:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        _schedule_guard(e, user)
        e.date = body.date or ""
        e.title = (body.title or "").strip()
        e.event_type = body.event_type or ""
        e.notes = body.notes or ""
        e.customer_id = body.customer_id or None
        s.commit()
        return {"ok": True, "id": e.id}
    finally:
        s.close()


@app.delete("/api/admin/schedule/{row_id}", dependencies=[Depends(require_token)])
def delete_schedule(row_id: int, user: dict = Depends(get_current_user)):
    s = get_session()
    try:
        e = s.query(ScheduleEvent).filter_by(id=row_id).first()
        if not e:
            raise HTTPException(status_code=404, detail="일정을 찾을 수 없습니다.")
        _schedule_guard(e, user)
        s.delete(e)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.get("/api/admin/customers", dependencies=[Depends(require_token)])
def customers():
    s = get_session()
    try:
        return [{"id": c.id, "name": c.name, "contact": c.contact or "",
                 "logo": getattr(c, "logo", None) or ""}
                for c in s.query(Customer).order_by(Customer.name).all()]
    finally:
        s.close()


@app.get("/api/admin/vendors", dependencies=[Depends(require_token)])
def vendors():
    s = get_session()
    try:
        return [{"id": v.id, "name": v.name, "email": v.email or "",
                 "logo": getattr(v, "logo", None) or ""}
                for v in s.query(Vendor).order_by(Vendor.name).all()]
    finally:
        s.close()


# ── Settings: master data (list + create) ─────────────────────────────────────
class CustomerCreate(BaseModel):
    name: str
    contact: str | None = ""
    contact_phone: str | None = ""
    email: str | None = ""
    country: str | None = ""
    address: str | None = ""
    tax_id: str | None = ""
    logo: str | None = ""    # 회사 로고 data URL(붙여넣기). None=변경 안 함(수정 시)


class VendorCreate(BaseModel):
    name: str
    contact: str | None = ""
    contact_phone: str | None = ""
    email: str | None = ""
    specialization: str | None = ""
    country: str | None = ""
    address: str | None = ""
    logo: str | None = ""    # 회사 로고 data URL(붙여넣기). None=변경 안 함(수정 시)


class VesselCreate(BaseModel):
    name: str
    imo: str | None = ""
    vessel_type: str | None = ""
    ais_flag: str | None = ""
    customer_id: int | None = None
    engine_type: str | None = ""
    hull_no: str | None = ""


class ItemMasterSave(BaseModel):
    part_no: str
    description: str | None = ""
    maker: str | None = ""
    origin: str | None = ""
    unit: str | None = "PCS"
    hs_code: str | None = ""
    std_price: float | None = 0.0


class UserSave(BaseModel):
    username: str
    email: str | None = ""
    password: str | None = None
    role: str = "sales"
    is_active: bool = True


class CompanyProfile(BaseModel):
    company_name_en: str | None = ""
    company_name_kr: str | None = ""
    address: str | None = ""
    business_no: str | None = ""
    phone: str | None = ""
    general_email: str | None = ""
    sales_email: str | None = ""
    tax_email: str | None = ""
    website: str | None = ""
    bank_name: str | None = ""
    bank_account: str | None = ""
    bank_holder: str | None = ""
    swift: str | None = ""
    tagline: str | None = ""


_COMPANY_CONFIG = ROOT / "config" / "company.json"


def _read_company_profile() -> dict:
    try:
        import json
        return json.loads(_COMPANY_CONFIG.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _write_company_profile(data: dict) -> None:
    import json
    _COMPANY_CONFIG.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


@app.get("/api/admin/settings/customers", dependencies=[Depends(require_token)])
def settings_customers():
    s = get_session()
    try:
        return [{"id": c.id, "name": c.name, "contact": c.contact or "",
                 "contact_phone": getattr(c, "contact_phone", None) or "",
                 "email": c.email or "", "country": c.country or "",
                 "address": c.address or "", "tax_id": c.tax_id or "",
                 "logo": getattr(c, "logo", None) or ""}
                for c in s.query(Customer).order_by(Customer.name).all()]
    finally:
        s.close()


@app.post("/api/admin/settings/customers", dependencies=[Depends(require_token)])
def create_customer(body: CustomerCreate):
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="이름을 입력하세요.")
    s = get_session()
    try:
        c = Customer(name=body.name.strip(), contact=body.contact or "",
                     contact_phone=body.contact_phone or "",
                     email=body.email or "", country=body.country or "",
                     address=body.address or "", tax_id=body.tax_id or "",
                     logo=body.logo or "")
        s.add(c)
        s.commit()
        return {"ok": True, "id": c.id}
    finally:
        s.close()


@app.put("/api/admin/settings/customers/{row_id}", dependencies=[Depends(require_token)])
def update_customer(row_id: int, body: CustomerCreate):
    s = get_session()
    try:
        c = s.query(Customer).filter_by(id=row_id).first()
        if not c:
            raise HTTPException(status_code=404, detail="Customer를 찾을 수 없습니다.")
        c.name = body.name.strip()
        c.contact = body.contact or ""
        c.contact_phone = body.contact_phone or ""
        c.email = body.email or ""
        c.country = body.country or ""
        c.address = body.address or ""
        c.tax_id = body.tax_id or ""
        if body.logo is not None:
            c.logo = body.logo
        s.commit()
        return {"ok": True, "id": c.id}
    finally:
        s.close()


@app.delete("/api/admin/settings/customers/{row_id}", dependencies=[Depends(require_token)])
def delete_customer(row_id: int):
    s = get_session()
    try:
        c = s.query(Customer).filter_by(id=row_id).first()
        if not c:
            raise HTTPException(status_code=404, detail="Customer를 찾을 수 없습니다.")
        s.delete(c)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.get("/api/admin/settings/vendors", dependencies=[Depends(require_token)])
def settings_vendors():
    s = get_session()
    try:
        return [{"id": v.id, "name": v.name, "contact": v.contact or "",
                 "contact_phone": getattr(v, "contact_phone", None) or "",
                 "email": v.email or "", "specialization": v.specialization or "",
                 "country": v.country or "", "address": v.address or "",
                 "logo": getattr(v, "logo", None) or ""}
                for v in s.query(Vendor).order_by(Vendor.name).all()]
    finally:
        s.close()


@app.post("/api/admin/settings/vendors", dependencies=[Depends(require_token)])
def create_vendor(body: VendorCreate):
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="이름을 입력하세요.")
    s = get_session()
    try:
        v = Vendor(name=body.name.strip(), contact=body.contact or "",
                   contact_phone=body.contact_phone or "",
                   email=body.email or "", specialization=body.specialization or "",
                   country=body.country or "", address=body.address or "",
                   logo=body.logo or "")
        s.add(v)
        s.commit()
        return {"ok": True, "id": v.id}
    finally:
        s.close()


@app.put("/api/admin/settings/vendors/{row_id}", dependencies=[Depends(require_token)])
def update_vendor(row_id: int, body: VendorCreate):
    s = get_session()
    try:
        v = s.query(Vendor).filter_by(id=row_id).first()
        if not v:
            raise HTTPException(status_code=404, detail="Vendor를 찾을 수 없습니다.")
        v.name = body.name.strip()
        v.contact = body.contact or ""
        v.contact_phone = body.contact_phone or ""
        v.email = body.email or ""
        v.specialization = body.specialization or ""
        v.country = body.country or ""
        v.address = body.address or ""
        if body.logo is not None:
            v.logo = body.logo
        s.commit()
        return {"ok": True, "id": v.id}
    finally:
        s.close()


@app.delete("/api/admin/settings/vendors/{row_id}", dependencies=[Depends(require_token)])
def delete_vendor(row_id: int):
    s = get_session()
    try:
        v = s.query(Vendor).filter_by(id=row_id).first()
        if not v:
            raise HTTPException(status_code=404, detail="Vendor를 찾을 수 없습니다.")
        s.delete(v)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.get("/api/admin/settings/vessels", dependencies=[Depends(require_token)])
def settings_vessels():
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        return [{"id": v.id, "name": v.name, "imo": v.imo or "",
                 "vessel_type": getattr(v, "vessel_type", None) or "",
                 "ais_flag": getattr(v, "ais_flag", None) or "",
                 "engine_type": v.engine_type or "", "hull_no": v.hull_no or "",
                 "customer_id": v.customer_id, "customer": cust_names.get(v.customer_id, "") if v.customer_id else ""}
                for v in s.query(Vessel).order_by(Vessel.name).all()]
    finally:
        s.close()


@app.post("/api/admin/settings/vessels", dependencies=[Depends(require_token)])
def create_vessel(body: VesselCreate):
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="선박명을 입력하세요.")
    s = get_session()
    try:
        v = Vessel(name=body.name.strip(), imo=body.imo or "",
                   vessel_type=body.vessel_type or "",
                   ais_flag=body.ais_flag or "",
                   customer_id=body.customer_id,
                   engine_type=body.engine_type or "",
                   hull_no=body.hull_no or "")
        s.add(v)
        s.commit()
        return {"ok": True, "id": v.id}
    finally:
        s.close()


@app.put("/api/admin/settings/vessels/{row_id}", dependencies=[Depends(require_token)])
def update_vessel(row_id: int, body: VesselCreate):
    s = get_session()
    try:
        v = s.query(Vessel).filter_by(id=row_id).first()
        if not v:
            raise HTTPException(status_code=404, detail="Vessel을 찾을 수 없습니다.")
        v.name = body.name.strip()
        v.imo = body.imo or ""
        v.vessel_type = body.vessel_type or ""
        v.ais_flag = body.ais_flag or ""
        v.customer_id = body.customer_id
        v.engine_type = body.engine_type or ""
        v.hull_no = body.hull_no or ""
        s.commit()
        return {"ok": True, "id": v.id}
    finally:
        s.close()


@app.delete("/api/admin/settings/vessels/{row_id}", dependencies=[Depends(require_token)])
def delete_vessel(row_id: int):
    s = get_session()
    try:
        v = s.query(Vessel).filter_by(id=row_id).first()
        if not v:
            raise HTTPException(status_code=404, detail="Vessel을 찾을 수 없습니다.")
        s.delete(v)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.get("/api/admin/settings/items", dependencies=[Depends(require_token)])
def settings_items():
    s = get_session()
    try:
        return [{
            "id": i.id, "part_no": i.part_no or "",
            "description": i.description or "", "maker": i.maker or "",
            "origin": i.origin or "", "unit": i.unit or "PCS",
            "hs_code": i.hs_code or "", "std_price": i.std_price or 0.0,
        } for i in s.query(ItemMaster).order_by(ItemMaster.part_no).all()]
    finally:
        s.close()


@app.post("/api/admin/settings/items", dependencies=[Depends(require_token)])
def create_item(body: ItemMasterSave):
    if not body.part_no.strip():
        raise HTTPException(status_code=400, detail="Part No.를 입력하세요.")
    s = get_session()
    try:
        item = ItemMaster(
            part_no=body.part_no.strip(), description=body.description or "",
            maker=body.maker or "", origin=body.origin or "",
            unit=body.unit or "PCS", hs_code=body.hs_code or "",
            std_price=body.std_price or 0.0,
        )
        s.add(item)
        s.commit()
        return {"ok": True, "id": item.id}
    finally:
        s.close()


@app.put("/api/admin/settings/items/{row_id}", dependencies=[Depends(require_token)])
def update_item(row_id: int, body: ItemMasterSave):
    s = get_session()
    try:
        item = s.query(ItemMaster).filter_by(id=row_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Item을 찾을 수 없습니다.")
        item.part_no = body.part_no.strip()
        item.description = body.description or ""
        item.maker = body.maker or ""
        item.origin = body.origin or ""
        item.unit = body.unit or "PCS"
        item.hs_code = body.hs_code or ""
        item.std_price = body.std_price or 0.0
        s.commit()
        return {"ok": True, "id": item.id}
    finally:
        s.close()


@app.delete("/api/admin/settings/items/{row_id}", dependencies=[Depends(require_token)])
def delete_item(row_id: int):
    s = get_session()
    try:
        item = s.query(ItemMaster).filter_by(id=row_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Item을 찾을 수 없습니다.")
        s.delete(item)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.get("/api/admin/settings/company", dependencies=[Depends(require_token)])
def settings_company():
    return _read_company_profile()


@app.put("/api/admin/settings/company", dependencies=[Depends(require_token)])
def update_company(body: CompanyProfile):
    data = body.dict()
    _write_company_profile(data)
    return {"ok": True}


class RolePermSave(BaseModel):
    role: str
    perms: dict
    scope: str = "all"


@app.get("/api/admin/settings/permissions", dependencies=[Depends(require_token)])
def settings_permissions(user: dict = Depends(get_current_user)):
    """역할별 권한 매트릭스(admin 전용 조회). admin 행은 전체 고정(편집 불가)."""
    if user.get("role") != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Admin only")
    editable = [UserRole.SALES.value, UserRole.VIEWER.value]
    roles = [{
        "role": UserRole.ADMIN.value, "perms": _full_perms(True),
        "scope": "all", "editable": False,
    }]
    for r in editable:
        roles.append({
            "role": r, "perms": _perms_for(r), "scope": _scope_for(r), "editable": True,
        })
    return {
        "roles": roles,
        "modules": PERM_MODULES,
        "actions": PERM_ACTIONS,
        "view_only": sorted(PERM_VIEW_ONLY),
    }


@app.put("/api/admin/settings/permissions", dependencies=[Depends(require_token)])
def update_permissions(body: RolePermSave, user: dict = Depends(get_current_user)):
    """역할 권한 저장(admin 전용). admin 역할은 변경 불가(잠금 방지)."""
    if user.get("role") != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Admin only")
    role = (body.role or "").strip()
    if role == UserRole.ADMIN.value:
        raise HTTPException(status_code=400, detail="admin 역할의 권한은 변경할 수 없습니다.")
    if role not in (UserRole.SALES.value, UserRole.VIEWER.value):
        raise HTTPException(status_code=400, detail=f"알 수 없는 역할: {role}")
    scope = "own" if body.scope == "own" else "all"
    perms = _normalize_perms(body.perms)
    s = get_session()
    try:
        rp = s.query(RolePermission).filter_by(role=role).first()
        if rp:
            rp.perms = perms
            rp.scope = scope
            rp.updated_at = datetime.utcnow()
        else:
            s.add(RolePermission(role=role, perms=perms, scope=scope))
        s.commit()
    finally:
        s.close()
    _reload_perms()
    return {"ok": True, "role": role, "perms": _perms_for(role), "scope": _scope_for(role)}


@app.get("/api/admin/assignable-users", dependencies=[Depends(require_token)])
def assignable_users():
    """담당자(PIC) 지정용 직원 목록 — id/username 만. (admin 외 편집자도 사용)"""
    s = get_session()
    try:
        return [{"id": u.id, "username": u.username}
                for u in s.query(User).filter_by(is_active=True)
                .order_by(User.username).all()]
    finally:
        s.close()


@app.get("/api/admin/settings/users", dependencies=[Depends(require_token)])
def settings_users(user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    s = get_session()
    try:
        return [{
            "id": u.id, "username": u.username,
            "email": u.email or "", "role": _enum_val(u.role),
            "is_active": bool(u.is_active),
        } for u in s.query(User).order_by(User.username).all()]
    finally:
        s.close()


@app.post("/api/admin/settings/users", dependencies=[Depends(require_token)])
def create_user(body: UserSave, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    if not body.username.strip() or not body.password:
        raise HTTPException(status_code=400, detail="사용자명과 비밀번호를 입력하세요.")
    s = get_session()
    try:
        u = User(
            username=body.username.strip(),
            email=body.email or "",
            password_hash=bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode(),
            role=UserRole(body.role),
            is_active=body.is_active,
        )
        s.add(u)
        s.commit()
        return {"ok": True, "id": u.id}
    finally:
        s.close()


@app.put("/api/admin/settings/users/{row_id}", dependencies=[Depends(require_token)])
def update_user(row_id: int, body: UserSave, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    s = get_session()
    try:
        u = s.query(User).filter_by(id=row_id).first()
        if not u:
            raise HTTPException(status_code=404, detail="User를 찾을 수 없습니다.")
        u.username = body.username.strip()
        u.email = body.email or ""
        u.role = UserRole(body.role)
        u.is_active = body.is_active
        if body.password:
            u.password_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
        s.commit()
        return {"ok": True, "id": u.id}
    finally:
        s.close()


@app.delete("/api/admin/settings/users/{row_id}", dependencies=[Depends(require_token)])
def delete_user(row_id: int, user: dict = Depends(get_current_user)):
    """사용자 삭제(admin 전용). 본인 계정과 마지막 활성 관리자 계정은 lockout
    방지를 위해 삭제를 막는다. 비활성화는 update_user 의 is_active 로 가능."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    if row_id == user.get("id"):
        raise HTTPException(status_code=400, detail="본인 계정은 삭제할 수 없습니다.")
    s = get_session()
    try:
        u = s.query(User).filter_by(id=row_id).first()
        if not u:
            raise HTTPException(status_code=404, detail="User를 찾을 수 없습니다.")
        if _enum_val(u.role) == "admin" and u.is_active:
            active_admins = (s.query(User)
                             .filter(User.role == UserRole.ADMIN, User.is_active.is_(True))
                             .count())
            if active_admins <= 1:
                raise HTTPException(status_code=400,
                    detail="마지막 활성 관리자 계정은 삭제할 수 없습니다.")
        s.delete(u)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


class PasswordChangeReq(BaseModel):
    old_password: str
    new_password: str


@app.post("/api/admin/me/password", dependencies=[Depends(require_token)])
def change_my_password(body: PasswordChangeReq, user: dict = Depends(get_current_user)):
    """로그인한 본인의 비밀번호 변경 — 현재 비밀번호 확인 후 교체
    (Streamlit 8_Settings.py 비밀번호 변경 패리티)."""
    if not body.new_password or len(body.new_password) < 4:
        raise HTTPException(status_code=400, detail="새 비밀번호는 4자 이상이어야 합니다.")
    uid = user.get("id")
    s = get_session()
    try:
        u = s.query(User).filter_by(id=uid).first()
        if not u:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
        if not bcrypt.checkpw(body.old_password.encode(), u.password_hash.encode()):
            raise HTTPException(status_code=400, detail="현재 비밀번호가 올바르지 않습니다.")
        u.password_hash = bcrypt.hashpw(body.new_password.encode(), bcrypt.gensalt()).decode()
        s.commit()
        return {"ok": True}
    finally:
        s.close()


# ── Write actions ─────────────────────────────────────────────────────────────




class VendorRfqCreate(BaseModel):
    vendor_id: int


class VendorRfqPreviewRequest(BaseModel):
    vendor_ids: list[int]
    lang: str = "en"
    notes: str = ""
    rfq_no_mode: str = "auto"   # 케이마리스 RFQ No. 발번: auto/manual
    rfq_no: str = ""            # manual 일 때 직접 입력값
    items: list[dict] | None = None   # 발신 화면에서 선택·편집한 품목(없으면 RFQ 원본)


@app.post("/api/admin/rfq/{rfq_id}/vendor-rfq-preview",
          dependencies=[Depends(require_token)])
def vendor_rfq_preview(rfq_id: int, body: VendorRfqPreviewRequest):
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        # Vendor RFQ 발신 단계 진입 시 케이마리스 RFQ No. 부여(미발급이면).
        _assign_rfq_no(s, rfq, body.rfq_no_mode, body.rfq_no)
        s.commit()
        cust = s.query(Customer).filter_by(id=rfq.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=rfq.vessel_id).first() if rfq.vessel_id else None
        lang = "ko" if body.lang == "ko" else "en"
        items = _sanitize_vendor_rfq_items(body.items) if body.items is not None else None
        previews = []
        for vid in body.vendor_ids:
            vendor = s.query(Vendor).filter_by(id=vid).first()
            if not vendor:
                continue
            safe_vname = "".join(c for c in vendor.name if c.isalnum() or c in "._- ")[:40]
            previews.append({
                "vendor_id": vendor.id,
                "vendor_name": vendor.name,
                "to": vendor.email or "",
                "subject": (
                    f"[K-MARIS] 견적 요청 — {rfq.rfq_no} / {vessel.name if vessel else rfq.rfq_no}"
                    if lang == "ko"
                    else f"[K-MARIS] Inquiry — {rfq.rfq_no} / {vessel.name if vessel else rfq.rfq_no}"
                ),
                "body": _vendor_rfq_email_body(rfq, cust, vessel, vendor, body.notes, lang, items),
                "xlsx_filename": f"{rfq.rfq_no}_VendorQuoteSheet_{safe_vname}.xlsx",
            })
        return {
            "previews": previews,
            "smtp_configured": bool(os.getenv("SMTP_USER") and os.getenv("SMTP_PASSWORD")),
        }
    finally:
        s.close()


@app.get("/api/admin/rfq/{rfq_id}/vendor-rfq-xlsx/{vendor_id}",
         dependencies=[Depends(require_token)])
def vendor_rfq_xlsx(rfq_id: int, vendor_id: int):
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        vendor = s.query(Vendor).filter_by(id=vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor를 찾을 수 없습니다.")
        cust = s.query(Customer).filter_by(id=rfq.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=rfq.vessel_id).first() if rfq.vessel_id else None
        xlsx = make_vendor_rfq_quote_xlsx(
            rfq_no=rfq.rfq_no,
            vessel_name=vessel.name if vessel else "—",
            customer_name=cust.name if cust else "—",
            enquiry_date=rfq.date or date.today().isoformat(),
            vendor_name=vendor.name,
            items=rfq.items or [],
        )
        safe_vname = "".join(c for c in vendor.name if c.isalnum() or c in "._- ")[:40]
        filename = f"{rfq.rfq_no}_VendorQuoteSheet_{safe_vname}.xlsx"
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        s.close()


class VendorRfqXlsxRequest(BaseModel):
    items: list[dict] | None = None   # 선택·편집한 품목(없으면 RFQ 원본)


@app.post("/api/admin/rfq/{rfq_id}/vendor-rfq-xlsx/{vendor_id}",
          dependencies=[Depends(require_token)])
def vendor_rfq_xlsx_post(rfq_id: int, vendor_id: int, body: VendorRfqXlsxRequest):
    """XLSX 견적 양식 — 발신 화면에서 선택·편집한 품목을 반영해 생성(POST)."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        vendor = s.query(Vendor).filter_by(id=vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=404, detail="Vendor를 찾을 수 없습니다.")
        cust = s.query(Customer).filter_by(id=rfq.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=rfq.vessel_id).first() if rfq.vessel_id else None
        items = (_sanitize_vendor_rfq_items(body.items)
                 if body.items is not None else (rfq.items or []))
        xlsx = make_vendor_rfq_quote_xlsx(
            rfq_no=rfq.rfq_no,
            vessel_name=vessel.name if vessel else "—",
            customer_name=cust.name if cust else "—",
            enquiry_date=rfq.date or date.today().isoformat(),
            vendor_name=vendor.name,
            items=items,
        )
        safe_vname = "".join(c for c in vendor.name if c.isalnum() or c in "._- ")[:40]
        filename = f"{rfq.rfq_no}_VendorQuoteSheet_{safe_vname}.xlsx"
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        s.close()


class VendorRfqSendItem(BaseModel):
    vendor_id: int
    to: str = ""
    subject: str
    body: str


class VendorRfqSendRequest(BaseModel):
    items: list[VendorRfqSendItem]
    rfq_no_mode: str = "auto"
    rfq_no: str = ""
    sent_at: str = ""        # 발신 일시 "YYYY-MM-DDTHH:MM"(비우면 현재)
    rfq_items: list[dict] | None = None   # 선택·편집한 품목(없으면 RFQ 원본을 그대로 저장)


@app.post("/api/admin/rfq/{rfq_id}/vendor-rfq-send",
          dependencies=[Depends(require_token)])
def vendor_rfq_send(rfq_id: int, body: VendorRfqSendRequest):
    """Vendor RFQ '발신 완료' 기록 — 시스템이 직접 이메일을 발송하지 않고, 선택한
    Vendor별 VendorRFQ 레코드를 저장(2단계 완료)한다. 케이마리스 RFQ No.도 부여한다.
    이메일은 '이메일 생성'에서 만든 초안을 사용자가 직접 발송한다."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        # 케이마리스 RFQ No. 부여(미발급이면)
        _assign_rfq_no(s, rfq, body.rfq_no_mode, body.rfq_no)
        sent_at = (body.sent_at or "").strip() or _kst_iso(datetime.utcnow())
        # 발신 화면에서 선택·편집한 품목이 오면 그것을, 없으면 RFQ 원본을 저장.
        sent_items = (_sanitize_vendor_rfq_items(body.rfq_items)
                      if body.rfq_items is not None else (rfq.items or []))
        saved = 0
        result_rows = []
        for item in body.items:
            vendor = s.query(Vendor).filter_by(id=item.vendor_id).first()
            if not vendor:
                continue
            vrfq = VendorRFQ(
                rfq_id=rfq.id,
                vendor_id=vendor.id,
                sent_date=sent_at[:10],
                sent_at=sent_at,
                sent_to_email=item.to or "",
                status="발신완료",
                items=sent_items,
            )
            s.add(vrfq)
            s.flush()
            saved += 1
            result_rows.append({"vendor": vendor.name})

        rfq.status = RFQStatus.SOURCING
        s.commit()
        return {
            "ok": True,
            "saved": saved,
            "rows": result_rows,
            "rfq_no": _rfq_no_disp(rfq.rfq_no),
        }
    finally:
        s.close()


@app.post("/api/admin/rfq/{rfq_id}/vendor-rfq",
          dependencies=[Depends(require_token)])
def create_vendor_rfq(rfq_id: int, body: VendorRfqCreate):
    """RFQ로부터 Vendor RFQ 발신(생성). 품목은 RFQ 품목을 그대로 이관한다."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        vendor = s.query(Vendor).filter_by(id=body.vendor_id).first()
        if not vendor:
            raise HTTPException(status_code=400, detail="Vendor를 선택하세요.")
        _assign_rfq_no(s, rfq)   # 미발급이면 케이마리스 RFQ No. 자동 부여

        # 요청 품목(가격 제외)만 이관
        req_items = [{
            "part_no": it.get("part_no", ""),
            "description": it.get("description", ""),
            "qty": it.get("qty", 1),
        } for it in (rfq.items or [])]

        vrfq = VendorRFQ(
            rfq_id=rfq.id,
            vendor_id=vendor.id,
            sent_date=date.today().strftime("%Y-%m-%d"),
            sent_to_email=vendor.email or "",
            status="발송됨",
            items=req_items,
        )
        s.add(vrfq)
        s.commit()
        return {"ok": True, "id": vrfq.id, "vendor": vendor.name}
    finally:
        s.close()


class VendorRfqUpdate(BaseModel):
    vendor_id: int | None = None
    sent_date: str | None = None
    sent_at: str | None = None
    sent_to_email: str | None = None
    status: str | None = None
    items: list[dict] | None = None


@app.get("/api/admin/vendor-rfq/{vrfq_id}", dependencies=[Depends(require_token)])
def vendor_rfq_detail(vrfq_id: int):
    """Vendor RFQ(발신) 1건 상세."""
    s = get_session()
    try:
        vr = s.query(VendorRFQ).filter_by(id=vrfq_id).first()
        if not vr:
            raise HTTPException(status_code=404, detail="Vendor RFQ를 찾을 수 없습니다.")
        vendor = s.query(Vendor).filter_by(id=vr.vendor_id).first()
        rfq = s.query(RFQ).filter_by(id=vr.rfq_id).first() if vr.rfq_id else None
        customer = s.query(Customer).filter_by(id=rfq.customer_id).first() if rfq else None
        vessel = s.query(Vessel).filter_by(id=rfq.vessel_id).first() if rfq and rfq.vessel_id else None
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        vendor_emails = {v.id: (v.email or "") for v in s.query(Vendor).all()}
        quote_count = s.query(VendorQuote).filter_by(vendor_rfq_id=vr.id).count()
        sibling_vrfqs = []
        if rfq:
            quote_counts: dict[int, int] = {}
            for q in s.query(VendorQuote).filter(VendorQuote.vendor_rfq_id.in_(
                [x.id for x in s.query(VendorRFQ).filter_by(rfq_id=rfq.id).all()]
            )).all():
                quote_counts[q.vendor_rfq_id] = quote_counts.get(q.vendor_rfq_id, 0) + 1
            for x in s.query(VendorRFQ).filter_by(rfq_id=rfq.id).order_by(VendorRFQ.id.desc()).all():
                sibling_vrfqs.append({
                    "id": x.id,
                    "vendor": vendor_names.get(x.vendor_id, "—"),
                    "vendor_email": x.sent_to_email or vendor_emails.get(x.vendor_id, "") or "",
                    "sent_at": x.sent_at or "",
                    "status": x.status or "",
                    "quote_count": quote_counts.get(x.id, 0),
                    "current": x.id == vr.id,
                })
        return {
            "id": vr.id,
            "rfq_id": vr.rfq_id,
            "assignee_id": (rfq.created_by or 0) if rfq else 0,
            "customer_rfq_no": rfq.customer_rfq_no if rfq else "",
            "kmaris_rfq_no": _rfq_no_disp(rfq.rfq_no) if rfq else "",
            "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
            "first_rfq_at": _first_rfq_iso(rfq) if rfq else "",
            "customer": customer.name if customer else "",
            "customer_contact": getattr(customer, "contact", "") if customer else "",
            "customer_email": getattr(customer, "email", "") if customer else "",
            "vessel": vessel.name if vessel else "—",
            "project_title": rfq.project_title if rfq else "",
            "work_type": rfq.work_type if rfq else "",
            "received_at": rfq.received_at if rfq else "",
            "vendor_id": vr.vendor_id or 0,
            "vendor": vendor.name if vendor else "—",
            "vendor_email": vr.sent_to_email or (vendor.email if vendor else "") or "",
            "sent_date": vr.sent_date or "",
            "sent_at": vr.sent_at or "",
            "status": vr.status or "",
            "quote_count": quote_count,
            "items": [_item_view(it) for it in (vr.items or [])],
            "project_vendor_rfqs": sibling_vrfqs,
        }
    finally:
        s.close()


@app.put("/api/admin/vendor-rfq/{vrfq_id}", dependencies=[Depends(require_token)])
def update_vendor_rfq(vrfq_id: int, body: VendorRfqUpdate):
    """Vendor RFQ 수정 — Vendor·발신정보·상태·품목 교체."""
    s = get_session()
    try:
        vr = s.query(VendorRFQ).filter_by(id=vrfq_id).first()
        if not vr:
            raise HTTPException(status_code=404, detail="Vendor RFQ를 찾을 수 없습니다.")
        if body.vendor_id is not None:
            vr.vendor_id = body.vendor_id
        if body.sent_to_email is not None:
            vr.sent_to_email = body.sent_to_email.strip()
        if body.status is not None:
            vr.status = body.status.strip() or vr.status
        if body.sent_at is not None:
            vr.sent_at = body.sent_at.strip()
            if body.sent_at.strip():
                vr.sent_date = body.sent_at.strip()[:10]
        if body.sent_date is not None:
            vr.sent_date = body.sent_date.strip()
        if body.items is not None:
            vr.items = [{
                "part_no": (it.get("part_no") or "").strip(),
                "description": (it.get("description") or "").strip(),
                "qty": it.get("qty", 1) or 1,
                "unit": (it.get("unit") or "").strip(),
                "remark": (it.get("remark") or "").strip(),
            } for it in body.items if (it.get("part_no") or it.get("description"))]
        s.commit()
        return {"ok": True, "id": vr.id}
    finally:
        s.close()


@app.delete("/api/admin/vendor-rfq/{vrfq_id}", dependencies=[Depends(require_token)])
def delete_vendor_rfq(vrfq_id: int):
    """Vendor RFQ 삭제 — 수신된 Vendor 견적이 있으면 거부한다."""
    s = get_session()
    try:
        vr = s.query(VendorRFQ).filter_by(id=vrfq_id).first()
        if not vr:
            raise HTTPException(status_code=404, detail="Vendor RFQ를 찾을 수 없습니다.")
        if s.query(VendorQuote).filter_by(vendor_rfq_id=vrfq_id).first():
            raise HTTPException(status_code=400,
                detail="수신된 Vendor 견적이 있는 Vendor RFQ 입니다. 먼저 견적을 삭제하세요.")
        s.query(VendorRFQ).filter_by(id=vrfq_id).delete(synchronize_session=False)
        s.commit()
        return {"ok": True, "id": vrfq_id}
    finally:
        s.close()


class VendorQuoteCreate(BaseModel):
    vendor_rfq_id: int
    vendor_quote_no: str
    amount: float | None = None
    currency: str = "USD"
    received_date: str | None = None
    received_at: str | None = None     # 견적 수신 일시 "YYYY-MM-DDTHH:MM"(비우면 현재)
    notes: str = ""
    items: list[dict] | None = None
    terms: dict | None = None


@app.post("/api/admin/vendor-quote-parse", dependencies=[Depends(require_token)])
def vendor_quote_parse(file: UploadFile = File(...)):
    """Vendor 견적 응답 파일(PDF/Excel/이미지) → 품목 리스트 자동 추출.

    정형 양식(KTMS 견적요청 시트)은 표 파서로 먼저 시도하고, 비정형 PDF는
    Claude 텍스트 파서로, 이미지/캡쳐는 Claude 비전으로 추출한다.
    """
    name = file.filename or ""
    lower = name.lower()
    img_media = _ocr_image_media_type(file)
    try:
        file.file.seek(0)
        raw = file.file.read()

        # 1) 이미지/캡쳐 → Claude 비전
        if img_media:
            return parse_vendor_quote_image(raw, img_media)

        # 2) Excel/정형 PDF → 표 파서 우선
        if lower.endswith((".xlsx", ".xls", ".pdf")):
            items = parse_vendor_quote_bytes(raw, name)
            if items:
                return {"items": items}

            # 3) 표 파서 실패 → Claude 폴백
            if lower.endswith(".pdf"):
                # 3a) 텍스트가 있으면 텍스트 파서
                text = extract_text_from_pdf(io.BytesIO(raw))
                if text:
                    result = parse_vendor_quote_text(text)
                    if result.get("items"):
                        return result
                # 3b) 텍스트 없음(스캔본)·텍스트 파서 실패 → PDF 비전 파서
                return parse_vendor_quote_pdf_document(raw)

            # Excel 비정형 → 셀 전체를 텍스트로 덤프해 Claude 텍스트 파서로 폴백
            xls_text = excel_to_text(raw)
            if xls_text:
                return parse_vendor_quote_text(xls_text)
            return {"items": []}

        raise HTTPException(
            status_code=400,
            detail="PDF·Excel 또는 이미지(PNG·JPG·WEBP) 파일만 업로드할 수 있습니다.",
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Vendor 견적 파싱 실패: {exc}") from exc


@app.post("/api/admin/rfq/{rfq_id}/vendor-quote",
          dependencies=[Depends(require_token)])
def create_vendor_quote(rfq_id: int, body: VendorQuoteCreate):
    """Vendor Quote 수신 등록. 품목 단위 items가 있으면 그대로 저장한다."""
    s = get_session()
    try:
        vrfq = s.query(VendorRFQ).filter_by(id=body.vendor_rfq_id, rfq_id=rfq_id).first()
        if not vrfq:
            raise HTTPException(status_code=400, detail="해당 RFQ의 Vendor RFQ를 선택하세요.")
        # Vendor 견적번호는 선택 입력(비워도 등록 가능).

        items = body.items
        if not items:
            amount = float(body.amount or 0)
            items = [{"cost_price": amount, "qty": 1, "amount": amount}]

        # 수신 일시: 수동 입력(received_at) 우선, 없으면 날짜만, 둘 다 없으면 현재(KST)
        received_at = (body.received_at or "").strip()
        if not received_at:
            received_at = _date_iso(body.received_date) or _kst_iso(datetime.utcnow())

        vq = VendorQuote(
            vendor_rfq_id=vrfq.id,
            vendor_quote_no=body.vendor_quote_no.strip(),
            received_date=received_at[:10],
            received_at=received_at,
            currency=body.currency or "USD",
            items=items,
            terms=body.terms or {},
            notes=body.notes or "",
        )
        s.add(vq)
        vrfq.status = "견적 수신완료"
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if rfq and rfq.status == RFQStatus.SOURCING:
            rfq.status = RFQStatus.QUOTING
        s.commit()
        return {"ok": True, "vendor_quote_no": vq.vendor_quote_no}
    finally:
        s.close()


@app.get("/api/admin/rfq/{rfq_id}/vendor-quotes",
         dependencies=[Depends(require_token)])
def rfq_vendor_quotes(rfq_id: int):
    """해당 RFQ의 Vendor 견적 목록(품목 포함). Customer Quotation 작성 시
    공급사 견적에서 cost_price/품목 정보를 불러오기 위한 selector 데이터."""
    s = get_session()
    try:
        vrfqs = s.query(VendorRFQ).filter_by(rfq_id=rfq_id).all()
        vrfq_map = {v.id: v for v in vrfqs}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        vqs = (s.query(VendorQuote)
               .filter(VendorQuote.vendor_rfq_id.in_(list(vrfq_map.keys())))
               .order_by(VendorQuote.id.desc()).all() if vrfq_map else [])
        out = []
        for q in vqs:
            vrfq = vrfq_map.get(q.vendor_rfq_id)
            out.append({
                "id": q.id,
                "vendor_quote_no": getattr(q, "vendor_quote_no", None) or "—",
                "vendor": vendor_names.get(vrfq.vendor_id, "—") if vrfq else "—",
                "received_date": q.received_date or "",
                "received_at": getattr(q, "received_at", None) or "",
                "currency": getattr(q, "currency", None) or "USD",
                "items": q.items or [],
                "terms": getattr(q, "terms", None) or {},
            })
        return {"vendor_quotes": out}
    finally:
        s.close()


class VendorQuoteUpdate(BaseModel):
    vendor_quote_no: str | None = None
    received_date: str | None = None
    received_at: str | None = None
    currency: str | None = None
    notes: str | None = None
    items: list[dict] | None = None
    terms: dict | None = None


@app.get("/api/admin/vendor-quote/{vq_id}", dependencies=[Depends(require_token)])
def vendor_quote_detail(vq_id: int):
    """Vendor Quote(수신 견적) 1건 상세 — 원본 품목(cost_price 등) 포함."""
    s = get_session()
    try:
        q = s.query(VendorQuote).filter_by(id=vq_id).first()
        if not q:
            raise HTTPException(status_code=404, detail="Vendor 견적을 찾을 수 없습니다.")
        vr = s.query(VendorRFQ).filter_by(id=q.vendor_rfq_id).first()
        vendor = s.query(Vendor).filter_by(id=vr.vendor_id).first() if vr else None
        rfq = s.query(RFQ).filter_by(id=vr.rfq_id).first() if vr and vr.rfq_id else None
        return {
            "id": q.id,
            "vendor_quote_no": q.vendor_quote_no or "",
            "vendor_rfq_id": q.vendor_rfq_id,
            "rfq_id": vr.rfq_id if vr else None,
            "assignee_id": (rfq.created_by or 0) if rfq else 0,
            "customer_rfq_no": _rfq_no_disp(rfq.rfq_no) if rfq else "",
            **_base_meta(s, rfq),   # 공통 기본정보(고객·선박·업무·Project No.·최초 RFQ)
            "vendor": vendor.name if vendor else "—",
            "received_date": q.received_date or "",
            "received_at": q.received_at or "",
            "notes": q.notes or "",
            "currency": getattr(q, "currency", None) or "USD",
            "items": q.items or [],
            "terms": getattr(q, "terms", None) or {},
        }
    finally:
        s.close()


@app.put("/api/admin/vendor-quote/{vq_id}", dependencies=[Depends(require_token)])
def update_vendor_quote(vq_id: int, body: VendorQuoteUpdate):
    """Vendor Quote 수정 — 견적번호·수신일시·비고·품목 교체."""
    s = get_session()
    try:
        q = s.query(VendorQuote).filter_by(id=vq_id).first()
        if not q:
            raise HTTPException(status_code=404, detail="Vendor 견적을 찾을 수 없습니다.")
        if body.vendor_quote_no is not None:
            q.vendor_quote_no = body.vendor_quote_no.strip()
        if body.notes is not None:
            q.notes = body.notes
        if body.currency is not None:
            cur = (body.currency or "USD").strip().upper() or "USD"
            q.currency = cur
            s.flush()
            s.execute(
                text("UPDATE vendor_quotes SET currency = :currency WHERE id = :id"),
                {"currency": cur, "id": vq_id},
            )
        if body.received_at is not None and body.received_at.strip():
            q.received_at = body.received_at.strip()
            q.received_date = body.received_at.strip()[:10]
        elif body.received_date is not None:
            q.received_date = body.received_date.strip()
        if body.items is not None:
            q.items = body.items
        if body.terms is not None:
            q.terms = body.terms
        s.commit()
        saved_currency = (
            s.execute(text("SELECT currency FROM vendor_quotes WHERE id = :id"), {"id": vq_id}).scalar()
            or "USD"
        )
        return {"ok": True, "vendor_quote_no": q.vendor_quote_no, "currency": saved_currency}
    finally:
        s.close()


@app.delete("/api/admin/vendor-quote/{vq_id}", dependencies=[Depends(require_token)])
def delete_vendor_quote(vq_id: int):
    """Vendor Quote 삭제."""
    s = get_session()
    try:
        q = s.query(VendorQuote).filter_by(id=vq_id).first()
        if not q:
            raise HTTPException(status_code=404, detail="Vendor 견적을 찾을 수 없습니다.")
        no = q.vendor_quote_no or ""
        s.query(VendorQuote).filter_by(id=vq_id).delete(synchronize_session=False)
        s.commit()
        return {"ok": True, "vendor_quote_no": no}
    finally:
        s.close()




# ── K-Maris RFQ No. 이연 발번 ──────────────────────────────────────────────
# 케이마리스 RFQ No.는 Vendor RFQ 발신 시점에 부여한다. 그 전까지는 임시 토큰
# (TMP-...)을 보유하며, 사용자에게는 "미발급"으로 표시된다.
_RFQ_TMP_PREFIX = "TMP-"


def _rfq_unassigned(rfq_no) -> bool:
    return (not rfq_no) or str(rfq_no).startswith(_RFQ_TMP_PREFIX)


def _rfq_no_disp(rfq_no) -> str:
    """사용자 표시용: 미발급(임시 토큰/빈값)이면 '-'."""
    return "-" if _rfq_unassigned(rfq_no) else rfq_no


def _new_tmp_rfq_no(session) -> str:
    while True:
        cand = _RFQ_TMP_PREFIX + secrets.token_hex(5)
        if not session.query(RFQ).filter_by(rfq_no=cand).first():
            return cand


def _assign_rfq_no(session, rfq, mode: str = "auto", manual: str = "") -> str:
    """RFQ 가 아직 미발급이면 수동 입력값만 부여한다. 비워두면 임시 토큰을 유지한다."""
    if not _rfq_unassigned(rfq.rfq_no):
        return rfq.rfq_no
    manual = (manual or "").strip()
    if mode == "manual" and manual:
        if session.query(RFQ).filter_by(rfq_no=manual).first():
            raise HTTPException(status_code=400, detail=f"이미 존재하는 RFQ No.입니다: {manual}")
        rfq.rfq_no = manual
    return rfq.rfq_no


class RfqItemIn(BaseModel):
    part_no: str = ""
    description: str = ""
    qty: float = 1
    remark: str | None = ""


class RfqCreate(BaseModel):
    customer_id: int
    vessel_id: int | None = None
    customer_rfq_no: str | None = ""
    contact_person: str | None = ""    # 고객 담당자
    rfq_no: str | None = None          # K-Maris RFQ No. 수동 지정(비우면 자동 채번)
    received_at: str | None = None     # RFQ 수신 일시 "YYYY-MM-DDTHH:MM"(비우면 현재)
    project_title: str | None = ""
    work_type: str | None = "부품공급"
    request_channel: str | None = ""   # 고객 요청 수단: Email/Phone/SMS/WhatsApp/WeChat 등
    notes: str | None = ""             # 내부 메모(자유 서술)
    items: list[RfqItemIn] = []


@app.post("/api/admin/rfq", dependencies=[Depends(require_token)])
def create_rfq(body: RfqCreate, user: dict = Depends(get_current_user)):
    """Customer RFQ 신규 등록. 케이마리스 RFQ No.는 기본적으로 미발급(임시) 상태로
    두고 Vendor RFQ 발신 시점에 부여한다. body.rfq_no 로 수동 선지정도 가능."""
    s = get_session()
    try:
        cust = s.query(Customer).filter_by(id=body.customer_id).first()
        if not cust:
            raise HTTPException(status_code=400, detail="Customer를 선택하세요.")
        items = [{
            "part_no": (it.part_no or "").strip(),
            "description": (it.description or "").strip(),
            "qty": it.qty or 1,
            "remark": (it.remark or "").strip(),
        } for it in body.items if (it.part_no or it.description)]

        try:
            work_type = WorkType(body.work_type) if body.work_type else WorkType.PARTS
        except ValueError:
            work_type = WorkType.PARTS

        manual_no = (body.rfq_no or "").strip()
        if manual_no:
            if s.query(RFQ).filter_by(rfq_no=manual_no).first():
                raise HTTPException(status_code=400, detail=f"이미 존재하는 RFQ No.입니다: {manual_no}")
            rfq_no = manual_no
        else:
            rfq_no = _new_tmp_rfq_no(s)   # 미발급 — Vendor RFQ 발신 시 부여

        received_at = (body.received_at or "").strip() or _kst_iso(datetime.utcnow())
        rfq = RFQ(
            rfq_no=rfq_no,
            customer_rfq_no=(body.customer_rfq_no or "").strip() or None,
            contact_person=(body.contact_person or "").strip() or None,
            project_title=(body.project_title or "").strip() or None,
            work_type=work_type,
            request_channel=(body.request_channel or "").strip() or None,
            notes=(body.notes or "").strip() or None,
            customer_id=cust.id,
            vessel_id=body.vessel_id,
            date=received_at[:10],
            received_at=received_at,
            status=RFQStatus.RECEIVED,
            items=items,
            created_by=(user.get("id") or None),   # 담당자 = 등록한 내부 직원
        )
        s.add(rfq)
        s.commit()
        return {"ok": True, "id": rfq.id, "rfq_no": _rfq_no_disp(rfq_no)}
    finally:
        s.close()


class RfqAssignNo(BaseModel):
    mode: str = "auto"     # auto/manual
    rfq_no: str = ""       # manual 일 때 직접 입력값


@app.post("/api/admin/rfq/{rfq_id}/assign-no", dependencies=[Depends(require_token)])
def assign_rfq_no_endpoint(rfq_id: int, body: RfqAssignNo):
    """케이마리스 RFQ No. 단독 발번(미발급이면 부여, 이미 발급됐으면 유지)."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        no = _assign_rfq_no(s, rfq, body.mode, body.rfq_no)
        s.commit()
        return {"ok": True, "rfq_no": _rfq_no_disp(no)}
    finally:
        s.close()


class RfqUpdate(BaseModel):
    """RFQ 헤더 필드 부분 수정. 보낸 필드만 반영(None=변경 안 함)."""
    customer_id: int | None = None
    vessel_id: int | None = None        # 0 → 선박 미지정으로 해제
    customer_rfq_no: str | None = None
    rfq_no: str | None = None           # K-Maris RFQ No. 수동 수정(빈값이면 변경 안 함)
    contact_person: str | None = None
    project_title: str | None = None
    work_type: str | None = None
    request_channel: str | None = None  # 고객 요청 수단
    notes: str | None = None            # 내부 메모(자유 서술)
    received_at: str | None = None      # "YYYY-MM-DDTHH:MM"
    assignee_id: int | None = None      # 담당자(PIC) = created_by. 0 → 미지정 해제
    items: list[RfqItemIn] | None = None  # 보내면 품목 전체 교체


@app.patch("/api/admin/rfq/{rfq_id}", dependencies=[Depends(require_token)])
def update_rfq(rfq_id: int, body: RfqUpdate):
    """업무 타입·고객사·선박·고객 RFQ No.·프로젝트 제목 등 RFQ 헤더 필드를 수정한다."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")

        if body.customer_id is not None:
            cust = s.query(Customer).filter_by(id=body.customer_id).first()
            if not cust:
                raise HTTPException(status_code=400, detail="Customer를 찾을 수 없습니다.")
            rfq.customer_id = body.customer_id
        if body.vessel_id is not None:
            # 0/음수 → 선박 미지정으로 해제
            if body.vessel_id <= 0:
                rfq.vessel_id = None
            else:
                vessel = s.query(Vessel).filter_by(id=body.vessel_id).first()
                if not vessel:
                    raise HTTPException(status_code=400, detail="선박을 찾을 수 없습니다.")
                rfq.vessel_id = body.vessel_id
        if body.customer_rfq_no is not None:
            rfq.customer_rfq_no = body.customer_rfq_no.strip() or None
        if body.rfq_no is not None:
            new_no = body.rfq_no.strip()
            if new_no and new_no != rfq.rfq_no:
                dup = s.query(RFQ).filter(RFQ.rfq_no == new_no, RFQ.id != rfq_id).first()
                if dup:
                    raise HTTPException(status_code=400, detail=f"이미 존재하는 RFQ No.입니다: {new_no}")
                rfq.rfq_no = new_no
        if body.contact_person is not None:
            rfq.contact_person = body.contact_person.strip() or None
        if body.project_title is not None:
            rfq.project_title = body.project_title.strip() or None
        if body.request_channel is not None:
            rfq.request_channel = body.request_channel.strip() or None
        if body.notes is not None:
            rfq.notes = body.notes.strip() or None
        if body.work_type is not None:
            wt = _coerce_work_type(body.work_type)
            if wt is None:
                raise HTTPException(status_code=400, detail="잘못된 업무 타입입니다.")
            rfq.work_type = wt
        if body.received_at is not None:
            recv = body.received_at.strip()
            if recv:
                rfq.received_at = recv
                rfq.date = recv[:10]
        if body.assignee_id is not None:
            # 0/음수 → 담당자 미지정 해제. 그 외엔 해당 직원이 존재할 때만 지정.
            if body.assignee_id <= 0:
                rfq.created_by = None
            elif s.query(User).filter_by(id=body.assignee_id).first():
                rfq.created_by = body.assignee_id
            else:
                raise HTTPException(status_code=400, detail="담당자(사용자)를 찾을 수 없습니다.")
        if body.items is not None:
            rfq.items = [{
                "part_no": (it.part_no or "").strip(),
                "description": (it.description or "").strip(),
                "qty": it.qty or 1,
                "remark": (it.remark or "").strip(),
            } for it in body.items if (it.part_no or it.description)]

        s.commit()
        return {"ok": True, "id": rfq.id}
    finally:
        s.close()


class RfqLevelUpdate(BaseModel):
    follow_up_level: str


@app.put("/api/admin/rfq/{rfq_id}/level", dependencies=[Depends(require_token)])
def update_rfq_level(rfq_id: int, body: RfqLevelUpdate):
    """RFQ Follow-up Level(A/B/C) 변경. 상태(12단계)는 진행에 따라 자동 반영되므로
    여기서는 Level 만 수정한다 (Streamlit 2_CRFQ.py render_rfq_detail 패리티)."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        try:
            rfq.follow_up_level = FollowUpLevel(body.follow_up_level)
        except ValueError:
            raise HTTPException(status_code=400, detail="잘못된 Level 값입니다.")
        s.commit()
        return {"ok": True, "follow_up_level": _enum_val(rfq.follow_up_level)}
    finally:
        s.close()


class StageDateUpdate(BaseModel):
    stage: int                 # 1~12
    value: str | None = None   # "YYYY-MM-DDTHH:MM" (KST) 또는 빈값/None → 해제


@app.put("/api/admin/rfq/{rfq_id}/stage-date", dependencies=[Depends(require_token)])
def update_rfq_stage_date(rfq_id: int, body: StageDateUpdate):
    """내부 12단계 중 한 단계의 완료 일시를 수동 입력/수정/해제한다."""
    if not (1 <= body.stage <= len(INTERNAL_STEPS)):
        raise HTTPException(status_code=400, detail="잘못된 단계 번호입니다.")
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        # JSON 컬럼은 새 dict 로 재할당해야 변경이 감지된다.
        dates = dict(getattr(rfq, "stage_dates", None) or {})
        key = str(body.stage)
        val = (body.value or "").strip()
        if val:
            dates[key] = val
        else:
            dates.pop(key, None)
        rfq.stage_dates = dates
        s.commit()
        return {"ok": True, "stage_dates": dates}
    finally:
        s.close()


class StageNoteAdd(BaseModel):
    stage: int                       # 1~12
    text: str
    datetime: str | None = None      # 활동 일시 "YYYY-MM-DDTHH:MM" (KST). 비우면 현재시각
    party: str | None = None         # 소통 상대: Customer / Vendor / 기타
    channel: str | None = None       # 소통 수단: 이메일 / 통화 / 문자 / 방문 / 기타


@app.post("/api/admin/rfq/{rfq_id}/stage-note", dependencies=[Depends(require_token)])
def add_rfq_stage_note(rfq_id: int, body: StageNoteAdd):
    """내부 12단계 중 한 단계에 코멘트/활동이력을 추가한다(누적 기록).
    날짜·시각·소통 상대(Customer/Vendor)·소통 수단(이메일/통화/문자 등)·내용을 함께 저장."""
    if not (1 <= body.stage <= len(INTERNAL_STEPS)):
        raise HTTPException(status_code=400, detail="잘못된 단계 번호입니다.")
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="활동 내용을 입력하세요.")
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        notes = dict(getattr(rfq, "stage_notes", None) or {})
        key = str(body.stage)
        log = list(notes.get(key, []))
        log.append({
            "text": text,
            "datetime": (body.datetime or "").strip() or _kst_iso(datetime.utcnow()),
            "party": (body.party or "").strip(),
            "channel": (body.channel or "").strip(),
            "at": _kst_iso(datetime.utcnow()),   # 기록 생성 시각(감사용)
        })
        notes[key] = log
        rfq.stage_notes = notes  # JSON 컬럼은 새 dict 재할당이 필요
        s.commit()
        return {"ok": True, "stage": body.stage, "notes": log}
    finally:
        s.close()


class StageNoteUpdate(BaseModel):
    stage: int
    index: int                       # 해당 단계 로그 내 인덱스
    text: str
    datetime: str | None = None
    party: str | None = None
    channel: str | None = None


@app.post("/api/admin/rfq/{rfq_id}/stage-note-update", dependencies=[Depends(require_token)])
def update_rfq_stage_note(rfq_id: int, body: StageNoteUpdate):
    """기존 활동 기록 1건을 수정한다. 생성 시각(at)은 유지한다."""
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="활동 내용을 입력하세요.")
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        notes = dict(getattr(rfq, "stage_notes", None) or {})
        key = str(body.stage)
        log = list(notes.get(key, []))
        if not (0 <= body.index < len(log)):
            raise HTTPException(status_code=400, detail="잘못된 기록 인덱스입니다.")
        old = log[body.index]
        log[body.index] = {
            "text": text,
            "datetime": (body.datetime or "").strip() or old.get("datetime") or _kst_iso(datetime.utcnow()),
            "party": (body.party or "").strip(),
            "channel": (body.channel or "").strip(),
            "at": old.get("at") or _kst_iso(datetime.utcnow()),  # 생성 시각 유지
        }
        notes[key] = log
        rfq.stage_notes = notes
        s.commit()
        return {"ok": True, "stage": body.stage, "notes": log}
    finally:
        s.close()


class StageNoteDelete(BaseModel):
    stage: int
    index: int                 # 해당 단계 로그 내 인덱스


@app.post("/api/admin/rfq/{rfq_id}/stage-note-delete", dependencies=[Depends(require_token)])
def delete_rfq_stage_note(rfq_id: int, body: StageNoteDelete):
    """단계 코멘트 1건 삭제."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")
        notes = dict(getattr(rfq, "stage_notes", None) or {})
        key = str(body.stage)
        log = list(notes.get(key, []))
        if 0 <= body.index < len(log):
            log.pop(body.index)
            if log:
                notes[key] = log
            else:
                notes.pop(key, None)
            rfq.stage_notes = notes
            s.commit()
        return {"ok": True, "stage": body.stage, "notes": notes.get(key, [])}
    finally:
        s.close()


@app.delete("/api/admin/rfq/{rfq_id}", dependencies=[Depends(require_token)])
def delete_rfq(rfq_id: int):
    """RFQ 삭제. 연결된 Vendor RFQ/Quote 도 함께 삭제한다. 단, 이미 Customer
    Quotation 이나 Order 로 진행된 건은 데이터 보호를 위해 삭제를 막는다."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")

        if s.query(Quotation).filter_by(rfq_id=rfq_id).first():
            raise HTTPException(status_code=400,
                detail="이미 Customer Quotation 이 연결된 RFQ 입니다. 먼저 견적을 정리하세요.")
        if s.query(Order).filter_by(rfq_id=rfq_id).first():
            raise HTTPException(status_code=400,
                detail="이미 Order 로 진행된 RFQ 입니다. 삭제할 수 없습니다.")

        rfq_no = rfq.rfq_no
        vrfq_ids = [v.id for v in s.query(VendorRFQ).filter_by(rfq_id=rfq_id).all()]
        if vrfq_ids:
            (s.query(VendorQuote)
             .filter(VendorQuote.vendor_rfq_id.in_(vrfq_ids))
             .delete(synchronize_session=False))
        s.query(VendorRFQ).filter_by(rfq_id=rfq_id).delete(synchronize_session=False)
        s.query(RFQ).filter_by(id=rfq_id).delete(synchronize_session=False)
        s.commit()
        return {"ok": True, "rfq_no": rfq_no}
    finally:
        s.close()




class CustomerQuoteCreate(BaseModel):
    qtn_no: str | None = None
    currency: str = "USD"
    cost_currency: str | None = None
    round_digits: int | None = None
    discount_pct: float | None = None
    amount: float | None = None
    items: list[dict] | None = None
    sent_at: str | None = None
    valid_until: str | None = None
    remarks: str = ""
    terms: dict | None = None


@app.post("/api/admin/rfq/{rfq_id}/customer-quote",
          dependencies=[Depends(require_token)])
def create_customer_quote(rfq_id: int, body: CustomerQuoteCreate,
                          user: dict = Depends(get_current_user)):
    """Customer Quote 발신. 품목 단위 items가 있으면 그대로 저장한다."""
    s = get_session()
    try:
        rfq = s.query(RFQ).filter_by(id=rfq_id).first()
        if not rfq:
            raise HTTPException(status_code=404, detail="RFQ를 찾을 수 없습니다.")

        items = body.items
        if not items:
            amount = float(body.amount or 0)
            items = [{"amount": amount, "qty": 1, "unit_price": amount}]

        terms = dict(body.terms or {})
        if body.remarks and not terms.get("remarks"):
            terms["remarks"] = body.remarks

        # 수동 입력. 비우면 번호 없이 저장(나중에 편집에서 채움).
        qtn_no = (body.qtn_no or "").strip() or None
        if qtn_no and s.query(Quotation).filter_by(qtn_no=qtn_no).first():
            raise HTTPException(status_code=409, detail="Quotation No. already exists.")
        sent_at = (body.sent_at or "").strip()
        qtn = Quotation(
            qtn_no=qtn_no,
            rfq_id=rfq.id,
            customer_id=rfq.customer_id,
            vessel_id=rfq.vessel_id,
            currency=(body.currency or "USD"),
            cost_currency=(body.cost_currency or None),
            round_digits=body.round_digits,
            discount_pct=(body.discount_pct or 0.0),
            status=QuotationStatus.SENT,
            valid_until=body.valid_until,
            items=items,
            terms=terms,
            date=date.today().strftime("%Y-%m-%d"),
            sent_date=(sent_at[:10] if sent_at else date.today().strftime("%Y-%m-%d")),
            sent_at=(sent_at or None),
            created_by=(user.get("id") or None),   # 담당자 = 발행한 내부 직원
        )
        s.add(qtn)
        s.commit()
        return {"ok": True, "id": qtn.id, "qtn_no": qtn.qtn_no or ""}
    finally:
        s.close()


class CustomerQuoteUpdate(BaseModel):
    qtn_no: str | None = None
    currency: str | None = None
    cost_currency: str | None = None
    round_digits: int | None = None
    discount_pct: float | None = None
    items: list[dict] | None = None
    sent_at: str | None = None
    valid_until: str | None = None
    status: str | None = None
    terms: dict | None = None


@app.get("/api/admin/quotation/{qtn_id}", dependencies=[Depends(require_token)])
def customer_quotation_detail(qtn_id: int):
    """Customer Quotation 1건 상세 — 원본 품목(cost/margin 등)·거래조건 포함."""
    s = get_session()
    try:
        qtn = s.query(Quotation).filter_by(id=qtn_id).first()
        if not qtn:
            raise HTTPException(status_code=404, detail="견적서를 찾을 수 없습니다.")
        cust = s.query(Customer).filter_by(id=qtn.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=qtn.vessel_id).first() if qtn.vessel_id else None
        rfq = s.query(RFQ).filter_by(id=qtn.rfq_id).first() if qtn.rfq_id else None
        return {
            "id": qtn.id,
            "qtn_no": qtn.qtn_no,
            "rfq_id": qtn.rfq_id,
            "assignee_id": (rfq.created_by or 0) if rfq else 0,
            "rfq_no": _rfq_no_disp(rfq.rfq_no) if rfq else "",
            **_base_meta(s, rfq),   # 공통 기본정보(고객·선박·업무·Project No.·최초 RFQ)
            "currency": qtn.currency or "USD",
            "cost_currency": getattr(qtn, "cost_currency", None) or "",
            "round_digits": getattr(qtn, "round_digits", None),
            "discount_pct": getattr(qtn, "discount_pct", 0) or 0,
            "amount": round(_quotation_total(qtn.items or [], getattr(qtn, "discount_pct", 0) or 0), 2),
            "valid_until": qtn.valid_until or "",
            "status": _enum_val(qtn.status),
            "level": _enum_val(qtn.follow_up_level) if qtn.follow_up_level else "",
            "sent_at": getattr(qtn, "sent_at", None) or "",
            "sent_date": getattr(qtn, "sent_at", None) or qtn.sent_date or "",
            "date": qtn.date or "",
            "terms": qtn.terms or {},
            "items": qtn.items or [],
        }
    finally:
        s.close()


@app.put("/api/admin/quotation/{qtn_id}", dependencies=[Depends(require_token)])
def update_customer_quotation(qtn_id: int, body: CustomerQuoteUpdate):
    """Customer Quotation 수정 — 통화·유효기간·상태·거래조건·품목 교체."""
    s = get_session()
    try:
        qtn = s.query(Quotation).filter_by(id=qtn_id).first()
        if not qtn:
            raise HTTPException(status_code=404, detail="견적서를 찾을 수 없습니다.")
        if body.qtn_no is not None:
            qtn_no = body.qtn_no.strip()
            if qtn_no:
                dup = s.query(Quotation).filter(Quotation.qtn_no == qtn_no, Quotation.id != qtn_id).first()
                if dup:
                    raise HTTPException(status_code=409, detail="Quotation No. already exists.")
                qtn.qtn_no = qtn_no
        if body.sent_at is not None:
            sent_at = body.sent_at.strip()
            qtn.sent_at = sent_at or None
            qtn.sent_date = sent_at[:10] if sent_at else None
        if body.currency is not None:
            qtn.currency = body.currency or "USD"
        if body.cost_currency is not None:
            qtn.cost_currency = body.cost_currency or None
        if body.round_digits is not None:
            qtn.round_digits = body.round_digits
        if body.discount_pct is not None:
            qtn.discount_pct = body.discount_pct
        if body.valid_until is not None:
            qtn.valid_until = body.valid_until or None
        if body.terms is not None:
            qtn.terms = body.terms
        if body.items is not None:
            qtn.items = body.items
        if body.status is not None and body.status.strip():
            try:
                qtn.status = QuotationStatus(body.status.strip())
            except ValueError:
                pass
        s.commit()
        return {"ok": True, "qtn_no": qtn.qtn_no}
    finally:
        s.close()


@app.delete("/api/admin/quotation/{qtn_id}", dependencies=[Depends(require_token)])
def delete_customer_quotation(qtn_id: int):
    """Customer Quotation 삭제 — 오더로 진행된 견적은 거부한다."""
    s = get_session()
    try:
        qtn = s.query(Quotation).filter_by(id=qtn_id).first()
        if not qtn:
            raise HTTPException(status_code=404, detail="견적서를 찾을 수 없습니다.")
        if s.query(Order).filter_by(quotation_id=qtn_id).first():
            raise HTTPException(status_code=400,
                detail="이미 오더로 진행된 견적입니다. 삭제할 수 없습니다.")
        qtn_no = qtn.qtn_no
        s.query(Quotation).filter_by(id=qtn_id).delete(synchronize_session=False)
        s.commit()
        return {"ok": True, "qtn_no": qtn_no}
    finally:
        s.close()


@app.get("/api/admin/quotations/{qtn_id}/pdf", dependencies=[Depends(require_token)])
def quotation_pdf(qtn_id: int, doc_type: str = "quotation"):
    s = get_session()
    try:
        qtn = s.query(Quotation).filter_by(id=qtn_id).first()
        if not qtn:
            raise HTTPException(status_code=404, detail="견적서를 찾을 수 없습니다.")
        cust = s.query(Customer).filter_by(id=qtn.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=qtn.vessel_id).first() if qtn.vessel_id else None
        payload = build_payload(
            doc_no=qtn.qtn_no,
            date=qtn.date or date.today().isoformat(),
            customer=cust,
            vessel=vessel,
            items=qtn.items or [],
            terms=qtn.terms or {},
            currency=qtn.currency or "USD",
            vat_rate=qtn.vat_rate or 0.0,
            valid_until=qtn.valid_until or "",
            discount_pct=getattr(qtn, "discount_pct", 0) or 0.0,
        )
        pdf = generate_pdf(doc_type, payload)
        return Response(
            content=pdf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{qtn.qtn_no}.pdf"'},
        )
    finally:
        s.close()


class QuotationEmailPreviewReq(BaseModel):
    lang: str = "en"


@app.post("/api/admin/quotations/{qtn_id}/email-preview", dependencies=[Depends(require_token)])
def quotation_email_preview(qtn_id: int, body: QuotationEmailPreviewReq):
    s = get_session()
    try:
        qtn = s.query(Quotation).filter_by(id=qtn_id).first()
        if not qtn:
            raise HTTPException(status_code=404, detail="견적서를 찾을 수 없습니다.")
        cust = s.query(Customer).filter_by(id=qtn.customer_id).first()
        lang = "kr" if body.lang in ("ko", "kr") else "en"
        return {
            "to": (cust.email if cust else "") or "",
            "subject": quotation_email_subject(qtn.qtn_no, lang),
            "body": quotation_email_body(cust.name if cust else "Customer", qtn.qtn_no, "", lang),
            "smtp_configured": bool(os.getenv("SMTP_USER") and os.getenv("SMTP_PASSWORD")),
        }
    finally:
        s.close()


class QuotationSendReq(BaseModel):
    to: str
    subject: str
    body: str
    doc_type: str = "quotation"


@app.post("/api/admin/quotations/{qtn_id}/send", dependencies=[Depends(require_token)])
def quotation_send(qtn_id: int, body: QuotationSendReq):
    s = get_session()
    try:
        qtn = s.query(Quotation).filter_by(id=qtn_id).first()
        if not qtn:
            raise HTTPException(status_code=404, detail="견적서를 찾을 수 없습니다.")
        if not body.to.strip():
            raise HTTPException(status_code=400, detail="수신자 이메일을 입력하세요.")
        cust = s.query(Customer).filter_by(id=qtn.customer_id).first()
        vessel = s.query(Vessel).filter_by(id=qtn.vessel_id).first() if qtn.vessel_id else None
        payload = build_payload(
            doc_no=qtn.qtn_no,
            date=qtn.date or date.today().isoformat(),
            customer=cust,
            vessel=vessel,
            items=qtn.items or [],
            terms=qtn.terms or {},
            currency=qtn.currency or "USD",
            vat_rate=qtn.vat_rate or 0.0,
            valid_until=qtn.valid_until or "",
            discount_pct=getattr(qtn, "discount_pct", 0) or 0.0,
        )
        pdf = generate_pdf(body.doc_type, payload)
        sent = send_email(
            to=body.to.strip(),
            subject=body.subject,
            body=body.body,
            attachments=[(f"{qtn.qtn_no}.pdf", pdf)],
        )
        if not sent:
            raise HTTPException(status_code=400, detail="이메일 발송 실패 — SMTP 설정 또는 서버 상태를 확인하세요.")
        qtn.status = QuotationStatus.SENT
        # 사용자가 발송일시를 직접 입력했으면 보존하고, 비어 있을 때만 발송 시각으로 채운다.
        if not (getattr(qtn, "sent_at", None) or qtn.sent_date):
            qtn.sent_at = _kst_iso(datetime.utcnow())
            qtn.sent_date = qtn.sent_at[:10]
        s.commit()
        return {"ok": True, "sent_date": qtn.sent_at or qtn.sent_date}
    finally:
        s.close()
