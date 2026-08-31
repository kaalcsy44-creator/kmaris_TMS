"""K-Maris TMS — ar routes (split from admin_api.py; behavior unchanged)."""
from __future__ import annotations

from _core import (
    APRecord,
    ARPayment,
    ARRecord,
    ARSave,
    ARStatus,
    CreditNote,
    Customer,
    Depends,
    HTTPException,
    Order,
    ProformaInvoice,
    PurchaseOrder,
    Response,
    User,
    Vendor,
    Vessel,
    _ar_outstanding,
    _ar_recalc_status,
    _ar_status_from_text,
    _enum_val,
    inbound_fee_in,
    _first_rfq_iso,
    _project_no_for_order,
    _project_no_map,
    _rfq_for_order,
    app,
    date,
    get_session,
    io,
    manual_stage_dates,
    require_token,
)



def _ap_progress(po_ids: list, ap_by_po: dict) -> dict:
    """이 오더의 매입(AP) 진척 — 벤더 P/O 총건수와 단계별 완료 건수.
    9(청구서 수취)·10(전자세금계산서 수취)·11(지급 완료)이 각각 총건수와 같아야
    그 단계가 완료된다(_core._deal_progress 의 ap_all_* 와 같은 판정)."""
    aps = [ap_by_po.get(pid) for pid in po_ids]
    return {
        "ap_total": len(po_ids),
        "ap_billed": sum(1 for a in aps if a is not None),
        "ap_tax": sum(1 for a in aps if a is not None and a.tax_received),
        "ap_paid": sum(1 for a in aps if a is not None and _enum_val(a.status) == "완납"),
    }


@app.get("/api/admin/ar-overview", dependencies=[Depends(require_token)])
def ar_overview():
    """미수금(AR) 현황 — 청구/수금/연체."""
    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        vessel_names = {v.id: v.name for v in s.query(Vessel).all()}
        vendor_names = {v.id: v.name for v in s.query(Vendor).all()}
        user_names = {u.id: u.username for u in s.query(User).all()}
        ord_map = {o.id: o for o in s.query(Order).all()}
        # 프로포마 인보이스를 발행한 오더 — 10단계(세금계산서 발행)를 PI 로 갈음한다
        # (_core._deal_progress 의 pi_covers_tax 와 같은 규칙). 오더당 최신 1건만 본다.
        pi_by_order = {p.order_id: p for p in s.query(ProformaInvoice).order_by(ProformaInvoice.id).all()}
        # order_id → 발주 Vendor 이름들(중복 제거, 발주 순) + 벤더 P/O id 목록
        po_vendors_by_order: dict[int, list[str]] = {}
        po_ids_by_order: dict[int, list[int]] = {}
        for po in s.query(PurchaseOrder).order_by(PurchaseOrder.id).all():
            po_ids_by_order.setdefault(po.order_id, []).append(po.id)
            nm = vendor_names.get(po.vendor_id)
            if not nm:
                continue
            lst = po_vendors_by_order.setdefault(po.order_id, [])
            if nm not in lst:
                lst.append(nm)
        # 매입(AP) 진행 — 9~11단계는 매출(AR)·매입(AP) 양쪽이 끝나야 완료라, 화면에서
        # "AR 은 끝났는데 왜 단계가 안 넘어가지?" 를 알 수 있게 벤더측 진척을 함께 내려보낸다.
        ap_by_po = {a.po_id: a for a in s.query(APRecord).all()}
        today_str = date.today().isoformat()

        rows = []
        out_usd = 0.0
        overdue_usd = 0.0
        for r in s.query(ARRecord).order_by(ARRecord.id.desc()).all():
            o = ord_map.get(r.order_id)
            cust = cust_names.get(o.customer_id, "—") if o else "—"
            # 받을 돈 = 청구액 − 수금액 − 크레딧(클레임 상계).
            outstanding = _ar_outstanding(r)
            overdue = (r.status != ARStatus.PAID and r.due_date
                       and r.due_date < today_str)
            status = "연체" if overdue else _enum_val(r.status)
            if (r.currency or "USD") == "USD" and r.status != ARStatus.PAID:
                out_usd += outstanding
                if overdue:
                    overdue_usd += outstanding
            # 10) 세금계산서 발행 · 11) 대금 결제 완료 — 이 고객 P/O(오더)의 수동 완료 표시.
            # 프로젝트가 아니라 오더 단위라, 같은 프로젝트의 P/O 라도 결제일이 다를 수 있다.
            rfq = _rfq_for_order(s, o) if o else None
            sd = manual_stage_dates(rfq, o)
            # 세금계산서 발행 = 수동 완료 또는 PI 발행(둘 중 먼저 있는 쪽이 완료 일시).
            pi = pi_by_order.get(r.order_id)
            rows.append({
                "id": r.id,
                "order_id": r.order_id,
                "assignee_id": (rfq.created_by or 0) if rfq else 0,
                "assignee": (user_names.get(rfq.created_by, "") or "") if rfq else "",
                "project_title": (getattr(rfq, "project_title", None) or "") if rfq else "",
                "contact_person": (getattr(rfq, "contact_person", None) or "") if rfq else "",
                "ci_no": r.ci_no or "",
                "customer": cust,
                "currency": r.currency or "USD",
                "invoice_amount": round(r.invoice_amount or 0, 2),
                "paid_amount": round(r.paid_amount or 0, 2),
                # 크레딧 노트(클레임 상계)로 깎아 준 금액 — 0 이면 상계 없음.
                "credit_amount": round(float(getattr(r, "credit_amount", None) or 0), 2),
                "outstanding": round(outstanding, 2),
                "due_date": r.due_date or "",
                "status": status,
                "overdue": bool(overdue),
                "notes": r.notes or "",
                # 세금계산서(대금청구서) 문서 필드
                "invoice_no": r.invoice_no or "",
                "invoice_date": r.invoice_date or "",
                "vat_rate": r.vat_rate if r.vat_rate is not None else 0.1,
                "items": r.items or [],
                "charges": r.charges or {},
                "remarks": r.remarks or "",
                # 청구처(BILL TO) 오버라이드 — 9단계 편집 폼에서 그대로 다시 편집하려면 필요.
                "bill_to_tax_id": r.bill_to_tax_id or "",
                "bill_to_contact": r.bill_to_contact or "",
                "bill_to_email": r.bill_to_email or "",
                "bill_to_phone": r.bill_to_phone or "",
                # 단계 재번호(구 8 제거) 후: 세금계산서 발행=10, 수금 완료=11.
                "tax_issued": bool(sd.get("10")) or pi is not None,
                "tax_issued_date": sd.get("10", "") or (pi.date or "" if pi else ""),
                # PI 로 갈음한 건 — 화면에서 "무엇으로 갈음했는지" 를 보여주고,
                # 되돌릴 수동 완료가 없다는 것도 함께 알려야 한다(Undo 가 먹지 않는다).
                "tax_covered_by_pi": pi is not None and not sd.get("10"),
                "tax_pi_no": (pi.pi_no or "") if pi else "",
                "paid_done": bool(sd.get("11")),
                "paid_date": sd.get("11", "") or "",
                # 매입(AP) 진척 — 이 오더의 벤더 P/O 총건수 대비 청구서 수취·세금계산서
                # 수취·지급 완료 건수. 셋 다 총건수와 같아야 9·10·11 단계가 완료된다.
                **_ap_progress(po_ids_by_order.get(r.order_id, []), ap_by_po),
                # 공통 식별 컬럼
                "vessel": (vessel_names.get(o.vessel_id, "") if o and o.vessel_id else ""),
                "trade_type": (o.trade_type or "수출") if o else "수출",
                "work_type": (_enum_val(rfq.work_type) if rfq and rfq.work_type else "부품공급"),
                "vendor": (lambda v: (v[0] + (f"  (외 {len(v) - 1}곳)" if len(v) > 1 else "")) if v else "")(po_vendors_by_order.get(r.order_id, [])),
                "first_rfq_at": _first_rfq_iso(rfq),
                "project_no": _project_no_map(s).get(rfq.id, "") if rfq else "",
            })
        return {
            "kpi": {
                "outstanding_usd": round(out_usd, 2),
                "overdue_usd": round(overdue_usd, 2),
                "count": len(rows),
            },
            "rows": rows,
        }
    finally:
        s.close()


@app.get("/api/admin/ar/soa.xlsx", dependencies=[Depends(require_token)])
def ar_soa_xlsx(status: str | None = None, currency: str | None = None):
    """Statement of Account (SOA) XLSX 내보내기 — AR 현황을 엑셀로 추출한다.
    AR 페이지의 status/currency 필터를 그대로 적용하고 통화별 합계를 덧붙인다."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    s = get_session()
    try:
        cust_names = {c.id: c.name for c in s.query(Customer).all()}
        ord_map = {o.id: o for o in s.query(Order).all()}
        today_str = date.today().isoformat()

        wb = Workbook()
        ws = wb.active
        ws.title = "SOA"
        ws.append(["Statement of Account (Accounts Receivable)"])
        ws.append([f"Generated: {today_str}"])
        active_filters = []
        if status:
            active_filters.append(f"Status={status}")
        if currency:
            active_filters.append(f"Currency={currency}")
        ws.append(["Filter: " + (", ".join(active_filters) if active_filters else "All")])
        ws.append([])

        headers = ["CI No.", "Customer", "Project No.", "Currency", "Invoice",
                   "Paid", "Outstanding", "Due Date", "Status"]
        ws.append(headers)
        head_row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=head_row, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3A5F")

        totals: dict[str, list[float]] = {}
        for r in s.query(ARRecord).order_by(ARRecord.id.desc()).all():
            o = ord_map.get(r.order_id)
            cust = cust_names.get(o.customer_id, "—") if o else "—"
            cur = r.currency or "USD"
            overdue = (r.status != ARStatus.PAID and r.due_date and r.due_date < today_str)
            st_label = "연체" if overdue else _enum_val(r.status)
            if status and status not in (st_label, _enum_val(r.status)):
                continue
            if currency and cur != currency:
                continue
            invoice = round(r.invoice_amount or 0, 2)
            paid = round(r.paid_amount or 0, 2)
            outstanding = round(invoice - paid, 2)
            ws.append([r.ci_no or "—", cust, _project_no_for_order(s, o) if o else "—", cur,
                       invoice, paid, outstanding, r.due_date or "—", st_label])
            t = totals.setdefault(cur, [0.0, 0.0, 0.0])
            t[0] += invoice
            t[1] += paid
            t[2] += outstanding

        ws.append([])
        for cur, (inv, paid, out) in sorted(totals.items()):
            ws.append([f"TOTAL ({cur})", "", "", cur, round(inv, 2),
                       round(paid, 2), round(out, 2), "", ""])
            total_row = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(row=total_row, column=c).font = Font(bold=True)

        for col, width in zip("ABCDEFGHI", [16, 22, 14, 9, 14, 14, 14, 12, 10]):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="SOA_{today_str}.xlsx"'},
        )
    finally:
        s.close()


@app.post("/api/admin/ar", dependencies=[Depends(require_token)])
def create_ar(body: ARSave):
    s = get_session()
    try:
        order = s.query(Order).filter_by(id=body.order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        ar = ARRecord(
            order_id=body.order_id,
            ci_no=body.ci_no or "",
            invoice_amount=body.invoice_amount or 0.0,
            paid_amount=body.paid_amount or 0.0,
            currency=body.currency or "USD",
            due_date=body.due_date,
            status=_ar_status_from_text(body.status, body.paid_amount or 0.0, body.invoice_amount or 0.0),
            notes=body.notes or "",
            invoice_no=body.invoice_no or "",
            invoice_date=body.invoice_date or "",
            vat_rate=body.vat_rate if body.vat_rate is not None else 0.1,
            items=body.items or [],
            charges=body.charges or {},
            remarks=body.remarks or "",
            bill_to_tax_id=body.bill_to_tax_id or "",
            bill_to_contact=body.bill_to_contact or "",
            bill_to_email=body.bill_to_email or "",
            bill_to_phone=body.bill_to_phone or "",
        )
        s.add(ar)
        s.commit()
        return {"ok": True, "id": ar.id}
    finally:
        s.close()


@app.put("/api/admin/ar/{ar_id}", dependencies=[Depends(require_token)])
def update_ar(ar_id: int, body: ARSave):
    s = get_session()
    try:
        ar = s.query(ARRecord).filter_by(id=ar_id).first()
        if not ar:
            raise HTTPException(status_code=404, detail="AR 레코드를 찾을 수 없습니다.")
        if not s.query(Order).filter_by(id=body.order_id).first():
            raise HTTPException(status_code=404, detail="Order를 찾을 수 없습니다.")
        ar.order_id = body.order_id
        ar.ci_no = body.ci_no or ""
        ar.invoice_amount = body.invoice_amount or 0.0
        ar.paid_amount = body.paid_amount or 0.0
        ar.currency = body.currency or "USD"
        ar.due_date = body.due_date
        ar.status = _ar_status_from_text(body.status, ar.paid_amount or 0.0, ar.invoice_amount or 0.0)
        # 크레딧 노트로 깎아 둔 청구서는 금액만 보면 여전히 '미수'로 보인다 — 상계까지
        # 합쳐 다시 매긴다(청구액을 고쳐 저장했을 때도 잔액이 맞게 따라온다).
        if float(getattr(ar, "credit_amount", None) or 0) > 0:
            _ar_recalc_status(ar)
        ar.notes = body.notes or ""
        # 문서 필드는 전달된 값만 갱신(미전달=기존 유지) — 수금 등록 등 부분 저장과 충돌 방지.
        if body.invoice_no is not None:
            ar.invoice_no = body.invoice_no
        if body.invoice_date is not None:
            ar.invoice_date = body.invoice_date
        if body.vat_rate is not None:
            ar.vat_rate = body.vat_rate
        if body.items is not None:
            ar.items = body.items
        if body.charges is not None:
            ar.charges = body.charges
        if body.remarks is not None:
            ar.remarks = body.remarks
        if body.bill_to_tax_id is not None:
            ar.bill_to_tax_id = body.bill_to_tax_id
        if body.bill_to_contact is not None:
            ar.bill_to_contact = body.bill_to_contact
        if body.bill_to_email is not None:
            ar.bill_to_email = body.bill_to_email
        if body.bill_to_phone is not None:
            ar.bill_to_phone = body.bill_to_phone
        s.commit()
        return {"ok": True, "id": ar.id, "status": _enum_val(ar.status)}
    finally:
        s.close()


@app.delete("/api/admin/ar/{ar_id}", dependencies=[Depends(require_token)])
def delete_ar(ar_id: int):
    s = get_session()
    try:
        ar = s.query(ARRecord).filter_by(id=ar_id).first()
        if not ar:
            raise HTTPException(status_code=404, detail="AR 레코드를 찾을 수 없습니다.")
        # 이 청구서를 깎아 둔 크레딧 노트가 있으면 막는다 — 지우면 상계의 상대가 사라져
        # 그 감액이 어디에 걸린 것인지 알 수 없게 된다(PostgreSQL 은 외래키로 어차피 막는다).
        cn = s.query(CreditNote).filter_by(ar_id=ar_id).count()
        if cn:
            raise HTTPException(
                status_code=400,
                detail=f"이 청구서에 걸린 크레딧 노트가 {cn}건 있습니다. 먼저 취소해 주세요.")
        s.delete(ar)
        s.commit()
        return {"ok": True}
    finally:
        s.close()


@app.post("/api/admin/ar/{ar_id}/payment", dependencies=[Depends(require_token)])
def ar_payment(ar_id: int, body: ARPayment):
    """수금 등록 — 상태 자동 갱신.

    set_total=True 면 amount 를 '지금까지 받은 총액'으로 그대로 설정한다(멱등). 화면의
    수금 칸이 이 방식이라, 완료 버튼을 다시 눌러도 같은 금액이 두 번 쌓이지 않고 잘못
    들어간 금액도 올바른 총액으로 고쳐 저장할 수 있다(0 을 보내면 수금 취소).
    set_total=False(기본)면 기존처럼 받은 금액을 누적한다.

    외화 입금은 은행이 수취수수료를 떼고 넣어 준다 — $8,200 을 청구했는데 통장에는
    $8,193.18 이 꽂힌다. 통장에 찍힌 그대로 적으면 늘 몇 달러가 모자라 영영 '부분수금'이
    되고, 그러면 완납일이 남지 않아 수수료를 계산할 날짜조차 없어진다(수수료가 수수료의
    기록을 막는 셈이다).

    그 차액은 못 받은 돈이 아니다 — 고객은 청구액 전부를 갚았고 은행이 중간에서 제 몫을
    가져간 것이다. 그래서 bank_fee 가 켜져 있으면 수금액을 청구액에 맞춰 채우고(받을 돈은
    0 이 된다) 차액은 우리 비용으로 넘긴다. 그 비용은 Outflow 에 수수료 행으로 선다
    (_inbound_fee_rows — 같은 입금일·같은 고시환율을 근거로 적는다).
    """
    s = get_session()
    try:
        ar = s.query(ARRecord).filter_by(id=ar_id).first()
        if not ar:
            raise HTTPException(status_code=404, detail="AR 레코드를 찾을 수 없습니다.")
        if body.set_total:
            if body.amount < 0:
                raise HTTPException(status_code=400, detail="수금 총액은 0 이상이어야 합니다.")
            ar.paid_amount = body.amount
        else:
            if body.amount <= 0:
                raise HTTPException(status_code=400, detail="수금액은 0보다 커야 합니다.")
            ar.paid_amount = (ar.paid_amount or 0) + body.amount
        if body.due_date:
            ar.due_date = body.due_date
        # 입금일 — 화면에서 받은 실제 입금일이 우선이고, 없을 때만 오늘로 본다. 이 날짜가
        # 곧 수수료 환율의 기준일이라 '적은 날'이 아니라 '들어온 날'이어야 한다.
        paid_on = (body.paid_on or "").strip()[:10] or date.today().isoformat()
        # 통장에 찍힌 금액이 청구액보다 수수료만큼 모자란 경우 — 고객은 전액을 갚았고 그
        # 차액은 은행이 가져간 것이다. 그러니 미수로 남겨 두면 안 된다: 받을 돈은 0 이 되고
        # (수금액을 청구액에 맞춘다), 그 차액은 우리 비용으로 Outflow 에 선다.
        fee = 0.0
        # 이미 크레딧 노트로 깎아 준 금액은 받을 필요가 없는 돈이라, 모자란 금액을 셀 때
        # 먼저 뺀다 — 안 그러면 상계액만큼이 늘 '수수료로 보기엔 너무 큰 차액'이 된다.
        shortfall = _ar_outstanding(ar)
        if body.bank_fee and (ar.paid_amount or 0) > 0 and shortfall > 0:
            est, _rate, _used = inbound_fee_in(ar.currency or "USD", paid_on)
            # 예상 수수료의 3배까지만 수수료로 본다 — 은행이 쓰는 환율이 고시와 조금 다르고
            # 센트에서 반올림도 되지만, 그 범위를 넘는 차액은 덜 받은 것이지 수수료가 아니다.
            if est and shortfall <= est * 3:
                fee = shortfall
                ar.paid_amount = round(ar.invoice_amount or 0, 2)
        # 실제로 떼인 금액을 남긴다 — 수금액을 청구액으로 채우고 나면 그 차액은 여기에만
        # 남고, Outflow 의 수수료 행이 추정 대신 이 값을 쓴다(두 화면 숫자가 갈리지 않게).
        ar.bank_fee = round(fee, 2)
        # 오차 1센트 — 은행이 센트에서 반올림하므로 정확히 맞아떨어지지 않는다.
        # 상계(크레딧 노트)까지 합쳐 잔액이 0 이면 완납이다(_ar_recalc_status).
        _ar_recalc_status(ar)
        if ar.status == ARStatus.PAID:
            # 완납일은 최초로 잔액이 0이 된 날. 이미 있으면 유지(재수정 시 날짜 밀림 방지).
            ar.paid_date = ar.paid_date or paid_on
        s.commit()
        return {"ok": True, "paid_amount": ar.paid_amount, "status": _enum_val(ar.status),
                # 화면이 "차액 얼마를 수수료로 보아 완납 처리했는지"를 그대로 알릴 수 있게.
                "bank_fee": round(fee, 2), "paid_date": ar.paid_date or ""}
    finally:
        s.close()
