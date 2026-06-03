from __future__ import annotations

import io
import json
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)

def _register_default_font() -> tuple[str, str]:
    candidates = [
        ("/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf", "/usr/share/fonts/truetype/nanum/NanumBarunGothicBold.ttf"),
        ("/usr/share/fonts/truetype/nanum/NanumGothic.ttf", "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"),
        ("C:/Windows/Fonts/malgun.ttf", "C:/Windows/Fonts/malgunbd.ttf"),
        ("/Library/Fonts/AppleGothic.ttf", "/Library/Fonts/AppleGothic.ttf"),
    ]
    for regular, bold in candidates:
        try:
            if Path(regular).exists():
                pdfmetrics.registerFont(TTFont("KMBaseFont", regular))
                if Path(bold).exists():
                    pdfmetrics.registerFont(TTFont("KMBoldFont", bold))
                    return "KMBaseFont", "KMBoldFont"
                return "KMBaseFont", "KMBaseFont"
        except Exception:
            continue
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
        return "HYGothic-Medium", "HYGothic-Medium"
    except Exception:
        return "Helvetica", "Helvetica-Bold"


DEFAULT_FONT, DEFAULT_BOLD_FONT = _register_default_font()

NAVY = colors.HexColor("#0B1D3A")
BLUE = colors.HexColor("#0055A8")
LIGHT_BLUE = colors.HexColor("#EAF3FF")
LIGHT_GRAY = colors.HexColor("#F4F6F8")
MID_GRAY = colors.HexColor("#D8DEE6")
DARK_GRAY = colors.HexColor("#3A3F44")

DOC_TITLES = {
    "quotation": "QUOTATION",
    "proforma_invoice": "PROFORMA INVOICE",
    "commercial_invoice": "COMMERCIAL INVOICE",
    "packing_list": "PACKING LIST",
    "shipping_advice": "SHIPPING ADVICE",
}

DOC_PREFIX = {
    "quotation": "QTN",
    "proforma_invoice": "PI",
    "commercial_invoice": "CI",
    "packing_list": "PL",
    "shipping_advice": "SA",
    "tax_invoice_data": "TAX",
}


def load_json(path: str | Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def make_doc_no(doc_type: str, sequence: int = 1, company_prefix: str = "KMS") -> str:
    year = date.today().year
    prefix = DOC_PREFIX.get(doc_type, "DOC")
    return f"{company_prefix}-{prefix}-{year}-{sequence:04d}"


def _money(value: Any, currency: str = "USD") -> str:
    try:
        q = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{currency} {q:,.2f}"
    except Exception:
        return f"{currency} 0.00"


def _num(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def normalize_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for i, raw in enumerate(items, start=1):
        qty = _num(raw.get("qty", 0))
        unit_price = _num(raw.get("unit_price", 0))
        amount = raw.get("amount")
        if amount in (None, "", 0):
            amount = qty * unit_price
        normalized.append(
            {
                "item_no": raw.get("item_no") or i,
                "part_no": raw.get("part_no", ""),
                "description": raw.get("description", ""),
                "maker": raw.get("maker", ""),
                "origin": raw.get("origin", ""),
                "qty": qty,
                "unit": raw.get("unit", "PCS"),
                "unit_price": unit_price,
                "amount": _num(amount),
                "lead_time": raw.get("lead_time", ""),
                "remark": raw.get("remark", ""),
                "gross_weight": raw.get("gross_weight", ""),
                "net_weight": raw.get("net_weight", ""),
                "package": raw.get("package", ""),
                "dimension": raw.get("dimension", ""),
                "hs_code": raw.get("hs_code", ""),
            }
        )
    return normalized


def calc_totals(items: List[Dict[str, Any]], vat_rate: float = 0.0) -> Dict[str, float]:
    subtotal = sum(_num(item.get("amount", 0)) for item in normalize_items(items))
    vat = subtotal * _num(vat_rate)
    total = subtotal + vat
    return {"subtotal": subtotal, "vat": vat, "total": total}


def _styles() -> Dict[str, ParagraphStyle]:
    styles = getSampleStyleSheet()
    base = ParagraphStyle(
        "KMBase",
        parent=styles["Normal"],
        fontName=DEFAULT_FONT,
        fontSize=8.2,
        leading=10.2,
        textColor=colors.black,
    )
    return {
        "base": base,
        "small": ParagraphStyle("KMSmall", parent=base, fontSize=7.3, leading=9.0),
        "tiny": ParagraphStyle("KMTiny", parent=base, fontSize=6.6, leading=8.2),
        "title": ParagraphStyle(
            "KMTitle",
            parent=base,
            fontName=DEFAULT_BOLD_FONT,
            fontSize=20,
            leading=24,
            alignment=TA_RIGHT,
            textColor=NAVY,
        ),
        "subtitle": ParagraphStyle(
            "KMSubtitle",
            parent=base,
            fontSize=8.8,
            leading=11,
            alignment=TA_RIGHT,
            textColor=DARK_GRAY,
        ),
        "section": ParagraphStyle(
            "KMSection",
            parent=base,
            fontName=DEFAULT_BOLD_FONT,
            fontSize=9,
            leading=11,
            textColor=colors.white,
        ),
        "right": ParagraphStyle("KMRight", parent=base, alignment=TA_RIGHT),
        "center": ParagraphStyle("KMCenter", parent=base, alignment=TA_CENTER),
    }


def _p(text: Any, style: ParagraphStyle) -> Paragraph:
    safe = "" if text is None else str(text)
    safe = (
        safe.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    # Allow only minimal internal markup used by this template.
    safe = safe.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
    return Paragraph(safe, style)


def _header(company: Dict[str, Any], doc_title: str, logo_path: Optional[str] = None):
    s = _styles()
    left_lines = [
        f"<b>{company.get('company_name_en', 'K-MARIS Energy & Solutions Co., Ltd.')}</b>",
        company.get("company_name_kr", ""),
        company.get("address", ""),
        f"Tel: {company.get('phone', '')} | Email: {company.get('general_email', '')}",
        f"Website: {company.get('website', '')}",
        f"{company.get('tagline', '')}",
    ]
    left = _p("\n".join([x for x in left_lines if x]), s["base"])
    if logo_path and Path(logo_path).exists():
        try:
            logo = Image(logo_path, width=32 * mm, height=18 * mm)
            left = Table([[logo, left]], colWidths=[36 * mm, 100 * mm])
            left.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        except Exception:
            pass
    right = [
        _p(f"<b>{doc_title}</b>", s["title"]),
        _p("Marine Equipment | Engine Parts | Bunkering | Technical Solutions", s["subtitle"]),
    ]
    table = Table([[left, right]], colWidths=[160 * mm, 110 * mm])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LINEBELOW", (0, 0), (-1, -1), 1, BLUE),
            ]
        )
    )
    return table


def _info_tables(data: Dict[str, Any], doc_type: str):
    s = _styles()
    customer = data.get("customer", {})
    vessel = data.get("vessel", {})
    terms = data.get("terms", {})
    shipping = data.get("shipping", {})

    left_rows = [
        [_p("<b>Customer / Buyer</b>", s["base"]), _p(customer.get("name", ""), s["base"])],
        [_p("Address", s["base"]), _p(customer.get("address", ""), s["base"])],
        [_p("Contact", s["base"]), _p(customer.get("contact", ""), s["base"])],
        [_p("Email", s["base"]), _p(customer.get("email", ""), s["base"])],
    ]
    mid_rows = [
        [_p("<b>Vessel</b>", s["base"]), _p(vessel.get("name", ""), s["base"])],
        [_p("IMO No.", s["base"]), _p(vessel.get("imo", ""), s["base"])],
        [_p("Engine Type", s["base"]), _p(vessel.get("engine_type", ""), s["base"])],
        [_p("Hull No.", s["base"]), _p(vessel.get("hull_no", ""), s["base"])],
    ]

    right_rows = [
        [_p("<b>Document No.</b>", s["base"]), _p(data.get("doc_no", ""), s["base"])],
        [_p("Date", s["base"]), _p(data.get("date", ""), s["base"])],
        [_p("Currency", s["base"]), _p(data.get("currency", "USD"), s["base"])],
        [_p("Incoterms", s["base"]), _p(terms.get("incoterms", ""), s["base"])],
    ]
    if doc_type == "quotation":
        right_rows.append([_p("Validity", s["base"]), _p(data.get("valid_until", ""), s["base"])])
    if doc_type in {"commercial_invoice", "packing_list", "shipping_advice"}:
        right_rows.extend(
            [
                [_p("PO No.", s["base"]), _p(shipping.get("po_no", ""), s["base"])],
                [_p("Export Ref.", s["base"]), _p(shipping.get("export_ref", ""), s["base"])],
            ]
        )

    def box(title, rows):
        table = Table(rows, colWidths=[28 * mm, 58 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.3, MID_GRAY),
                    ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        return table

    outer = Table(
        [[box("Customer", left_rows), box("Vessel", mid_rows), box("Doc", right_rows)]],
        colWidths=[90 * mm, 90 * mm, 90 * mm],
    )
    outer.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    return outer


def _items_table(data: Dict[str, Any], doc_type: str):
    s = _styles()
    currency = data.get("currency", "USD")
    items = normalize_items(data.get("items", []))

    if doc_type == "packing_list":
        headers = ["No.", "Part No.", "Description", "Qty", "Unit", "Package", "N.W.", "G.W.", "Dimension", "Remark"]
        widths = [10, 30, 65, 15, 16, 28, 20, 20, 34, 32]
        rows = [[_p(h, s["tiny"]) for h in headers]]
        for item in items:
            rows.append(
                [
                    _p(item["item_no"], s["tiny"]),
                    _p(item["part_no"], s["tiny"]),
                    _p(item["description"], s["tiny"]),
                    _p(str(item["qty"]), s["tiny"]),
                    _p(item["unit"], s["tiny"]),
                    _p(item.get("package", ""), s["tiny"]),
                    _p(item.get("net_weight", ""), s["tiny"]),
                    _p(item.get("gross_weight", ""), s["tiny"]),
                    _p(item.get("dimension", ""), s["tiny"]),
                    _p(item.get("remark", ""), s["tiny"]),
                ]
            )
    else:
        headers = ["No.", "Part No.", "Description", "Maker", "Origin", "Qty", "Unit", "Unit Price", "Amount", "Lead Time / Remark"]
        widths = [10, 30, 58, 35, 24, 15, 16, 25, 28, 47]
        rows = [[_p(h, s["tiny"]) for h in headers]]
        for item in items:
            lead_remark = f"{item.get('lead_time', '')}\n{item.get('remark', '')}".strip()
            rows.append(
                [
                    _p(item["item_no"], s["tiny"]),
                    _p(item["part_no"], s["tiny"]),
                    _p(item["description"], s["tiny"]),
                    _p(item["maker"], s["tiny"]),
                    _p(item["origin"], s["tiny"]),
                    _p(str(item["qty"]), s["tiny"]),
                    _p(item["unit"], s["tiny"]),
                    _p(_money(item["unit_price"], currency), s["tiny"]),
                    _p(_money(item["amount"], currency), s["tiny"]),
                    _p(lead_remark, s["tiny"]),
                ]
            )

    table = Table(rows, colWidths=[w * mm for w in widths], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (5, 1), (8, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for r in range(1, len(rows)):
        if r % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FAFBFC")))
    table.setStyle(TableStyle(style_cmds))
    return table


def _totals_table(data: Dict[str, Any]):
    s = _styles()
    currency = data.get("currency", "USD")
    totals = calc_totals(data.get("items", []), _num(data.get("vat_rate", 0)))
    rows = [
        [_p("Subtotal", s["base"]), _p(_money(totals["subtotal"], currency), s["right"])],
        [_p("VAT", s["base"]), _p(_money(totals["vat"], currency), s["right"])],
        [_p("Total", s["base"]), _p(f"<b>{_money(totals['total'], currency)}</b>", s["right"])],
    ]
    table = Table(rows, colWidths=[35 * mm, 45 * mm])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, MID_GRAY),
                ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                ("BACKGROUND", (0, -1), (-1, -1), LIGHT_BLUE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return Table([["", table]], colWidths=[190 * mm, 80 * mm])


def _terms_block(data: Dict[str, Any], doc_type: str):
    s = _styles()
    terms = data.get("terms", {})
    company = data.get("company", {})
    shipping = data.get("shipping", {})

    rows = []
    if doc_type in {"quotation", "proforma_invoice", "commercial_invoice"}:
        rows.extend(
            [
                ["Payment Terms", terms.get("payment_terms", "")],
                ["Delivery Place", terms.get("delivery_place", "")],
                ["Shipment Method", terms.get("shipment_method", "")],
                ["Packing", terms.get("packing", "")],
                ["Warranty", terms.get("warranty", "")],
                ["Remarks", terms.get("remarks", "")],
            ]
        )
        if doc_type == "proforma_invoice":
            rows.extend(
                [
                    ["Bank", company.get("bank_name", "")],
                    ["Account", company.get("bank_account", "")],
                    ["Account Holder", company.get("bank_holder", "")],
                    ["SWIFT", company.get("swift", "")],
                ]
            )
    if doc_type in {"packing_list", "shipping_advice"}:
        rows.extend(
            [
                ["Port of Loading", shipping.get("port_loading", "")],
                ["Port of Discharge", shipping.get("port_discharge", "")],
                ["Carrier", shipping.get("carrier", "")],
                ["B/L or AWB No.", shipping.get("bl_awb_no", "")],
                ["ETD", shipping.get("etd", "")],
                ["ETA", shipping.get("eta", "")],
                ["Shipping Marks", shipping.get("shipping_marks", "")],
            ]
        )

    if not rows:
        return Spacer(1, 1)

    table = Table([[_p("<b>Terms / Instructions</b>", s["section"]), ""]] + [[_p(a, s["small"]), _p(b, s["small"])] for a, b in rows], colWidths=[42 * mm, 228 * mm])
    table.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (-1, 0)),
                ("BACKGROUND", (0, 0), (-1, 0), BLUE),
                ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
                ("BACKGROUND", (0, 1), (0, -1), LIGHT_GRAY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(MID_GRAY)
    canvas.setLineWidth(0.5)
    canvas.line(12 * mm, 11 * mm, 285 * mm, 11 * mm)
    canvas.setFont(DEFAULT_FONT, 7)
    canvas.setFillColor(DARK_GRAY)
    canvas.drawString(12 * mm, 7 * mm, "K-MARIS Energy & Solutions Co., Ltd. | This document is system-generated.")
    canvas.drawRightString(285 * mm, 7 * mm, f"Page {doc.page}")
    canvas.restoreState()


def make_pdf(doc_type: str, data: Dict[str, Any], company: Optional[Dict[str, Any]] = None, logo_path: Optional[str] = None) -> bytes:
    if doc_type not in DOC_TITLES:
        raise ValueError(f"Unsupported document type: {doc_type}")
    payload = dict(data)
    payload["company"] = company or data.get("company", {})
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=15 * mm,
        title=DOC_TITLES[doc_type],
        author="K-MARIS Energy & Solutions Co., Ltd.",
    )
    s = _styles()
    story = []
    story.append(_header(payload["company"], DOC_TITLES[doc_type], logo_path))
    story.append(Spacer(1, 5 * mm))
    story.append(_info_tables(payload, doc_type))
    story.append(Spacer(1, 5 * mm))
    story.append(_items_table(payload, doc_type))
    if doc_type not in {"packing_list", "shipping_advice"}:
        story.append(Spacer(1, 4 * mm))
        story.append(_totals_table(payload))
    story.append(Spacer(1, 5 * mm))
    story.append(_terms_block(payload, doc_type))
    story.append(Spacer(1, 2 * mm))
    sign = Table(
        [[
            _p("Prepared by\n\n________________", s["base"]),
            _p("Approved by\n\n________________", s["base"]),
            _p("For and on behalf of K-MARIS Energy & Solutions Co., Ltd.\n\nAuthorized Signature", s["base"]),
        ]],
        colWidths=[70 * mm, 70 * mm, 130 * mm],
        rowHeights=[12 * mm],
    )
    sign.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.35, MID_GRAY),
                ("BACKGROUND", (0, 0), (-1, 0), LIGHT_GRAY),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(sign)
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def make_tax_invoice_xlsx(data: Dict[str, Any], company: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Invoice Data"
    currency = data.get("currency", "KRW")
    items = normalize_items(data.get("items", []))
    vat_rate = _num(data.get("vat_rate", 0.1))
    totals = calc_totals(items, vat_rate)
    customer = data.get("customer", {})
    tax = data.get("tax_invoice", {})

    title_fill = PatternFill("solid", fgColor="0B1D3A")
    section_fill = PatternFill("solid", fgColor="EAF3FF")
    header_fill = PatternFill("solid", fgColor="D8DEE6")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    thin = Side(style="thin", color="D8DEE6")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.merge_cells("A1:J1")
    ws["A1"] = "K-MARIS TAX INVOICE DATA SHEET / 세금계산서 발행용 데이터"
    ws["A1"].fill = title_fill
    ws["A1"].font = white_font
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Issue Date / 작성일자", tax.get("issue_date", data.get("date", "")), "Document No.", data.get("doc_no", "")),
        ("Supply Type / 공급유형", tax.get("supply_type", ""), "Currency", currency),
        ("Supplier / 공급자", company.get("company_name_kr", ""), "Supplier Business No.", tax.get("supplier_business_no", company.get("business_no", ""))),
        ("Supplier Email", company.get("tax_email", company.get("general_email", "")), "Supplier Address", company.get("address", "")),
        ("Buyer / 공급받는 자", customer.get("name", ""), "Buyer Business No.", tax.get("buyer_business_no", customer.get("tax_id", ""))),
        ("Buyer Email", customer.get("email", ""), "Buyer Address", customer.get("address", "")),
    ]
    start = 3
    for r_idx, row in enumerate(rows, start=start):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx in [1, 3]:
                cell.fill = section_fill
                cell.font = bold
    item_start = start + len(rows) + 2
    headers = ["No.", "Part No.", "Description", "Maker", "Qty", "Unit", "Unit Price", "Amount", "HS Code", "Remark"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(item_start, c_idx, h)
        cell.fill = header_fill
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for r_offset, item in enumerate(items, start=1):
        values = [
            item["item_no"],
            item["part_no"],
            item["description"],
            item["maker"],
            item["qty"],
            item["unit"],
            item["unit_price"],
            item["amount"],
            item.get("hs_code", ""),
            item.get("remark", ""),
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(item_start + r_offset, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx in [7, 8]:
                cell.number_format = "#,##0.00"

    total_row = item_start + len(items) + 2
    total_values = [
        ("Supply Amount / 공급가액", totals["subtotal"]),
        ("VAT / 부가세", totals["vat"]),
        ("Total / 합계", totals["total"]),
    ]
    for i, (label, value) in enumerate(total_values):
        r = total_row + i
        ws.cell(r, 7, label).fill = section_fill
        ws.cell(r, 7).font = bold
        ws.cell(r, 7).border = border
        ws.cell(r, 8, value).border = border
        ws.cell(r, 8).number_format = "#,##0.00"

    note_row = total_row + 5
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=10)
    ws.cell(note_row, 1).value = (
        "Note: This sheet is a data-preparation document only. Actual electronic tax invoice issuance must be processed "
        "through Hometax or an authorized e-tax invoice provider/ERP after tax review. / 본 시트는 발행용 데이터이며 실제 전자세금계산서 발행은 홈택스 또는 공인 발급 시스템에서 세무 검토 후 진행하십시오."
    )
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(note_row, 1).border = border

    widths = [10, 20, 35, 24, 10, 10, 16, 18, 16, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
